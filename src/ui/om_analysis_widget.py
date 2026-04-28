from __future__ import annotations
import json
import os
import re
import ast

try:
    import numpy as np
except ImportError:
    np = None

from PySide6.QtCore import Qt, QRectF, Signal, QUrl, QMimeData, QThread
from PySide6.QtGui import QColor, QPainter, QPen, QBrush, QIcon, QDrag
from PySide6.QtMultimedia import QAudioOutput, QMediaPlayer
from PySide6.QtMultimediaWidgets import QVideoWidget
from PySide6.QtWidgets import (
    QWidget,
    QVBoxLayout,
    QHBoxLayout,
    QLabel,
    QPushButton,
    QComboBox,
    QDoubleSpinBox,
    QLineEdit,
    QTabWidget,
    QMessageBox,
    QTreeWidget,
    QTreeWidgetItemIterator,
    QAbstractItemView,
    QRadioButton,
    QTableWidget,
    QTableWidgetItem,
    QHeaderView,
    QTextEdit,
)

from src.core.utils import resource_path
from src.core.audio_analysis import find_first_valid_event
from src.ui.analysis_widget import AnalysisWidget, AnalysisScanner, LogicTabWidget
from src.ui.styles import IDIADA_ORANGE
from src.ui.widgets import setup_tab_icon_switching


class OMReportGeneratorWorker(QThread):
    """Worker that generates OM reports using OMReportBuilder off the main thread."""
    finished = Signal(str)
    error = Signal(str)

    def __init__(self, config, output_path, dpi=300):
        super().__init__()
        self.config = config
        self.output_path = output_path
        self.dpi = dpi

    def run(self):
        try:
            from src.core.om_report_builder import OMReportBuilder
            builder = OMReportBuilder(self.config)
            builder.generate(self.output_path, dpi=self.dpi)
            self.finished.emit(self.output_path)
        except Exception as e:
            import traceback
            traceback.print_exc()
            self.error.emit(str(e))


OM_AUDIO_SIGNAL_ALIASES = [
    "SoundPressure",
    "MySound PressureTask.Sound Pressure",
]

OM_CATEGORY_FOLDER_MAP = {
    "correct belt routing": "Correct Belt Routing",
    "out of position": "Out of Position",
    "occupant stature": "Occupant Stature",
}


def _normalize_folder_name(name: str) -> str:
    text = str(name or "").strip().lower()
    text = text.replace("_", " ").replace("-", " ")
    text = re.sub(r"\s+", " ", text)
    return text


def _normalize_path_for_match(path_text: str) -> str:
    norm = os.path.normpath(str(path_text or ""))
    norm = norm.replace('\\', '/').strip()
    return norm.lower()


def _extract_mark_time_from_entry(entry):
    if isinstance(entry, dict):
        mark = entry.get("mark_time_s")
        if isinstance(mark, (int, float)):
            return float(mark)
        return None
    if isinstance(entry, (int, float)):
        return float(entry)
    if isinstance(entry, list):
        for v in entry:
            if isinstance(v, (int, float)):
                return float(v)
    return None


def _is_om_correct_belt_routing(config: dict) -> bool:
    if not isinstance(config, dict):
        return False

    target_category = str(config.get('target_category', '') or '').strip().lower()
    if 'correct belt routing' in target_category:
        return True

    variant = str(config.get('om_report_variant', '') or '').strip().lower()
    if variant == 'correct_belt_routing':
        return True

    signals = config.get('signals') or {}
    if isinstance(signals, dict):
        for sig_data in signals.values():
            if not isinstance(sig_data, dict):
                continue
            sig_category = str(sig_data.get('category', '') or '').strip().lower()
            if 'correct belt routing' in sig_category:
                return True

    return False


def _is_om_out_of_position(config: dict) -> bool:
    if not isinstance(config, dict):
        return False

    target_category = str(config.get('target_category', '') or '').strip().lower()
    if 'out of position' in target_category:
        return True

    variant = str(config.get('om_report_variant', '') or '').strip().lower()
    if variant == 'out_of_position':
        return True

    signals = config.get('signals') or {}
    if isinstance(signals, dict):
        for sig_data in signals.values():
            if not isinstance(sig_data, dict):
                continue
            sig_category = str(sig_data.get('category', '') or '').strip().lower()
            if 'out of position' in sig_category:
                return True

    return False


def _is_om_warning_category(config: dict) -> bool:
    return _is_om_correct_belt_routing(config) or _is_om_out_of_position(config)


def _infer_om_category_from_path(path_text: str):
    if not path_text:
        return None

    normalized_path = os.path.normpath(str(path_text)).replace('\\', '/')
    parts = [p for p in normalized_path.split('/') if p]
    # Evaluate folders from closest ancestor to farthest.
    for part in reversed(parts):
        token = _normalize_folder_name(part)
        if token in OM_CATEGORY_FOLDER_MAP:
            return OM_CATEGORY_FOLDER_MAP[token]
    return None


def _guess_video_for_mf4_path(mf4_path: str, camera_mode: str = ".avi", preferred_video_name: str = None):
    if not mf4_path:
        return None

    base_name = os.path.splitext(os.path.basename(mf4_path))[0]
    clean_base = base_name.replace("_tracking", "")
    directory = os.path.dirname(mf4_path)
    parent = os.path.dirname(directory)

    if isinstance(preferred_video_name, str) and preferred_video_name.strip():
        preferred_video_name = preferred_video_name.strip()
        for folder in (directory, parent):
            candidate = os.path.join(folder, preferred_video_name)
            if os.path.exists(candidate):
                return candidate

    for ext in [".avi", ".mp4", ".mov", ".mkv"]:
        if camera_mode == "_m0.avi":
            for folder in (directory, parent):
                candidate = os.path.join(folder, clean_base + "_m0" + ext)
                if os.path.exists(candidate):
                    return candidate
        else:
            for folder in (directory, parent):
                candidate = os.path.join(folder, clean_base + ext)
                if os.path.exists(candidate):
                    return candidate

    candidates = []
    base_candidates = [
        base_name,
        clean_base,
        f"{clean_base}_c2",
        f"{clean_base}_c_2",
        f"{clean_base}c2",
        f"{clean_base}c_2",
        f"{clean_base}_cam1",
        f"{clean_base}_cam2",
    ]
    for base in base_candidates:
        for ext in [".avi", ".mp4", ".mov", ".mkv"]:
            candidates.append(os.path.join(directory, base + ext))
            candidates.append(os.path.join(parent, base + ext))

    for candidate in candidates:
        if os.path.exists(candidate):
            return candidate
    return None


def _guess_secondary_video_for_mf4_path(mf4_path: str, primary_video_path: str = None):
    if not mf4_path:
        return None

    base_name = os.path.splitext(os.path.basename(mf4_path))[0]
    clean_base = base_name.replace("_tracking", "")
    directory = os.path.dirname(mf4_path)
    parent = os.path.dirname(directory)

    def _first_existing(base):
        for ext in [".avi", ".mp4", ".mov", ".mkv"]:
            for folder in (directory, parent):
                candidate = os.path.join(folder, base + ext)
                if os.path.exists(candidate):
                    return candidate
        return None

    normal_video = _first_existing(clean_base)
    m0_video = _first_existing(clean_base + "_m0")

    if not normal_video or not m0_video:
        return None

    if isinstance(primary_video_path, str) and primary_video_path:
        p_norm = _normalize_path_for_match(primary_video_path)
        if p_norm == _normalize_path_for_match(normal_video):
            return m0_video
        if p_norm == _normalize_path_for_match(m0_video):
            return normal_video

    return m0_video


def _find_audio_clusters(mask: np.ndarray, timestamps: np.ndarray, max_gap: float = 5.0):
    if mask.size == 0 or timestamps.size == 0 or not np.any(mask):
        return []

    padded = np.concatenate(([False], mask, [False]))
    diff = np.diff(padded.astype(int))
    starts = np.flatnonzero(diff == 1)
    ends = np.flatnonzero(diff == -1)

    if starts.size == 0:
        return []

    clusters = []
    current_start = int(starts[0])
    current_end = int(ends[0])

    for i in range(1, len(starts)):
        previous_end_idx = int(ends[i - 1]) - 1
        current_start_idx = int(starts[i])
        gap = float(timestamps[current_start_idx] - timestamps[previous_end_idx])

        if gap <= max_gap:
            current_end = int(ends[i])
        else:
            start_t = float(timestamps[current_start])
            end_t = float(timestamps[current_end - 1])
            clusters.append({
                'start_idx': current_start,
                'end_idx': current_end,
                'start_time': start_t,
                'end_time': end_t,
                'duration': max(0.0, end_t - start_t),
            })
            current_start = current_start_idx
            current_end = int(ends[i])

    start_t = float(timestamps[current_start])
    end_t = float(timestamps[current_end - 1])
    clusters.append({
        'start_idx': current_start,
        'end_idx': current_end,
        'start_time': start_t,
        'end_time': end_t,
        'duration': max(0.0, end_t - start_t),
    })
    return clusters


def _calculate_max_gap_in_cluster(mask: np.ndarray, timestamps: np.ndarray, start_idx: int, end_idx: int) -> float:
    if end_idx <= start_idx:
        return 0.0

    cluster_mask = mask[start_idx:end_idx]
    cluster_time = timestamps[start_idx:end_idx]
    if cluster_mask.size == 0 or cluster_time.size == 0:
        return 0.0

    gap_mask = ~cluster_mask
    padded = np.concatenate(([False], gap_mask, [False]))
    diff = np.diff(padded.astype(int))
    starts = np.flatnonzero(diff == 1)
    ends = np.flatnonzero(diff == -1)

    max_gap = 0.0
    for s, e in zip(starts, ends):
        if e <= s:
            continue
        t0 = float(cluster_time[s])
        t1 = float(cluster_time[e - 1])
        max_gap = max(max_gap, max(0.0, t1 - t0))
    return max_gap


def _find_first_signal_activation(
    samples,
    timestamps,
    threshold,
    operator='>',
    reference_start=None,
    snap_to_reference_if_first_match=False,
):
    """Return the first timestamp where the signal satisfies the configured criterion.

    This is intentionally a direct activation check, not a clustered/noise-resistant
    audio_event detector. For CBR mark=0 the requested behavior is to use the selected
    signal activation time as-is, and an activation at t=0.0 is valid.
    """
    timestamps = np.asarray(timestamps, dtype=float)
    samples = np.asarray(samples)

    if samples.size == 0 or timestamps.size == 0 or samples.size != timestamps.size:
        return None

    finite_time_mask = np.isfinite(timestamps)
    samples = samples[finite_time_mask]
    timestamps = timestamps[finite_time_mask]
    if samples.size == 0 or timestamps.size == 0:
        return None

    is_numeric = np.issubdtype(samples.dtype, np.number)

    if is_numeric:
        samples = np.asarray(samples, dtype=float)
        finite_sample_mask = np.isfinite(samples)
        samples = samples[finite_sample_mask]
        timestamps = timestamps[finite_sample_mask]
        if samples.size == 0 or timestamps.size == 0:
            return None

        try:
            threshold_value = float(threshold)
        except Exception:
            return None

        if operator == '>':
            mask = samples > threshold_value
        elif operator == '<':
            mask = samples < threshold_value
        elif operator == '>=':
            mask = samples >= threshold_value
        elif operator == '<=':
            mask = samples <= threshold_value
        elif operator == '==':
            mask = np.abs(samples - threshold_value) < 1e-6
        elif operator == '!=':
            mask = np.abs(samples - threshold_value) >= 1e-6
        else:
            return None
    else:
        def _normalize_value(value):
            if isinstance(value, bytes):
                try:
                    return value.decode('utf-8', errors='ignore')
                except Exception:
                    return str(value)
            if isinstance(value, np.bytes_):
                try:
                    return bytes(value).decode('utf-8', errors='ignore')
                except Exception:
                    return str(value)
            if isinstance(value, str):
                text = value.strip()
                # QComboBox values for byte channels may arrive as the literal string "b'Abuse'".
                if (len(text) >= 3 and text[0] == 'b' and text[1] in ('\'', '"') and text[-1] == text[1]):
                    try:
                        parsed = ast.literal_eval(text)
                        if isinstance(parsed, bytes):
                            return parsed.decode('utf-8', errors='ignore')
                    except Exception:
                        pass
                return text
            return str(value)

        samples_norm = np.array([_normalize_value(v) for v in samples], dtype=object)
        threshold_value = _normalize_value(threshold)

        if operator == '==':
            mask = samples_norm == threshold_value
        elif operator == '!=':
            mask = samples_norm != threshold_value
        elif operator == '>':
            mask = samples_norm > threshold_value
        elif operator == '<':
            mask = samples_norm < threshold_value
        elif operator == '>=':
            mask = samples_norm >= threshold_value
        elif operator == '<=':
            mask = samples_norm <= threshold_value
        else:
            return None

    idx = np.flatnonzero(mask)
    if idx.size == 0:
        return None

    first_idx = int(idx[0])
    first_time = float(timestamps[first_idx])

    # If the first available sample is already active and caller asks for it,
    # snap to the known reference start (e.g. movement_start=0.0).
    if snap_to_reference_if_first_match and reference_start is not None and first_idx == 0:
        try:
            ref = float(reference_start)
            if np.isfinite(ref) and ref <= first_time:
                return ref
        except Exception:
            pass

    return first_time


def _normalize_om_audio_signal_config(config: dict) -> dict:
    if not isinstance(config, dict):
        return config

    signals = config.get('signals') or {}
    if not isinstance(signals, dict) or not signals:
        return config

    is_correct_belt_routing = _is_om_correct_belt_routing(config)
    is_out_of_position = _is_om_out_of_position(config)

    movement_start = config.get('movement_start', config.get('tgaze'))
    try:
        movement_start = float(movement_start) if movement_start is not None else None
    except Exception:
        movement_start = None

    cbr_det_name = config.get('cbr_detection_signal')
    print(
        f"[OM DEBUG] normalize start | cbr={is_correct_belt_routing} oop={is_out_of_position} "
        f"movement_start={movement_start} cbr_detection_signal={cbr_det_name!r} "
        f"signals={list(signals.keys())[:8]}"
    )

    audio_key = None
    for key in list(signals.keys()):
        if key in OM_AUDIO_SIGNAL_ALIASES:
            audio_key = key
            break

    if audio_key is None:
        print("[OM DEBUG] no OM audio signal alias found; audio warning timing will be unavailable")
        signal_data = {}
        signal_times = config.get('signal_times') if isinstance(config.get('signal_times'), dict) else None
    else:
        signal_data = signals.get(audio_key, {})
        if not isinstance(signal_data, dict):
            signal_data = {}

    audio_params = config.get('audio_params') or {}
    micro_threshold = audio_params.get('threshold', 0)
    try:
        micro_threshold = float(micro_threshold)
    except Exception:
        micro_threshold = 0

    if audio_key is not None:
        signal_data['threshold'] = micro_threshold
        signal_data['operator'] = '>='
        signal_data.setdefault('alias', audio_key)

        if audio_key != "SoundPressure":
            signals["SoundPressure"] = signal_data
            del signals[audio_key]
        else:
            signals["SoundPressure"] = signal_data

        signal_times = config.get('signal_times')
        if isinstance(signal_times, dict):
            if audio_key in signal_times:
                signal_times["SoundPressure"] = signal_times[audio_key]
                if audio_key != "SoundPressure":
                    del signal_times[audio_key]

        pass_signal_name = config.get('pass_signal_name')
        if pass_signal_name in OM_AUDIO_SIGNAL_ALIASES:
            config['pass_signal_name'] = "SoundPressure"

    config['show_thresholds'] = True
    config['om_debug_audio'] = True
    config['om_audio_simple_crossing'] = True

    warning_start = None
    audio_duration = None
    audio_max_gap = None
    audio_span = None
    detection_delay = None
    cbr_det_activation = None

    try:
        import numpy as np
        ts = np.array(signal_data.get('timestamps', []), dtype=float)
        ys = np.array(signal_data.get('samples', []), dtype=float)
        print(f"[OM DEBUG] audio_key={audio_key} samples={ys.size} timestamps={ts.size} threshold={micro_threshold}")
        if ts.size > 1 and ys.size == ts.size and micro_threshold > 0:
            min_f = float(audio_params.get('min_freq', 0) or 0)
            max_f = float(audio_params.get('max_freq', 0) or 0)
            print(f"[OM DEBUG] filter params min_f={min_f} max_f={max_f}")

            ys_filtered = ys
            if min_f > 0 or max_f > 0:
                try:
                    from scipy import signal as scipy_signal
                    dt = np.mean(np.diff(ts))
                    if dt > 0:
                        fs = 1.0 / dt
                        nyq = 0.5 * fs
                        low = min_f / nyq if nyq > 0 else 0
                        high = max_f / nyq if nyq > 0 else 0
                        if 0 < low < 1 and 0 < high < 1 and low < high:
                            b, a = scipy_signal.butter(4, [low, high], btype='band')
                            ys_filtered = scipy_signal.filtfilt(b, a, ys)
                            print(f"[OM DEBUG] filtering applied fs={fs:.3f} low={low:.5f} high={high:.5f}")
                        else:
                            print(f"[OM DEBUG] filtering skipped invalid band low={low:.5f} high={high:.5f}")
                except Exception:
                    print("[OM DEBUG] filtering failed")

            first_cross = None

            def _has_activation_in_previous_second(candidate_t: float) -> bool:
                window_start = candidate_t - 1.0
                prev_mask = (ts >= window_start) & (ts < candidate_t)
                if not np.any(prev_mask):
                    return False
                return bool(np.any(ys_filtered[prev_mask] >= micro_threshold))

            selected_cluster = None

            if is_correct_belt_routing:
                print(f"[OM DEBUG] Correct Belt Routing mode active. movement_start={movement_start}")
                active_idx = [
                    i for i, (t, v) in enumerate(zip(ts, ys_filtered))
                    if (movement_start is None or t >= movement_start) and v >= micro_threshold
                ]
                print(f"[OM DEBUG] candidate activations after movement_start: {len(active_idx)}")

                for pos, idx in enumerate(active_idx):
                    t0 = float(ts[idx])
                    next_dt = None
                    if pos + 1 < len(active_idx):
                        next_dt = float(ts[active_idx[pos + 1]] - t0)

                    if _has_activation_in_previous_second(t0):
                        continue

                    if next_dt is not None and next_dt < 1.0:
                        first_cross = t0
                        print(f"[OM DEBUG] sustained activation accepted at t={first_cross:.4f}s (next dt={next_dt:.4f}s < 1.0s)")
                        break
                    else:
                        print(f"[OM DEBUG] activation rejected at t={t0:.4f}s (next dt={next_dt})")
            elif is_out_of_position:
                print(f"[OM DEBUG] Out of Position mode active. movement_start={movement_start}")
                for t, v in zip(ts, ys_filtered):
                    if movement_start is not None and t < movement_start:
                        continue
                    if v >= micro_threshold:
                        if _has_activation_in_previous_second(float(t)):
                            continue
                        first_cross = float(t)
                        print(f"[OM DEBUG] OOP first activation accepted at t={first_cross:.4f}s value={float(v):.6f}")
                        break
            else:
                # Default OM audio behavior: first sample above threshold on filtered signal,
                # also requiring a clean previous second and respecting movement_start when present.
                for t, v in zip(ts, ys_filtered):
                    if movement_start is not None and t < movement_start:
                        continue
                    if v >= micro_threshold:
                        if _has_activation_in_previous_second(float(t)):
                            continue
                        first_cross = float(t)
                        print(f"[OM DEBUG] first crossing found t={first_cross:.4f}s value={float(v):.6f} >= threshold={micro_threshold}")
                        break

            if first_cross is None:
                y_min = float(np.min(ys_filtered)) if ys_filtered.size else float('nan')
                y_max = float(np.max(ys_filtered)) if ys_filtered.size else float('nan')
                print(f"[OM DEBUG] no crossing found; filtered min={y_min:.6f} max={y_max:.6f} threshold={micro_threshold}")

            if isinstance(signal_times, dict):
                signal_times['SoundPressure'] = first_cross
                signal_times['MySound PressureTask.Sound Pressure'] = first_cross
                print(f"[OM DEBUG] signal_times['SoundPressure']={first_cross}")

            ts_eval = ts
            ys_eval = ys_filtered
            if movement_start is not None:
                time_mask = ts_eval >= movement_start
                ts_eval = ts_eval[time_mask]
                ys_eval = ys_eval[time_mask]

            warning_start = first_cross
            if ts_eval.size > 1 and ys_eval.size == ts_eval.size and micro_threshold > 0:
                mask_eval = ys_eval >= micro_threshold
                clusters = _find_audio_clusters(mask_eval, ts_eval, max_gap=5.0)
                if clusters:
                    if warning_start is not None:
                        for c in clusters:
                            if c['start_time'] <= warning_start <= c['end_time']:
                                selected_cluster = c
                                break
                    if selected_cluster is None:
                        selected_cluster = clusters[0]

                    audio_duration = float(selected_cluster['duration'])
                    audio_max_gap = float(
                        _calculate_max_gap_in_cluster(
                            mask_eval,
                            ts_eval,
                            selected_cluster['start_idx'],
                            selected_cluster['end_idx']
                        )
                    )
                    audio_span = (
                        float(selected_cluster['start_time']),
                        float(selected_cluster['end_time'])
                    )

            if warning_start is not None and movement_start is not None:
                detection_delay = float(warning_start - movement_start)
        else:
            print("[OM DEBUG] crossing calc skipped (insufficient data or threshold <= 0)")
    except Exception:
        print("[OM DEBUG] exception while normalizing audio config")

    if (is_correct_belt_routing and movement_start is not None and float(movement_start) == 0.0):
        print(
            f"[OM DEBUG] CBR mark==0 path enabled | selected signal={cbr_det_name!r} "
            f"available={cbr_det_name in signals if cbr_det_name else False}"
        )
        if not cbr_det_name:
            print("[OM DEBUG] CBR mark==0: no detection signal selected in new table")
        elif cbr_det_name not in signals:
            print(f"[OM DEBUG] CBR mark==0: selected signal '{cbr_det_name}' not present in config['signals']")
        else:
            import numpy as np
            det_data = signals.get(cbr_det_name) or {}
            det_ts_arr = np.array(det_data.get('timestamps') or [], dtype=float)
            det_ys_arr = np.asarray(det_data.get('samples') or [])
            det_op = str(det_data.get('operator') or '>').strip() or '>'
            raw_thresh = det_data.get('threshold')
            det_thresh = raw_thresh if raw_thresh is not None else 0.0

            det_dtype = str(det_ys_arr.dtype) if hasattr(det_ys_arr, 'dtype') else type(det_ys_arr).__name__

            print(
                f"[OM DEBUG] CBR detection signal data | name={cbr_det_name!r} "
                f"samples={det_ys_arr.size} timestamps={det_ts_arr.size} "
                f"dtype={det_dtype} operator={det_op!r} threshold={det_thresh!r}"
            )
            if det_ys_arr.size and det_ts_arr.size == det_ys_arr.size:
                preview_count = min(5, det_ys_arr.size)
                print(
                    f"[OM DEBUG] CBR detection preview | "
                    f"t={det_ts_arr[:preview_count].tolist()} "
                    f"y={det_ys_arr[:preview_count].tolist()}"
                )

            cbr_det_activation = _find_first_signal_activation(
                det_ys_arr,
                det_ts_arr,
                det_thresh,
                operator=det_op,
                reference_start=movement_start,
                snap_to_reference_if_first_match=True,
            )
            print(f"[OM DEBUG] CBR detection activation result | activation={cbr_det_activation}")

            if cbr_det_activation is not None:
                detection_delay = float(cbr_det_activation - movement_start)
                print(
                    f"[OM DEBUG] CBR mark==0 override applied | "
                    f"movement_start={movement_start} activation={cbr_det_activation} "
                    f"detection_delay={detection_delay}"
                )
            else:
                print("[OM DEBUG] CBR mark==0 override not applied because no activation matched the configured criterion")

    config['om_audio_metrics'] = {
        'movement_start': movement_start,
        'warning_start': warning_start,
        'detection_delay': detection_delay,
        'audio_duration': audio_duration,
        'max_audio_gap': audio_max_gap,
    }
    if cbr_det_name:
        config['om_audio_metrics']['cbr_detection_signal'] = cbr_det_name
    if cbr_det_activation is not None:
        config['om_audio_metrics']['cbr_detection_signal_start'] = float(cbr_det_activation)
    config['om_audio_event_span'] = audio_span

    return config


def _format_seconds(seconds: float) -> str:
    if seconds is None or seconds < 0:
        return "00:00"
    total = int(seconds)
    minutes = total // 60
    secs = total % 60
    return f"{minutes:02d}:{secs:02d}"


class OMAnalysisScanner(AnalysisScanner):
    def run(self):
        if not self.source_dir or not os.path.exists(self.source_dir):
            self.log.emit(f"Invalid source directory: {self.source_dir}")
            self.finished.emit([])
            return

        self.log.emit(f"Scanning source directory (OM): {self.source_dir}")

        marks_keys = set()
        try:
            marks_path = self.marks_path or os.path.join(self.source_dir, 'OM_marks.json')
            if os.path.exists(marks_path):
                with open(marks_path, 'r', encoding='utf-8') as f:
                    marks_data = json.load(f)
                    marks_keys = set(marks_data.keys())
        except Exception:
            pass

        results = []
        try:
            fusion_results_dir = os.path.join(self.source_dir, "_FUSION_RESULTS")
            scan_root = fusion_results_dir if os.path.exists(fusion_results_dir) else self.source_dir

            parts = [
                d for d in os.listdir(scan_root)
                if os.path.isdir(os.path.join(scan_root, d))
                and re.search(r'(?:^|\D)(0[1-9]|[1-9][0-9])(?:\D|$)', d)
            ]
            parts.sort()

            for p in parts:
                p_path = os.path.join(scan_root, p)
                data = self._scan_recursive(p_path, marks_keys)
                total = data["total_mf4"]
                tracking_done = data["total_tracking"]
                analysis_done = data["total_analysis"]

                color = "#d1242f"
                if total > 0:
                    if tracking_done == total:
                        color = "#2da44e"
                    elif tracking_done > 0:
                        color = IDIADA_ORANGE
                elif total == 0:
                    color = "gray"

                results.append({
                    "name": p,
                    "type": "participant",
                    "path": p_path,
                    "children": data["children"],
                    "tracking_stats": (tracking_done, total),
                    "marks_stats": (data.get("total_marks", 0), total),
                    "analysis_stats": (analysis_done, total),
                    "color": color
                })
        except Exception:
            pass

        self.finished.emit(results)
        self.log.emit("OM scan finished.")

    def _scan_recursive(self, path, marks_keys=None):
        data = super()._scan_recursive(path, marks_keys)
        try:
            avi_files = [
                f for f in os.listdir(path)
                if os.path.isfile(os.path.join(path, f)) and f.lower().endswith('.avi') and not f.startswith('._')
            ]
            avi_files.sort(key=lambda s: [int(t) if t.isdigit() else t.lower() for t in re.split('([0-9]+)', s)])

            existing_paths = {c.get("path") for c in data.get("children", []) if isinstance(c, dict)}
            for avi in avi_files:
                avi_path = os.path.join(path, avi)
                if avi_path in existing_paths:
                    continue
                data["children"].append({
                    "name": avi,
                    "type": "video",
                    "path": avi_path,
                })
        except Exception:
            pass
        return data


class OMParticipantTreeWidget(QTreeWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setDragEnabled(True)
        self.setSelectionMode(QAbstractItemView.SingleSelection)
        self.setDragDropMode(QAbstractItemView.DragOnly)
        self.setDefaultDropAction(Qt.CopyAction)

    def startDrag(self, supportedActions):
        item = self.currentItem()
        if item is None:
            return

        fpath = item.data(0, Qt.UserRole)
        if not fpath or not isinstance(fpath, str) or not os.path.exists(fpath):
            return

        mime_data = QMimeData()
        mime_data.setUrls([QUrl.fromLocalFile(fpath)])

        drag = QDrag(self)
        drag.setMimeData(mime_data)
        drag.exec(Qt.CopyAction)


class OMLogicTabWidget(LogicTabWidget):
    MAX_SIGNALS_PER_CATEGORY = 6

    CATEGORIES = [
        "Correct Belt Routing",
        "Out of Position",
        "Occupant Stature",
    ]

    def __init__(self, parent=None):
        super().__init__(parent)
        self._suppress_limit_enforcement = False
        self._connect_category_limit_handlers()
        self._set_fixed_protocol_indicator()
        self._add_generate_reports_button()
        self._setup_cbr_detection_signal_column()

    def _setup_cbr_detection_signal_column(self):
        """Insert a narrow 'Signal for detection' table to the left of the
        Correct Belt Routing signals table.  Only one signal can be checked
        at a time (enforced by _on_cbr_detection_item_changed).
        """
        cbr_table = self.category_tables.get("Correct Belt Routing")
        if cbr_table is None:
            return
        gb = cbr_table.parent()
        if gb is None:
            return
        l_cat = gb.layout()
        if l_cat is None:
            return

        cbr_idx = l_cat.indexOf(cbr_table)
        l_cat.removeWidget(cbr_table)

        self.cbr_detection_table = QTableWidget(0, 1)
        self.cbr_detection_table.setHorizontalHeaderLabels(["Signal for\ndetection"])
        self.cbr_detection_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.cbr_detection_table.verticalHeader().setVisible(False)
        self.cbr_detection_table.setFixedWidth(145)
        self.cbr_detection_table.setMinimumHeight(80)
        self.cbr_detection_table.setStyleSheet(cbr_table.styleSheet())
        self.cbr_detection_table.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.cbr_detection_table.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.cbr_detection_table.itemChanged.connect(self._on_cbr_detection_item_changed)

        container = QWidget()
        container.setObjectName("cbr_signals_container")
        h_lay = QHBoxLayout(container)
        h_lay.setContentsMargins(0, 0, 0, 0)
        h_lay.setSpacing(4)
        h_lay.addWidget(self.cbr_detection_table, 0)
        h_lay.addWidget(cbr_table, 1)

        if cbr_idx >= 0:
            l_cat.insertWidget(cbr_idx, container)
        else:
            l_cat.insertWidget(0, container)

        # Keep detection table scrolled in sync with the main CBR table.
        cbr_table.verticalScrollBar().valueChanged.connect(self._on_cbr_main_scroll)

    def _on_cbr_main_scroll(self, value: int):
        if hasattr(self, 'cbr_detection_table'):
            self.cbr_detection_table.verticalScrollBar().blockSignals(True)
            self.cbr_detection_table.verticalScrollBar().setValue(value)
            self.cbr_detection_table.verticalScrollBar().blockSignals(False)

    def _on_cbr_detection_item_changed(self, item):
        """Enforce single-selection: unchecks all other rows when one is checked."""
        if item is None or item.column() != 0:
            return
        if item.checkState() != Qt.Checked:
            return
        table = self.cbr_detection_table
        table.blockSignals(True)
        try:
            for row in range(table.rowCount()):
                if row != item.row():
                    chk = table.item(row, 0)
                    if chk:
                        chk.setCheckState(Qt.Unchecked)
        finally:
            table.blockSignals(False)

    def _sync_cbr_detection_table(self):
        """Mirror the CBR main table's signal rows into cbr_detection_table.
        Preserves the currently selected signal across rebuilds.
        """
        if not hasattr(self, 'cbr_detection_table'):
            return
        cbr_table = self.category_tables.get("Correct Belt Routing")
        if cbr_table is None:
            return
        det_table = self.cbr_detection_table

        # Remember which signal is currently selected
        selected_name = None
        for row in range(det_table.rowCount()):
            chk = det_table.item(row, 0)
            if chk and chk.checkState() == Qt.Checked:
                selected_name = chk.data(Qt.UserRole)
                break

        det_table.blockSignals(True)
        det_table.setRowCount(0)
        try:
            for row in range(cbr_table.rowCount()):
                name_item = cbr_table.item(row, 1)
                if not name_item:
                    continue
                sig_name = name_item.text()
                det_row = det_table.rowCount()
                det_table.insertRow(det_row)

                chk = QTableWidgetItem(sig_name)
                chk.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
                chk.setCheckState(Qt.Checked if sig_name == selected_name else Qt.Unchecked)
                chk.setData(Qt.UserRole, sig_name)
                chk.setToolTip(sig_name)
                det_table.setItem(det_row, 0, chk)
                det_table.setRowHeight(det_row, cbr_table.rowHeight(row))
        finally:
            det_table.blockSignals(False)

    def _connect_category_limit_handlers(self):
        for table in self.category_tables.values():
            if getattr(table, '_om_limit_connected', False):
                continue
            table.itemChanged.connect(self._on_category_table_item_changed)
            table._om_limit_connected = True

    def _checked_count_for_table(self, table) -> int:
        count = 0
        for row in range(table.rowCount()):
            chk = table.item(row, 0)
            if chk and chk.checkState() == Qt.Checked:
                count += 1
        return count

    def _enforce_table_limit(self, table):
        checked_rows = []
        for row in range(table.rowCount()):
            chk = table.item(row, 0)
            if chk and chk.checkState() == Qt.Checked:
                checked_rows.append(row)

        if len(checked_rows) <= self.MAX_SIGNALS_PER_CATEGORY:
            return

        table.blockSignals(True)
        try:
            for row in checked_rows[self.MAX_SIGNALS_PER_CATEGORY:]:
                chk = table.item(row, 0)
                if chk:
                    chk.setCheckState(Qt.Unchecked)
        finally:
            table.blockSignals(False)

    def _enforce_category_selection_limits(self):
        for table in self.category_tables.values():
            self._enforce_table_limit(table)

    def _on_category_table_item_changed(self, item):
        if self._suppress_limit_enforcement:
            return
        if item is None or item.column() != 0:
            return
        table = item.tableWidget()
        if table is None:
            return
        if item.checkState() != Qt.Checked:
            return

        if self._checked_count_for_table(table) > self.MAX_SIGNALS_PER_CATEGORY:
            table.blockSignals(True)
            try:
                item.setCheckState(Qt.Unchecked)
            finally:
                table.blockSignals(False)
            QMessageBox.warning(
                self,
                "Selection limit",
                f"You can select up to {self.MAX_SIGNALS_PER_CATEGORY} signals in each OM category.",
            )

    def _set_fixed_protocol_indicator(self):
        if not hasattr(self, 'combo_report_type') or self.combo_report_type is None:
            return

        # Keep report generation value deterministic for OM.
        self.combo_report_type.blockSignals(True)
        self.combo_report_type.clear()
        self.combo_report_type.addItem("Euro NCAP")
        self.combo_report_type.setCurrentIndex(0)
        self.combo_report_type.blockSignals(False)

        container = self.combo_report_type.parentWidget()
        layout = container.layout() if container is not None else None
        if layout is None:
            self.combo_report_type.hide()
            return

        idx = layout.indexOf(self.combo_report_type)
        self.combo_report_type.hide()

        self.lbl_protocol_fixed = QLabel("Euro NCAP")
        self.lbl_protocol_fixed.setStyleSheet("color: #ddd; font-weight: bold;")
        if idx >= 0:
            layout.insertWidget(idx, self.lbl_protocol_fixed, 1)
        else:
            layout.addWidget(self.lbl_protocol_fixed, 1)

    def _add_generate_reports_button(self):
        if not hasattr(self, 'btn_generate') or self.btn_generate is None:
            return

        parent_widget = self.btn_generate.parentWidget()
        parent_layout = parent_widget.layout() if parent_widget is not None else None
        if parent_layout is None:
            return

        index = parent_layout.indexOf(self.btn_generate)
        if index < 0:
            return

        self.btn_generate_reports = QPushButton("GENERATE REPORTS")
        self.btn_generate_reports.setStyleSheet(
            f"background-color: {IDIADA_ORANGE}; color: black; border: 1px solid {IDIADA_ORANGE}; padding: 12px; font-weight: bold; font-size: 12px;"
        )
        self.btn_generate_reports.setCursor(Qt.PointingHandCursor)
        self.btn_generate_reports.clicked.connect(self._on_generate_reports_clicked)

        self.btn_stop_generating = QPushButton("STOP GENERATING")
        self.btn_stop_generating.setStyleSheet(
            "background-color: transparent; color: #ff8a80; border: 1px solid #ff8a80; padding: 12px; font-weight: bold; font-size: 12px;"
        )
        self.btn_stop_generating.setCursor(Qt.PointingHandCursor)
        self.btn_stop_generating.clicked.connect(self._on_stop_generating_reports_clicked)

        row = QHBoxLayout()
        row.setContentsMargins(0, 0, 0, 0)
        row.setSpacing(8)

        parent_layout.removeWidget(self.btn_generate)
        row.addWidget(self.btn_generate, 1)
        row.addWidget(self.btn_generate_reports, 1)
        row.addWidget(self.btn_stop_generating, 1)
        parent_layout.insertLayout(index, row)

    def _on_generate_reports_clicked(self):
        parent_aw = self._get_parent_analysis_widget()
        if parent_aw is None or not hasattr(parent_aw, 'generate_reports_from_time_selector'):
            QMessageBox.warning(self, "Generate Reports", "OM Analysis context not available.")
            return
        parent_aw.generate_reports_from_time_selector()

    def _on_stop_generating_reports_clicked(self):
        parent_aw = self._get_parent_analysis_widget()
        if parent_aw is None or not hasattr(parent_aw, 'request_stop_report_generation'):
            QMessageBox.warning(self, "Stop Generating", "OM Analysis context not available.")
            return
        parent_aw.request_stop_report_generation()

    def _make_report_worker(self, config, output_path, dpi=300):
        """Use OMReportGeneratorWorker for preview reports so they go through OMReportBuilder."""
        return OMReportGeneratorWorker(config, output_path, dpi)

    def _lock_audio_rows(self):
        parent = self._get_parent_analysis_widget()
        micro_threshold = None
        if parent and hasattr(parent, 'spin_threshold'):
            try:
                micro_threshold = float(parent.spin_threshold.value())
            except Exception:
                micro_threshold = None

        for _, table in self.category_tables.items():
            for row in range(table.rowCount()):
                name_item = table.item(row, 1)
                sig_name = name_item.text() if name_item else ""
                is_audio = sig_name in OM_AUDIO_SIGNAL_ALIASES

                op_widget = table.cellWidget(row, 2)
                if op_widget is not None:
                    op_widget.setEnabled(not is_audio)
                    if is_audio and hasattr(op_widget, 'setCurrentText'):
                        op_widget.setCurrentText('>=')

                val_widget = table.cellWidget(row, 3)
                if val_widget is not None:
                    val_widget.setEnabled(not is_audio)
                    if is_audio and micro_threshold is not None:
                        try:
                            if hasattr(val_widget, 'setValue'):
                                val_widget.setValue(micro_threshold)
                            elif hasattr(val_widget, 'setText'):
                                val_widget.setText(str(micro_threshold))
                        except Exception:
                            pass

    def update_signals(self, signal_names, mdf):
        super().update_signals(signal_names, mdf)
        self._suppress_limit_enforcement = True
        try:
            self._enforce_category_selection_limits()
        finally:
            self._suppress_limit_enforcement = False
        self._lock_audio_rows()
        self._sync_cbr_detection_table()

    def _apply_logic_config(self, config: dict):
        super()._apply_logic_config(config)
        self._suppress_limit_enforcement = True
        try:
            self._enforce_category_selection_limits()
        finally:
            self._suppress_limit_enforcement = False
        self._lock_audio_rows()
        # Sync detection table and restore saved selection
        self._sync_cbr_detection_table()
        cbr_det = config.get('cbr_detection_signal')
        if cbr_det and hasattr(self, 'cbr_detection_table'):
            det_table = self.cbr_detection_table
            det_table.blockSignals(True)
            try:
                for row in range(det_table.rowCount()):
                    chk = det_table.item(row, 0)
                    if chk:
                        chk.setCheckState(
                            Qt.Checked if chk.data(Qt.UserRole) == cbr_det else Qt.Unchecked
                        )
            finally:
                det_table.blockSignals(False)

    def update_audio_params(self, min_f, max_f, thresh):
        super().update_audio_params(min_f, max_f, thresh)
        self._lock_audio_rows()

    def _collect_report_config(self) -> dict:
        config = super()._collect_report_config()

        parent_aw = self._get_parent_analysis_widget()

        # Resolve active OM category from the currently analyzed case path.
        candidate_paths = []
        if getattr(self, 'mdf', None) is not None:
            candidate_paths.append(getattr(self.mdf, 'name', None))
        if parent_aw is not None and hasattr(parent_aw, 'tab_time'):
            candidate_paths.append(getattr(parent_aw.tab_time, 'current_tracking_path', None))
            candidate_paths.append(getattr(parent_aw.tab_time, 'current_video_path', None))

        target_category = None
        for path in candidate_paths:
            inferred = _infer_om_category_from_path(path)
            if inferred in self.category_tables:
                target_category = inferred
                break

        if target_category is None:
            cfg_target = config.get('target_category')
            if isinstance(cfg_target, str) and cfg_target in self.category_tables:
                target_category = cfg_target

        if target_category:
            config['target_category'] = target_category

        # OM preview must always reflect the current ticks in Logic tables.
        # Rebuild selected signals from the target OM category at preview time.
        if self.mdf is not None and isinstance(target_category, str) and target_category in self.category_tables:
            table = self.category_tables[target_category]
            rebuilt_signals = {}
            for row in range(table.rowCount()):
                chk = table.item(row, 0)
                name_item = table.item(row, 1)
                if not (chk and name_item and chk.checkState() == Qt.Checked):
                    continue

                sig_name = name_item.text()
                try:
                    sig = self.mdf.get(sig_name)

                    op_widget = table.cellWidget(row, 2)
                    operator = op_widget.currentText() if isinstance(op_widget, QComboBox) else "None"

                    val_widget = table.cellWidget(row, 3)
                    value = None
                    if isinstance(val_widget, QDoubleSpinBox):
                        value = val_widget.value()
                    elif isinstance(val_widget, QComboBox):
                        value = val_widget.currentText()
                    elif isinstance(val_widget, QLineEdit):
                        try:
                            value = float(val_widget.text()) if val_widget.text() else None
                        except Exception:
                            value = val_widget.text()

                    units_item = table.item(row, 4)
                    units = units_item.text() if units_item else ""

                    alias_item = table.item(row, 5)
                    alias = alias_item.text() if alias_item else sig_name

                    rebuilt_signals[sig_name] = {
                        'timestamps': list(sig.timestamps),
                        'samples': list(sig.samples),
                        'threshold': value,
                        'operator': operator,
                        'unit': units or getattr(sig, 'unit', 'Value'),
                        'category': target_category,
                        'alias': alias,
                    }
                except Exception as e:
                    print(f"Error getting signal {sig_name}: {e}")

            # Only replace if we actually rebuilt at least one selected signal.
            # This prevents empty previews when category inference/path mapping is off.
            if rebuilt_signals:
                config['signals'] = rebuilt_signals
            else:
                fallback = config.get('signals') if isinstance(config.get('signals'), dict) else {}
                if fallback:
                    filtered = {
                        k: v for k, v in fallback.items()
                        if isinstance(v, dict) and str(v.get('category', '') or '') == target_category
                    }
                    if filtered:
                        config['signals'] = filtered

            pass_table = self.pass_criteria_tables.get(target_category)
            if pass_table and pass_table.rowCount() > 0:
                if hasattr(pass_table, 'signal_combo'):
                    pass_sig = pass_table.signal_combo.currentText()
                else:
                    signal_combo = pass_table.cellWidget(0, 0)
                    pass_sig = signal_combo.currentText() if hasattr(signal_combo, 'currentText') else ""
                if pass_sig and pass_sig != "-- Select Signal --":
                    config['pass_signal_name'] = pass_sig

        movement_start = None
        matched_video_name = None
        try:
            mdf_path = getattr(self.mdf, 'name', None) if getattr(self, 'mdf', None) is not None else None
            if parent_aw and mdf_path:
                if hasattr(parent_aw, '_resolve_om_mark_for_mf4'):
                    mark_info = parent_aw._resolve_om_mark_for_mf4(mdf_path)
                    movement_start = mark_info.get('mark_time')
                    matched_video_name = mark_info.get('matched_video')
                    if movement_start is not None:
                        parent_aw.log(
                            f"[OM MARKS][PREVIEW] match for {os.path.basename(mdf_path)} "
                            f"-> {mark_info.get('matched_key')} | mark_time_s={movement_start:.6f}"
                        )
                    else:
                        parent_aw.log(
                            f"[OM MARKS][PREVIEW] no match for {os.path.basename(mdf_path)} "
                            f"({mark_info.get('reason')})"
                        )
                elif hasattr(parent_aw, '_find_om_mark_time_for_mf4'):
                    movement_start = parent_aw._find_om_mark_time_for_mf4(mdf_path)
        except Exception:
            movement_start = None

        # Preview fallback only: do not persist, just use 0 when no mark match is found.
        if movement_start is None:
            if parent_aw is not None and hasattr(parent_aw, 'log'):
                mdf_name = os.path.basename(getattr(self.mdf, 'name', '') or '')
                parent_aw.log(f"[OM MARKS][PREVIEW] fallback to default movement_start=0.0 for {mdf_name}")
            movement_start = 0.0

        config['tgaze'] = movement_start
        config['movement_start'] = movement_start
        config['movement_start_label'] = 'Movement Start'

        is_cbr = _is_om_correct_belt_routing(config)
        is_oop = _is_om_out_of_position(config)
        if is_cbr or is_oop:
            config['om_report_variant'] = 'correct_belt_routing' if is_cbr else 'out_of_position'
            config['om_use_video_frame'] = True
            config['om_hide_date_time_vin'] = True
            config['om_plot_show_marks'] = True
            config['om_plot_show_shading'] = True

            current_mdf_path = getattr(self.mdf, 'name', None) if getattr(self, 'mdf', None) is not None else None
            video_path = None
            if parent_aw is not None and hasattr(parent_aw, 'tab_time'):
                video_path = getattr(parent_aw.tab_time, 'current_video_path', None)

            if not video_path and getattr(self, 'mdf', None) is not None:
                mdf_name = getattr(self.mdf, 'name', None)
                if isinstance(mdf_name, str):
                    camera_mode = ".avi"
                    if parent_aw is not None and hasattr(parent_aw, 'tab_time'):
                        combo = getattr(parent_aw.tab_time, 'combo_camera_index', None)
                        if combo is not None:
                            camera_mode = combo.currentData() or ".avi"
                    video_path = _guess_video_for_mf4_path(
                        mdf_name,
                        camera_mode=camera_mode,
                        preferred_video_name=matched_video_name,
                    )

            if video_path and os.path.exists(video_path):
                config['om_video_path'] = video_path
                secondary_video = _guess_secondary_video_for_mf4_path(current_mdf_path, video_path)
                if secondary_video and os.path.exists(secondary_video):
                    config['om_video_path_secondary'] = secondary_video

            pass_sig = config.get('pass_signal_name')
            warn_start = None
            if pass_sig:
                warn_start = (config.get('signal_times') or {}).get(pass_sig)
            capture_t = warn_start if isinstance(warn_start, (int, float)) else movement_start
            config['om_video_time_s'] = float(capture_t) if isinstance(capture_t, (int, float)) else 0.0

        # CBR: attach selected detection signal name (used in _normalize when mark==0)
        if is_cbr and hasattr(self, 'cbr_detection_table'):
            for row in range(self.cbr_detection_table.rowCount()):
                chk = self.cbr_detection_table.item(row, 0)
                if chk and chk.checkState() == Qt.Checked:
                    det_sig_name = chk.data(Qt.UserRole)
                    print(f"[OM DEBUG] _collect_report_config selected CBR detection signal: {det_sig_name!r}")
                    config['cbr_detection_signal'] = det_sig_name
                    # Ensure signal data is included even if not checked in the main table
                    if (det_sig_name and self.mdf is not None
                            and isinstance(config.get('signals'), dict)
                            and det_sig_name not in config['signals']):
                        try:
                            sig = self.mdf.get(det_sig_name)
                            op, val = '>', None
                            cbr_table = self.category_tables.get('Correct Belt Routing')
                            if cbr_table is not None:
                                for r in range(cbr_table.rowCount()):
                                    ni = cbr_table.item(r, 1)
                                    if ni and ni.text() == det_sig_name:
                                        ow = cbr_table.cellWidget(r, 2)
                                        vw = cbr_table.cellWidget(r, 3)
                                        if isinstance(ow, QComboBox):
                                            op = ow.currentText()
                                        if isinstance(vw, QDoubleSpinBox):
                                            val = vw.value()
                                        elif isinstance(vw, QLineEdit):
                                            try:
                                                val = float(vw.text())
                                            except Exception:
                                                val = vw.text()
                                        break
                            config['signals'][det_sig_name] = {
                                'timestamps': list(sig.timestamps),
                                'samples': list(sig.samples),
                                'threshold': val,
                                'operator': op,
                                'unit': getattr(sig, 'unit', 'Value'),
                                'category': 'Correct Belt Routing',
                                'alias': det_sig_name,
                            }
                            print(
                                f"[OM DEBUG] _collect_report_config injected CBR detection signal data: "
                                f"name={det_sig_name!r} operator={op!r} threshold={val} samples={len(sig.samples)}"
                            )
                        except Exception as _e:
                            print(f"[OM CBR] detection signal data fetch failed: {_e}")
                    break

        return _normalize_om_audio_signal_config(config)


class OMDualTimelineWidget(QWidget):
    position_changed = Signal(float)
    mark_changed = Signal(object)

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setMinimumHeight(120)
        self.setMouseTracking(True)
        self.duration_s = 0.0
        self.position_s = 0.0
        self.mark_s = None
        self.guide_x = None
        self._dragging = False

    def set_duration(self, duration_s: float):
        self.duration_s = max(0.0, float(duration_s or 0.0))
        if self.position_s > self.duration_s:
            self.position_s = self.duration_s
        if self.mark_s is not None and self.mark_s > self.duration_s:
            self.mark_s = self.duration_s
            self.mark_changed.emit(self.mark_s)
        self.update()

    def set_position(self, position_s: float):
        self.position_s = max(0.0, min(float(position_s or 0.0), self.duration_s if self.duration_s > 0 else 0.0))
        self.update()

    def set_mark(self, mark_s):
        if mark_s is None:
            self.mark_s = None
        else:
            self.mark_s = max(0.0, min(float(mark_s), self.duration_s if self.duration_s > 0 else float(mark_s)))
        self.update()

    def clear_mark(self):
        self.mark_s = None
        self.mark_changed.emit(None)
        self.update()

    def _bars_rects(self):
        margin_x = 12
        top = 26
        bar_h = 16
        gap = 24
        width = max(20, self.width() - (margin_x * 2))
        duration_rect = QRectF(margin_x, top, width, bar_h)
        mark_rect = QRectF(margin_x, top + bar_h + gap, width, bar_h)
        return duration_rect, mark_rect

    def _x_to_time(self, x: float) -> float:
        duration_rect, _ = self._bars_rects()
        left = duration_rect.left()
        right = duration_rect.right()
        if right <= left or self.duration_s <= 0:
            return 0.0
        clamped = min(max(x, left), right)
        ratio = (clamped - left) / (right - left)
        return ratio * self.duration_s

    def _time_to_x(self, t: float) -> float:
        duration_rect, _ = self._bars_rects()
        if self.duration_s <= 0:
            return duration_rect.left()
        ratio = min(max(t / self.duration_s, 0.0), 1.0)
        return duration_rect.left() + ratio * duration_rect.width()

    def mouseMoveEvent(self, event):
        self.guide_x = event.position().x()
        if self._dragging:
            t = self._x_to_time(self.guide_x)
            self.position_changed.emit(t)
        self.update()
        super().mouseMoveEvent(event)

    def leaveEvent(self, event):
        self.guide_x = None
        self._dragging = False
        self.update()
        super().leaveEvent(event)

    def mousePressEvent(self, event):
        duration_rect, mark_rect = self._bars_rects()
        x = event.position().x()
        y = event.position().y()

        if event.button() == Qt.LeftButton and duration_rect.contains(x, y):
            t = self._x_to_time(x)
            self._dragging = True
            self.position_changed.emit(t)
            self.update()
            return

        if event.button() == Qt.LeftButton and mark_rect.contains(x, y):
            t = self._x_to_time(x)
            self.mark_s = t
            self.mark_changed.emit(t)
            self.update()
            return

        if event.button() == Qt.RightButton and mark_rect.contains(x, y):
            t = self._x_to_time(x)
            self.mark_s = t
            self.mark_changed.emit(t)
            self.update()
            return

        super().mousePressEvent(event)

    def mouseReleaseEvent(self, event):
        self._dragging = False
        super().mouseReleaseEvent(event)

    def paintEvent(self, event):
        super().paintEvent(event)
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)

        duration_rect, mark_rect = self._bars_rects()
        bar_bg = QColor("#2b2b2b")
        border = QColor("#515151")

        painter.setPen(QPen(border, 1.0))
        painter.setBrush(QBrush(bar_bg))
        painter.drawRoundedRect(duration_rect, 6, 6)
        painter.drawRoundedRect(mark_rect, 6, 6)

        if self.duration_s > 0:
            progress_ratio = min(max(self.position_s / self.duration_s, 0.0), 1.0)
            progress_rect = QRectF(duration_rect.left(), duration_rect.top(), duration_rect.width() * progress_ratio, duration_rect.height())
            painter.setPen(Qt.NoPen)
            painter.setBrush(QColor(IDIADA_ORANGE))
            painter.drawRoundedRect(progress_rect, 6, 6)

        painter.setPen(QPen(QColor("#bdbdbd")))
        painter.drawText(QRectF(duration_rect.left(), 4, duration_rect.width(), 18), Qt.AlignLeft | Qt.AlignVCenter, "Duration / Seek")
        painter.drawText(QRectF(mark_rect.left(), mark_rect.bottom() + 6, mark_rect.width(), 18), Qt.AlignLeft | Qt.AlignVCenter, "Right click to mark")
        painter.drawText(
            QRectF(duration_rect.left(), duration_rect.bottom() + 4, duration_rect.width(), 18),
            Qt.AlignRight | Qt.AlignVCenter,
            f"{_format_seconds(self.position_s)} / {_format_seconds(self.duration_s)}",
        )

        if self.mark_s is not None:
            mark_x = self._time_to_x(self.mark_s)
            mark_pen = QPen(QColor(IDIADA_ORANGE), 2)
            painter.setPen(mark_pen)
            painter.drawLine(mark_x, duration_rect.top() - 2, mark_x, duration_rect.bottom() + 2)
            painter.drawLine(mark_x, mark_rect.top() - 2, mark_x, mark_rect.bottom() + 2)

        if self.guide_x is not None:
            guide_pen = QPen(QColor("#cfcfcf"), 1, Qt.DashLine)
            painter.setPen(guide_pen)
            top = duration_rect.top() - 8
            bottom = mark_rect.bottom() + 8
            painter.drawLine(self.guide_x, top, self.guide_x, bottom)


class OMTimeSelectorWidget(QWidget):
    tracking_loaded = Signal(str)
    log_message = Signal(str)
    marks_saved = Signal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)
        self.setFocusPolicy(Qt.StrongFocus)
        self.current_tracking_path = None
        self.current_video_path = None
        self._all_case_files = []
        self._source_root = None
        self.markers = []

        layout = QVBoxLayout(self)
        layout.setContentsMargins(8, 8, 8, 8)
        layout.setSpacing(8)

        top = QHBoxLayout()
        top.addWidget(QLabel("Subject:"))
        self.combo_subject = QComboBox()
        self.combo_subject.setFixedWidth(120)
        self.combo_subject.currentIndexChanged.connect(self._on_subject_changed)
        top.addWidget(self.combo_subject)

        top.addWidget(QLabel("Case:"))
        self.combo_cases = QComboBox()
        self.combo_cases.setMinimumWidth(200)
        self.combo_cases.currentIndexChanged.connect(self._on_case_selector_changed)
        top.addWidget(self.combo_cases, 1)

        self.btn_play = QPushButton("Play")
        self.btn_play.setCursor(Qt.PointingHandCursor)
        self.btn_play.clicked.connect(self._play_video)
        top.addWidget(self.btn_play)

        self.btn_pause = QPushButton("Pause")
        self.btn_pause.setCursor(Qt.PointingHandCursor)
        self.btn_pause.clicked.connect(self._pause_video)
        top.addWidget(self.btn_pause)

        top.addWidget(QLabel("Select Camera Index:"))
        self.combo_camera_index = QComboBox()
        self.combo_camera_index.addItem(".avi", ".avi")
        self.combo_camera_index.addItem("_m0.avi", "_m0.avi")
        self.combo_camera_index.setFixedWidth(100)
        top.addWidget(self.combo_camera_index)
        layout.addLayout(top)

        self.video_widget = QVideoWidget()
        self.video_widget.setMinimumHeight(360)
        self.video_widget.setStyleSheet("background-color: #000;")
        layout.addWidget(self.video_widget, 1)

        self.player = QMediaPlayer(self)
        self.audio_output = QAudioOutput(self)
        self.audio_output.setVolume(0.0)
        self.player.setAudioOutput(self.audio_output)
        self.player.setVideoOutput(self.video_widget)
        self.player.durationChanged.connect(self._on_duration_changed)
        self.player.positionChanged.connect(self._on_position_changed)

        self.timeline = OMDualTimelineWidget()
        self.timeline.position_changed.connect(self._on_seek_requested)
        self.timeline.mark_changed.connect(self._on_mark_changed)
        layout.addWidget(self.timeline)

        bottom = QHBoxLayout()
        self.lbl_status = QLabel("No video loaded.")
        self.lbl_status.setStyleSheet("color: #999;")
        bottom.addWidget(self.lbl_status, 1)

        self.btn_prev = QPushButton("◀ Previous")
        self.btn_prev.setCursor(Qt.PointingHandCursor)
        self.btn_prev.clicked.connect(self._navigate_prev)
        bottom.addWidget(self.btn_prev)

        self.btn_next = QPushButton("Next ▶")
        self.btn_next.setCursor(Qt.PointingHandCursor)
        self.btn_next.clicked.connect(self._navigate_next)
        bottom.addWidget(self.btn_next)

        self.btn_clear_mark = QPushButton("Clear Last Mark")
        self.btn_clear_mark.setCursor(Qt.PointingHandCursor)
        self.btn_clear_mark.clicked.connect(self._clear_last_mark)
        bottom.addWidget(self.btn_clear_mark)

        self.btn_mark_current = QPushButton("Select this frame as Mark")
        self.btn_mark_current.setCursor(Qt.PointingHandCursor)
        self.btn_mark_current.clicked.connect(self._select_current_frame_as_mark)
        bottom.addWidget(self.btn_mark_current)

        self.btn_mark_start = QPushButton("Select Start as Mark")
        self.btn_mark_start.setCursor(Qt.PointingHandCursor)
        self.btn_mark_start.clicked.connect(self._select_start_as_mark)
        bottom.addWidget(self.btn_mark_start)
        layout.addLayout(bottom)

    def set_case_list(self, files: list, source_root: str = None):
        self._all_case_files = files or []
        self._source_root = source_root
        subjects = []
        for f in self._all_case_files:
            subj = self._extract_subject(f)
            if subj and subj not in subjects:
                subjects.append(subj)

        self.combo_subject.blockSignals(True)
        self.combo_subject.clear()
        for subject in subjects:
            self.combo_subject.addItem(subject)
        self.combo_subject.blockSignals(False)

        if subjects:
            self._populate_cases_for_subject(subjects[0])
        else:
            self.combo_cases.clear()

        self.update_navigation_labels()

    def _extract_subject(self, filepath):
        try:
            normalized = os.path.normpath(filepath)
            if self._source_root:
                root_norm = os.path.normpath(self._source_root)
                rel = os.path.relpath(normalized, root_norm)
                rel_parts = rel.split(os.sep)
                if rel_parts and rel_parts[0] == '_FUSION_RESULTS' and len(rel_parts) > 1:
                    return rel_parts[1]
                if rel_parts:
                    return rel_parts[0]
        except Exception:
            pass

        parts = filepath.replace("\\", "/").split("/")
        for part in parts:
            if re.match(r'^[A-Z]\d{2}$', part):
                return part
        return os.path.basename(os.path.dirname(filepath)) or "Unknown"

    def _populate_cases_for_subject(self, subject):
        self.combo_cases.blockSignals(True)
        self.combo_cases.clear()
        for fpath in self._all_case_files:
            if self._extract_subject(fpath) == subject:
                if fpath.lower().endswith('.avi'):
                    name = os.path.basename(fpath)
                else:
                    name = os.path.basename(fpath).replace("_tracking.mf4", "").replace(".mf4", "")
                self.combo_cases.addItem(name, fpath)
        self.combo_cases.blockSignals(False)

    def _on_subject_changed(self, index):
        if index < 0:
            return
        subject = self.combo_subject.itemText(index)
        self._populate_cases_for_subject(subject)
        if self.combo_cases.count() > 0:
            self.combo_cases.setCurrentIndex(0)
            first = self.combo_cases.itemData(0)
            if first:
                self.load_tracking_file(first)
        self.update_navigation_labels()

    def _on_case_selector_changed(self, index):
        if index < 0:
            return
        fpath = self.combo_cases.itemData(index)
        if fpath:
            if isinstance(fpath, str) and fpath.lower().endswith('.avi'):
                related_tracking = self._guess_tracking_from_video(fpath)
                self.load_video_file(fpath, related_tracking)
            elif fpath != self.current_tracking_path:
                self.load_tracking_file(fpath)
        self.update_navigation_labels()

    def update_navigation_labels(self):
        pass

    def previous_plot(self):
        count = self.combo_cases.count()
        if count == 0:
            return
        idx = self.combo_cases.currentIndex()
        if idx > 0:
            self.combo_cases.setCurrentIndex(idx - 1)

    def next_plot(self):
        count = self.combo_cases.count()
        if count == 0:
            return
        idx = self.combo_cases.currentIndex()
        if idx < count - 1:
            self.combo_cases.setCurrentIndex(idx + 1)

    def load_tracking_file(self, fpath_or_mdf):
        if not isinstance(fpath_or_mdf, str):
            return

        self.current_tracking_path = fpath_or_mdf
        self.tracking_loaded.emit(fpath_or_mdf)

        idx = self.combo_cases.findData(fpath_or_mdf)
        if idx >= 0:
            self.combo_cases.blockSignals(True)
            self.combo_cases.setCurrentIndex(idx)
            self.combo_cases.blockSignals(False)

        video_path = self._guess_video_path(fpath_or_mdf)
        if video_path and os.path.exists(video_path):
            self.load_video_file(video_path)
        else:
            self.player.setSource(QUrl())
            self.current_video_path = None
            self.timeline.set_duration(0)
            self.timeline.set_position(0)
            self.lbl_status.setText("No matching video found for this case.")
            self.log_message.emit(f"OM video not found for: {os.path.basename(fpath_or_mdf)}")

        self._restore_mark()

    def _guess_video_path(self, tracking_path: str):
        base_name = os.path.splitext(os.path.basename(tracking_path))[0]
        clean_base = base_name.replace("_tracking", "")
        directory = os.path.dirname(tracking_path)
        parent = os.path.dirname(directory)

        suffix = getattr(self, 'combo_camera_index', None)
        mode = suffix.currentData() if suffix is not None else ".avi"

        video_exts = [".avi", ".mp4", ".mov", ".mkv"]

        if mode == "_m0.avi":
            # Strict: video filename must be clean_base + "_m0" + ext
            for ext in video_exts:
                for folder in [directory, parent]:
                    candidate = os.path.join(folder, clean_base + "_m0" + ext)
                    if os.path.exists(candidate):
                        return candidate
        else:
            # Strict: video filename must be clean_base + ext (exact match, no suffix)
            for ext in video_exts:
                for folder in [directory, parent]:
                    candidate = os.path.join(folder, clean_base + ext)
                    if os.path.exists(candidate):
                        return candidate
        return None

    def _guess_tracking_from_video(self, video_path: str):
        base = os.path.splitext(os.path.basename(video_path))[0]

        # Check active camera mode and strip the corresponding suffix from the
        # video name to recover the canonical MF4 base name.
        mode = ".avi"
        combo = getattr(self, 'combo_camera_index', None)
        if combo is not None:
            mode = combo.currentData() or ".avi"

        if mode == "_m0.avi":
            if not re.search(r'_m0$', base, flags=re.IGNORECASE):
                return None
            base = re.sub(r'_m0$', '', base, flags=re.IGNORECASE)
        else:
            # Default mode: strip generic _cam<n> suffixes only
            base = re.sub(r'_cam\d+$', '', base, flags=re.IGNORECASE)

        directory = os.path.dirname(video_path)
        parent = os.path.dirname(directory)

        candidates = [
            os.path.join(directory, f"{base}_tracking.mf4"),
            os.path.join(directory, f"{base}.mf4"),
            os.path.join(parent, f"{base}_tracking.mf4"),
            os.path.join(parent, f"{base}.mf4"),
        ]
        for candidate in candidates:
            if os.path.exists(candidate):
                if candidate.lower().endswith('_tracking.mf4'):
                    return candidate
                return candidate[:-4] + '_tracking.mf4'
        return None

    def _on_duration_changed(self, duration_ms: int):
        self.timeline.set_duration(duration_ms / 1000.0)

    def _on_position_changed(self, position_ms: int):
        self.timeline.set_position(position_ms / 1000.0)

    def _on_seek_requested(self, seconds: float):
        self.player.setPosition(int(max(0.0, seconds) * 1000.0))

    def _on_mark_changed(self, seconds):
        if self.current_tracking_path is None:
            self.markers = []
            try:
                self.timeline.blockSignals(True)
                self.timeline.clear_mark()
            finally:
                self.timeline.blockSignals(False)
            self.lbl_status.setText("No MF4 associated with this video.")
            QMessageBox.information(self, "Mark", "No MF4 associated with this video.\nCheck the Camera Index selector matches the video suffix.")
            return

        if seconds is None:
            self.markers = []
            self._update_om_marks_file(None)
            self.lbl_status.setText("Mark cleared.")
            self.log_message.emit("OM mark cleared")
            return

        mark_time = float(seconds)
        self.markers = [(None, None, mark_time), (None, None, mark_time)]
        self._update_om_marks_file(mark_time)
        self.lbl_status.setText(f"Mark at {_format_seconds(mark_time)}")
        self.log_message.emit(f"OM mark added at T={mark_time:.3f}s")

    def _clear_last_mark(self):
        self.timeline.clear_mark()

    def _set_mark_from_seconds(self, seconds: float):
        if self.current_tracking_path is None:
            QMessageBox.information(self, "Mark", "No case selected to save mark.")
            return

        mark_time = max(0.0, float(seconds or 0.0))
        self.timeline.set_mark(mark_time)
        self._on_mark_changed(mark_time)

    def _select_current_frame_as_mark(self):
        current_seconds = self.player.position() / 1000.0
        self._set_mark_from_seconds(current_seconds)

    def _select_start_as_mark(self):
        self._set_mark_from_seconds(0.0)

    def keyPressEvent(self, event):
        key = event.key()
        if key == Qt.Key_Space:
            from PySide6.QtMultimedia import QMediaPlayer as _QMP
            if self.player.playbackState() == _QMP.PlayingState:
                self._pause_video()
            else:
                self._play_video()
            event.accept()
        elif key in (Qt.Key_Return, Qt.Key_Enter):
            self._select_current_frame_as_mark()
            event.accept()
        else:
            super().keyPressEvent(event)

    def _play_video(self):
        if self.current_video_path and os.path.exists(self.current_video_path):
            self.player.play()

    def _pause_video(self):
        self.player.pause()

    def load_video_file(self, video_path: str, related_tracking: str = None):
        if not video_path or not os.path.exists(video_path):
            return
        self.current_video_path = video_path
        self.current_tracking_path = related_tracking
        self.player.setSource(QUrl.fromLocalFile(video_path))
        self.player.pause()
        self.lbl_status.setText(os.path.basename(video_path))
        self.log_message.emit(f"OM video loaded: {os.path.basename(video_path)}")
        if related_tracking:
            self.tracking_loaded.emit(related_tracking)
        self._sync_combos_to_video(video_path)
        self._restore_mark()

    def _sync_combos_to_video(self, video_path: str):
        """Update Subject and Case combos to reflect the given video path."""
        if not video_path or not self._all_case_files:
            return
        subj = self._extract_subject(video_path)
        # Update subject combo if needed
        subj_idx = self.combo_subject.findText(subj)
        if subj_idx < 0:
            return
        if self.combo_subject.currentIndex() != subj_idx:
            self.combo_subject.blockSignals(True)
            self.combo_subject.setCurrentIndex(subj_idx)
            self.combo_subject.blockSignals(False)
            self._populate_cases_for_subject(subj)
        # Update case combo
        case_idx = self.combo_cases.findData(video_path)
        if case_idx < 0:
            # fallback: match by basename
            vbase = os.path.basename(video_path).lower()
            for i in range(self.combo_cases.count()):
                d = self.combo_cases.itemData(i)
                if isinstance(d, str) and os.path.basename(d).lower() == vbase:
                    case_idx = i
                    break
        if case_idx >= 0 and self.combo_cases.currentIndex() != case_idx:
            self.combo_cases.blockSignals(True)
            self.combo_cases.setCurrentIndex(case_idx)
            self.combo_cases.blockSignals(False)

    def _navigate_prev(self):
        p = self.parent()
        while p is not None:
            if hasattr(p, 'previous_plot'):
                p.previous_plot()
                return
            p = p.parent()

    def _navigate_next(self):
        p = self.parent()
        while p is not None:
            if hasattr(p, 'next_plot'):
                p.next_plot()
                return
            p = p.parent()

    def dragEnterEvent(self, event):
        if event.mimeData().hasUrls():
            urls = event.mimeData().urls()
            if urls:
                fpath = urls[0].toLocalFile().lower()
                if fpath.endswith('.avi'):
                    event.acceptProposedAction()
                    return
        super().dragEnterEvent(event)

    def dropEvent(self, event):
        if event.mimeData().hasUrls():
            urls = event.mimeData().urls()
            if urls:
                fpath = urls[0].toLocalFile()
                if fpath.lower().endswith('.avi'):
                    self.load_video_file(fpath)
                    event.acceptProposedAction()
                    return
        super().dropEvent(event)

    def _get_om_marks_path(self):
        parent = self.parent()
        while parent is not None:
            if hasattr(parent, '_current_project_source') and parent._current_project_source:
                return os.path.join(parent._current_project_source, 'OM_marks.json')
            parent = parent.parent()
        return os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '..', 'OM_marks.json')

    def _get_logical_key(self):
        if not self.current_tracking_path:
            return None
        normalized = os.path.normpath(self.current_tracking_path).replace('\\', '/')
        parts = normalized.split('/')
        filename = os.path.basename(normalized)
        # Store marks with canonical mf4 names (without implicit _tracking suffix).
        filename = re.sub(r'_tracking(\.mf4)$', r'\1', filename, flags=re.IGNORECASE)
        if len(parts) >= 3:
            return '/'.join(parts[-3:-1] + [filename])
        return filename

    def _load_om_marks(self):
        marks_path = self._get_om_marks_path()
        if not os.path.exists(marks_path):
            return {}
        try:
            with open(marks_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return data if isinstance(data, dict) else {}
        except Exception:
            return {}

    def _save_om_marks(self, marks):
        marks_path = self._get_om_marks_path()
        try:
            os.makedirs(os.path.dirname(os.path.abspath(marks_path)), exist_ok=True)
            with open(marks_path, 'w', encoding='utf-8') as f:
                json.dump(marks, f, indent=2)
            self.marks_saved.emit()
        except Exception:
            pass

    def _update_om_marks_file(self, mark_time):
        key = self._get_logical_key()
        if not key:
            return
        marks = self._load_om_marks()
        if mark_time is None:
            if key in marks:
                del marks[key]
        else:
            marks[key] = {
                "mark_time_s": round(float(mark_time), 6),
                "video": os.path.basename(self.current_video_path) if self.current_video_path else None,
            }
        self._save_om_marks(marks)

    def _restore_mark(self):
        key = self._get_logical_key()
        if not key:
            self.timeline.set_mark(None)
            self.markers = []
            return

        marks = self._load_om_marks()
        entry = marks.get(key)
        if isinstance(entry, dict):
            mark_time = entry.get("mark_time_s")
        elif isinstance(entry, (float, int)):
            mark_time = float(entry)
        else:
            mark_time = None

        if mark_time is None:
            self.timeline.set_mark(None)
            self.markers = []
            return

        self.timeline.set_mark(float(mark_time))
        self.markers = [(None, None, float(mark_time)), (None, None, float(mark_time))]


class OMAnalysisWidget(AnalysisWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        setup_tab_icon_switching(self.tabs_right, [
            ("point_scan_16dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png", "point_scan_16dp_FFFFFF_FILL1_wght400_GRAD0_opsz20.png"),
            ("flowchart_16dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png", "flowchart_16dp_FFFFFF_FILL1_wght400_GRAD0_opsz20.png"),
            ("terminal_2_16dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png", "terminal_2_16dp_FFFFFF_FILL1_wght400_GRAD0_opsz20.png"),
        ])

    def _make_report_worker(self, config, output_path, dpi=300):
        """Use OMReportGeneratorWorker so all OM reports go through OMReportBuilder."""
        return OMReportGeneratorWorker(config, output_path, dpi)

    def create_tabs_right(self):
        self.tabs_right = QTabWidget()
        self.tabs_right.setIconSize(self.tabs_left.iconSize())

        icon_selector = QIcon(resource_path("assets/icons/point_scan_16dp_FFFFFF_FILL1_wght400_GRAD0_opsz20.png"))
        self.tab_time = OMTimeSelectorWidget()
        self.tabs_right.addTab(self.tab_time, icon_selector, "Time Selector")

        icon_logic = QIcon(resource_path("assets/icons/flowchart_16dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png"))
        self.tab_logic = OMLogicTabWidget()
        self.tabs_right.addTab(self.tab_logic, icon_logic, "Logic")

        self.tab_time.tracking_loaded.connect(self._on_tracking_loaded)
        self.tab_time.log_message.connect(self.log)

        self.tab_log = QWidget()
        self.l_log = QVBoxLayout(self.tab_log)
        self.l_log.setContentsMargins(0,0,0,0)
        self.txt_log = QTextEdit()
        self.txt_log.setReadOnly(True)
        self.txt_log.setStyleSheet("background-color: #1e1e1e; color: #ccc; font-family: Consolas; font-size: 10pt;")
        self.l_log.addWidget(self.txt_log)
        icon_log = QIcon(resource_path("assets/icons/terminal_2_16dp_FFFFFF_FILL1_wght400_GRAD0_opsz20.png"))
        self.tabs_right.addTab(self.tab_log, icon_log, "Log")

        self.right_layout.addWidget(self.tabs_right)

    def create_group_participants(self):
        super().create_group_participants()

        if hasattr(self, 'rb_inc_tracking'):
            self.rb_inc_tracking.hide()
            self.rb_inc_tracking.setEnabled(False)

        # Add CBR and OOP filter radio buttons
        if hasattr(self, 'grp_radios') and hasattr(self, 'row_radios_layout'):
            self.rb_cbr = QRadioButton("CBR")
            self.rb_oop = QRadioButton("OOP")
            self.grp_radios.addButton(self.rb_cbr)
            self.grp_radios.addButton(self.rb_oop)
            self.rb_cbr.clicked.connect(self.select_cbr_only)
            self.rb_oop.clicked.connect(self.select_oop_only)
            self.row_radios_layout.addWidget(self.rb_cbr)
            self.row_radios_layout.addWidget(self.rb_oop)

        if hasattr(self, 'tree_participants'):
            old_tree = self.tree_participants
            if hasattr(self.grp_participants, 'content_widget') and self.grp_participants.content_widget is not None:
                parent_layout = self.grp_participants.content_widget.layout()
            else:
                parent_layout = self.grp_participants.layout()
            index = parent_layout.indexOf(old_tree)

            self.tree_participants = OMParticipantTreeWidget()
            self.tree_participants.setHeaderLabels(["Structure", "Tracking", "Marks", "Report"])
            hdr = self.tree_participants.header()
            hdr.setSectionResizeMode(0, hdr.sectionResizeMode(0))
            hdr.setSectionResizeMode(1, hdr.sectionResizeMode(1))
            hdr.setSectionResizeMode(2, hdr.sectionResizeMode(2))
            hdr.setSectionResizeMode(3, hdr.sectionResizeMode(3))
            self.tree_participants.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
            self.tree_participants.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
            self.tree_participants.itemChanged.connect(self._on_participant_check_changed)
            self.tree_participants.itemExpanded.connect(self.on_tree_item_expanded)
            self.tree_participants.itemCollapsed.connect(self.on_tree_item_expanded)
            self.tree_participants.itemDoubleClicked.connect(self._on_om_tree_item_clicked)

            if index >= 0:
                parent_layout.insertWidget(index, self.tree_participants)
                parent_layout.removeWidget(old_tree)
                old_tree.deleteLater()
            else:
                # Fallback: keep UI functional if layout index cannot be resolved
                parent_layout.addWidget(self.tree_participants)
                old_tree.hide()

            self.tree_participants.setColumnHidden(1, True)

    def select_cbr_only(self):
        """Check only items under Correct Belt Routing folders."""
        self._select_by_om_category(("correct belt routing",))

    def select_oop_only(self):
        """Check only items under Out of Position folders."""
        self._select_by_om_category(("out of position",))

    def _select_by_om_category(self, keywords: tuple):
        """Check tree items that are under a category folder whose normalized
        name contains any of the given keywords; uncheck the rest."""
        def recursive(item, inside_category):
            if not inside_category:
                inside_category = any(
                    kw in _normalize_folder_name(item.text(0))
                    for kw in keywords
                )
            item.setCheckState(0, Qt.Checked if inside_category else Qt.Unchecked)
            for i in range(item.childCount()):
                recursive(item.child(i), inside_category)

        for i in range(self.tree_participants.topLevelItemCount()):
            recursive(self.tree_participants.topLevelItem(i), False)

    def _on_om_tree_item_clicked(self, item, column):
        try:
            if not item or item.childCount() != 0:
                return
            fpath = item.data(0, Qt.UserRole)
            if not fpath or not isinstance(fpath, str):
                return

            low = fpath.lower()
            if low.endswith('.avi') and hasattr(self, 'tab_time'):
                self.tabs_right.setCurrentWidget(self.tab_time)
                related_tracking = self.tab_time._guess_tracking_from_video(fpath)
                self.tab_time.load_video_file(fpath, related_tracking)
            elif low.endswith('.mf4'):
                # In the new architecture, logic/signals are handled within tab_logic (OMLogicTabWidget)
                # and do not use a separate Signal Filter tab.
                pass
        except Exception:
            pass

    def _get_checked_videos(self) -> list:
        """Return all checked leaf video files from the participant tree."""
        video_exts = ('.avi', '.mp4', '.mov', '.mkv')
        video_files = []
        try:
            it = QTreeWidgetItemIterator(self.tree_participants)
            while it.value():
                item = it.value()
                if item.checkState(0) == Qt.Checked and item.childCount() == 0:
                    fpath = item.data(0, Qt.UserRole)
                    if isinstance(fpath, str) and fpath.lower().endswith(video_exts):
                        video_files.append(fpath)
                it += 1
        except Exception:
            pass
        return video_files

    def auto_load_first_tracking(self):
        try:
            video_files = self._get_checked_videos()
            if video_files and hasattr(self, 'tab_time'):
                source_root = getattr(self, '_current_project_source', None)
                self.tab_time.set_case_list(video_files, source_root=source_root)
                first_video = video_files[0]
                related_tracking = self.tab_time._guess_tracking_from_video(first_video)
                self.tab_time.load_video_file(first_video, related_tracking)
                return
            super().auto_load_first_tracking()
        except Exception:
            super().auto_load_first_tracking()

    def next_plot(self):
        """Navigate to the next checked video in the participant tree."""
        if not hasattr(self, 'tab_time'):
            super().next_plot()
            return
        video_files = self._get_checked_videos()
        if not video_files:
            return
        current = getattr(self.tab_time, 'current_video_path', None)
        try:
            idx = video_files.index(current)
        except ValueError:
            # current not in list – find nearest by basename
            current_base = os.path.basename(current or '').lower()
            idx = next(
                (i for i, f in enumerate(video_files) if os.path.basename(f).lower() == current_base),
                -1,
            )
        if idx < len(video_files) - 1:
            next_video = video_files[idx + 1]
            self.tabs_right.setCurrentWidget(self.tab_time)
            related = self.tab_time._guess_tracking_from_video(next_video)
            self.tab_time.load_video_file(next_video, related)

    def previous_plot(self):
        """Navigate to the previous checked video in the participant tree."""
        if not hasattr(self, 'tab_time'):
            super().previous_plot()
            return
        video_files = self._get_checked_videos()
        if not video_files:
            return
        current = getattr(self.tab_time, 'current_video_path', None)
        try:
            idx = video_files.index(current)
        except ValueError:
            current_base = os.path.basename(current or '').lower()
            idx = next(
                (i for i, f in enumerate(video_files) if os.path.basename(f).lower() == current_base),
                len(video_files),
            )
        if idx > 0:
            prev_video = video_files[idx - 1]
            self.tabs_right.setCurrentWidget(self.tab_time)
            related = self.tab_time._guess_tracking_from_video(prev_video)
            self.tab_time.load_video_file(prev_video, related)

    def log(self, msg):
        try:
            if hasattr(self, 'tab_time') and hasattr(self.tab_time, 'lbl_status'):
                self.tab_time.lbl_status.setText(str(msg))
        except Exception:
            pass

    def on_source_changed(self, text):
        if not text:
            return
        om_marks_path = os.path.join(text, 'OM_marks.json')
        if not os.path.exists(om_marks_path):
            try:
                with open(om_marks_path, 'w', encoding='utf-8') as f:
                    json.dump({}, f, indent=2)
            except Exception:
                pass
        super().on_source_changed(text)

    def on_micro_include_toggled(self, checked):
        if checked:
            accepted_micro_channels = [
                "SoundPressure",
                "MySound PressureTask.Sound Pressure",
            ]

            found_channel = None
            available = getattr(self, 'available_signals', [])
            for channel_name in accepted_micro_channels:
                if available and channel_name in available:
                    found_channel = channel_name
                    break

            if not found_channel:
                if hasattr(self, 'toggle_micro_include'):
                    self.toggle_micro_include.blockSignals(True)
                    self.toggle_micro_include.setChecked(False)
                    self.toggle_micro_include.blockSignals(False)
                QMessageBox.warning(
                    self,
                    "Signal Missing",
                    "Include Micro requires one of these signals:\n"
                    "- SoundPressure\n"
                    "- MySound PressureTask.Sound Pressure"
                )
                return

        self.log(f"Micro Include Toggled: {checked}.")
        if hasattr(self, 'enable_micro_controls'):
            self.enable_micro_controls(checked)

    def scan_participants(self, folder):
        self.tree_participants.clear()
        self.busy_changed.emit(True)
        self.tree_participants.setEnabled(False)

        if hasattr(self, 'scanner_thread') and self.scanner_thread is not None:
            self.scanner_thread.quit()
            self.scanner_thread.wait()

        marks_path = os.path.join(folder, 'OM_marks.json')
        self.scanner_thread = OMAnalysisScanner(folder, marks_path=marks_path)
        self.scanner_thread.finished.connect(self.on_scan_finished)
        self.scanner_thread.log.connect(self.log)
        self.scanner_thread.start()

    def _get_marks_path(self):
        source = getattr(self, '_current_project_source', None)
        if source and os.path.isdir(source):
            return os.path.join(source, 'OM_marks.json')
        return os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '..', 'OM_marks.json')

    def _resolve_om_mark_for_mf4(self, mf4_path: str):
        info = {
            'mark_time': None,
            'matched_key': None,
            'matched_video': None,
            'score': -1,
            'reason': 'no-match',
        }

        if not mf4_path:
            info['reason'] = 'empty-mf4-path'
            return info

        marks_map = self._load_marks_map()
        if not isinstance(marks_map, dict) or not marks_map:
            info['reason'] = 'empty-om-marks'
            return info

        target_path = _normalize_path_for_match(mf4_path)
        target_name = os.path.basename(target_path)
        target_dir = os.path.dirname(target_path)

        best_score = -1
        best_mark = None
        best_key = None
        best_entry = None

        for raw_key, entry in marks_map.items():
            if not isinstance(raw_key, str):
                continue
            key_path = _normalize_path_for_match(raw_key)
            key_name = os.path.basename(key_path)

            # Keys in OM_marks.json may use _tracking.mf4 suffix; strip it for comparison.
            key_name_canonical = re.sub(r'_tracking(\.mf4)$', r'\1', key_name)

            if key_name != target_name and key_name_canonical != target_name:
                continue

            key_dir = os.path.dirname(key_path)
            if not key_dir:
                continue

            score = -1
            if key_path == target_path:
                score = 10_000
            elif target_dir.endswith(key_dir):
                score = len(key_dir)
            elif key_dir.endswith(target_dir):
                score = len(target_dir)
            else:
                continue

            mark_time = _extract_mark_time_from_entry(entry)
            if mark_time is None:
                continue

            if score > best_score:
                best_score = score
                best_mark = mark_time
                best_key = raw_key
                best_entry = entry

        if best_mark is not None:
            info['mark_time'] = best_mark
            info['matched_key'] = best_key
            if isinstance(best_entry, dict):
                vid = best_entry.get('video')
                if isinstance(vid, str) and vid.strip():
                    info['matched_video'] = vid.strip()
            info['score'] = best_score
            info['reason'] = 'ok'
        else:
            info['reason'] = f'no-key-for-name-route:{target_name}'

        return info

    def _find_om_mark_time_for_mf4(self, mf4_path: str):
        return self._resolve_om_mark_for_mf4(mf4_path).get('mark_time')

    def _get_marks_for_mf4(self, mf4_path: str):
        mark_info = self._resolve_om_mark_for_mf4(mf4_path)
        mark_time = mark_info.get('mark_time')
        if isinstance(mark_time, (int, float)):
            self.log(
                f"[OM MARKS][BATCH] match for {os.path.basename(mf4_path)} "
                f"-> {mark_info.get('matched_key')} | mark_time_s={mark_time:.6f}"
            )
            return [float(mark_time)]
        self.log(
            f"[OM MARKS][BATCH] no match for {os.path.basename(mf4_path)} "
            f"({mark_info.get('reason')})"
        )
        return None

    def _build_report_config_for_mf4(self, mf4_path: str, driver_marks: list):
        config = super()._build_report_config_for_mf4(mf4_path, driver_marks)

        def _get_cbr_operator_threshold(sig_name: str):
            op, val = '>', None
            if not hasattr(self, 'tab_logic'):
                return op, val
            cbr_table = self.tab_logic.category_tables.get('Correct Belt Routing') if hasattr(self.tab_logic, 'category_tables') else None
            if cbr_table is None:
                return op, val
            for r in range(cbr_table.rowCount()):
                ni = cbr_table.item(r, 1)
                if not (ni and ni.text() == sig_name):
                    continue
                ow = cbr_table.cellWidget(r, 2)
                vw = cbr_table.cellWidget(r, 3)
                if isinstance(ow, QComboBox):
                    op = ow.currentText()
                if isinstance(vw, QDoubleSpinBox):
                    val = vw.value()
                elif isinstance(vw, QLineEdit):
                    try:
                        val = float(vw.text())
                    except Exception:
                        val = vw.text()
                elif isinstance(vw, QComboBox):
                    val = vw.currentText()
                break
            return op, val

        def _rebuild_batch_signals_from_logic_table(target_category: str):
            if not hasattr(self, 'tab_logic'):
                return None
            category_tables = getattr(self.tab_logic, 'category_tables', None)
            if not isinstance(category_tables, dict) or target_category not in category_tables:
                return None

            table = category_tables[target_category]
            rebuilt_signals = {}
            try:
                from asammdf import MDF
                with MDF(mf4_path) as mdf:
                    for row in range(table.rowCount()):
                        chk = table.item(row, 0)
                        name_item = table.item(row, 1)
                        if not (chk and name_item and chk.checkState() == Qt.Checked):
                            continue

                        sig_name = name_item.text()
                        try:
                            sig = mdf.get(sig_name)

                            op_widget = table.cellWidget(row, 2)
                            operator = op_widget.currentText() if isinstance(op_widget, QComboBox) else "None"

                            val_widget = table.cellWidget(row, 3)
                            value = None
                            if isinstance(val_widget, QDoubleSpinBox):
                                value = val_widget.value()
                            elif isinstance(val_widget, QComboBox):
                                value = val_widget.currentText()
                            elif isinstance(val_widget, QLineEdit):
                                try:
                                    value = float(val_widget.text()) if val_widget.text() else None
                                except Exception:
                                    value = val_widget.text()

                            units_item = table.item(row, 4)
                            units = units_item.text() if units_item else ""

                            alias_item = table.item(row, 5)
                            alias = alias_item.text() if alias_item else sig_name

                            rebuilt_signals[sig_name] = {
                                'timestamps': list(sig.timestamps),
                                'samples': list(sig.samples),
                                'threshold': value,
                                'operator': operator,
                                'unit': units or getattr(sig, 'unit', 'Value'),
                                'category': target_category,
                                'alias': alias,
                            }
                        except Exception as e:
                            print(f"[OM DEBUG][BATCH] Error getting signal {sig_name}: {e}")
            except Exception as e:
                print(f"[OM DEBUG][BATCH] failed rebuilding signals from logic table: {e}")
                return None

            return rebuilt_signals or None

        inferred = _infer_om_category_from_path(mf4_path)
        if inferred:
            config['target_category'] = inferred

            rebuilt_signals = _rebuild_batch_signals_from_logic_table(inferred)
            if rebuilt_signals:
                config['signals'] = rebuilt_signals

            # Super may not infer OM categories from filename; enforce category filtering here.
            signals = config.get('signals')
            if isinstance(signals, dict) and signals:
                filtered_signals = {
                    k: v for k, v in signals.items()
                    if isinstance(v, dict) and str(v.get('category', '') or '') == inferred
                }
                if filtered_signals:
                    config['signals'] = filtered_signals
                    signal_times = config.get('signal_times')
                    if isinstance(signal_times, dict):
                        config['signal_times'] = {k: signal_times.get(k) for k in filtered_signals.keys()}

                    pass_sig = config.get('pass_signal_name')
                    if pass_sig not in filtered_signals:
                        pass_sig = None
                        if hasattr(self, 'tab_logic') and hasattr(self.tab_logic, 'pass_criteria_tables'):
                            pass_table = self.tab_logic.pass_criteria_tables.get(inferred)
                            if pass_table and pass_table.rowCount() > 0:
                                if hasattr(pass_table, 'signal_combo'):
                                    candidate = pass_table.signal_combo.currentText()
                                else:
                                    signal_combo = pass_table.cellWidget(0, 0)
                                    candidate = signal_combo.currentText() if isinstance(signal_combo, QComboBox) else ""
                                if candidate and candidate != "-- Select Signal --":
                                    pass_sig = candidate
                        if pass_sig:
                            config['pass_signal_name'] = pass_sig

        movement_start = None
        if driver_marks and len(driver_marks) > 0 and isinstance(driver_marks[0], (int, float)):
            movement_start = float(driver_marks[0])

        # Keep batch behavior aligned with preview: when mark is missing, use 0.0.
        if movement_start is None:
            movement_start = 0.0
            self.log(f"[OM MARKS][BATCH] fallback to movement_start=0.0 for {os.path.basename(mf4_path)}")

        config['tgaze'] = movement_start
        config['movement_start'] = movement_start
        config['movement_start_label'] = 'Movement Start'

        is_cbr = _is_om_correct_belt_routing(config)
        is_oop = _is_om_out_of_position(config)
        if is_cbr or is_oop:
            config['om_report_variant'] = 'correct_belt_routing' if is_cbr else 'out_of_position'
            config['om_use_video_frame'] = True
            config['om_hide_date_time_vin'] = True
            config['om_plot_show_marks'] = True
            config['om_plot_show_shading'] = True

            camera_mode = ".avi"
            if hasattr(self, 'tab_time'):
                combo = getattr(self.tab_time, 'combo_camera_index', None)
                if combo is not None:
                    camera_mode = combo.currentData() or ".avi"

            mark_info = self._resolve_om_mark_for_mf4(mf4_path)
            matched_video_name = mark_info.get('matched_video') if isinstance(mark_info, dict) else None

            video_path = _guess_video_for_mf4_path(
                mf4_path,
                camera_mode=camera_mode,
                preferred_video_name=matched_video_name,
            )
            if video_path and os.path.exists(video_path):
                config['om_video_path'] = video_path
                secondary_video = _guess_secondary_video_for_mf4_path(mf4_path, video_path)
                if secondary_video and os.path.exists(secondary_video):
                    config['om_video_path_secondary'] = secondary_video

            pass_sig = config.get('pass_signal_name')
            warn_start = None
            if pass_sig:
                warn_start = (config.get('signal_times') or {}).get(pass_sig)
            capture_t = warn_start if isinstance(warn_start, (int, float)) else movement_start
            config['om_video_time_s'] = float(capture_t) if isinstance(capture_t, (int, float)) else 0.0

        # Batch PNG path: mirror preview behavior for CBR detection signal selection.
        if is_cbr and hasattr(self, 'tab_logic') and hasattr(self.tab_logic, 'cbr_detection_table'):
            det_table = self.tab_logic.cbr_detection_table
            for row in range(det_table.rowCount()):
                chk = det_table.item(row, 0)
                if chk and chk.checkState() == Qt.Checked:
                    det_sig_name = chk.data(Qt.UserRole)
                    if not det_sig_name:
                        break
                    config['cbr_detection_signal'] = det_sig_name
                    print(f"[OM DEBUG][BATCH] selected CBR detection signal: {det_sig_name!r}")

                    signals = config.get('signals')
                    if isinstance(signals, dict) and det_sig_name not in signals:
                        try:
                            from asammdf import MDF
                            with MDF(mf4_path) as mdf:
                                sig = mdf.get(det_sig_name)
                            op, val = _get_cbr_operator_threshold(det_sig_name)
                            signals[det_sig_name] = {
                                'timestamps': list(sig.timestamps),
                                'samples': list(sig.samples),
                                'threshold': val,
                                'operator': op,
                                'unit': getattr(sig, 'unit', 'Value'),
                                'category': 'Correct Belt Routing',
                                'alias': det_sig_name,
                            }
                            print(
                                f"[OM DEBUG][BATCH] injected detection signal data: "
                                f"name={det_sig_name!r} operator={op!r} threshold={val!r} samples={len(sig.samples)}"
                            )
                        except Exception as e:
                            print(f"[OM DEBUG][BATCH] failed to inject detection signal data for {det_sig_name!r}: {e}")
                        else:
                            # Signal already present but operator/threshold may be wrong — always override from UI table
                            signals[det_sig_name]['operator'] = op
                            signals[det_sig_name]['threshold'] = val
                            print(
                                f"[OM DEBUG][BATCH] updated operator/threshold on existing detection signal: "
                                f"name={det_sig_name!r} operator={op!r} threshold={val!r}"
                            )
                    break

        print(
            f"[OM DEBUG][BATCH] final config before normalize | "
            f"file={os.path.basename(mf4_path)!r} "
            f"variant={config.get('om_report_variant')!r} "
            f"movement_start={config.get('movement_start')!r} "
            f"cbr_detection_signal={config.get('cbr_detection_signal')!r}"
        )

        return _normalize_om_audio_signal_config(config)

    def _update_excel_results(self, config: dict, mf4_path: str):
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment

        try:
            participant_dir = os.path.dirname(mf4_path)
            excel_path = os.path.join(participant_dir, "Analysis_Results.xlsx")

            if os.path.exists(excel_path):
                wb = openpyxl.load_workbook(excel_path)
                ws = wb.active
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Results"
                headers = ["Folder Name", "File Name", "Movement Start", "Warning Start", "Warning Timer", "Score"]
                ws.append(headers)

                header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                header_font = Font(bold=True)
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")

            if ws.max_row >= 1:
                if str(ws.cell(row=1, column=3).value).strip().lower() in ["distraction start", "t_gaze", "tgaze"]:
                    ws.cell(row=1, column=3).value = "Movement Start"

            folder_name = config.get('target_category', '--')
            file_name = config.get('filename', os.path.basename(mf4_path))

            movement_start = config.get('movement_start', config.get('tgaze'))
            pass_sig = config.get('pass_signal_name')
            warn_start = config.get('signal_times', {}).get(pass_sig) if pass_sig else None
            warn_timer = config.get('t_event')
            score = "PASS" if config.get('t_event_color') == "green" else "FAIL"

            row_data = [
                folder_name,
                file_name,
                movement_start if movement_start is not None else "",
                warn_start if warn_start is not None else "nan",
                warn_timer if isinstance(warn_timer, (int, float)) else "",
                score,
            ]

            found_row = -1
            for r in range(2, ws.max_row + 1):
                if ws.cell(row=r, column=2).value == file_name:
                    found_row = r
                    break

            if found_row != -1:
                for idx, val in enumerate(row_data, 1):
                    ws.cell(row=found_row, column=idx).value = val
            else:
                ws.append(row_data)

            target_row = found_row if found_row != -1 else ws.max_row
            for cell in ws[target_row]:
                cell.alignment = Alignment(horizontal="center")

            for i, col in enumerate(ws.columns, 1):
                max_len = 0
                col_letter = openpyxl.utils.get_column_letter(i)
                for cell in col:
                    try:
                        max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = min(max_len + 2, 35)

            wb.save(excel_path)
        except Exception as e:
            self.log(f"OM Excel update error: {e}")
