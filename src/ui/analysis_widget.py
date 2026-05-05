from __future__ import annotations
import os
import re
import traceback
from datetime import datetime
from PySide6.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QSplitter, 
                               QTabWidget, QTreeWidget, QTreeWidgetItem, QLabel,
                               QSizePolicy, QAbstractItemView,
                               QLineEdit, QPushButton, QSlider, QSpinBox, QDoubleSpinBox,
                               QComboBox, QFileDialog, QRadioButton, QButtonGroup, 
                               QHeaderView, QTextEdit, QMessageBox, QProgressBar,
                               QTreeWidgetItemIterator, QStackedWidget, QScrollArea,
                               QFrame, QGridLayout, QGroupBox, QFormLayout, 
                               QTableWidget, QTableWidgetItem, QCheckBox, QDialog,
                               QGraphicsOpacityEffect, QGraphicsDropShadowEffect)
from PySide6.QtWidgets import QMenu
from PySide6.QtCore import (Qt, QSize, QThread, Signal, QPoint, QRect, QPropertyAnimation, QVariantAnimation,
                             QEasingCurve, QParallelAnimationGroup, QSequentialAnimationGroup, QTimer)
from PySide6.QtGui import QIcon, QColor, QFont, QImage, QPixmap, QDragEnterEvent, QDropEvent, QPainter, QAction, QPen, QTransform, QCursor
from PySide6.QtMultimedia import QMediaPlayer, QVideoSink, QVideoFrame
from src.core.ai_analyzer import AIWorker
from src.ui.ai_brain_widget import AIBrainWidget
from PySide6.QtMultimediaWidgets import QVideoWidget
# Scientific and processing libraries are imported locally where needed
# to ensure application startup robustness.
try:
    import numpy as np
except ImportError:
    np = None

try:
    from asammdf import MDF
except Exception:
    MDF = None

try:
    import matplotlib
    matplotlib.use('QtAgg')
    from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
    from matplotlib.figure import Figure
except ImportError:
    matplotlib = None
    FigureCanvas = None
    Figure = None

try:
    import pyqtgraph as pg
except ImportError:
    pg = None

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
except ImportError:
    openpyxl = None
    PatternFill = Font = Alignment = Border = Side = None

try:
    from scipy.signal import butter, filtfilt, hilbert
except ImportError:
    butter = filtfilt = hilbert = None

from src.core.utils import resource_path
from src.ui.widgets import IconGroupBox, AnimatedToggle, setup_tab_icon_switching
from src.ui.styles import IDIADA_ORANGE
from src.core.audio_analysis import obtain_peak_frequency, find_first_valid_event
from src.core.chronos_worker import ChronosWorker
from src.core.chronos_manager import ChronosManager





class FadeStackedWidget(QStackedWidget):
    """A version of QStackedWidget that cross-fades between widgets."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self._fade_duration = 300
        self._current_idx = -1

    def setCurrentIndex(self, index):
        if index == self._current_idx:
            return
        
        old_widget = self.currentWidget()
        new_widget = self.widget(index)
        self._current_idx = index
        
        if not old_widget:
            super().setCurrentIndex(index)
            return

        # Prepare new widget
        new_widget.setGraphicsEffect(None)
        opacity_effect = QGraphicsOpacityEffect()
        new_widget.setGraphicsEffect(opacity_effect)
        opacity_effect.setOpacity(0)
        
        # Super call to switch index immediately but hidden
        super().setCurrentIndex(index)
        
        # Animate fade in
        self.anim = QPropertyAnimation(opacity_effect, b"opacity")
        self.anim.setDuration(self._fade_duration)
        self.anim.setStartValue(0)
        self.anim.setEndValue(1)
        self.anim.setEasingCurve(QEasingCurve.InOutQuad)
        self.anim.start()


class AutodetectWorker(QThread):
    finished = Signal(float)
    error = Signal(str)
    log = Signal(str)

    def __init__(self, file_path):
        super().__init__()
        self.file_path = file_path

    def run(self):
        try:
            self.log.emit(f"Starting audio autodetection on: {os.path.basename(self.file_path)}")
            res, err = obtain_peak_frequency(self.file_path)
            if err:
                self.error.emit(err)
            else:
                self.finished.emit(res)
        except Exception as e:
            self.error.emit(str(e))

LOGIC_INPUT_STYLE = """
    QSpinBox, QDoubleSpinBox {
        background-color: #3a3a3a;
        color: #ddd;
        border: 1px solid #555;
        border-radius: 3px;
        padding-right: 4px;
        font-size: 9pt;
    }
    QSpinBox::up-button, QDoubleSpinBox::up-button {
        subcontrol-origin: border;
        subcontrol-position: top right;
        width: 16px;
        background-color: #444;
        border-left: 1px solid #555;
        border-bottom: 0.5px solid #555;
    }
    QSpinBox::down-button, QDoubleSpinBox::down-button {
        subcontrol-origin: border;
        subcontrol-position: bottom right;
        width: 16px;
        background-color: #444;
        border-left: 1px solid #555;
        border-top: 0.5px solid #555;
    }
    QSpinBox::up-arrow, QDoubleSpinBox::up-arrow {
        image: none;
        width: 0; height: 0;
        border-left: 4px solid transparent;
        border-right: 4px solid transparent;
        border-bottom: 5px solid #aaa;
    }
    QSpinBox::down-arrow, QDoubleSpinBox::down-arrow {
        image: none;
        width: 0; height: 0;
        border-left: 4px solid transparent;
        border-right: 4px solid transparent;
        border-top: 5px solid #aaa;
    }
    QSpinBox::up-button:pressed, QDoubleSpinBox::up-button:pressed { background-color: #555; }
    QSpinBox::down-button:pressed, QDoubleSpinBox::down-button:pressed { background-color: #555; }
"""

class MdfLoaderWorker(QThread):
    """Generic worker to load an MDF file off the main thread.
    
    Pre-extracts all signal data so the main thread never calls mdf.get().
    Emits: loaded(mdf_object, signal_names_list, signal_data_dict)
      where signal_data_dict = {name: (timestamps_ndarray, samples_ndarray)}
    """
    loaded = Signal(object, list, object)   # (mdf, signal_names, signal_data)
    error = Signal(str)

    def __init__(self, file_path, numeric_only=False):
        super().__init__()
        self.file_path = file_path
        self.numeric_only = numeric_only

    def run(self):
        try:
            mdf = MDF(self.file_path)
            signals = []
            signal_data = {}
            
            # Iterate once and extract data directly from the yielded Signal object
            # This avoids calling mdf.get(name) later which fails for duplicate names
            for ch in mdf.iter_channels():
                name = ch.name
                is_valid = False
                
                if self.numeric_only:
                    try:
                        if ch.samples is None:
                            continue
                        kind = getattr(ch.samples, 'dtype', None)
                        if kind is not None and getattr(ch.samples.dtype, 'kind', 'f') in 'fiudb':
                            is_valid = True
                    except Exception:
                        try:
                            float(ch.samples[0])
                            is_valid = True
                        except Exception:
                            pass
                else:
                    if name.lower() not in ['t', 'time', 't1', 'timestamps']:
                        is_valid = True
                
                if is_valid:
                    # If name already exists, we might overwrite it. 
                    # Complex handling of duplicates (e.g. name(1)) could be added,
                    # but for now we just accept the Last one or avoid errors.
                    # Appending to list is fine.
                    if name not in signal_data:
                        signals.append(name)
                    
                    # DIRECT EXTRACTION: ch is already the Signal object
                    try:
                        signal_data[name] = (
                            np.array(ch.timestamps),
                            np.array(ch.samples, dtype=float)
                        )
                    except Exception:
                        pass

            signals.sort()

            
            self.loaded.emit(mdf, signals, signal_data)
        except Exception as e:
            self.error.emit(str(e))

class ReportGeneratorWorker(QThread):
    """Worker to generate a report PNG off the main thread."""
    finished = Signal(str)    # output_path
    error = Signal(str)

    def __init__(self, config, output_path, dpi=300):
        super().__init__()
        self.config = config
        self.output_path = output_path
        self.dpi = dpi

    def run(self):
        try:
            from src.core.report_builder import MatplotlibReportBuilder
            builder = MatplotlibReportBuilder(self.config)
            builder.generate(self.output_path, dpi=self.dpi)
            self.finished.emit(self.output_path)
        except Exception as e:
            import traceback
            traceback.print_exc()
            self.error.emit(str(e))

class AnalysisScanner(QThread):
    finished = Signal(list)
    log = Signal(str)
    
    def __init__(self, source_dir, marks_path=None):
        super().__init__()
        self.source_dir = source_dir
        self.marks_path = marks_path
        
    def run(self):
        if not self.source_dir or not os.path.exists(self.source_dir):
            self.log.emit(f"Invalid source directory: {self.source_dir}")
            self.finished.emit([])
            return

        self.log.emit(f"Scanning source directory: {self.source_dir}")
        
        # Load marks.json once
        marks_keys = set()
        try:
            marks_path = self.marks_path or os.path.join(self.source_dir, 'marks.json')
            if os.path.exists(marks_path):
                import json
                with open(marks_path, 'r', encoding='utf-8') as f:
                    marks_data = json.load(f)
                    marks_keys = set(marks_data.keys())
        except Exception:
            pass
            
        results = []
        try:
            # Look for _FUSION_RESULTS
            fusion_results_dir = os.path.join(self.source_dir, "_FUSION_RESULTS")
            scan_root = fusion_results_dir if os.path.exists(fusion_results_dir) else self.source_dir
            
            parts = [d for d in os.listdir(scan_root) 
                     if os.path.isdir(os.path.join(scan_root, d)) 
                     and re.match(r"^[A-Z][0-9]{2}$", d)]
            parts.sort()
            
            for p in parts:
                p_path = os.path.join(scan_root, p)
                data = self._scan_recursive(p_path, marks_keys)
                total = data["total_mf4"]
                tracking_done = data["total_tracking"]
                analysis_done = data["total_analysis"]
                
                # Color based on tracking status
                color = "#d1242f"
                if total > 0:
                    if tracking_done == total: color = "#2da44e"
                    elif tracking_done > 0: color = IDIADA_ORANGE
                elif total == 0: color = "gray"

                results.append({
                    "name": p,
                    "type": "participant",
                    "path": p_path,
                    "children": data["children"],
                    "tracking_stats": (tracking_done, total),
                    "marks_stats": (data.get("total_marks", 0), total),
                    "analysis_stats": (analysis_done, total),
                    "color": color
                })
                
        except Exception as e:
            pass
            
        self.finished.emit(results)
        self.log.emit("Scan finished.")

    def _scan_recursive(self, path, marks_keys=None):
        children = []
        total_mf4 = 0
        total_tracking = 0
        total_marks = 0
        total_analysis = 0
        
        if marks_keys is None:
            marks_keys = set()
        
        try:
            entries = os.listdir(path)
            import re
            entries.sort(key=lambda s: [int(t) if t.isdigit() else t.lower() for t in re.split('([0-9]+)', s)])
            
            dirs = []
            files = []
            
            for e in entries:
                full = os.path.join(path, e)
                if os.path.isdir(full): dirs.append(e)
                else: files.append(e)
            
            for d in dirs:
                sub_res = self._scan_recursive(os.path.join(path, d), marks_keys)
                if sub_res["total_mf4"] > 0:
                    children.append({
                        "name": d,
                        "type": "folder",
                        "children": sub_res["children"],
                        "path": os.path.join(path, d),
                        "tracking_stats": (sub_res["total_tracking"], sub_res["total_mf4"]),
                        "marks_stats": (sub_res.get("total_marks", 0), sub_res["total_mf4"]),
                        "analysis_stats": (sub_res["total_analysis"], sub_res["total_mf4"])
                    })
                    total_mf4 += sub_res["total_mf4"]
                    total_tracking += sub_res["total_tracking"]
                    total_marks += sub_res.get("total_marks", 0)
                    total_analysis += sub_res["total_analysis"]
            
            # Read MF4 files and resolve tracking vs normal 
            mf4_files_raw = [f for f in files 
                             if f.lower().endswith('.mf4') 
                             and not f.startswith('._')]
            
            base_map = {}
            for f in mf4_files_raw:
                is_tracking = f.lower().endswith('_tracking.mf4')
                base = f[:-13] if is_tracking else os.path.splitext(f)[0]
                
                if base not in base_map:
                    base_map[base] = {'has_tracking': False, 'file': None, 'path': None, 'tracking_path': None}
                
                if is_tracking:
                    base_map[base]['has_tracking'] = True
                    base_map[base]['tracking_path'] = os.path.join(path, f)
                    # Only set path if not yet set by a non-tracking file
                    if base_map[base]['file'] is None:
                        base_map[base]['file'] = f
                        base_map[base]['path'] = os.path.join(path, f)
                else:
                    # Always prefer non-tracking file as the main reference
                    base_map[base]['file'] = f
                    base_map[base]['path'] = os.path.join(path, f)
                        
            mf4_bases = list(base_map.keys())
            
            # Sort naturally by filename so D1, D2, D10... and D1_1, D1_2, D1_10... are sorted logically
            def natural_sort_key(basename):
                import re
                return [int(text) if text.isdigit() else text.lower() for text in re.split('([0-9]+)', basename)]
            
            mf4_bases.sort(key=natural_sort_key)
            
            for base in mf4_bases:
                info = base_map[base]
                f = info['file']
                fpath = info['path']
                has_tracking = info['has_tracking']
                
                # Check for analysis PNG in Reports subfolder
                reports_dir = os.path.join(path, "Reports")
                file_base_name = os.path.splitext(f)[0]
                analysis_png_path = os.path.join(reports_dir, f"{file_base_name}.png")
                has_analysis = os.path.exists(analysis_png_path)
                
                # Fallback to the group base name if they differ
                if not has_analysis and file_base_name != base:
                    analysis_png_path_alt = os.path.join(reports_dir, f"{base}.png")
                    has_analysis = os.path.exists(analysis_png_path_alt)
                
                total_mf4 += 1
                if has_tracking: total_tracking += 1
                if has_analysis: total_analysis += 1
                
                children.append({
                    "name": f,
                    "type": "file",
                    "path": fpath,
                    "has_tracking": has_tracking,
                    "has_analysis": has_analysis,
                    "has_marks": False,
                    "has_report": has_analysis # Alias for now
                })
                
                # Check for marks
                # Key logic: if inside a participant folder, key is "Participant/Filename.mf4" or similar
                # We need the full path
                try:
                    paths_to_check = [fpath]
                    if info.get('tracking_path'):
                        paths_to_check.append(info['tracking_path'])
                        
                    for p_path in paths_to_check:
                        p = os.path.normpath(p_path)
                        # Normalize slashes specifically for matching marks.json format regardless of platform
                        p_unix = p.replace('\\', '/')
                        parts = p_unix.split('/')
                        if len(parts) >= 3:
                            key = '/'.join(parts[-3:])
                            if key in marks_keys:
                                 children[-1]["has_marks"] = True
                        
                        # Also check basename just in case
                        if children[-1]["has_marks"] == False:
                            if os.path.basename(p_path) in marks_keys:
                                 children[-1]["has_marks"] = True
                                 
                        if children[-1]["has_marks"]:
                            break
                    
                    if children[-1]["has_marks"]:
                        total_marks += 1
                except: pass
                
                # Separate Report check (already done as has_analysis, but explicit for clarity)
                children[-1]["has_report"] = has_analysis
        except: pass
        return {"children": children, "total_mf4": total_mf4, "total_tracking": total_tracking, "total_marks": total_marks, "total_analysis": total_analysis}


class GaugeRulesEditorDialog(QDialog):
    def __init__(self, known_paths, active_path, parent=None):
        super().__init__(parent)
        self.setWindowTitle("Gauge Rules Configuration")
        self.resize(600, 600)
        self.setStyleSheet("""
            QDialog { background-color: #2a2a2a; color: #ddd; }
            QPushButton { background-color: #444; color: #fff; border: 1px solid #666; padding: 4px; border-radius: 3px; }
            QPushButton:hover { background-color: #555; }
            QTableWidget { background-color: #1e1e1e; color: #ddd; border: 1px solid #555; gridline-color: #444; }
            QTableWidget::item:selected { background-color: #ff9800; color: black; }
            QHeaderView::section { background-color: #333; color: white; padding: 4px; border: 1px solid #444; }
            QScrollArea { background-color: #2a2a2a; border: 1px solid #444; }
            QRadioButton { color: #ddd; background-color: transparent; }
            QRadioButton::indicator { border: 1px solid #777; border-radius: 6px; width: 12px; height: 12px; background-color: #333; }
            QRadioButton::indicator:checked { background-color: #ff9800; border-color: #ff9800; }
            QLabel { color: #ddd; background-color: transparent; }
        """)
        
        from src.core.utils import resource_path
        self.default_path = resource_path('config/gauge_rules.json')
        
        self.known_paths = [self.default_path] + [p for p in known_paths if p != self.default_path and os.path.exists(p)]
        self.active_path = active_path if (active_path and os.path.exists(active_path)) else self.default_path
        
        self.layout = QVBoxLayout(self)
        
        # Top buttons: New, Import (left) | Edit (right)
        h_top = QHBoxLayout()
        self.btn_new = QPushButton(" New")
        self.btn_import = QPushButton(" Import")
        self.btn_edit_top = QPushButton(" Edit")
        
        icon_add = resource_path('assets/icons/add_16dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png')
        if os.path.exists(icon_add): self.btn_new.setIcon(QIcon(icon_add))
        icon_import = resource_path('assets/icons/download_16dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png')
        if os.path.exists(icon_import): self.btn_import.setIcon(QIcon(icon_import))
        icon_edit = resource_path('assets/icons/edit_16dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png')
        if os.path.exists(icon_edit): self.btn_edit_top.setIcon(QIcon(icon_edit))
        
        h_top.addWidget(self.btn_new)
        h_top.addWidget(self.btn_import)
        h_top.addStretch()
        h_top.addWidget(self.btn_edit_top)
        self.layout.addLayout(h_top)
        
        # List of radio buttons
        self.scroll = QScrollArea()
        self.scroll.setWidgetResizable(True)
        self.scroll_content = QWidget()
        self.scroll_content.setStyleSheet("background-color: #2a2a2a;")
        self.scroll_layout = QVBoxLayout(self.scroll_content)
        self.scroll.setWidget(self.scroll_content)
        self.layout.addWidget(self.scroll)
        
        self.btn_group = QButtonGroup(self)
        self.current_editing_path = None
        
        self._build_list()
        
        # Editor section (hidden by default)
        self.editor_widget = QWidget()
        self.editor_layout = QVBoxLayout(self.editor_widget)
        self._is_modified = False
        
        h_ed_top = QHBoxLayout()
        self.lbl_editing = QLabel("Editing:")
        self.btn_save = QPushButton("Save")
        self.btn_revert = QPushButton("Undo")
        
        icon_save = resource_path('assets/icons/save_16dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png')
        if os.path.exists(icon_save): self.btn_save.setIcon(QIcon(icon_save))
        icon_refresh = resource_path('assets/icons/refresh_16dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png')
        if os.path.exists(icon_refresh): self.btn_revert.setIcon(QIcon(icon_refresh))
        
        h_ed_top.addWidget(self.lbl_editing)
        h_ed_top.addStretch()
        h_ed_top.addWidget(self.btn_save)
        h_ed_top.addWidget(self.btn_revert)
        
        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels(["Category", "Min", "Max", "Green Min", "Green Max", "Ticks (comma separated)"])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        self.table.cellChanged.connect(self._on_cell_changed)
        
        self.editor_layout.addLayout(h_ed_top)
        self.editor_layout.addWidget(self.table)
        self.editor_widget.hide()
        self.layout.addWidget(self.editor_widget)
        
        # Bottom Accept button
        h_bot = QHBoxLayout()
        self.btn_accept = QPushButton("Accept")
        self.btn_accept.setStyleSheet("background-color: #ff9800; color: #000; font-weight: bold; padding: 6px;")
        h_bot.addStretch()
        h_bot.addWidget(self.btn_accept)
        self.layout.addLayout(h_bot)
        
        # Connections
        self.btn_new.clicked.connect(self._on_new)
        self.btn_import.clicked.connect(self._on_import)
        self.btn_edit_top.clicked.connect(self._on_edit_selected)
        self.btn_save.clicked.connect(self._on_save)
        self.btn_revert.clicked.connect(self._on_revert)
        self.btn_accept.clicked.connect(self._on_accept)
        

    def _build_list(self):
        # clear existing
        for i in reversed(range(self.scroll_layout.count())): 
            widget = self.scroll_layout.itemAt(i).widget()
            if widget:
                widget.setParent(None)
                
        for path in self.known_paths:
            row = QWidget()
            h = QHBoxLayout(row)
            h.setContentsMargins(0, 0, 0, 0)
            
            rb = QRadioButton(os.path.basename(path))
            rb.setProperty("path", path)
            if path == self.active_path:
                rb.setChecked(True)
            self.btn_group.addButton(rb)
            h.addWidget(rb)
            
            lbl_path = QLabel(f"({os.path.dirname(path)})")
            lbl_path.setStyleSheet("color: #777; font-size: 10px; margin-left: 5px;")
            h.addWidget(lbl_path)
            
            if path == self.default_path:
                lbl_default = QLabel("(Default)")
                lbl_default.setStyleSheet("color: #888;")
                h.addWidget(lbl_default)
                
            h.addStretch()
            self.scroll_layout.addWidget(row)
        self.scroll_layout.addStretch()

    def _on_edit_selected(self):
        """Edit the currently selected gauge rules file."""
        checked = self.btn_group.checkedButton()
        if not checked:
            QMessageBox.warning(self, "No Selection", "Please select a gauge rules file first.")
            return
        path = checked.property("path")
        if path == self.default_path:
            QMessageBox.warning(self, "Default File", "The default gauge rules file cannot be edited.\nCreate a new one or import an existing file.")
            return
        self._open_editor(path)

    def _on_new(self):
        file_path, _ = QFileDialog.getSaveFileName(self, "Create New Gauge Rules", "", "JSON Files (*.json)")
        if file_path:
            # Check if name is already in list (basename)
            basenames = [os.path.basename(p) for p in self.known_paths]
            if os.path.basename(file_path) in basenames:
                QMessageBox.warning(self, "Error", "A configuration with this name already exists in the list.")
                return
            
            try:
                import shutil
                shutil.copy(self.default_path, file_path)
                self.known_paths.append(file_path)
                self.active_path = file_path
                self._build_list()
                self._open_editor(file_path)
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to create file:\n{e}")

    def _on_import(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Import Gauge Rules", "", "JSON Files (*.json)")
        if file_path:
            if file_path not in self.known_paths:
                self.known_paths.append(file_path)
            self.active_path = file_path
            self._build_list()

    def _open_editor(self, path):
        if self._is_modified:
            ans = QMessageBox.question(self, "Unsaved Changes", f"You have unsaved changes in '{os.path.basename(self.current_editing_path)}'. Save before switching?", QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel)
            if ans == QMessageBox.Yes:
                if not self._on_save(): return
            elif ans == QMessageBox.Cancel:
                return

        self.current_editing_path = path
        self.lbl_editing.setText(f"Editing: {os.path.basename(path)}")
        try:
            import json
            with open(path, 'r', encoding='utf-8') as f:
                data = json.load(f)
            
            self.table.blockSignals(True)
            self.table.setRowCount(len(data))
            for i, (cat, rules) in enumerate(data.items()):
                self.table.setItem(i, 0, QTableWidgetItem(cat))
                self.table.setItem(i, 1, QTableWidgetItem(str(rules.get('min', 0))))
                self.table.setItem(i, 2, QTableWidgetItem(str(rules.get('max', 10))))
                gr = rules.get('green_range', [0, 0])
                self.table.setItem(i, 3, QTableWidgetItem(str(gr[0])))
                self.table.setItem(i, 4, QTableWidgetItem(str(gr[1])))
                
                ticks = rules.get('ticks', [])
                ticks_str = ", ".join(str(t) for t in ticks)
                self.table.setItem(i, 5, QTableWidgetItem(ticks_str))
            self.table.blockSignals(False)
            
            self._is_modified = False
            self.editor_widget.show()
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to open file for editing:\n{e}")

    def _on_cell_changed(self):
        self._is_modified = True

    def _on_save(self):
        if not self.current_editing_path: return False
        
        try:
            import json
            # Build dict from table
            data = {}
            for i in range(self.table.rowCount()):
                cat = self.table.item(i, 0).text()
                try:
                    mi = float(self.table.item(i, 1).text())
                    ma = float(self.table.item(i, 2).text())
                    g1 = float(self.table.item(i, 3).text())
                    g2 = float(self.table.item(i, 4).text())
                    
                    # Parse ticks
                    raw_ticks = self.table.item(i, 5).text()
                    if raw_ticks.strip():
                        ticks = [float(t.strip()) for t in raw_ticks.split(",") if t.strip()]
                    else:
                        # Auto-generate if empty
                        step = (ma - mi) / 5
                        ticks = [mi + step*j for j in range(6)]
                except ValueError:
                    QMessageBox.warning(self, "Invalid Data", f"Numerical values expected for row '{cat}'. For Ticks, use comma-separated numbers.")
                    return False
                
                data[cat] = {
                    "min": mi,
                    "max": ma,
                    "ticks": ticks,
                    "green_range": [g1, g2]
                }
                
            with open(self.current_editing_path, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=2)
            
            self._is_modified = False
            QMessageBox.information(self, "Success", "Gauge rules saved successfully.")
            return True
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save file:\n{e}")
            return False

    def _on_revert(self):
        if not self.current_editing_path: return
        ans = QMessageBox.question(self, "Undo Changes", "Are you sure you want to revert to the default template? This will overwrite your current edits in the table.", QMessageBox.Yes | QMessageBox.No)
        if ans == QMessageBox.Yes:
            try:
                import json
                with open(self.default_path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                
                self.table.blockSignals(True)
                self.table.setRowCount(len(data))
                for i, (cat, rules) in enumerate(data.items()):
                    self.table.setItem(i, 0, QTableWidgetItem(cat))
                    self.table.setItem(i, 1, QTableWidgetItem(str(rules.get('min', 0))))
                    self.table.setItem(i, 2, QTableWidgetItem(str(rules.get('max', 10))))
                    gr = rules.get('green_range', [0, 0])
                    self.table.setItem(i, 3, QTableWidgetItem(str(gr[0])))
                    self.table.setItem(i, 4, QTableWidgetItem(str(gr[1])))
                    
                    ticks = rules.get('ticks', [])
                    ticks_str = ", ".join(str(t) for t in ticks)
                    self.table.setItem(i, 5, QTableWidgetItem(ticks_str))
                self.table.blockSignals(False)
                self._is_modified = True
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to load default template:\n{e}")

    def _on_accept(self):
        if self._is_modified:
            ans = QMessageBox.question(self, "Unsaved Changes", "You have unsaved changes. Save before closing?", QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel)
            if ans == QMessageBox.Yes:
                if self._on_save():
                    self.accept()
            elif ans == QMessageBox.No:
                self.accept()
        else:
            self.accept()

    def closeEvent(self, event):
        if self._is_modified:
            ans = QMessageBox.question(self, "Unsaved Changes", "You have unsaved changes. Save before closing?", QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel)
            if ans == QMessageBox.Yes:
                if self._on_save(): event.accept()
                else: event.ignore()
            elif ans == QMessageBox.No:
                event.accept()
            else:
                event.ignore()
        else:
            event.accept()
                
    def get_selected_path(self):
        if self.btn_group.checkedButton():
            return self.btn_group.checkedButton().property("path")
        return self.default_path


class GSRImagesEditorDialog(QDialog):
    def __init__(self, mapping, parent=None):
        super().__init__(parent)
        self.setWindowTitle("GSR ADDW Case Images Configuration")
        self.resize(800, 700)
        self.setStyleSheet("""
            QDialog { background-color: #2a2a2a; color: #ddd; }
            QPushButton { background-color: #444; color: #fff; border: 1px solid #666; padding: 4px; border-radius: 3px; }
            QPushButton:hover { background-color: #555; }
            QTableWidget { background-color: #1e1e1e; color: #ddd; border: 1px solid #555; }
            QComboBox { background-color: #333; color: #fff; border: 1px solid #555; }
            QLabel { color: #ddd; }
        """)
        
        # We work on a copy to allow Cancel
        self.mapping = mapping.copy() if mapping else {}
        # Paths
        self.gsr_assets_path = resource_path("assets/gsr")
        self.local_assets_path = os.path.join(self.gsr_assets_path, "local")
        
        # Ensure directories exist
        for d in [self.gsr_assets_path, self.local_assets_path]:
            if not os.path.exists(d):
                try: os.makedirs(d)
                except: pass

        layout = QVBoxLayout(self)
        
        # Header with Add Case and Import Image
        h_top = QHBoxLayout()
        self.btn_add_case = QPushButton(" Add Case (+)")
        self.btn_import_img = QPushButton(" Import New Image")
        
        icon_add = resource_path('assets/icons/add_16dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png')
        if os.path.exists(icon_add): self.btn_add_case.setIcon(QIcon(icon_add))
        icon_import = resource_path('assets/icons/image_16dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png')
        if os.path.exists(icon_import): self.btn_import_img.setIcon(QIcon(icon_import))
        
        h_top.addWidget(self.btn_add_case)
        h_top.addWidget(self.btn_import_img)
        h_top.addStretch()
        layout.addLayout(h_top)
        
        # Table
        self.table = QTableWidget(0, 4)
        self.table.setHorizontalHeaderLabels(["Case Name", "Assigned Image", "Preview", "Actions"])
        self.table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
        self.table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Stretch)
        self.table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Fixed)
        self.table.horizontalHeader().setSectionResizeMode(3, QHeaderView.Fixed)
        self.table.setColumnWidth(2, 120)
        self.table.setColumnWidth(3, 50)
        self.table.verticalHeader().setVisible(False)
        self.table.verticalHeader().setDefaultSectionSize(80)
        layout.addWidget(self.table)
        
        # Bottom buttons
        h_bot = QHBoxLayout()
        self.btn_save = QPushButton("Save Mapping")
        self.btn_save.setStyleSheet(f"background-color: {IDIADA_ORANGE}; color: black; font-weight: bold; padding: 6px 15px;")
        self.btn_cancel = QPushButton("Cancel")
        h_bot.addStretch()
        h_bot.addWidget(self.btn_cancel)
        h_bot.addWidget(self.btn_save)
        layout.addLayout(h_bot)
        
        # Connections
        self.btn_add_case.clicked.connect(self._on_add_case)
        self.btn_import_img.clicked.connect(self._on_import_image)
        self.btn_save.clicked.connect(self.accept)
        self.btn_cancel.clicked.connect(self.reject)
        
        self._refresh_table()

    def _get_available_images(self):
        imgs = []
        # Core images
        if os.path.exists(self.gsr_assets_path):
            imgs += [f for f in os.listdir(self.gsr_assets_path) 
                     if os.path.isfile(os.path.join(self.gsr_assets_path, f)) 
                     and f.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp'))]
        # Local images
        if os.path.exists(self.local_assets_path):
            local_imgs = [f"local/{f}" for f in os.listdir(self.local_assets_path) 
                          if os.path.isfile(os.path.join(self.local_assets_path, f)) 
                          and f.lower().endswith(('.png', '.jpg', '.jpeg', '.bmp'))]
            imgs += local_imgs
            
        # Sort naturally (1.png, 2.png, local/15.png...)
        imgs.sort(key=lambda s: [int(t) if t.isdigit() else t.lower() for t in re.split('([0-9]+)', s)])
        return imgs

    def _refresh_table(self):
        self.table.setRowCount(0)
        available_imgs = self._get_available_images()
        
        # Sort cases naturally
        sorted_cases = sorted(self.mapping.keys(), key=lambda s: [int(t) if t.isdigit() else t.lower() for t in re.split('([0-9]+)', s)])
        
        for case in sorted_cases:
            row = self.table.rowCount()
            self.table.insertRow(row)
            
            # 0. Case Name
            self.table.setItem(row, 0, QTableWidgetItem(case))
            
            # 1. Image Selector
            combo = QComboBox()
            combo.addItems(["None"] + available_imgs)
            current = self.mapping.get(case, "")
            if current in available_imgs:
                combo.setCurrentText(current)
            else:
                combo.setCurrentIndex(0)
            
            combo.currentTextChanged.connect(lambda text, c=case, r=row: self._on_image_selected(c, text, r))
            self.table.setCellWidget(row, 1, combo)
            
            # 2. Preview (1:1 aspect ratio)
            lbl_preview = QLabel()
            lbl_preview.setFixedSize(70, 70)
            lbl_preview.setScaledContents(True)
            lbl_preview.setAlignment(Qt.AlignCenter)
            lbl_preview.setStyleSheet("border: 1px solid #444; background-color: #000; border-radius: 4px;")
            self.table.setCellWidget(row, 2, lbl_preview)
            
            # 3. Actions (Delete button only for > 14)
            match = re.search(r'ADDW(\d+)', case)
            if match and int(match.group(1)) > 14:
                btn_del = QPushButton()
                btn_del.setFixedSize(30, 30)
                btn_del.setCursor(Qt.PointingHandCursor)
                btn_del.setToolTip("Delete this custom case")
                icon_del = resource_path("assets/icons/delete_16dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png")
                if os.path.exists(icon_del):
                    btn_del.setIcon(QIcon(icon_del))
                btn_del.setStyleSheet("background-color: #d1242f; border: none; border-radius: 4px;")
                btn_del.clicked.connect(lambda _, c=case: self._on_delete_case(c))
                
                # Center the button in cell
                container = QWidget()
                l_del = QHBoxLayout(container)
                l_del.setContentsMargins(0, 0, 0, 0)
                l_del.setAlignment(Qt.AlignCenter)
                l_del.addWidget(btn_del)
                self.table.setCellWidget(row, 3, container)
            else:
                self.table.setItem(row, 3, QTableWidgetItem(""))

            self._update_preview(row, combo.currentText())

    def _update_preview(self, row, img_name):
        lbl = self.table.cellWidget(row, 2)
        if not lbl: return
        if img_name == "None" or not img_name:
            lbl.clear()
            lbl.setText("No Img")
        else:
            path = os.path.join(self.gsr_assets_path, img_name)
            if os.path.exists(path):
                lbl.setPixmap(QPixmap(path))
            else:
                lbl.setText("Error")

    def _on_image_selected(self, case, img_name, row):
        self.mapping[case] = img_name if img_name != "None" else ""
        self._update_preview(row, img_name)

    def _on_delete_case(self, case):
        if case in self.mapping:
            del self.mapping[case]
            self._refresh_table()

    def _on_add_case(self):
        # Find highest ADDW number
        max_num = 0
        for k in self.mapping.keys():
            match = re.search(r'ADDW(\d+)', k)
            if match:
                max_num = max(max_num, int(match.group(1)))
        
        next_num = max_num + 1
        new_case = f"ADDW{next_num}"
        
        # Check if corresponding image exists (e.g. 15.png)
        available = self._get_available_images()
        default_img = f"{next_num}.png"
        if default_img not in available:
            # Maybe search for any format?
            found = False
            for ext in ['.jpg', '.jpeg', '.bmp']:
                if f"{next_num}{ext}" in available:
                    default_img = f"{next_num}{ext}"
                    found = True
                    break
            if not found: default_img = ""

        self.mapping[new_case] = default_img
        self._refresh_table()

    def _on_import_image(self):
        file_path, _ = QFileDialog.getOpenFileName(self, "Import Local Image for GSR", "", "Images (*.png *.jpg *.jpeg *.bmp)")
        if file_path:
            # Suggest a name based on next available number in local or core
            all_imgs = self._get_available_images()
            max_num = 0
            for f in all_imgs:
                # Handle local/ prefix if present
                base = os.path.basename(f)
                match = re.search(r'^(\d+)\.', base)
                if match:
                    max_num = max(max_num, int(match.group(1)))
            
            next_num = max_num + 1
            ext = os.path.splitext(file_path)[1]
            dest_name = f"{next_num}{ext}"
            dest_path = os.path.join(self.local_assets_path, dest_name)
            
            try:
                import shutil
                shutil.copy(file_path, dest_path)
                QMessageBox.information(self, "Success", f"Image imported as local/{dest_name}")
                self._refresh_table()
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to import image:\n{e}")

    def get_mapping(self):
        return self.mapping


class ThumbnailWorker(QThread):
    finished = Signal(list, float) # (thumbnails, duration)
    
    def __init__(self, video_path, num_thumbnails=80):
        super().__init__()
        self.video_path = video_path
        self.num_thumbnails = num_thumbnails
        
    def run(self):
        try:
            import cv2
            import numpy as np
            cap = cv2.VideoCapture(self.video_path)
            if not cap.isOpened(): return
            
            fps = cap.get(cv2.CAP_PROP_FPS)
            total_frames = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
            if total_frames <= 0 or fps <= 0: return
            
            duration = total_frames / fps
            
            # Select num_thumbnails evenly spaced indices across the whole video
            indices = np.linspace(0, total_frames - 1, self.num_thumbnails, dtype=int)
            thumbnails = []
            
            for idx in indices:
                cap.set(cv2.CAP_PROP_POS_FRAMES, int(idx))
                ret, frame = cap.read()
                if ret:
                    frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                    h, w, ch = frame.shape
                    image = QImage(frame.data, w, h, ch * w, QImage.Format_RGB888).copy()
                    thumbnails.append((int(idx) / fps, image))
            cap.release()
            self.finished.emit(thumbnails, duration)
        except Exception:
            pass

class VideoFilmstripWidget(QWidget):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFixedHeight(50)
        self.thumbnails = [] # List of (timestamp, QImage)
        self.current_x_range = (0, 100)
        self.cursor_pos = -1
        self.left_margin = 60 
        self.right_margin = 10
        self.video_duration = 0 # To know where the last frame ends
        self.setStyleSheet("background-color: #111; border-top: 1px solid #333; border-bottom: 1px solid #333;")
        self.setToolTip("Video Timeline (Matches Graph Scale)")
        
    def set_thumbnails(self, thumbnails, duration=None):
        self.thumbnails = thumbnails
        if duration is not None:
            self.video_duration = duration
        elif thumbnails and not self.video_duration:
            self.video_duration = thumbnails[-1][0]
        self.update()
        
    def update_range(self, x_min, x_max):
        self.current_x_range = (x_min, x_max)
        self.update()
        
    def set_cursor(self, x):
        self.cursor_pos = x
        self.update()
        
    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        rect = self.rect()
        
        # Drawing area adjusted for plot axes
        draw_rect = rect.adjusted(self.left_margin, 2, -self.right_margin, -2)
        painter.setClipRect(draw_rect)
        
        if not self.thumbnails:
            # Draw placeholder
            painter.setPen(QPen(QColor("#222"), 1, Qt.DashLine))
            painter.drawRect(draw_rect)
        else:
            t_start, t_end = self.current_x_range
            t_duration = t_end - t_start
            if t_duration > 0:
                draw_width = draw_rect.width()
                for i in range(len(self.thumbnails)):
                    t, img = self.thumbnails[i]
                    
                    # Calculate next timestamp to fill the gap
                    if i + 1 < len(self.thumbnails):
                        t_next = self.thumbnails[i+1][0]
                    else:
                        # Use video_duration for the last frame
                        t_next = max(t + 0.1, self.video_duration)
                    
                    # Map time to x coordinate
                    x = draw_rect.left() + (t - t_start) / t_duration * draw_width
                    x_next = draw_rect.left() + (t_next - t_start) / t_duration * draw_width
                    w = max(1, x_next - x)
                    
                    # Only draw if roughly in view
                    if x_next > draw_rect.left() and x < draw_rect.right():
                        h = draw_rect.height()
                        # Draw image stretched to fill the time slot
                        painter.drawImage(QRect(int(x), draw_rect.top(), int(w), int(h)), img)
        
        # Draw current timestamp line (cursor) - Unclipped for better visibility
        painter.setClipping(False)
        if self.cursor_pos != -1:
            t_start, t_end = self.current_x_range
            t_duration = t_end - t_start
            if t_duration > 0:
                cx = draw_rect.left() + (self.cursor_pos - t_start) / t_duration * draw_rect.width()
                if draw_rect.left() <= cx <= draw_rect.right():
                    painter.setPen(QPen(QColor("#FFFFFF"), 2))
                    painter.drawLine(int(cx), 0, int(cx), rect.height())

class AnimatedPlayButton(QPushButton):
    """Play/Pause button with smooth hover scaling and icon swap."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFixedSize(50, 50)
        self.setCursor(Qt.PointingHandCursor)
        self.is_playing = False
        self._hover = False
        self.setStyleSheet("background: transparent; border: none;")
        self.update_icon()

    def set_playing(self, playing):
        self.is_playing = playing
        self.update_icon()

    def update_icon(self):
        if self.is_playing:
            asset = "pause_circle_32dp_F39200_FILL1_wght400_GRAD0_opsz40.png" if self._hover else "pause_circle_24dp_FFFFFF_FILL1_wght400_GRAD0_opsz24.png"
            size = 40 if self._hover else 24
        else:
            asset = "play_circle_32dp_F39200_FILL1_wght400_GRAD0_opsz40.png" if self._hover else "play_circle_24dp_FFFFFF_FILL1_wght400_GRAD0_opsz24.png"
            size = 40 if self._hover else 24
        
        from src.core.utils import resource_path
        icon_path = resource_path(f"assets/icons/{asset}")
        self.setIcon(QIcon(icon_path))
        self.setIconSize(QSize(size, size))

    def enterEvent(self, event):
        self._hover = True
        self.update_icon()
        super().enterEvent(event)

    def leaveEvent(self, event):
        self._hover = False
        self.update_icon()
        super().leaveEvent(event)

class VideoOverlayControls(QFrame):
    """Floating translucent overlay for camera and speed controls."""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setObjectName("VideoOverlay")
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.setStyleSheet("""
            #VideoOverlay {
                background-color: rgba(0, 0, 0, 200);
                border: 1px solid rgba(255, 255, 255, 45);
                border-radius: 20px;
            }
            QLabel { 
                color: #FFFFFF; 
                font-size: 12px; 
                font-weight: 800; 
                background: transparent; 
            }
            QComboBox { 
                background: rgba(255, 255, 255, 30); 
                border: 1px solid rgba(255, 255, 255, 60); 
                border-radius: 8px;
                color: white; 
                padding: 4px 10px;
                min-width: 90px;
                font-weight: 600;
            }
            QComboBox::drop-down { border: none; }
            QComboBox QAbstractItemView {
                background-color: #1A1A1A;
                color: white;
                selection-background-color: #F39200;
            }
        """)
        
        # Add a subtle glow/shadow to the whole panel to separate it from video
        shadow = QGraphicsDropShadowEffect(self)
        shadow.setBlurRadius(15)
        shadow.setColor(QColor(0, 0, 0, 180))
        shadow.setOffset(0, 2)
        self.setGraphicsEffect(shadow)
        
        self.layout = QHBoxLayout(self)
        self.layout.setContentsMargins(25, 5, 25, 5)
        self.layout.setSpacing(15)
        self.setFixedHeight(55)
        self.setMinimumWidth(380) 
        self.raise_() # Ensure it stays on top of other widgets in container

class ReportDropZoneLabel(QLabel):
    """Area to drag and drop participants to generate report previews."""
    case_dropped = Signal(str, str)
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)
        self.setAlignment(Qt.AlignCenter)
        self.setText("Drag & Drop cases from 'Participant Status'\nhere to Generate Report Preview")
        self.setWordWrap(True)
        self.setStyleSheet("""
            QLabel {
                color: #888; 
                font-size: 16px; 
                font-style: italic;
                padding: 50px;
                border: 2px dashed #555;
                border-radius: 10px;
                background-color: transparent;
            }
        """)

    def dragEnterEvent(self, event: QDragEnterEvent):
        event.acceptProposedAction()
        text = "Drop file to Generate"
        source = event.source()
        if source and hasattr(source, 'currentItem'):
            item = source.currentItem()
            if item:
                text = item.text(0)
                
        icon_path = resource_path("assets/icons/file_export_48dp_FFFFFF_FILL0_wght400_GRAD0_opsz48.png")
        self.setText(f'<html><center><img src="{icon_path}" width="80" height="80"><br><br><b>{text}</b></center></html>')
        
        self.setStyleSheet("""
            QLabel {
                color: #ff9800; 
                font-size: 16px; 
                font-style: italic;
                padding: 30px;
                border: 2px solid #ff9800;
                border-radius: 10px;
                background-color: #2a2a2a;
            }
        """)
                
    def dragLeaveEvent(self, event):
        self.setText("Drag & Drop cases from 'Participant Status'\nhere to Generate Report Preview")
        self.setStyleSheet("""
            QLabel {
                color: #888; 
                font-size: 16px; 
                font-style: italic;
                padding: 50px;
                border: 2px dashed #555;
                border-radius: 10px;
                background-color: transparent;
            }
        """)

    def dropEvent(self, event: QDropEvent):
        self.dragLeaveEvent(None)
        event.acceptProposedAction()
        
        name, path = "", ""
        source = event.source()
        if source and hasattr(source, 'currentItem'):
            item = source.currentItem()
            if item:
                name = item.text(0)
                path = item.data(0, Qt.UserRole)
                
        self.case_dropped.emit(name, path if isinstance(path, str) else "")


class SignalDropZone(QLabel):
    file_loaded = Signal(str)
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setAcceptDrops(True)
        self.setAlignment(Qt.AlignCenter)
        self.setText("Drag & Drop Master MF4 here\nto filter signals")
        self.setWordWrap(True)
        self.setMinimumHeight(100)
        self.setStyleSheet("""
            QLabel {
                color: #888; 
                font-size: 14px; 
                font-style: italic;
                padding: 20px;
                border: 2px dashed #555;
                border-radius: 8px;
                background-color: #252525;
            }
        """)

    def dragEnterEvent(self, event: QDragEnterEvent):
        if event.mimeData().hasUrls():
            urls = event.mimeData().urls()
            if urls and urls[0].toLocalFile().lower().endswith(".mf4"):
                event.acceptProposedAction()
                self.setStyleSheet(self.styleSheet().replace("#555", IDIADA_ORANGE).replace("#252525", "#333"))
        
    def dragLeaveEvent(self, event):
        self.setStyleSheet(self.styleSheet().replace(IDIADA_ORANGE, "#555").replace("#333", "#252525"))

    def dropEvent(self, event: QDropEvent):
        self.dragLeaveEvent(None)
        if event.mimeData().hasUrls():
            urls = event.mimeData().urls()
            if urls:
                path = urls[0].toLocalFile()
                if path.lower().endswith(".mf4"):
                    self.file_loaded.emit(path)
                    event.acceptProposedAction()

class SignalSelectorWidget(QWidget):
    selection_changed = Signal(list, object) # names, mdf
    limit_reached = Signal()
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.max_signals = 0
        self.mdf = None
        self.layout = QVBoxLayout(self)
        self.layout.setContentsMargins(0, 5, 0, 0)
        
        self.txt_search = QLineEdit()
        self.txt_search.setPlaceholderText("Search signals...")
        self.txt_search.textChanged.connect(self._filter_signals)
        self.layout.addWidget(self.txt_search)
        
        self.tree = QTreeWidget()
        self.tree.setHeaderLabels(["Signal", "Group"])
        self.tree.setAlternatingRowColors(True)
        self.tree.header().setStretchLastSection(False)
        self.tree.header().setSectionResizeMode(0, QHeaderView.Stretch)
        self.tree.itemChanged.connect(self._on_item_changed)
        self.layout.addWidget(self.tree)
        
        h_footer = QHBoxLayout()
        self.lbl_counter = QLabel("0 selected")
        self.lbl_counter.setStyleSheet("color: #aaa; font-size: 11px;")
        self.btn_clear = QPushButton("Clear")
        self.btn_clear.setFixedWidth(60)
        self.btn_clear.setCursor(Qt.PointingHandCursor)
        self.btn_clear.setStyleSheet("font-size: 10px; padding: 2px;")
        
        h_footer.addWidget(self.lbl_counter)
        h_footer.addStretch()
        h_footer.addWidget(self.btn_clear)
        self.layout.addLayout(h_footer)

    def load_file(self, path):
        if not path or not os.path.exists(path): return
        try:
            self.mdf = MDF(path)
            self.tree.clear()
            self.tree.blockSignals(True)
            for ch in self.mdf.iter_channels():
                if ch.name.lower() in ["t", "time"]: continue
                item = QTreeWidgetItem(self.tree)
                item.setText(0, ch.name)
                item.setText(1, str(getattr(ch, 'group_index', '0')))
                item.setCheckState(0, Qt.Unchecked)
            self.tree.blockSignals(False)
            self.update_counter_text()
        except Exception as e:
            print(f"Error loading SignalSelector MF4: {e}")

    def _filter_signals(self, text):
        text = text.lower()
        it = QTreeWidgetItemIterator(self.tree)
        while it.value():
            item = it.value()
            item.setHidden(text not in item.text(0).lower())
            it += 1

    def _on_item_changed(self, item, col):
        if col != 0: return
        checked = []
        it = QTreeWidgetItemIterator(self.tree)
        while it.value():
            it_item = it.value()
            if it_item.checkState(0) == Qt.Checked:
                checked.append(it_item.text(0))
            it += 1
        
        if self.max_signals > 0 and len(checked) > self.max_signals:
            self.tree.blockSignals(True)
            item.setCheckState(0, Qt.Unchecked)
            self.tree.blockSignals(False)
            self.limit_reached.emit()
            return

        self.update_counter_text()
        self.selection_changed.emit(checked, self.mdf)

    def update_counter_text(self):
        count = 0
        it = QTreeWidgetItemIterator(self.tree)
        while it.value():
            if it.value().checkState(0) == Qt.Checked: count += 1
            it += 1
        limit_str = f" / {self.max_signals}" if self.max_signals > 0 else ""
        self.lbl_counter.setText(f"{count}{limit_str} selected")

    def has_signal(self, name):
        it = QTreeWidgetItemIterator(self.tree)
        while it.value():
            if it.value().text(0) == name: return True
            it += 1
        return False

    def set_signal_checked(self, name, checked):
        it = QTreeWidgetItemIterator(self.tree)
        while it.value():
            item = it.value()
            if item.text(0) == name:
                self.tree.blockSignals(True)
                item.setCheckState(0, Qt.Checked if checked else Qt.Unchecked)
                self.tree.blockSignals(False)
                self._on_item_changed(item, 0)
                break
            it += 1

class LogicTabWidget(QWidget):
    """
    Tab for defining analysis logic by distraction categories.
    Left: Configuration with tables for each distraction category.
    Right: Report Preview (generated PNG displayed after clicking Generate).
    """
    EURO_NCAP_CATEGORIES = [
        "Long Distraction (NDT)",
        "Long Distraction (DT)",
        "Short Distraction (NDT)",
        "Short Distraction (DT)",
        "Microsleep",
        "Sleep",
        "Drowsiness",
        "Unresponsive driver"
    ]
    
    GSR_ADDW_CATEGORIES = [
        "High Speed",
        "Low Speed"
    ]
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.known_gauge_rules = []
        self.active_gauge_rules_path = None
        self.category_counters = {}
        self.load_mf4_buttons = {}
        self._config_modified = False
        self._current_config_name = "Save config"
        
        # Pulsing animation for Save button (smooth yoyo)
        self.save_btn_anim = QSequentialAnimationGroup(self)
        
        self._save_anim_up = QVariantAnimation(self)
        self._save_anim_up.setDuration(1000)
        self._save_anim_up.setStartValue(QColor("#444444"))
        self._save_anim_up.setEndValue(QColor("#FFFFFF"))
        self._save_anim_up.setEasingCurve(QEasingCurve.InOutSine)
        self._save_anim_up.valueChanged.connect(self._update_save_btn_style)
        
        self._save_anim_down = QVariantAnimation(self)
        self._save_anim_down.setDuration(1000)
        self._save_anim_down.setStartValue(QColor("#FFFFFF"))
        self._save_anim_down.setEndValue(QColor("#444444"))
        self._save_anim_down.setEasingCurve(QEasingCurve.InOutSine)
        self._save_anim_down.valueChanged.connect(self._update_save_btn_style)
        
        self.save_btn_anim.addAnimation(self._save_anim_up)
        self.save_btn_anim.addAnimation(self._save_anim_down)
        self.save_btn_anim.setLoopCount(-1)
        
        # GSR ADDW Image Mapping (Case Name -> Image File Name)
        self.gsr_image_mapping = {f"ADDW{i}": f"{i}.png" for i in range(1, 15)}
        
        self.layout = QVBoxLayout(self)
        self.layout.setContentsMargins(10, 10, 10, 10)
        
        self.splitter = QSplitter(Qt.Horizontal)
        self.splitter.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        
        # --- Left Panel: Configuration ---
        self.panel_config = QWidget()
        l_conf = QVBoxLayout(self.panel_config)
        l_conf.setContentsMargins(5, 5, 5, 5)
        
        # 0. Config Save/Load + Gauges + Marks
        gb_config = QGroupBox("Configuration")
        l_config = QHBoxLayout(gb_config)
        l_config.setContentsMargins(10, 5, 10, 5)
        l_config.setSpacing(10)
        
        self.btn_save_config = QPushButton("Save config")
        self.btn_save_config.setIcon(QIcon(resource_path("assets/icons/save_16dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png")))
        self.btn_save_config.setToolTip("Save current settings to a .json file")
        self.btn_save_config.setFixedHeight(32)
        self.btn_save_config.setStyleSheet("padding: 0 10px; font-weight: normal;")
        self.btn_save_config.setCursor(Qt.PointingHandCursor)
        self.btn_save_config.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        
        self.btn_load_config = QPushButton(" Import settings")
        self.btn_load_config.setIcon(QIcon(resource_path("assets/icons/download_16dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png")))
        self.btn_load_config.setToolTip("Load settings from a .json file")
        self.btn_load_config.setFixedHeight(32)
        self.btn_load_config.setStyleSheet("padding: 0 10px; font-weight: normal;")
        self.btn_load_config.setCursor(Qt.PointingHandCursor)
        self.btn_load_config.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        
        self.btn_save_config.clicked.connect(self.save_config)
        self.btn_load_config.clicked.connect(self.load_config)
        
        l_config.addWidget(self.btn_save_config)
        l_config.addWidget(self.btn_load_config)
        
        # Consolidation: Single Gauges button
        self.btn_gauges = QPushButton(" Gauges")
        self.btn_gauges.setIcon(QIcon(resource_path("assets/icons/speed_16dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png")))
        self.btn_gauges.setToolTip("Configure global Gauge Rules")
        self.btn_gauges.setFixedHeight(32)
        self.btn_gauges.setStyleSheet("padding: 0 10px; font-weight: normal;")
        self.btn_gauges.setCursor(Qt.PointingHandCursor)
        self.btn_gauges.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.btn_gauges.clicked.connect(lambda: self.open_gauge_rules_dialog(None))
        l_config.addWidget(self.btn_gauges)
        
        self.btn_auto_load = QPushButton("Auto-Load data")
        self.btn_auto_load.setToolTip("Automatically load corresponding MF4 files for each category from the selected source directory")
        self.btn_auto_load.setIcon(QIcon(resource_path("assets/icons/folder_match_16dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png")))
        self.btn_auto_load.setStyleSheet("padding: 0 10px; font-weight: normal;")
        self.btn_auto_load.setFixedHeight(32)
        self.btn_auto_load.setCursor(Qt.PointingHandCursor)
        self.btn_auto_load.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.btn_auto_load.clicked.connect(self.auto_load_project_mdfs)
        l_config.addWidget(self.btn_auto_load)
        
        # GSR Images Mapping Button (Hidden by default, shown for GSR)
        self.btn_gsr_images = QPushButton(" Images")
        self.btn_gsr_images.setIcon(QIcon(resource_path("assets/icons/image_16dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png")))
        self.btn_gsr_images.setToolTip("Configure images associated with each GSR ADDW case")
        self.btn_gsr_images.setFixedHeight(32)
        self.btn_gsr_images.setStyleSheet("padding: 0 10px; font-weight: normal;")
        self.btn_gsr_images.setCursor(Qt.PointingHandCursor)
        self.btn_gsr_images.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.btn_gsr_images.clicked.connect(self.open_gsr_images_dialog)
        self.btn_gsr_images.setVisible(False)
        l_config.addWidget(self.btn_gsr_images)
        
        # New Marks Toggle Button (aligned left with others)
        # REMOVED: system now always uses marks
        
        self.btn_generate_report = QPushButton(" Generate Reports")
        self.btn_generate_report.setToolTip("Start or stop report generation")
        icon_report = QIcon(resource_path("assets/icons/play_circle_16dp_000000_FILL0_wght400_GRAD0_opsz20.png"))
        if not icon_report.isNull():
            self.btn_generate_report.setIcon(icon_report)
        self.btn_generate_report.setFixedHeight(32)
        self.btn_generate_report.setStyleSheet(f"background-color: {IDIADA_ORANGE}; color: black; font-weight: bold; padding: 0 10px; border-radius: 3px;")
        self.btn_generate_report.setCursor(Qt.PointingHandCursor)
        self.btn_generate_report.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        l_config.addWidget(self.btn_generate_report)
        
        l_conf.addWidget(gb_config)
        
        # New Protocol Selection Box
        gb_protocol = QGroupBox("Protocol Selection")
        l_protocol = QHBoxLayout(gb_protocol)
        l_protocol.setContentsMargins(10, 5, 10, 5)
        
        self.combo_protocol = QComboBox()
        self.combo_protocol.addItems(["Euro NCAP", "GSR ADDW"])
        self.combo_protocol.setStyleSheet("""
            QComboBox {
                background-color: #333; color: white; border: 1px solid #555;
                padding: 4px; border-radius: 3px; font-weight: bold;
            }
        """)
        self.combo_protocol.currentTextChanged.connect(self._on_protocol_changed)
        l_protocol.addWidget(self.combo_protocol)
        
        l_conf.addWidget(gb_protocol)
        
        # 2. Signal Configuration Tables by Category
        self.gb_signals = QGroupBox("Signal Configuration by Category")
        self.l_sig_conf = QVBoxLayout(self.gb_signals)
        self.l_sig_conf.setContentsMargins(10, 10, 10, 10)
        
        self.active_categories = self.EURO_NCAP_CATEGORIES.copy()
        self.category_tables = {}
        self.pass_criteria_tables = {}
        self._pass_criteria_table_refs = {}
        self.category_tab_buttons = {}
        
        self.tabs_container = None
        self.category_stack = None
        
        self._last_selected_tabs = {
            "Euro NCAP": "Long Distraction (NDT)",
            "GSR ADDW": "High Speed"
        }
        
        self._init_categories_ui(self.combo_protocol.currentText())
        
        l_conf.addWidget(self.gb_signals, 1)
        
        # Data References
        self.mdf = None
        # Audio params
        self.audio_min_freq = 0
        self.audio_max_freq = 0
        self.audio_threshold = 0

        # --- Right Panel: Report Preview (PNG) ---
        self.scroll_preview = QScrollArea()
        self.scroll_preview.setStyleSheet("background-color: #444;")
        self.scroll_preview.setWidgetResizable(True)
        self.scroll_preview.setAlignment(Qt.AlignCenter)
        
        # Placeholder drop zone
        self.lbl_preview = ReportDropZoneLabel()
        self.lbl_preview.case_dropped.connect(self.generate_report)
        self.scroll_preview.setWidget(self.lbl_preview)
        
        self.splitter.addWidget(self.panel_config)
        self.splitter.addWidget(self.scroll_preview)
        self.splitter.setStretchFactor(0, 1)
        self.splitter.setStretchFactor(1, 1)
        
        self.layout.addWidget(self.splitter)

    def showEvent(self, event):
        super().showEvent(event)
        self._update_indicator_geometry(animate=False)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self._update_indicator_geometry(animate=False)

    def _update_indicator_geometry(self, animate=True):
        if not hasattr(self, 'category_tab_buttons') or not hasattr(self, 'tab_indicator'):
            return
            
        for btn in self.category_tab_buttons.values():
            if btn.isChecked():
                if animate and self.tab_indicator.isVisible():
                    self.anim_indicator = QPropertyAnimation(self.tab_indicator, b"geometry")
                    self.anim_indicator.setDuration(300)
                    self.anim_indicator.setStartValue(self.tab_indicator.geometry())
                    self.anim_indicator.setEndValue(btn.geometry())
                    self.anim_indicator.setEasingCurve(QEasingCurve.OutQuint)
                    self.anim_indicator.start()
                else:
                    self.tab_indicator.show()
                    self.tab_indicator.setGeometry(btn.geometry())
                break

    def _on_table_item_changed(self, item, category, table):
        if item.column() != 0:
            return
            
        checked_count = 0
        for row in range(table.rowCount()):
            chk = table.item(row, 0)
            if chk and chk.checkState() == Qt.Checked:
                checked_count += 1
                
        if checked_count > 5:
            table.blockSignals(True)
            item.setCheckState(Qt.Unchecked)
            table.blockSignals(False)
            QMessageBox.warning(self, "Signal Limit", "You can only select a maximum of 5 signals per category.")
            checked_count = 5
            
        if hasattr(self, 'category_counters') and category in self.category_counters:
            lbl = self.category_counters[category]
            lbl.setText(f"{checked_count}/5 signals selected")
            if checked_count == 5:
                lbl.setStyleSheet("color: #ff3333; font-weight: bold; font-size: 11px;")
            else:
                lbl.setStyleSheet("color: #ff9800; font-weight: bold; font-size: 11px;")
        
        self._mark_modified()

    def open_gauge_rules_dialog(self, category):
        dialog = GaugeRulesEditorDialog(self.known_gauge_rules, self.active_gauge_rules_path, self)
        if dialog.exec():
            self.active_gauge_rules_path = dialog.get_selected_path()
            self.known_gauge_rules = dialog.known_paths

    def generate_report(self, name="", path=""):
        """Generate report using MatplotlibReportBuilder and display it asynchronously using MdfLoaderWorker."""
        if not path:
            return
            
        parent = self._get_parent_analysis_widget()
        if parent and hasattr(parent, 'busy_changed'):
            parent.busy_changed.emit(True)
            
        self.lbl_preview.setText(f"Loading {name}...")
        
        # We must keep a reference to the loader so it doesn't get garbage collected
        self._report_mdf_loader = MdfLoaderWorker(path)
        self._report_mdf_loader.loaded.connect(self._on_mdf_loaded_for_report)
        self._report_mdf_loader.error.connect(self._on_mdf_error_for_report)
        self._report_mdf_loader.start()

    def _on_mdf_loaded_for_report(self, mdf, sig_names, sig_data):
        try:
            # Collect configuration utilizing the asynchronously loaded mdf
            config = self._collect_report_config(mdf_source=mdf)
            
            # Start report generation background worker
            if hasattr(self, '_report_worker') and self._report_worker is not None:
                self._report_worker.quit()
                self._report_worker.wait()
            
            self._report_worker = ReportGeneratorWorker(config, "report_preview.png", dpi=200)
            self._report_worker.finished.connect(self._on_preview_report_done)
            self._report_worker.error.connect(self._on_preview_report_error)
            self._report_worker.start()
            self.lbl_preview.setText("Generating preview...")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to collect report config:\n{str(e)}")
            import traceback
            traceback.print_exc()
            self._emit_not_busy()

        self._mark_modified()


    def _on_protocol_changed(self, protocol_name):
        self._init_categories_ui(protocol_name)
        # Toggle GSR Images button visibility
        if hasattr(self, 'btn_gsr_images'):
            self.btn_gsr_images.setVisible(protocol_name == "GSR ADDW")
        self._mark_modified()

    def _init_categories_ui(self, protocol_name):
        # Clear all items from l_sig_conf (widgets and spacers)
        while self.l_sig_conf.count():
            item = self.l_sig_conf.takeAt(0)
            if item.widget():
                item.widget().setParent(None)
                item.widget().deleteLater()
            
        self.category_tab_buttons.clear()
        self.category_tables.clear()
        self.pass_criteria_tables.clear()
        self._pass_criteria_table_refs.clear()
        self.load_mf4_buttons.clear()
        self.category_counters.clear()
            
        if protocol_name == "GSR ADDW":
            self.active_categories = self.GSR_ADDW_CATEGORIES.copy()
        else:
            self.active_categories = self.EURO_NCAP_CATEGORIES.copy()
            
        # CATEGORY TAB BAR (Custom Navigation)
        self.tabs_container = QFrame()
        self.tabs_container.setStyleSheet("background-color: #222; border-radius: 6px; border: 1px solid #444;")
        self.l_tabs = QHBoxLayout(self.tabs_container)
        self.l_tabs.setContentsMargins(5, 5, 5, 5) 
        self.l_tabs.setSpacing(8)
        
        # Sliding Indicator (Animated Box)
        self.tab_indicator = QFrame(self.tabs_container)
        self.tab_indicator.setStyleSheet("""
            QFrame {
                background-color: #2a2a2a;
                border: 1px solid #444;
                border-radius: 6px;
            }
        """)
        self.tab_indicator.lower() # Place behind buttons
        
        self.category_stack = FadeStackedWidget()
        self.category_stack.setStyleSheet("background-color: transparent;")
        
        # Create a table for each category
        for category in self.active_categories:
            # Create Navigation Button
            btn = QPushButton(category)
            btn.setCheckable(True)
            btn.setCursor(Qt.PointingHandCursor)
            btn.clicked.connect(lambda checked, c=category: self._on_category_tab_clicked(c))
            self.l_tabs.addWidget(btn)
            self.category_tab_buttons[category] = btn
            
            table = QTableWidget(0, 5)
            table.setHorizontalHeaderLabels(["✓", "Signal", "Operator", "Value", "Alias"])
            table.horizontalHeader().setSectionResizeMode(0, QHeaderView.ResizeToContents)
            table.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
            table.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
            table.horizontalHeader().setSectionResizeMode(3, QHeaderView.Stretch)
            table.horizontalHeader().setSectionResizeMode(4, QHeaderView.Stretch)
            table.verticalHeader().setVisible(False)
            table.setStyleSheet("""
                QTableWidget { 
                    font-size: 9px; 
                    background-color: #333; 
                    gridline-color: #444;
                    alternate-background-color: #2a2a2a;
                } 
                QHeaderView::section { 
                    background-color: #444; 
                    color: #fff; 
                    padding: 3px; 
                    border: 1px solid #555; 
                    font-weight: bold;
                }
                QTableWidget::item {
                    padding: 2px;
                    color: #ddd;
                }
            """)
            table.setMinimumHeight(80)
            
            # Create frame group for each category (title-less since we have tabs)
            gb_category = QGroupBox()
            gb_category.setStyleSheet("""
                QGroupBox {
                    border: 2px solid #555;
                    border-radius: 5px;
                    margin-top: 5px;
                    padding-top: 15px;
                    padding-left: 10px;
                    padding-right: 10px;
                    padding-bottom: 10px;
                }
            """)
            l_cat = QVBoxLayout(gb_category)
            l_cat.setContentsMargins(5, 5, 5, 5)
            
            # Add load MF4 button per category
            h_title = QHBoxLayout()
            h_title.setContentsMargins(0, 0, 0, 0)
            
            lbl_counter = QLabel("0/5 signals selected")
            lbl_counter.setStyleSheet("color: #ff9800; font-weight: bold; font-size: 11px;")
            h_title.addWidget(lbl_counter)
            
            self.category_counters[category] = lbl_counter
            
            case_ids_map = {
                "Long Distraction (NDT)": "Case IDs: D1-D9",
                "Long Distraction (DT)": "Case IDs: D10-D15",
                "Short Distraction (NDT)": "Case IDs: D16-D19, D28, D29-D42 (Includes Phone Use)",
                "Short Distraction (DT)": "Case IDs: D20-D27",
                "High Speed": "Case IDs: ADDW High Speed",
                "Low Speed": "Case IDs: ADDW Low Speed"
            }
            btn_info = QPushButton()
            btn_info.setIcon(QIcon(resource_path("assets/icons/info_16dp_666666_FILL0_wght400_GRAD0_opsz20.png")))
            btn_info.setToolTip(case_ids_map.get(category, f"{category}"))
            btn_info.setFixedSize(20, 20)
            btn_info.setStyleSheet("""
                QPushButton {
                    background-color: transparent;
                    border: none;
                }
                QPushButton:hover {
                    background-color: #444;
                    border-radius: 10px;
                }
            """)
            btn_info.setCursor(Qt.PointingHandCursor)
            h_title.addWidget(btn_info)
            
            h_title.addStretch()
            
            btn_load_mf4 = QPushButton()
            btn_load_mf4.setToolTip("Import signals from MF4 for this category")
            btn_load_mf4.setIcon(QIcon(resource_path("assets/icons/download_16dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png")))
            btn_load_mf4.setMinimumHeight(24)
            btn_load_mf4.setStyleSheet("""
                QPushButton { background-color: transparent; border: 1px solid #555; border-radius: 3px; padding-left: 5px; padding-right: 5px; color: #aaa; font-size: 10px; }
                QPushButton:hover { background-color: #444; color: white; }
            """)
            btn_load_mf4.setCursor(Qt.PointingHandCursor)
            btn_load_mf4.clicked.connect(lambda _, c=category: self._on_load_mf4_for_category(c))
            self.load_mf4_buttons[category] = btn_load_mf4
            h_title.addWidget(btn_load_mf4)
            
            l_cat.addLayout(h_title)
            
            l_cat.addWidget(table)
            
            table.itemChanged.connect(lambda item, c=category, t=table: self._on_table_item_changed(item, c, t))
            self.category_tables[category] = table
            
            # PASS Criteria Table for this category
            pass_criteria_table = QTableWidget(1, 6)
            pass_criteria_table.setHorizontalHeaderLabels(["PASS Criteria Signal", "Operator", "Value", "Operator 2", "Value 2", "Mask"])
            pass_criteria_table.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
            pass_criteria_table.horizontalHeader().setSectionResizeMode(1, QHeaderView.Interactive)
            pass_criteria_table.horizontalHeader().setSectionResizeMode(2, QHeaderView.Interactive)
            pass_criteria_table.horizontalHeader().setSectionResizeMode(3, QHeaderView.Interactive)
            pass_criteria_table.horizontalHeader().setSectionResizeMode(4, QHeaderView.Interactive)
            pass_criteria_table.horizontalHeader().setSectionResizeMode(5, QHeaderView.Interactive)
            pass_criteria_table.verticalHeader().setVisible(False)
            pass_criteria_table.setMaximumHeight(90)
            pass_criteria_table.setStyleSheet("""
                QTableWidget { 
                    font-size: 9px; 
                    background-color: #2a2a2a; 
                    gridline-color: #444;
                } 
                QHeaderView::section { 
                    background-color: #555; 
                    color: #fff; 
                    padding: 2px; 
                    border: 1px solid #666; 
                    font-weight: bold;
                }
                QTableWidget::item {
                    padding: 2px;
                    color: #ddd;
                }
            """)
            
            self._init_pass_criteria_row(pass_criteria_table, category)
            self._pass_criteria_table_refs[category] = pass_criteria_table
            l_cat.addWidget(pass_criteria_table)
            
            self.category_stack.addWidget(gb_category)
            self.pass_criteria_tables[category] = pass_criteria_table
            
        self.l_sig_conf.addWidget(self.tabs_container)
        self.l_sig_conf.addSpacing(10)
        self.l_sig_conf.addWidget(self.category_stack)
        
        # Select first category visually, geometry updated on showEvent
        if self.active_categories:
            default_tab = self._last_selected_tabs.get(protocol_name, self.active_categories[0])
            if default_tab not in self.active_categories:
                default_tab = self.active_categories[0]
            # Use singleShot so layout calculates sizes before updating indicator geometry
            QTimer.singleShot(0, lambda: self._on_category_tab_clicked(default_tab, animate=False))

    def showEvent(self, event):
        super().showEvent(event)
        self._update_indicator_geometry(animate=False)

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self._update_indicator_geometry(animate=False)

    def _update_indicator_geometry(self, animate=True):
        if not hasattr(self, 'category_tab_buttons') or not hasattr(self, 'tab_indicator'):
            return
            
        for btn in self.category_tab_buttons.values():
            if btn.isChecked():
                if animate and self.tab_indicator.isVisible():
                    self.anim_indicator = QPropertyAnimation(self.tab_indicator, b"geometry")
                    self.anim_indicator.setDuration(300)
                    self.anim_indicator.setStartValue(self.tab_indicator.geometry())
                    self.anim_indicator.setEndValue(btn.geometry())
                    self.anim_indicator.setEasingCurve(QEasingCurve.OutQuint)
                    self.anim_indicator.start()
                else:
                    self.tab_indicator.show()
                    self.tab_indicator.setGeometry(btn.geometry())
                break

    def _on_table_item_changed(self, item, category, table):
        if item.column() != 0:
            return
            
        checked_count = 0
        for row in range(table.rowCount()):
            chk = table.item(row, 0)
            if chk and chk.checkState() == Qt.Checked:
                checked_count += 1
                
        if checked_count > 5:
            table.blockSignals(True)
            item.setCheckState(Qt.Unchecked)
            table.blockSignals(False)
            QMessageBox.warning(self, "Signal Limit", "You can only select a maximum of 5 signals per category.")
            checked_count = 5
            
        if hasattr(self, 'category_counters') and category in self.category_counters:
            lbl = self.category_counters[category]
            lbl.setText(f"{checked_count}/5 signals selected")
            if checked_count == 5:
                lbl.setStyleSheet("color: #ff3333; font-weight: bold; font-size: 11px;")
            else:
                lbl.setStyleSheet("color: #ff9800; font-weight: bold; font-size: 11px;")
        
        self._mark_modified()

    def open_gauge_rules_dialog(self, category):
        dialog = GaugeRulesEditorDialog(self.known_gauge_rules, self.active_gauge_rules_path, self)
        if dialog.exec():
            self.active_gauge_rules_path = dialog.get_selected_path()
            self.known_gauge_rules = dialog.known_paths

    def generate_report(self, name="", path=""):
        """Generate report using MatplotlibReportBuilder and display it asynchronously using MdfLoaderWorker."""
        if not path:
            return
            
        parent = self._get_parent_analysis_widget()
        if parent and hasattr(parent, 'busy_changed'):
            parent.busy_changed.emit(True)
            
        self.lbl_preview.setText(f"Loading {name}...")
        
        # We must keep a reference to the loader so it doesn't get garbage collected
        self._report_mdf_loader = MdfLoaderWorker(path)
        self._report_mdf_loader.loaded.connect(self._on_mdf_loaded_for_report)
        self._report_mdf_loader.error.connect(self._on_mdf_error_for_report)
        self._report_mdf_loader.start()

    def _on_mdf_loaded_for_report(self, mdf, sig_names, sig_data):
        try:
            # Collect configuration utilizing the asynchronously loaded mdf
            config = self._collect_report_config(mdf_source=mdf)
            
            # Start report generation background worker
            if hasattr(self, '_report_worker') and self._report_worker is not None:
                self._report_worker.quit()
                self._report_worker.wait()
            
            self._report_worker = ReportGeneratorWorker(config, "report_preview.png", dpi=200)
            self._report_worker.finished.connect(self._on_preview_report_done)
            self._report_worker.error.connect(self._on_preview_report_error)
            self._report_worker.start()
            self.lbl_preview.setText("Generating preview...")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to collect report config:\n{str(e)}")
            import traceback
            traceback.print_exc()
            self._emit_not_busy()

        self._mark_modified()

    def _on_marks_toggled(self, checked):
        """Update Marks toggle button style and icon."""
        if checked:
            self.btn_marks.setIcon(QIcon(resource_path("assets/icons/file_json_20dp_000000_FILL1_wght400_GRAD0_opsz20.png")))
            self.btn_marks.setStyleSheet(f"background-color: {IDIADA_ORANGE}; color: black; font-weight: bold; border-top-right-radius: 4px; border-bottom-right-radius: 4px; border-left: 1px solid #555; padding: 0 15px;")
        else:
            self.btn_marks.setIcon(QIcon(resource_path("assets/icons/file_json_20dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png")))
            self.btn_marks.setStyleSheet("padding: 0 15px; font-weight: normal;")
        self._mark_modified()


    def _on_category_tab_clicked(self, category, animate=True):
        """Update tab selection with sliding animation and switch view."""
        target_btn = self.category_tab_buttons.get(category)
        if not target_btn:
            return

        for cat, btn in self.category_tab_buttons.items():
            btn.blockSignals(True)
            if cat == category:
                btn.setChecked(True)
                btn.setStyleSheet("""
                    QPushButton {
                        color: #FFFFFF;
                        border: none;
                        font-weight: 600;
                        background-color: transparent;
                        padding: 6px 10px;
                        font-size: 11px;
                    }
                """)
            else:
                btn.setChecked(False)
                btn.setStyleSheet("""
                    QPushButton {
                        color: #777;
                        border: none;
                        background: transparent;
                        font-weight: normal;
                        padding: 6px 10px;
                        font-size: 11px;
                    }
                """)
            btn.blockSignals(False)
        
        self._update_indicator_geometry(animate)
        
        idx = self.active_categories.index(category)
        self.category_stack.setCurrentIndex(idx)
        
        protocol_name = self.combo_protocol.currentText()
        self._last_selected_tabs[protocol_name] = category

    def _mark_modified(self):
        """Highlight Save button if there are unsaved changes."""
        if not self._config_modified:
            self._config_modified = True
            if self.save_btn_anim.state() != QVariantAnimation.Running:
                self.save_btn_anim.start()

    def _update_save_btn_style(self, color):
        """Update save button stylesheet with a pulsing border color."""
        self.btn_save_config.setStyleSheet(f"padding: 0 10px; font-weight: normal; border: 2px solid {color.name()};")

    def _on_mdf_error_for_report(self, e_msg):
        QMessageBox.critical(self, "Load Error", f"Failed to load dropped MF4:\n{e_msg}")
        self.lbl_preview.setText("Drag Report Item from Source Tree here")
        self._emit_not_busy()

    def _emit_not_busy(self):
        parent = self._get_parent_analysis_widget()
        if parent and hasattr(parent, 'busy_changed'):
            parent.busy_changed.emit(False)

    def _on_preview_report_done(self, output_path):
        """Callback when preview report is generated."""
        # Load and display in scroll area
        pixmap = QPixmap(output_path)
        if not pixmap.isNull():
            available_width = self.scroll_preview.width() - 30
            if pixmap.width() > available_width:
                pixmap = pixmap.scaledToWidth(available_width, Qt.SmoothTransformation)
            self.lbl_preview.setPixmap(pixmap)
            self.lbl_preview.setStyleSheet("")
        
        parent = self._get_parent_analysis_widget()
        if parent:
            parent._notify("Preview report generated successfully.", "success")
        
        self._emit_not_busy()

    def _on_preview_report_error(self, error_msg):
        """Callback when preview report generation fails."""
        QMessageBox.critical(self, "Error", f"Failed to generate report:\n{error_msg}")
        self._emit_not_busy()
    
    def _collect_report_config(self, mdf_source=None) -> dict:
        """Collect report configuration from UI elements."""
        from datetime import datetime
        
        # Determine protocol
        protocol = self.combo_protocol.currentText()
        if protocol == "GSR ADDW":
            # Keep "GSR ADDW" for title logic but also support regulatory code
            pass
        
        # Get OEM and vehicle from parent AnalysisWidget
        oem_name = ""
        vehicle = ""
        engineer = ""
        analyst = ""
        track = ""
        
        parent = self.parent()
        while parent:
            if hasattr(parent, 'combo_oem'):
                oem_name = parent.combo_oem.currentText() if parent.combo_oem else ""
            if hasattr(parent, 'txt_vehicle'):
                vehicle = parent.txt_vehicle.text() if parent.txt_vehicle else ""
            if hasattr(parent, 'txt_engineer'):
                engineer = parent.txt_engineer.text() if parent.txt_engineer else ""
            if hasattr(parent, 'txt_analyst'):
                analyst = parent.txt_analyst.text() if parent.txt_analyst else ""
            if hasattr(parent, 'combo_track'):
                track = parent.combo_track.currentText() if parent.combo_track else ""
            if hasattr(parent, 'combo_oem'):
                break
            parent = parent.parent()
        
        # Collect signals data from all categories
        signals = {}
        
        current_mdf = mdf_source if mdf_source else self.mdf
        
        # Determine target category from filename
        target_category = None
        if current_mdf and hasattr(current_mdf, 'name'):
            target_category = self._determine_category_from_filename(str(current_mdf.name))
            
        pass_signal_name = None
        if hasattr(self, 'pass_criteria_tables') and target_category:
            pass_table = self.pass_criteria_tables.get(target_category)
            if pass_table and pass_table.rowCount() > 0:
                if hasattr(pass_table, 'signal_combo'):
                    pass_sig = pass_table.signal_combo.currentText()
                else:
                    signal_combo = pass_table.cellWidget(0, 0)
                    pass_sig = signal_combo.currentText() if hasattr(signal_combo, 'currentText') else ""
                if pass_sig and pass_sig != "-- Select Signal --":
                    pass_signal_name = pass_sig

        if current_mdf:
            for category, table in self.category_tables.items():
                # If we determined a category from filename, only use that one
                if target_category and category != target_category:
                    continue
                
                for row in range(table.rowCount()):
                    chk = table.item(row, 0)
                    name_item = table.item(row, 1)
                    if chk and name_item and chk.checkState() == Qt.Checked:
                        sig_name = name_item.text()
                        try:
                            sig = current_mdf.get(sig_name)
                            
                            # Get operator
                            op_widget = table.cellWidget(row, 2)
                            operator = op_widget.currentText() if isinstance(op_widget, QComboBox) else "None"
                            
                            # Get value
                            val_widget = table.cellWidget(row, 3)
                            value = None
                            if isinstance(val_widget, QDoubleSpinBox):
                                value = val_widget.value()
                            elif isinstance(val_widget, QComboBox):
                                value = val_widget.currentText()
                            elif isinstance(val_widget, QLineEdit):
                                try:
                                    value = float(val_widget.text()) if val_widget.text() else None
                                except:
                                    value = val_widget.text()
                            
                            # Get alias
                            alias_item = table.item(row, 4)
                            alias = alias_item.text() if alias_item else sig_name
                            
                            # Override for SoundPressure
                            if sig_name == "SoundPressure":
                                operator = ">="
                                
                            signals[sig_name] = {
                                'timestamps': list(sig.timestamps),
                                'samples': list(sig.samples),
                                'threshold': value,
                                'operator': operator,
                                'unit': getattr(sig, 'unit', 'Value'),
                                'category': category,
                                'alias': alias
                            }
                        except Exception as e:
                            print(f"Error getting signal {sig_name}: {e}")
                            # Attempt to use sig_name manually for sound pressure if missing
                            if sig_name == "SoundPressure":
                                pass
        
        t_event = "No warn"
        
        filename = ""
        if current_mdf and hasattr(current_mdf, 'name'):
            filename = os.path.basename(str(current_mdf.name))

        # Load driver marks (always use marks if file exists)
        driver_marks = []
        if filename:
            parent_scope = self._get_parent_analysis_widget()
            if parent_scope:
                # Use standard logic from AnalysisWidget
                driver_marks = parent_scope._get_marks_for_mf4(str(current_mdf.name)) or []
            else:
                print("WARNING: parent_scope not found for marks loading in preview")
        # Calculate relative path for the report banner
        relative_path = filename
        parent_scope = self._get_parent_analysis_widget()
        if parent_scope:
            source_root = getattr(parent_scope, '_current_project_source', None)
            if source_root:
                try:
                    # Find the 'Pxx' or 'Exx' participant folder level
                    abs_mf4 = os.path.abspath(str(current_mdf.name))
                    parts = abs_mf4.split(os.sep)
                    # Look for a part that looks like a participant ID
                    for i, part in enumerate(parts):
                        if re.match(r'^[PE]\d+', part, re.IGNORECASE):
                            relative_path = os.sep.join(parts[i:])
                            break
                except Exception:
                    pass

        # Evaluate metrics using unified computation block to extract tgaze, t_event, times and pass_signal_name
        if parent_scope:
            metrics = parent_scope._compute_times_and_events(signals, target_category, driver_marks, parent_scope=parent_scope)
        else:
            # Fallback or error handling if parent is not found
            print("WARNING: parent_scope not found in LogicTabWidget._collect_report_config")
            metrics = {
                'tgaze': 6.0,
                't_event': "N/A",
                't_event_color': "gray",
                'signal_times': {},
                'pass_signal_name': pass_signal_name
            }

        config = {
            'oem_name': oem_name,
            'vehicle': vehicle,
            'protocol': protocol,
            'driver_marks': driver_marks,
            'engineer': engineer,
            'analyst': analyst,
            'track': track,
            'test_date': datetime.now(),
            'signals': signals,
            'camera_image_path': None,
            'filename': filename,
            'tgaze': metrics['tgaze'],
            't_event': metrics['t_event'],
            't_event_color': metrics['t_event_color'],
            'signal_times': metrics['signal_times'],
            'pass_signal_name': metrics['pass_signal_name'],
            'mask': metrics.get('mask', 6.0),
            'audio_params': {
                'min_freq': getattr(self, 'audio_min_freq', 0),
                'max_freq': getattr(self, 'audio_max_freq', 0),
                'threshold': getattr(self, 'audio_threshold', 0)
            },
            'gauge_rules_path': self.active_gauge_rules_path,
            'gsr_image_mapping': self.gsr_image_mapping,
            'relative_path': relative_path
        }

        # Resolve camera_image_path for GSR if protocol is active
        if protocol == "2023/2590" or protocol == "GSR ADDW":
            config['camera_image_path'] = self._resolve_gsr_image_path(filename)

        return config

    def _resolve_gsr_image_path(self, filename: str) -> Optional[str]:
        """Resolves the absolute path for a GSR case image from mapping."""
        import re
        match = re.search(r'(ADDW\d+)', filename, re.IGNORECASE)
        if match:
            case_key = match.group(1).upper()
            image_name = self.gsr_image_mapping.get(case_key)
            if image_name and image_name != "None":
                # Check local first (user custom)
                local_path = resource_path(os.path.join("assets/gsr/local", image_name))
                if os.path.exists(local_path):
                    return local_path
                # Check core
                core_path = resource_path(os.path.join("assets/gsr", image_name))
                if os.path.exists(core_path):
                    return core_path
        return None

    def update_oem_logo(self, oem_name):
        """Compatibility stub - logos are handled in matplotlib report."""
        pass

    def update_vehicle_text(self, text):
        """Compatibility stub - vehicle text is handled in matplotlib report."""
        pass

    def update_audio_params(self, min_f, max_f, thresh):
        """Store audio filter parameters for use in report generation."""
        self.audio_min_freq = min_f
        self.audio_max_freq = max_f
        self.audio_threshold = thresh

    def _get_parent_analysis_widget(self):
        parent = self.parent()
        while parent:
            if hasattr(parent, 'combo_oem') and hasattr(parent, 'spin_min_freq'):
                return parent
            parent = parent.parent()
        return None

    def open_gsr_images_dialog(self):
        dialog = GSRImagesEditorDialog(self.gsr_image_mapping, self)
        if dialog.exec():
            self.gsr_image_mapping = dialog.get_mapping()
            self._mark_modified()

    def save_config(self):
        """Save protocol, category tables, micro, and report settings to file."""
        import json
        from PySide6.QtWidgets import QFileDialog
        file_path, _ = QFileDialog.getSaveFileName(
            self,
            "Save Config",
            "",
            "Config Files (*.json)"
        )
        if not file_path:
            return
        if not file_path.lower().endswith('.json'):
            file_path += '.json'
        try:
            config = self._collect_logic_config()
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(config, f, indent=2)
            
            # Reset modified state and stop animation
            self._config_modified = False
            self.save_btn_anim.stop()
            self.btn_save_config.setStyleSheet("padding: 0 10px; font-weight: normal;")
            
            # Show notification instead of messagebox
            parent = self._get_parent_analysis_widget()
            if parent:
                parent._notify("Config saved successfully.", "success")
        except Exception as e:
            QMessageBox.critical(self, "Config", f"Failed to save config:\n{e}")

    def load_config(self):
        """Load protocol, category tables, micro, and report settings from file."""
        import json
        from PySide6.QtWidgets import QFileDialog
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "Load Config",
            "",
            "Config Files (*.json)"
        )
        if not file_path:
            return
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                config = json.load(f)

            # If no MF4 is loaded, try to auto-load from project source first
            any_loaded = any(btn.text().strip() for btn in self.load_mf4_buttons.values())
            if not any_loaded:
                self.auto_load_project_mdfs()

            self._apply_logic_config(config)
            
            # Reset modified state and stop animation
            self._config_modified = False
            self.save_btn_anim.stop()
            self.btn_save_config.setStyleSheet("padding: 0 10px; font-weight: normal;")
            
            # Show notification instead of messagebox
            parent = self._get_parent_analysis_widget()
            if parent:
                parent._notify("Config loaded successfully.", "success")
        except Exception as e:
            QMessageBox.critical(self, "Config", f"Failed to load config:\n{e}")

    def _collect_logic_config(self) -> dict:
        """Collect Logic tab config, including micro, report settings and PASS Criteria."""
        config = {
            'version': 1,
            'protocol': "Euro NCAP",
            'categories': {},
            'pass_criteria': {},
            'micro': {},
            'report': {}
        }

        for category, table in self.category_tables.items():
            rows = []
            for row in range(table.rowCount()):
                name_item = table.item(row, 1)
                if not name_item:
                    continue
                sig_name = name_item.text()
                chk = table.item(row, 0)
                checked = chk.checkState() == Qt.Checked if chk else False

                op_widget = table.cellWidget(row, 2)
                operator = op_widget.currentText() if isinstance(op_widget, QComboBox) else "None"

                val_widget = table.cellWidget(row, 3)
                value = None
                if isinstance(val_widget, QDoubleSpinBox):
                    value = val_widget.value()
                elif isinstance(val_widget, QComboBox):
                    value = val_widget.currentText()
                elif isinstance(val_widget, QLineEdit):
                    value = val_widget.text()

                alias_item = table.item(row, 4)
                alias = alias_item.text() if alias_item else ""

                rows.append({
                    'signal': sig_name,
                    'checked': checked,
                    'operator': operator,
                    'value': value,
                    'alias': alias
                })
            config['categories'][category] = rows
        
        # Collect PASS Criteria
        for category, pass_table in self.pass_criteria_tables.items():
            if pass_table.rowCount() > 0:
                signal_combo = pass_table.cellWidget(0, 0)
                op1_widget = pass_table.cellWidget(0, 1)
                value1_widget = pass_table.cellWidget(0, 2)
                op2_widget = pass_table.cellWidget(0, 3)
                value2_widget = pass_table.cellWidget(0, 4)
                mask_widget = pass_table.cellWidget(0, 5)
                
                signal_name = signal_combo.currentText() if isinstance(signal_combo, QComboBox) else ""
                value1 = value1_widget.value() if isinstance(value1_widget, QDoubleSpinBox) else 0
                op1 = op1_widget.currentText() if isinstance(op1_widget, QComboBox) else ""
                value2 = value2_widget.value() if isinstance(value2_widget, QDoubleSpinBox) else 0
                op2 = op2_widget.currentText() if isinstance(op2_widget, QComboBox) else ""
                mask_val = mask_widget.value() if isinstance(mask_widget, QDoubleSpinBox) else 6.0
                
                config['pass_criteria'][category] = {
                    'signal': signal_name,
                    'value1': value1,
                    'operator1': op1,
                    'value2': value2,
                    'operator2': op2,
                    'mask': mask_val
                }

        parent = self._get_parent_analysis_widget()
        if parent:
            config['micro'] = {
                'min_freq': parent.spin_min_freq.value() if hasattr(parent, 'spin_min_freq') else 0,
                'max_freq': parent.spin_max_freq.value() if hasattr(parent, 'spin_max_freq') else 0,
                'threshold': parent.spin_threshold.value() if hasattr(parent, 'spin_threshold') else 0
            }
            config['report'] = {
                'oem': parent.combo_oem.currentText() if hasattr(parent, 'combo_oem') else "",
                'vehicle': parent.txt_vehicle.text() if hasattr(parent, 'txt_vehicle') else "",
                'engineer': parent.txt_engineer.text() if hasattr(parent, 'txt_engineer') else "",
                'analyst': parent.txt_analyst.text() if hasattr(parent, 'txt_analyst') else "",
                'track': parent.combo_track.currentText() if hasattr(parent, 'combo_track') else "",
                'ncap': parent.toggle_ncap.isChecked() if hasattr(parent, 'toggle_ncap') else False
            }

        if hasattr(self, 'active_gauge_rules_path') and self.active_gauge_rules_path:
            config['gauge_rules_path'] = self.active_gauge_rules_path

        return config

    def _apply_logic_config(self, config: dict):
        """Apply Logic tab config, including micro and report settings."""
        if not isinstance(config, dict):
            return

        # Protocol removal handling
        pass

        categories = config.get('categories', {})
        
        # Handle backward compatibility: older configs have "Microsleep & Sleep" which is now split.
        if "Microsleep & Sleep" in categories and "Microsleep" not in categories and "Sleep" not in categories:
            # Duplicate the shared settings to both new split categories
            categories["Microsleep"] = categories["Microsleep & Sleep"]
            categories["Sleep"] = categories["Microsleep & Sleep"]
            
        for category, rows in categories.items():
            table = self.category_tables.get(category)
            if not table:
                continue
            table.blockSignals(True)
            for entry in rows:
                sig_name = entry.get('signal', '')
                if not sig_name:
                    continue
                row_idx = self._ensure_row_for_signal(table, sig_name)

                chk_item = table.item(row_idx, 0)
                if chk_item:
                    chk_item.setCheckState(Qt.Checked if entry.get('checked', False) else Qt.Unchecked)

                op_widget = table.cellWidget(row_idx, 2)
                if isinstance(op_widget, QComboBox):
                    op_widget.setCurrentText(entry.get('operator', 'None') or 'None')

                val_widget = table.cellWidget(row_idx, 3)
                value = entry.get('value', None)
                if isinstance(val_widget, QDoubleSpinBox) and value is not None:
                    try:
                        val_widget.setValue(float(value))
                    except Exception:
                        pass
                elif isinstance(val_widget, QComboBox):
                    if value is not None:
                        text_val = str(value)
                        if val_widget.findText(text_val) < 0:
                            val_widget.addItem(text_val)
                        val_widget.setCurrentText(text_val)
                elif isinstance(val_widget, QLineEdit):
                    if value is not None:
                        val_widget.setText(str(value))

                alias_item = table.item(row_idx, 4)
                if alias_item is None:
                    alias_item = QTableWidgetItem("")
                    table.setItem(row_idx, 4, alias_item)
                alias_item.setText(entry.get('alias', '') or sig_name)
            table.blockSignals(False)
            self._adjust_table_height(table)
        
        # Apply PASS Criteria
        pass_criteria = config.get('pass_criteria', {})
        
        if "Microsleep & Sleep" in pass_criteria and "Microsleep" not in pass_criteria and "Sleep" not in pass_criteria:
            pass_criteria["Microsleep"] = pass_criteria["Microsleep & Sleep"]
            pass_criteria["Sleep"] = pass_criteria["Microsleep & Sleep"]
            
        for category, pass_data in pass_criteria.items():
            pass_table = self.pass_criteria_tables.get(category)
            if not pass_table or pass_table.rowCount() == 0:
                continue
            
            signal_combo = pass_table.cellWidget(0, 0)
            op1_widget = pass_table.cellWidget(0, 1)
            value1_widget = pass_table.cellWidget(0, 2)
            op2_widget = pass_table.cellWidget(0, 3)
            value2_widget = pass_table.cellWidget(0, 4)
            
            if isinstance(signal_combo, QComboBox) and pass_data.get('signal'):
                signal_text = pass_data.get('signal', '')
                if signal_combo.findText(signal_text) >= 0:
                    signal_combo.setCurrentText(signal_text)
            
            if isinstance(op1_widget, QComboBox) and pass_data.get('operator1'):
                op1_widget.setCurrentText(pass_data.get('operator1'))
            
            if isinstance(value1_widget, QDoubleSpinBox):
                value1_widget.setValue(pass_data.get('value1', 0))
            
            if isinstance(op2_widget, QComboBox) and pass_data.get('operator2'):
                op2_widget.setCurrentText(pass_data.get('operator2'))
            
            if isinstance(value2_widget, QDoubleSpinBox):
                value2_widget.setValue(pass_data.get('value2', 0))
                
            mask_widget = pass_table.cellWidget(0, 5)
            if isinstance(mask_widget, QDoubleSpinBox):
                mask_widget.setValue(pass_data.get('mask', 6.0))

        parent = self._get_parent_analysis_widget()
        if parent:
            micro = config.get('micro', {})
            report = config.get('report', {})

            try:
                parent.spin_min_freq.setValue(int(micro.get('min_freq', parent.spin_min_freq.value())))
                parent.spin_max_freq.setValue(int(micro.get('max_freq', parent.spin_max_freq.value())))
                parent.spin_threshold.setValue(float(micro.get('threshold', parent.spin_threshold.value())))
                parent.on_audio_params_changed()
            except Exception:
                pass

            try:
                if 'oem' in report and hasattr(parent, 'combo_oem'):
                    parent.combo_oem.setCurrentText(report.get('oem', ''))
                if 'vehicle' in report and hasattr(parent, 'txt_vehicle'):
                    parent.txt_vehicle.setText(report.get('vehicle', ''))
                if 'engineer' in report and hasattr(parent, 'txt_engineer'):
                    parent.txt_engineer.setText(report.get('engineer', ''))
                if 'analyst' in report and hasattr(parent, 'txt_analyst'):
                    parent.txt_analyst.setText(report.get('analyst', ''))
                if 'track' in report and hasattr(parent, 'combo_track'):
                    parent.combo_track.setCurrentText(report.get('track', ''))
                if 'ncap' in report and hasattr(parent, 'toggle_ncap'):
                    parent.toggle_ncap.blockSignals(True)
                    parent.toggle_ncap.setChecked(bool(report.get('ncap', False)))
                    parent.toggle_ncap.blockSignals(False)
            except Exception:
                pass

        if 'gauge_rules_path' in config and config['gauge_rules_path']:
            gauge_path = config['gauge_rules_path']
            if os.path.exists(gauge_path):
                self.active_gauge_rules_path = gauge_path
                self.known_gauge_rules.append(gauge_path)

        # Update category counters
        for category, table in self.category_tables.items():
            checked_count = 0
            for row in range(table.rowCount()):
                chk = table.item(row, 0)
                if chk and chk.checkState() == Qt.Checked:
                    checked_count += 1
            if hasattr(self, 'category_counters') and category in self.category_counters:
                lbl = self.category_counters[category]
                lbl.setText(f"{checked_count}/5 signals selected")
                if checked_count == 5:
                    lbl.setStyleSheet("color: #ff3333; font-weight: bold; font-size: 11px;")
                else:
                    lbl.setStyleSheet("color: #ff9800; font-weight: bold; font-size: 11px;")

    def _ensure_row_for_signal(self, table: QTableWidget, sig_name: str) -> int:
        for row in range(table.rowCount()):
            name_item = table.item(row, 1)
            if name_item and name_item.text() == sig_name:
                return row

        row = table.rowCount()
        table.insertRow(row)

        chk = QTableWidgetItem()
        chk.setFlags(Qt.ItemIsUserCheckable | Qt.ItemIsEnabled)
        chk.setCheckState(Qt.Unchecked)
        table.setItem(row, 0, chk)

        name_item = QTableWidgetItem(sig_name)
        name_item.setFlags(Qt.ItemIsEnabled)
        table.setItem(row, 1, name_item)

        combo_op = QComboBox()
        combo_op.setStyleSheet("""
            QComboBox {
                background-color: #3a3a3a;
                color: #ddd;
                border: 1px solid #555;
                border-radius: 3px;
                padding: 2px;
                font-size: 9pt;
            }
            QComboBox::drop-down {
                border: none;
                background-color: #3a3a3a;
            }
            QComboBox::down-arrow {
                image: none;
                color: #ddd;
            }
            QComboBox QAbstractItemView {
                background-color: #3a3a3a;
                color: #ddd;
                selection-background-color: #ff9800;
                selection-color: #000;
            }
        """)
        combo_op.addItems(["None", ">", "<", ">=", "<=", "==", "!="])
        combo_op.setCurrentText("None")

        if sig_name == "SoundPressure":
            # Display empty cells for SoundPressure's Operator and Value
            table.setCellWidget(row, 2, QLabel(""))
            table.setCellWidget(row, 3, QLabel(""))
        else:
            table.setCellWidget(row, 2, combo_op)
            value_widget = self._create_value_widget(sig_name)
            if value_widget:
                if isinstance(value_widget, QComboBox):
                    value_widget.setStyleSheet("""
                        QComboBox {
                            background-color: #3a3a3a;
                            color: #ddd;
                            border: 1px solid #555;
                            border-radius: 3px;
                            padding: 2px;
                            font-size: 9pt;
                        }
                        QComboBox::drop-down {
                            border: none;
                            background-color: #3a3a3a;
                        }
                        QComboBox QAbstractItemView {
                            background-color: #3a3a3a;
                            color: #ddd;
                            selection-background-color: #ff9800;
                            selection-color: #000;
                        }
                    """)
                elif isinstance(value_widget, QDoubleSpinBox):
                    value_widget.setStyleSheet(LOGIC_INPUT_STYLE)
                elif isinstance(value_widget, QLineEdit):
                    value_widget.setStyleSheet("""
                        QLineEdit {
                            background-color: #3a3a3a;
                            color: #ddd;
                            border: 1px solid #555;
                            border-radius: 3px;
                            padding: 2px;
                            font-size: 9pt;
                        }
                    """)
            table.setCellWidget(row, 3, value_widget)
            # Connect signals to mark modified
            combo_op.currentTextChanged.connect(self._mark_modified)
            if isinstance(value_widget, QLineEdit):
                value_widget.textChanged.connect(self._mark_modified)
            elif isinstance(value_widget, (QComboBox, QDoubleSpinBox)):
                if hasattr(value_widget, 'currentTextChanged'):
                    value_widget.currentTextChanged.connect(self._mark_modified)
                if hasattr(value_widget, 'valueChanged'):
                    value_widget.valueChanged.connect(self._mark_modified)

        alias_item = QTableWidgetItem(sig_name)
        table.setItem(row, 4, alias_item)

        return row

    def update_signals(self, signal_names, mdf):
        """Populate all category tables with signals from Signal Filter."""
        self.mdf = mdf
        
        # Update tables iteratively to preserve user inputs (aliases, operators, etc.)
        for category, table in self.category_tables.items():
            table.blockSignals(True)
            
            # 1. Remove rows for signals that are no longer selected
            for row in reversed(range(table.rowCount())):
                name_item = table.item(row, 1)
                if name_item and name_item.text() not in signal_names:
                    table.removeRow(row)
                    
            # 2. Add properties for newly selected signals
            for sig_name in signal_names:
                self._ensure_row_for_signal(table, sig_name)
                
            table.blockSignals(False)
            self._adjust_table_height(table)
        
        # Update PASS Criteria dropdowns with new signals
        self._update_pass_criteria_dropdowns()
    
    def update_mdf_only(self, mdf):
        """Update only the MDF reference without resetting table configurations."""
        self.mdf = mdf

    def _create_value_widget(self, sig_name):
        """Create appropriate widget for value column based on signal type."""
        if not self.mdf:
            return QLineEdit()
            
        try:
            sig = self.mdf.get(sig_name)
            if sig.samples.dtype.kind in 'SUa':
                combo = QComboBox()
                unique_vals = sorted(list(set([str(v) for v in sig.samples if v])))
                combo.addItems(unique_vals)  # Limit removed to show all options
                return combo
            else:
                spin = QDoubleSpinBox()
                spin.setRange(-999999, 999999)
                spin.setDecimals(2)
                return spin
        except:
            return QLineEdit()
    
    def get_selected_signals_by_category(self) -> dict:
        """Get list of selected signals organized by category."""
        signals_by_category = {}
        
        for category, table in self.category_tables.items():
            signals = []
            for row in range(table.rowCount()):
                chk = table.item(row, 0)
                name_item = table.item(row, 1)
                if chk and name_item and chk.checkState() == Qt.Checked:
                    signals.append(name_item.text())
            signals_by_category[category] = signals
        
        return signals_by_category
    
    def get_selected_signals(self):
        """Get all selected signals from all categories (for backward compatibility)."""
        all_signals = set()
        for signals in self.get_selected_signals_by_category().values():
            all_signals.update(signals)
        return list(all_signals)

    def get_alias_for_signal(self, signal_name: str) -> str:
        """Return alias for a signal if configured in any category table, else return the original name."""
        for category, table in self.category_tables.items():
            for row in range(table.rowCount()):
                name_item = table.item(row, 1)
                if name_item and name_item.text() == signal_name:
                    alias_item = table.item(row, 5)
                    if alias_item and alias_item.text().strip():
                        return alias_item.text().strip()
        return signal_name
    
    def _determine_category_from_filename(self, filename):
        """
        Determine the distraction category based on filename pattern.
        Returns the category name or None if pattern doesn't match.
        """
        import re
        import os
        
        # Extract base name without extension
        filename_str = str(filename)
        basename = os.path.splitext(os.path.basename(filename_str))[0]
        
        # Check for ADDW
        if "ADDW" in basename.upper():
            path_str = filename_str.replace('\\', '/').lower()
            if "high speed" in path_str:
                return "High Speed"
            elif "low speed" in path_str:
                return "Low Speed"
            else:
                return "High Speed" # Default fallback
                
        # Try to match D + number pattern
        match_d = re.match(r'^D(\d+)', basename)
        if match_d:
            num = int(match_d.group(1))
            if 1 <= num <= 9:
                return "Long Distraction (NDT)"
            elif 10 <= num <= 15:
                return "Long Distraction (DT)"
            elif (16 <= num <= 19) or num == 28 or (29 <= num <= 42):
                return "Short Distraction (NDT)"
            elif 20 <= num <= 27:
                return "Short Distraction (DT)"
        
        # Try to match F + number pattern
        match_f = re.match(r'^F(\d+)', basename)
        if match_f:
            num = int(match_f.group(1))
            if num == 1:
                return "Microsleep"
            elif num == 2:
                return "Sleep"
            elif num == 3:
                return "Drowsiness"
            elif num in [4, 5]:
                return "Unresponsive driver"
        
        return None
    def _adjust_table_height(self, table):
        """Adjust table height based on number of rows."""
        # Calculate height: header + rows + margins
        header_height = table.horizontalHeader().height()
        row_height = table.rowHeight(0) if table.rowCount() > 0 else 25
        total_height = header_height + (table.rowCount() * row_height) + 10
        
        # Set height with minimum and reasonable maximum
        total_height = max(80, min(total_height, 400))  # Min 80, Max 400
        table.setFixedHeight(total_height)

    def _init_pass_criteria_row(self, pass_criteria_table: QTableWidget, category: str):
        """Initialize PASS Criteria row with defaults."""
        # Get available signals from the corresponding category table
        category_table = self.category_tables.get(category)
        available_signals = []
        if category_table:
            for row in range(category_table.rowCount()):
                name_item = category_table.item(row, 1)
                if name_item:
                    available_signals.append(name_item.text())
        
        # Column 0: Signal dropdown
        signal_combo = QComboBox()
        signal_combo.setStyleSheet("""
            QComboBox {
                background-color: #333;
                color: #ddd;
                border: 1px solid #555;
                padding: 2px;
            }
            QComboBox::drop-down {
                border: none;
            }
            QComboBox QAbstractItemView {
                background-color: #333;
                color: #ddd;
                selection-background-color: #ff9800;
            }
        """)
        signal_combo.addItem("-- Select Signal --")  # Placeholder
        signal_combo.addItems(available_signals)
        signal_combo.currentTextChanged.connect(self._mark_modified)
        
        # Set SoundPressure as default if it exists and is in the list
        if "SoundPressure" in available_signals:
            signal_combo.setCurrentText("SoundPressure")
        
            signal_combo.setCurrentText("SoundPressure")
        
        pass_criteria_table.setCellWidget(0, 0, signal_combo)
        # Store signal combo for later updates
        # Store signal combo for later updates
        if not hasattr(pass_criteria_table, 'signal_combo'):
            pass_criteria_table.signal_combo = signal_combo
        
        # Column 1: Operator 1
        op1_combo = QComboBox()
        op1_combo.setStyleSheet("""
            QComboBox {
                background-color: #333;
                color: #ddd;
                border: 1px solid #555;
                padding: 2px;
            }
            QComboBox QAbstractItemView {
                background-color: #333;
                color: #ddd;
                selection-background-color: #ff9800;
            }
        """)
        op1_combo.addItems([">=", "<=", ">", "<", "==", "!="])
        op1_combo.currentTextChanged.connect(self._mark_modified)
        pass_criteria_table.setCellWidget(0, 1, op1_combo)
        
        # Column 2: Value 1 (QDoubleSpinBox)
        value1 = QDoubleSpinBox()
        value1.setStyleSheet(LOGIC_INPUT_STYLE)
        value1.setRange(-1000, 1000)
        value1.setDecimals(2)
        value1.setValue(0)
        value1.valueChanged.connect(self._mark_modified)
        pass_criteria_table.setCellWidget(0, 2, value1)
        
        # Column 3: Operator 2
        op2_combo = QComboBox()
        op2_combo.setStyleSheet("""
            QComboBox {
                background-color: #333;
                color: #ddd;
                border: 1px solid #555;
                padding: 2px;
            }
            QComboBox QAbstractItemView {
                background-color: #333;
                color: #ddd;
                selection-background-color: #ff9800;
            }
        """)
        op2_combo.addItems([">=", "<=", ">", "<", "==", "!="])
        op2_combo.currentTextChanged.connect(self._mark_modified)
        pass_criteria_table.setCellWidget(0, 3, op2_combo)
        
        # Column 4: Value 2 (QDoubleSpinBox)
        value2 = QDoubleSpinBox()
        value2.setStyleSheet(LOGIC_INPUT_STYLE)
        value2.setRange(-1000, 1000)
        value2.setDecimals(2)
        value2.setValue(0)
        value2.valueChanged.connect(self._mark_modified)
        pass_criteria_table.setCellWidget(0, 4, value2)
        
        # Column 5: Mask (QDoubleSpinBox)
        mask_val = QDoubleSpinBox()
        mask_val.setStyleSheet(LOGIC_INPUT_STYLE)
        mask_val.setRange(0, 100)
        mask_val.setDecimals(2)
        mask_val.setValue(6.0)
        mask_val.setToolTip("Ignored time at the beginning of the signal (seconds)")
        mask_val.valueChanged.connect(self._mark_modified)
        pass_criteria_table.setCellWidget(0, 5, mask_val)
        
        # Set defaults based on category
        defaults = {
            "Long Distraction": {"op1": ">=", "val1": 3, "op2": "<=", "val2": 4},
            "Short Distractions": {"op1": "<=", "val1": 10, "op2": "", "val2": 0},
            "Phone Use": {"op1": "<=", "val1": 10, "op2": "", "val2": 0},
            "Microsleep": {"op1": "<=", "val1": 3.5, "op2": "", "val2": 0},
            "Sleep": {"op1": "<=", "val1": 3.5, "op2": "", "val2": 0},
            "Drowsiness": {"op1": "", "val1": 0, "op2": "", "val2": 0},
            "Unresponsive driver": {"op1": ">=", "val1": 3, "op2": "", "val2": 0}
        }
        
        if category in defaults:
            default = defaults[category]
            if default["op1"]:
                op1_combo.setCurrentText(default["op1"])
                value1.setValue(default["val1"])
            if default["op2"]:
                op2_combo.setCurrentText(default["op2"])
                value2.setValue(default["val2"])

    def _update_pass_criteria_dropdowns(self):
        """Update PASS Criteria signal dropdowns after signals are loaded."""
        for category, pass_table in self.pass_criteria_tables.items():
            # Get signal combo from this category's PASS Criteria table
            if hasattr(pass_table, 'signal_combo'):
                signal_combo = pass_table.signal_combo
            else:
                signal_combo = pass_table.cellWidget(0, 0)
                
            if not isinstance(signal_combo, QComboBox):
                continue
            
            # Get available signals from this category's signal table
            category_table = self.category_tables.get(category)
            if not category_table:
                continue
            
            available_signals = []
            for row in range(category_table.rowCount()):
                name_item = category_table.item(row, 1)
                if name_item:
                    available_signals.append(name_item.text())
            
            # Block signals to prevent triggering events
            signal_combo.blockSignals(True)
            
            # Clear and repopulate combo
            signal_combo.clear()
            signal_combo.addItem("-- Select Signal --")
            signal_combo.addItems(available_signals)
            
            # Set SoundPressure as default if it exists
            if "SoundPressure" in available_signals:
                signal_combo.setCurrentText("SoundPressure")
            
            signal_combo.blockSignals(False)

    def _on_load_mf4_for_category(self, category):
        from PySide6.QtWidgets import QFileDialog, QMessageBox
        import os
        
        file_path, _ = QFileDialog.getOpenFileName(
            self, f"Select MF4 for {category}", "", "MDF Files (*.mf4)"
        )
        if not file_path:
            return

        parent = self._get_parent_analysis_widget()
        if parent:
            parent.busy_changed.emit(True)

        self.setCursor(Qt.WaitCursor)
        try:
            count = self._load_mf4_file_to_category(file_path, category)
            self.setCursor(Qt.ArrowCursor)
            if parent:
                parent._notify(f"Loaded {count} signals for {category}.", "success")
        except Exception as e:
            self.setCursor(Qt.ArrowCursor)
            QMessageBox.critical(self, "Error", f"Failed to load file:\n{str(e)}")
        finally:
            if parent:
                parent.busy_changed.emit(False)

    def _load_mf4_file_to_category(self, file_path, category) -> int:
        from asammdf import MDF
        temp_mdf = MDF(file_path)
        sigs = []
        for ch in temp_mdf.iter_channels():
            name = ch.name
            try:
                if ch.samples is None: continue
                if getattr(ch.samples.dtype, 'kind', 'f') in 'fiudbSUa':
                    sigs.append(name)
            except Exception:
                try:
                    float(ch.samples[0])
                    sigs.append(name)
                except Exception:
                    pass
        signal_names = sorted(list(set(sigs)))
        
        table = self.category_tables.get(category)
        if table:
            # Update button text with filename
            if category in self.load_mf4_buttons:
                fname = os.path.basename(file_path)
                btn = self.load_mf4_buttons[category]
                btn.setText(f" {fname}")
                btn.setToolTip(f"File: {file_path}")
                # Adjust style to show it's loaded (White edge, regular text)
                btn.setStyleSheet("""
                    QPushButton { background-color: #333; border: 1px solid white; border-radius: 3px; padding-left: 5px; padding-right: 5px; color: white; font-size: 10px; font-weight: normal; }
                    QPushButton:hover { background-color: #444; }
                """)

            table.blockSignals(True)
            # Clear existing rows to ensure fresh load/substitution
            table.setRowCount(0)
            
            prev_mdf = self.mdf
            self.mdf = temp_mdf
            
            for sig_name in signal_names:
                self._ensure_row_for_signal(table, sig_name)
                
            self.mdf = prev_mdf
            
            table.blockSignals(False)
            self._adjust_table_height(table)
            self._update_pass_criteria_dropdowns()
            
        return len(signal_names)

    def auto_load_project_mdfs(self):
        """Automatically find and load the corresponding MF4 for each category based on the first participant in the tree."""
        from PySide6.QtWidgets import QMessageBox
        
        parent = self._get_parent_analysis_widget()
        if not parent or not parent.txt_source.text():
            if parent:
                parent._notify("No project source selected. Please select a folder first.", "error")
            return
            
        root = parent.tree_participants.invisibleRootItem()
        p_item = root.child(0) if root.childCount() > 0 else None
        
        if not p_item:
            if parent:
                parent._notify("No participants loaded in the Source tree.", "warning")
            return

        # Gather all MF4 paths for this participant
        mdf_files = []
        def _get_mf4s(item):
            path = item.data(0, Qt.UserRole)
            if path and isinstance(path, str) and path.lower().endswith('.mf4') and 'tracking' not in path.lower():
                mdf_files.append(path)
            for i in range(item.childCount()):
                _get_mf4s(item.child(i))
                
        _get_mf4s(p_item)
        
        if not mdf_files:
            if parent:
                parent._notify(f"No valid MF4 cases found for the first participant ({p_item.text(0)}).", "warning")
            return
            
        if parent:
            parent.busy_changed.emit(True)

        try:
            self.setCursor(Qt.WaitCursor)
                        
            loaded_count = 0
            for category in self.category_tables.keys():
                matching_file = None
                for fpath in mdf_files:
                    cat = self._determine_category_from_filename(fpath)
                    if cat == category:
                        matching_file = fpath
                        break
                
                if matching_file:
                    try:
                        self._load_mf4_file_to_category(matching_file, category)
                        loaded_count += 1
                    except Exception as e:
                        print(f"Error auto-loading {matching_file} for {category}: {e}")
                    
        finally:
            self.setCursor(Qt.ArrowCursor)
            if parent:
                parent.busy_changed.emit(False)

        if loaded_count > 0:
            if parent:
                parent._notify(f"Auto-loaded MF4 data for {loaded_count} categories.", "success")
        else:
            if parent:
                parent._notify("No matching MF4 files found in the source directory.", "warning")

class VideoFrameWidget(QWidget):
    """
    Replaces QVideoWidget using QVideoSink + QPainter.
    Being a pure QWidget, it doesn't create a native HWND on Windows,
    resolving the z-order issue with overlays.
    """
    def __init__(self, parent=None):
        super().__init__(parent)
        self._frame = QVideoFrame()
        self._sink = QVideoSink(self)
        self._sink.videoFrameChanged.connect(self._on_frame)
        self.setMinimumSize(640, 360)
        self.setAttribute(Qt.WA_OpaquePaintEvent)
        self.setStyleSheet("background-color: black;")

    @property
    def videoSink(self) -> QVideoSink:
        return self._sink

    def _on_frame(self, frame: QVideoFrame):
        self._frame = frame
        self.update()

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.fillRect(self.rect(), Qt.black)

        if self._frame.isValid():
            img = self._frame.toImage()
            if not img.isNull():
                scaled = img.scaled(
                    self.size(),
                    Qt.KeepAspectRatio,
                    Qt.SmoothTransformation
                )
                x = (self.width() - scaled.width()) // 2
                y = (self.height() - scaled.height()) // 2
                painter.drawImage(x, y, scaled)
        painter.end()

    def sizeHint(self) -> QSize:
        return QSize(640, 360)

class TimeSelectorWidget(QWidget):
    def _get_marks_path(self):
        parent = self.parent()
        while parent:
            if hasattr(parent, '_get_marks_path'):
                context_file = getattr(self, 'current_tracking_path', None)
                return parent._get_marks_path(context_file)
            parent = parent.parent()
        # Fallback to app root
        import os
        return os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '..', 'marks.json')

    def _get_logical_key(self):
        import os
        if not self.current_tracking_path:
            return None
        p = os.path.normpath(self.current_tracking_path)
        parts = p.split(os.sep)
        if len(parts) < 3:
            return os.path.basename(p)
        return '/'.join(parts[-3:])

    def _find_matching_mark_key(self, marks, exact_key):
        if exact_key in marks:
            return exact_key

        import os
        basename = os.path.basename(self.current_tracking_path).lower() if self.current_tracking_path else ''
        if basename:
            alt_basename = basename.replace('_tracking.mf4', '.mf4')
            candidates = [k for k in marks if k.lower().endswith(basename) or k.lower().endswith(alt_basename)]
            if len(candidates) == 1:
                return candidates[0]

            # Try matching last two segments if exact key fails
            suffix = '/'.join(exact_key.split('/')[-2:]).lower()
            candidates = [k for k in marks if k.lower().endswith(suffix)]
            if len(candidates) == 1:
                return candidates[0]

        return None

    def _load_marks(self):
        import json
        marks_path = self._get_marks_path()
        if not os.path.exists(marks_path):
            return {}
        try:
            with open(marks_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            return {}

    def _save_marks(self, marks):
        import json
        marks_path = self._get_marks_path()
        try:
            with open(marks_path, 'w', encoding='utf-8') as f:
                json.dump(marks, f, indent=2)
            self.marks_saved.emit()
        except Exception:
            pass

    def _restore_marks(self):
        key = self._get_logical_key()
        if not key:
            return
        marks = self._load_marks()
        actual_key = self._find_matching_mark_key(marks, key)
        if not actual_key:
            return
        times = sorted(marks.get(actual_key, []))
        if times:
            self._clear_markers(persist=False)
            pen = pg.mkPen('#ff9800', width=2)
            hover_pen = pg.mkPen('#ffb74d', width=3)
            for t in times:
                line1 = pg.InfiniteLine(pos=t, angle=90, movable=True, pen=pen, hoverPen=hover_pen)
                line2 = pg.InfiniteLine(pos=t, angle=90, movable=True, pen=pen, hoverPen=hover_pen)
                
                # Ensure cursor changes correctly over lines
                line1.setCursor(Qt.SizeHorCursor)
                line2.setCursor(Qt.SizeHorCursor)
                
                # Add context menu for deletion
                line1.sigClicked.connect(self._on_line_clicked)
                line2.sigClicked.connect(self._on_line_clicked)
                
                self.plot_top.addItem(line1, ignoreBounds=True)
                self.plot_bottom.addItem(line2, ignoreBounds=True)
                self.markers.append((line1, line2, t))
                line1.sigPositionChanged.connect(lambda *args, l=line1: self._on_marker_moved(l))
                line2.sigPositionChanged.connect(lambda *args, l=line2: self._on_marker_moved(l))
            
            self.selected_marker_index = -1
            self._ensure_marker_visibility()

    def _ensure_marker_visibility(self):
        try:
            if not self.markers:
                return
            x_vals = [m[2] for m in self.markers]
            x_min, x_max = min(x_vals), max(x_vals)
            vb = self.plot_top.getPlotItem().vb
            current_range = vb.viewRange()[0]
            if x_min < current_range[0] or x_max > current_range[1]:
                vb.setXRange(min(x_min, current_range[0]), max(x_max, current_range[1]), padding=0.02)
        except Exception:
            pass

    def _restore_shaded_regions(self):
        """Restore shaded regions for pairs of restored markers."""
        import pyqtgraph as pg
        from pyqtgraph import LinearRegionItem
        
        # Clear existing shaded regions first
        for region_top, region_bottom in self.shaded_regions:
            try:
                self.plot_top.removeItem(region_top)
            except Exception:
                pass
            try:
                self.plot_bottom.removeItem(region_bottom)
            except Exception:
                pass
        self.shaded_regions = []
        
        # Create shaded regions for each pair
        num_markers = len(self.markers)
        for i in range(1, num_markers, 2):  # Process pairs: (0,1), (2,3), (4,5)...
            start_x = self.markers[i-1][2]
            end_x = self.markers[i][2]
            
            region_top = LinearRegionItem(
                values=[start_x, end_x],
                movable=False,
                brush=pg.mkBrush(255, 152, 0, 50)
            )
            region_bottom = LinearRegionItem(
                values=[start_x, end_x],
                movable=False,
                brush=pg.mkBrush(255, 152, 0, 50)
            )
            self.plot_top.addItem(region_top)
            self.plot_bottom.addItem(region_bottom)
            self.shaded_regions.append((region_top, region_bottom))

    def _update_marks_file(self):
        key = self._get_logical_key()
        if not key:
            return
        marks = self._load_marks()
        times = sorted([x[2] for x in self.markers])
        if times:
            marks[key] = times
        elif key in marks:
            del marks[key]
        self._save_marks(marks)
    def _clear_last_marker(self, from_undo=False):
        if not self.markers:
            return
            
        # Remove last marker lines
        l1, l2, _ = self.markers.pop()
        try:
            self.plot_top.removeItem(l1)
        except Exception:
            pass
        try:
            self.plot_bottom.removeItem(l2)
        except Exception:
            pass
            
        # Remove associated shaded region if it was a pair completion
        # Logic: if we just removed the 2nd marker of a pair (even index before pop -> odd index after pop? No.)
        # If we had 2 markers, len=2. Pop -> len=1. We removed the 2nd one. which completed the region.
        # Regions are stored in self.shaded_regions.
        # If we removed a marker that completed a region, we must remove the last region.
        # BUT: markers list is now shorter. 
        # If len(markers) is odd (meaning we had even, now odd), we removed the closing marker.
        if len(self.markers) % 2 != 0 and self.shaded_regions:
             region_top, region_bottom = self.shaded_regions.pop()
             try:
                 self.plot_top.removeItem(region_top)
             except Exception:
                 pass
             try:
                 self.plot_bottom.removeItem(region_bottom)
             except Exception:
                 pass
        
        self.log_message.emit("Last marker removed.")
        self._update_marks_file()
        
        if not from_undo and self.undo_stack:
            # If user manually clears last, we should pop from undo stack to keep sync?
            # Or treat "Clear Last" as an action that can itself be undone?
            # User request: "Undo deshara las ultimas acciones".
            # For simplicity: Clear Last *is* an undo of "Add Marker".
            # So we should pop the "add_marker" action from stack.
            if self.undo_stack[-1]['type'] == 'add_marker':
                self.undo_stack.pop()
            
            if not self.undo_stack:
                self.btn_undo.setEnabled(False)


    from PySide6.QtCore import Signal
    tracking_loaded = Signal(str)
    log_message = Signal(str)
    marks_saved = Signal()

    def __init__(self, parent=None):
        super().__init__(parent)

        from PySide6.QtWidgets import QVBoxLayout, QHBoxLayout, QLabel, QPushButton, QComboBox
        self.layout = QVBoxLayout(self)
        self.layout.setContentsMargins(8,8,8,8)

        # Initialize media player early
        self.media_player = QMediaPlayer()
        self.current_tracking_path = None

        # Top controls: Grid Layout
        # Top controls: Single Row Layout
        ctrl = QHBoxLayout()
        ctrl.setContentsMargins(0, 0, 0, 0)
        
        # Top
        ctrl.addWidget(QLabel("Top:"))
        self.cb_top = QComboBox()
        self.cb_top.setFixedWidth(130)
        ctrl.addWidget(self.cb_top)
        
        # Bottom
        ctrl.addWidget(QLabel("Bottom:"))
        self.cb_bottom = QComboBox()
        self.cb_bottom.setFixedWidth(130)
        ctrl.addWidget(self.cb_bottom)
        
        # Subject Selector
        ctrl.addWidget(QLabel("Subject:"))
        self.combo_subject = QComboBox()
        self.combo_subject.setFixedWidth(70)
        self.combo_subject.currentIndexChanged.connect(self._on_subject_changed)
        ctrl.addWidget(self.combo_subject)
        
        # Case Selector
        ctrl.addWidget(QLabel("Case:"))
        self.combo_cases = QComboBox()
        self.combo_cases.setFixedWidth(130)
        self.combo_cases.currentIndexChanged.connect(self._on_case_selector_changed)
        ctrl.addWidget(self.combo_cases)
        
        # Internal: store all files for filtering
        self._all_case_files = []
        
        # Refresh/Reset View Button
        self.btn_reset_view = QPushButton()
        self.btn_reset_view.setToolTip("Reset Graph View (AutoRange)")
        self.btn_reset_view.setFixedSize(24, 24)
        self.btn_reset_view.setCursor(Qt.PointingHandCursor)
        icon_refresh = QIcon(resource_path("assets/icons/center_focus_strong_16dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png"))
        if not icon_refresh.isNull():
             self.btn_reset_view.setIcon(icon_refresh)
        else:
             self.btn_reset_view.setText("R")
        self.btn_reset_view.clicked.connect(self.reset_graph_view)
        ctrl.addWidget(self.btn_reset_view)
        
        # AI Smart Marker Button
        self.btn_ai_mark = QPushButton()
        self.btn_ai_mark.setToolTip("AI Smart Marker (Auto-detect Gaze Fixation)")
        self.btn_ai_mark.setFixedSize(24, 24)
        self.btn_ai_mark.setCursor(Qt.PointingHandCursor)
        icon_ai = QIcon(resource_path("assets/icons/star_shine_16dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png"))
        if not icon_ai.isNull():
            self.btn_ai_mark.setIcon(icon_ai)
        else:
            self.btn_ai_mark.setText("AI")
        self.btn_ai_mark.clicked.connect(self._run_ai_analysis)
        ctrl.addWidget(self.btn_ai_mark)
        
        ctrl.addStretch()

        # Buttons (Icons)
        btn_layout = QHBoxLayout()
        
        # Undo Button
        self.btn_undo = QPushButton()
        self.btn_undo.setToolTip("Undo Last Action")
        self.btn_undo.setCursor(Qt.PointingHandCursor)
        self.btn_undo.setFixedHeight(28)
        icon_undo = QIcon(resource_path("assets/icons/undo_16dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png"))
        if not icon_undo.isNull():
             self.btn_undo.setIcon(icon_undo)
        else:
             self.btn_undo.setText("Undo")
        self.btn_undo.clicked.connect(self._undo_last_action)
        self.btn_undo.setEnabled(False) 
             
        # Clear Last (D)
        self.btn_clear_last_marker = QPushButton()
        self.btn_clear_last_marker.setStyleSheet("font-weight: normal;")
        self.btn_clear_last_marker.setToolTip("Remove last marker (D)")
        self.btn_clear_last_marker.setCursor(Qt.PointingHandCursor)
        self.btn_clear_last_marker.setFixedHeight(28)
        self.btn_clear_last_marker.clicked.connect(self._clear_last_marker)
        icon_d = QIcon(resource_path("assets/icons/d_key.png"))
        if not icon_d.isNull():
            self.btn_clear_last_marker.setIcon(icon_d)
            self.btn_clear_last_marker.setText(" Clear Last")
        else:
            self.btn_clear_last_marker.setText("Clear Last (D)")
        
        # Clear All (Space)
        self.btn_clear_markers = QPushButton()
        self.btn_clear_markers.setStyleSheet("font-weight: normal;")
        self.btn_clear_markers.setToolTip("Remove all markers (Space)")
        self.btn_clear_markers.setCursor(Qt.PointingHandCursor)
        self.btn_clear_markers.setFixedHeight(28)
        self.btn_clear_markers.clicked.connect(self._clear_markers)
        icon_space = QIcon(resource_path("assets/icons/space_key.png"))
        if not icon_space.isNull():
             self.btn_clear_markers.setIcon(icon_space)
             self.btn_clear_markers.setText(" Clear All")
        else:
             self.btn_clear_markers.setText("Clear All (Space)")
        
        btn_layout.addWidget(self.btn_undo)
        btn_layout.addWidget(self.btn_clear_last_marker)
        btn_layout.addWidget(self.btn_clear_markers)
        
        ctrl.addLayout(btn_layout)
        
        self.layout.addLayout(ctrl)

        # Plots (pyqtgraph)
        import pyqtgraph as pg
        self.pg = pg
        self.plot_top = pg.PlotWidget(background='#1e1e1e')
        self.plot_bottom = pg.PlotWidget(background='#1e1e1e')
        self.plot_top.getPlotItem().hideButtons()
        self.plot_bottom.getPlotItem().hideButtons()
        self.plot_top.showGrid(x=True, y=True, alpha=0.3)
        self.plot_bottom.showGrid(x=True, y=True, alpha=0.3)
        
        # Cursor: Crosshair for precision
        self.plot_top.setCursor(Qt.CrossCursor)
        self.plot_bottom.setCursor(Qt.CrossCursor)
        
        self.layout.addWidget(self.plot_top)
        self.layout.addWidget(self.plot_bottom)
        
        # Video Filmstrip (Timeline)
        self.filmstrip = VideoFilmstripWidget()
        self.filmstrip.setVisible(False)
        self.layout.addWidget(self.filmstrip)
        
        # Video widget with scroll area for zoom isolation
        self.video_scroll_area = QScrollArea()
        self.video_scroll_area.setFixedSize(640, 360)
        self.video_scroll_area.setWidgetResizable(True)
        self.video_scroll_area.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.video_scroll_area.setVerticalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        self.video_scroll_area.setStyleSheet("""
            QScrollArea {
                background-color: #1e1e1e;
                border: 1px solid #444;
            }
            QScrollArea QWidget {
                background-color: #1e1e1e;
            }
        """)
        
        self.video_widget = VideoFrameWidget()
        self.video_widget.setMinimumSize(640, 360)
        self.video_widget.setVisible(False)  # Hidden by default
        
        # Container to center the video
        self.video_container = QWidget()
        self.video_container.setStyleSheet("background-color: black;")
        self.video_container_layout = QGridLayout(self.video_container)
        self.video_container_layout.setContentsMargins(0, 0, 0, 0)
        self.video_container_layout.addWidget(self.video_widget, 0, 0, Qt.AlignCenter)
        
        # Floating Overlay for Speed and Camera
        self.video_overlay = VideoOverlayControls(self.video_container)
        self.video_overlay.setVisible(False)
        
        self.video_scroll_area.setWidget(self.video_container)
        self.video_scroll_area.setAlignment(Qt.AlignCenter)
        
        self.layout.addWidget(self.video_scroll_area)
        self.video_scroll_area.setVisible(False)  # Hidden by default

        # Video zoom functionality
        self.video_zoom_factor = 1.0
        self.video_base_size = (640, 360)  # Base size for zoom calculations
        self.video_widget.installEventFilter(self)
        
        # Video pan functionality
        self.pan_active = False
        self.pan_start_pos = None
        self.pan_start_scroll = None
        
        # Marker selection functionality
        self.selected_marker_index = -1  # No marker selected initially

        # Video controls
        self.video_controls = QWidget()
        self.video_controls.setVisible(False)
        video_ctrl_layout = QHBoxLayout(self.video_controls)
        video_ctrl_layout.setContentsMargins(0, 5, 0, 5)
        
        self.video_progress = QSlider(Qt.Horizontal)
        self.video_progress.setRange(0, 100)
        self.video_progress.sliderMoved.connect(self._seek_video)
        self.media_player.positionChanged.connect(self._update_progress)
        self.media_player.durationChanged.connect(self._update_duration)
        video_ctrl_layout.addWidget(self.video_progress, 1)
        
        self.lbl_video_time = QLabel("00:00 / 00:00")
        self.lbl_video_time.setStyleSheet("color: #ccc; font-size: 10pt;")
        video_ctrl_layout.addWidget(self.lbl_video_time)
        
        self.layout.addWidget(self.video_controls)

        self.media_player.setVideoOutput(self.video_widget.videoSink)

        # Video controls (moved from top)
        self.chk_video = AnimatedToggle(self)
        self.chk_video.setChecked(False)
        self.chk_video.toggled.connect(self._toggle_video_display)
        
        self.lbl_video_icon = QLabel()
        cam_icon_path = resource_path("assets/icons/camera_video_24dp_FFFFFF_FILL0_wght400_GRAD0_opsz24.png")
        if os.path.exists(cam_icon_path):
            self.lbl_video_icon.setPixmap(QPixmap(cam_icon_path))
            self.lbl_video_icon.setToolTip("Video")
        else:
            self.lbl_video_icon.setText("Video:")
            
        self.lbl_speed = QLabel("Speed:")
        self.cb_video_speed = QComboBox()
        self.cb_video_speed.addItems(["0.25x", "0.5x", "1x", "1.5x", "2x"])
        self.cb_video_speed.setCurrentText("1x")
        self.cb_video_speed.currentTextChanged.connect(self._on_video_speed_changed)
        
        self.lbl_camera = QLabel("Camera:")
        self.cb_camera_id = QComboBox()
        self.cb_camera_id.currentTextChanged.connect(self._on_camera_id_changed)
        
        # Container for extra controls to allow smooth animation
        self.video_extra_controls = QWidget()
        self.video_extra_controls.setContentsMargins(0, 0, 0, 0)
        # Speed and Camera moved to Overlay
        self.lbl_speed = QLabel("Speed")
        self.cb_video_speed = QComboBox()
        self.cb_video_speed.addItems(["0.1x", "0.25x", "0.5x", "1x", "2x", "4x"])
        self.cb_video_speed.setCurrentText("1x")
        self.cb_video_speed.currentTextChanged.connect(self._on_video_speed_changed)
        
        self.lbl_camera = QLabel("Camera")
        self.cb_camera_id = QComboBox()
        self.cb_camera_id.currentTextChanged.connect(self._on_camera_id_changed)
        
        overlay_layout = self.video_overlay.layout
        overlay_layout.addStretch()
        overlay_layout.addWidget(self.lbl_speed)
        overlay_layout.addWidget(self.cb_video_speed)
        overlay_layout.addSpacing(15)
        overlay_layout.addWidget(self.lbl_camera)
        overlay_layout.addWidget(self.cb_camera_id)
        overlay_layout.addStretch()
        
        # Bottom Bar: Only active video controls
        extra_layout = QHBoxLayout(self.video_extra_controls)
        extra_layout.setContentsMargins(0, 0, 0, 0)
        extra_layout.setSpacing(10)
        
        # New Video Controls re-located here
        self.btn_sync_video = QPushButton()
        self.btn_sync_video.setCheckable(True)
        self.btn_sync_video.setFixedSize(32, 32)
        self.btn_sync_video.setIcon(QIcon(resource_path("assets/icons/mouse_lock_off_24dp_FFFFFF_FILL0_wght400_GRAD0_opsz24.png")))
        self.btn_sync_video.setIconSize(QSize(20, 20))
        self.btn_sync_video.setToolTip("Sync Video with Mouse Cursor")
        self.btn_sync_video.setStyleSheet("QPushButton { background: transparent; border: none; } QPushButton:hover { background: rgba(255,255,255,0.1); border-radius: 4px; }")
        self.btn_sync_video.clicked.connect(self._toggle_video_sync)
        extra_layout.addWidget(self.btn_sync_video)
        
        extra_layout.addSpacing(5)
        
        btn_style = "QPushButton { background: transparent; border: none; } QPushButton:hover { background: rgba(255,255,255,0.1); border-radius: 4px; }"
        
        self.btn_first_page = QPushButton()
        self.btn_first_page.setFixedSize(32, 32)
        self.btn_first_page.setIcon(QIcon(resource_path("assets/icons/first_page_24dp_FFFFFF_FILL0_wght400_GRAD0_opsz24.png")))
        self.btn_first_page.setIconSize(QSize(24, 24))
        self.btn_first_page.setToolTip("Go to start")
        self.btn_first_page.setStyleSheet(btn_style)
        self.btn_first_page.clicked.connect(lambda: self.media_player.setPosition(0))
        extra_layout.addWidget(self.btn_first_page)
        
        self.btn_replay_5 = QPushButton()
        self.btn_replay_5.setFixedSize(32, 32)
        self.btn_replay_5.setIcon(QIcon(resource_path("assets/icons/replay_5_24dp_FFFFFF_FILL0_wght400_GRAD0_opsz24.png")))
        self.btn_replay_5.setIconSize(QSize(24, 24))
        self.btn_replay_5.setToolTip("Jump back 5s")
        self.btn_replay_5.setStyleSheet(btn_style)
        self.btn_replay_5.clicked.connect(lambda: self.media_player.setPosition(max(0, self.media_player.position() - 5000)))
        extra_layout.addWidget(self.btn_replay_5)
        
        self.btn_forward_5 = QPushButton()
        self.btn_forward_5.setFixedSize(32, 32)
        self.btn_forward_5.setIcon(QIcon(resource_path("assets/icons/forward_5_24dp_FFFFFF_FILL0_wght400_GRAD0_opsz24.png")))
        self.btn_forward_5.setIconSize(QSize(24, 24))
        self.btn_forward_5.setToolTip("Jump forward 5s")
        self.btn_forward_5.setStyleSheet(btn_style)
        self.btn_forward_5.clicked.connect(lambda: self.media_player.setPosition(min(self.media_player.duration(), self.media_player.position() + 5000)))
        extra_layout.addWidget(self.btn_forward_5)
        
        extra_layout.addSpacing(5)
        
        self.btn_play_pause = AnimatedPlayButton()
        self.btn_play_pause.clicked.connect(self._toggle_play_pause)
        extra_layout.addWidget(self.btn_play_pause)
        
        # Initial state: hidden and zero width
        self.video_extra_controls.setMaximumWidth(0)
        self.video_extra_controls.setOpacityEffect = QGraphicsOpacityEffect(self.video_extra_controls)
        self.video_extra_controls.setGraphicsEffect(self.video_extra_controls.setOpacityEffect)
        self.video_extra_controls.setOpacityEffect.setOpacity(0)
        
        self.video_extra_anim_group = QParallelAnimationGroup(self)
        self.video_extra_width_anim = QPropertyAnimation(self.video_extra_controls, b"maximumWidth")
        self.video_extra_width_anim.setDuration(400)
        self.video_extra_width_anim.setEasingCurve(QEasingCurve.OutQuint)
        self.video_extra_opacity_anim = QPropertyAnimation(self.video_extra_controls.setOpacityEffect, b"opacity")
        self.video_extra_opacity_anim.setDuration(400)
        self.video_extra_opacity_anim.setEasingCurve(QEasingCurve.OutQuint)
        self.video_extra_anim_group.addAnimation(self.video_extra_width_anim)
        self.video_extra_anim_group.addAnimation(self.video_extra_opacity_anim)

        # Previous/Next Plot buttons below the plots
        # Next/Previous Plot buttons
        from PySide6.QtWidgets import QHBoxLayout
        next_plot_layout = QHBoxLayout()
        
        # Previous
        # Previous Button with 2 icons + text (Centered)
        self.btn_prev_plot = QPushButton()
        self.btn_prev_plot.setStyleSheet("font-weight: normal;")
        self.btn_prev_plot.setFixedWidth(140)
        self.btn_prev_plot.setToolTip("Previous Case (Shift+Tab)")
        self.btn_prev_plot.setCursor(Qt.PointingHandCursor)
        self.btn_prev_plot.setFocusPolicy(Qt.NoFocus)
        
        l_prev = QHBoxLayout(self.btn_prev_plot)
        l_prev.setContentsMargins(4, 2, 4, 2)
        l_prev.setSpacing(4)
        
        l_prev.addStretch() # Push to center
        
        # Icons
        icon_shift = QIcon(resource_path("assets/icons/shift_key.png"))
        icon_tab = QIcon(resource_path("assets/icons/tab_key.png"))
        
        # Helper to add icon label
        def add_icon_label(layout, icon):
            if not icon.isNull():
                 lbl = QLabel()
                 lbl.setPixmap(icon.pixmap(16, 16))
                 layout.addWidget(lbl)
        
        add_icon_label(l_prev, icon_shift)
        add_icon_label(l_prev, icon_tab)
        
        l_prev.addWidget(QLabel("Previous"))
        l_prev.addStretch() # Push to center

        # Remove default text to rely on layout
        self.btn_prev_plot.setText("")
             
        self.lbl_prev_case = QLabel("")
        self.lbl_prev_case.setStyleSheet("color: #888; font-size: 9pt;")
        
        # Next
        self.btn_next_plot = QPushButton()
        self.btn_next_plot.setStyleSheet("font-weight: normal;")
        self.btn_next_plot.setFixedWidth(140)
        self.btn_next_plot.setToolTip("Next Case (Tab)")
        self.btn_next_plot.setCursor(Qt.PointingHandCursor)
        self.btn_next_plot.setFocusPolicy(Qt.NoFocus)
        try:
             # Check for tab_key.png, fallback to text
             icon_next = QIcon(resource_path("assets/icons/tab_key.png"))
             if icon_next.isNull():
                 self.btn_next_plot.setText("Next Case")
             else:
                 self.btn_next_plot.setIcon(icon_next)
                 self.btn_next_plot.setText(" Next")
        except:
             self.btn_next_plot.setText("Next Case")

        self.lbl_next_case = QLabel("")
        self.lbl_next_case.setStyleSheet("color: #888; font-size: 9pt;")

        next_plot_layout.addWidget(self.btn_prev_plot)
        next_plot_layout.addWidget(self.lbl_prev_case)
        
        next_plot_layout.addStretch()
        
        # Video controls
        next_plot_layout.addWidget(self.lbl_video_icon)
        next_plot_layout.addWidget(self.chk_video)
        next_plot_layout.addSpacing(5)
        next_plot_layout.addWidget(self.video_extra_controls)
        
        next_plot_layout.addStretch()
        
        next_plot_layout.addWidget(self.lbl_next_case)
        next_plot_layout.addWidget(self.btn_next_plot)
        
        self.layout.addLayout(next_plot_layout)

        # Link X axes
        try:
            self.plot_bottom.getPlotItem().vb.setXLink(self.plot_top.getPlotItem().vb)
        except Exception:
            pass

        # Crosshair lines
        self.vline_top = pg.InfiniteLine(angle=90, movable=False, pen=pg.mkPen('#AAAAAA'))
        self.vline_bottom = pg.InfiniteLine(angle=90, movable=False, pen=pg.mkPen('#AAAAAA'))
        self.plot_top.addItem(self.vline_top, ignoreBounds=True)
        self.plot_bottom.addItem(self.vline_bottom, ignoreBounds=True)
        
        # Sync filmstrip with plot range
        def on_range_changed():
            vb = self.plot_top.getPlotItem().vb
            vr = vb.viewRect()
            self.filmstrip.update_range(vr.left(), vr.right())
            
            # Pixel-perfect margin synchronization
            # Map view coordinates to scene, then to viewport, then to global, then to filmstrip local coordinates
            p_left_scene = vb.mapViewToScene(pg.Point(vr.left(), 0))
            p_left_viewport = self.plot_top.mapFromScene(p_left_scene)
            p_left_global = self.plot_top.viewport().mapToGlobal(p_left_viewport)
            p_left_local = self.filmstrip.mapFromGlobal(p_left_global)
            
            p_right_scene = vb.mapViewToScene(pg.Point(vr.right(), 0))
            p_right_viewport = self.plot_top.mapFromScene(p_right_scene)
            p_right_global = self.plot_top.viewport().mapToGlobal(p_right_viewport)
            p_right_local = self.filmstrip.mapFromGlobal(p_right_global)
            
            self.filmstrip.left_margin = p_left_local.x()
            self.filmstrip.right_margin = self.filmstrip.width() - p_right_local.x()
            self.filmstrip.update()
            
        self._on_range_changed_manual = on_range_changed # Store for manual triggering
        self.plot_top.getPlotItem().vb.sigXRangeChanged.connect(on_range_changed)
        
        # Set zoom limit to prevent frames from getting too separated/pixelated
        # (Minimum 1 second visible on screen)
        self.plot_top.getPlotItem().vb.setLimits(minXRange=1.0)
        self.plot_bottom.getPlotItem().vb.setLimits(minXRange=1.0)

        # Persistent markers list
        self.drag_mode = False
        self.delete_mode = False
        self.markers = []
        # Shaded regions list (for even-numbered markers)
        self.shaded_regions = []

        # Tooltip: use QToolTip on mouse move
        from PySide6.QtWidgets import QToolTip
        from PySide6.QtGui import QCursor
        self._tooltip = QToolTip

        # Events
        self.proxy_top = self.plot_top.scene().sigMouseMoved.connect(self._on_mouse_moved)
        self.plot_bottom.scene().sigMouseMoved.connect(self._on_mouse_moved)
        
        # Handling Left Click for markers + ignoring drag
        self.plot_top.scene().sigMouseClicked.connect(self._on_mouse_clicked)
        self.plot_bottom.scene().sigMouseClicked.connect(self._on_mouse_clicked)

        self.cb_top.currentIndexChanged.connect(self._plot_signals)
        self.cb_bottom.currentIndexChanged.connect(self._plot_signals)
        # NOTE: btn_prev/next_plot connections are done in AnalysisWidget.create_tabs_right
        # to avoid double-firing. Do NOT connect them here.

        # Keyboard shortcut: 'd' to clear last marker
        from PySide6.QtGui import QShortcut, QKeySequence
        # Keyboard shortcut: 'd' to clear last marker
        from PySide6.QtGui import QShortcut, QKeySequence
        self.shortcut_clear_last = QShortcut(QKeySequence('d'), self)
        self.shortcut_clear_last.activated.connect(self._clear_last_marker)
        
        # Keyboard shortcut: Space to clear all markers
        self.shortcut_clear_all = QShortcut(QKeySequence(Qt.Key_Space), self)
        self.shortcut_clear_all.activated.connect(self._clear_markers)
        
        # Shortcut: Shift+Tab for previous is handled in AnalysisWidget
        # (using Key_Backtab) to ensure marker validation happens first.
        
        # Undo history
        self.undo_stack = [] # Stack of states/actions
        self.markers = []
        self.drag_mode = True # Always True now
        self.delete_mode = False

    def _run_ai_analysis(self):
        """Triggers the AI detection logic."""
        if not self.current_tracking_path:
            self._notify_parent("Please load a case first.", "error")
            return
            
        # USER REQUEST: Ensure video is enabled and camera selected
        if not self.chk_video.isChecked():
            self._notify_parent("Please enable 'Camera' toggle first.", "warning")
            return
            
        if self.cb_camera_id.currentIndex() < 0:
            self._notify_parent("Please select a Camera ID (cam1, cam2...) first.", "warning")
            return

        # Find video path
        mf4_dir = os.path.dirname(self.current_tracking_path)
        base_name = os.path.splitext(os.path.basename(self.current_tracking_path))[0].replace("_tracking", "")
        selected_cam = self.cb_camera_id.currentText()
        video_path = os.path.join(mf4_dir, f"{base_name}_{selected_cam}.avi")
        if not os.path.exists(video_path):
             video_path = os.path.join(os.path.dirname(mf4_dir), f"{base_name}_{selected_cam}.avi")

        if not os.path.exists(video_path):
            self._notify_parent(f"Video file not found: {os.path.basename(video_path)}", "error")
            return

        self.btn_ai_mark.setEnabled(False)
        self._notify_parent("AI Analysis started... analyzing signals and video frames.", "info")
        
        self.ai_worker = AIWorker(self.current_tracking_path, video_path)
        # Add a small delay to progress updates to make it readable
        self.ai_worker.signals.progress.connect(lambda p: self._notify_parent(f"AI Analysis: {int(p*100)}%", "info"))
        self.ai_worker.signals.log.connect(lambda m: self.log_message.emit(m))
        self.ai_worker.signals.finished.connect(self._on_ai_finished)
        self.ai_worker.signals.error.connect(self._on_ai_error)
        self.ai_worker.start()

    def _on_ai_finished(self, timestamps):
        self.btn_ai_mark.setEnabled(True)
        if not timestamps:
            self._notify_parent("AI finished: No clear event detected.", "warning")
            return
            
        # Add markers
        # Sort and deduplicate
        timestamps = sorted(list(set(timestamps)))
        
        for t in timestamps:
            self._add_marker(t)
            
        self._notify_parent(f"AI Success: {len(timestamps)} marker(s) added.", "success")
        self._update_marks_file()

    def _on_ai_error(self, message):
        self.btn_ai_mark.setEnabled(True)
        self._notify_parent(f"AI Error: {message}", "error")

    def _notify_parent(self, message, type_="info"):
        parent = self._get_parent_analysis_widget()
        if parent:
            parent._notify(message, type_)
        else:
            self.log_message.emit(f"[{type_.upper()}] {message}")

    def _get_parent_analysis_widget(self):
        curr = self.parent()
        while curr:
            if hasattr(curr, '_notify'):
                return curr
            curr = curr.parent()
        return None

    def _on_line_clicked(self, line, ev):
        """Handle context menu for markers."""
        from PySide6.QtGui import QCursor
        if ev.button() == Qt.RightButton:
            ev.accept() # Prevent pyqtgraph context menu
            menu = QMenu(self)
            menu.setStyleSheet("QMenu { background-color: #222; color: white; border: 1px solid #444; }")
            action_delete = menu.addAction("Delete Marker")
            action_delete.setIcon(QIcon(resource_path("assets/icons/delete_16dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png")))
            
            action = menu.exec(QCursor.pos())
            if action == action_delete:
                # Find index of this line
                for i, (l1, l2, _) in enumerate(self.markers):
                    if l1 is line or l2 is line:
                        self._remove_marker_by_index(i)
                        break

    def previous_plot(self):
        """Go to previous case."""
        count = self.combo_cases.count()
        if count == 0: return
        idx = self.combo_cases.currentIndex()
        if idx > 0:
            self.combo_cases.setCurrentIndex(idx - 1)
            
    def next_plot(self):
        """Go to next case."""
        count = self.combo_cases.count()
        if count == 0: return
        idx = self.combo_cases.currentIndex()
        if idx < count - 1:
             self.combo_cases.setCurrentIndex(idx + 1)

    def update_navigation_labels(self):
        """Update Next/Prev case labels based on currently selected case in combo."""
        count = self.combo_cases.count()
        if count == 0:
            if hasattr(self, 'lbl_next_case'): self.lbl_next_case.setText("Next: None")
            if hasattr(self, 'lbl_prev_case'): self.lbl_prev_case.setText("Prev: None")
            return
            
        idx = self.combo_cases.currentIndex()
        
        # Determine Next/Prev
        prev_name = "None"
        next_name = "None"
        
        if idx > 0:
            prev_name = self.combo_cases.itemText(idx - 1)
        if idx < count - 1:
            next_name = self.combo_cases.itemText(idx + 1)
        
        # Strip extension for cleaner display
        if prev_name != "None":
             prev_name = prev_name.replace(".mf4", "").replace("_tracking", "")
        if next_name != "None":
             next_name = next_name.replace(".mf4", "").replace("_tracking", "")

        # Update Labels
        if hasattr(self, 'lbl_next_case'):
             self.lbl_next_case.setText(f"Next: {next_name[:20]}..." if next_name != "None" else "Next: None")
        
        if hasattr(self, 'lbl_prev_case'):
             self.lbl_prev_case.setText(f"Prev: {prev_name[:20]}..." if prev_name != "None" else "Prev: None")

    def _on_case_selector_changed(self, index):
        if index >= 0:
             fpath = self.combo_cases.itemData(index)
             if fpath and fpath != self.current_tracking_path:
                 self.load_tracking_file(fpath)
        # Ensure labels are updated
        self.update_navigation_labels()

    def set_case_list(self, files: list):
        """Populate the Subject and Case Selector combos."""
        self._all_case_files = files
        
        # Extract unique subjects from file paths
        subjects = []
        for f in files:
            subj = self._extract_subject(f)
            if subj and subj not in subjects:
                subjects.append(subj)
        
        self.combo_subject.blockSignals(True)
        self.combo_subject.clear()
        for s in subjects:
            self.combo_subject.addItem(s)
        self.combo_subject.blockSignals(False)
        
        # Populate cases for first subject
        if subjects:
            self._populate_cases_for_subject(subjects[0])
        
        self.update_navigation_labels()
    
    def _extract_subject(self, filepath):
        """Extract subject (e.g. P01, E01) from file path."""
        import re
        parts = filepath.replace("\\", "/").split("/")
        for part in parts:
            if re.match(r'^[A-Z]\d{2}$', part):
                return part
        return "Unknown"
    
    def _populate_cases_for_subject(self, subject):
        """Populate case combo with only files belonging to the given subject."""
        self.combo_cases.blockSignals(True)
        self.combo_cases.clear()
        
        for f in self._all_case_files:
            subj = self._extract_subject(f)
            if subj == subject:
                name = os.path.basename(f).replace("_tracking.mf4", "").replace(".mf4", "")
                self.combo_cases.addItem(name, f)
        
        self.combo_cases.blockSignals(False)
    
    def _on_subject_changed(self, index):
        """When subject changes, update cases combo and load first case."""
        if index < 0:
            return
        subject = self.combo_subject.itemText(index)
        self._populate_cases_for_subject(subject)
        
        # Auto-load first case of this subject
        if self.combo_cases.count() > 0:
            self.combo_cases.setCurrentIndex(0)
            fpath = self.combo_cases.itemData(0)
            if fpath:
                self.load_tracking_file(fpath)
        self.update_navigation_labels()

    def _undo_last_action(self):
        """Revert the last marker addition."""
        if not self.undo_stack:
            return
            
        action = self.undo_stack.pop()
        type_ = action.get('type')
        
        if type_ == 'add_marker':
            # Remove the specific marker items
            # In our case, simple way is pop from list and remove from plot
            # But since _clear_last_marker removes the *last* one in list, 
            # and stack pushes on add, it matches.
            self._clear_last_marker(from_undo=True) # Modifying _clear_last_marker to not pop stack?
            # Actually, standard _clear_last_marker logic is what we want, 
            # but we need to prevent it from clearing the stack further if we call it.
            # Simpler: just implement undo logic directly here or reuse.
            pass
        
        if not self.undo_stack:
            self.btn_undo.setEnabled(False)

    def _toggle_drag_mode(self):
        self.drag_mode = not self.drag_mode
        if self.drag_mode and self.delete_mode:
            self.delete_mode = False
            self.btn_delete.setChecked(False)
        pen = pg.mkPen('#00FF00' if self.drag_mode else '#ff9800', width=2)
        # Update all existing markers
        for line1, line2, _ in self.markers:
            line1.setMovable(self.drag_mode)
            line2.setMovable(self.drag_mode)
            line1.setPen(pen)
            line2.setPen(pen)
        self.log_message.emit(f"Drag mode {'enabled' if self.drag_mode else 'disabled'}")

    def _toggle_delete_mode(self):
        self.delete_mode = not self.delete_mode
        if self.delete_mode and self.drag_mode:
            self.drag_mode = False
            self.btn_drag_mode.setChecked(False)
        self.log_message.emit(f"Delete mode {'enabled' if self.delete_mode else 'disabled'}")

    def resizeEvent(self, event):
        super().resizeEvent(event)
        if hasattr(self, 'video_overlay') and self.video_overlay.isVisible():
            self._update_overlay_position()

    def _update_overlay_position(self):
        if not hasattr(self, 'video_overlay'): return
        
        # Position relative to video_widget (mapping its local geometry to video_container)
        vw_pos = self.video_widget.pos()
        vw_width = self.video_widget.width()
        vw_height = self.video_widget.height()
        
        if vw_width <= 10: return # Avoid positioning when hidden/tiny
        
        overlay_width = self.video_overlay.width()
        overlay_height = self.video_overlay.height()
        
        # Center horizontally relative to video_widget
        x = vw_pos.x() + (vw_width - overlay_width) // 2
        # Bottom of video_widget with margin
        y = vw_pos.y() + vw_height - overlay_height - 20
        
        self.video_overlay.move(x, y)
        self.video_overlay.show()
        self.video_overlay.raise_()

    def _toggle_video_display(self):
        show_video = self.chk_video.isChecked()
        self.video_scroll_area.setVisible(show_video)
        self.video_widget.setVisible(show_video) # Ensure the actual video widget is shown
        self.video_controls.setVisible(show_video)
        self.filmstrip.setVisible(show_video)
        self.video_overlay.setVisible(show_video)
        
        # Smooth animation for speed/camera controls
        self.video_extra_anim_group.stop()
        if show_video:
            # Expand to accommodate all controls centered
            self.video_extra_width_anim.setEndValue(650) 
            self.video_extra_opacity_anim.setEndValue(1.0)
            # Center overlay at bottom
            self._update_overlay_position()
            # Update position again after animation finishes to be sure
            QTimer.singleShot(410, self._update_overlay_position)
            # Delay heavy work to avoid lag during animation
            QTimer.singleShot(400, self._delayed_video_setup)
        else:
            self.video_extra_width_anim.setEndValue(0)
            self.video_extra_opacity_anim.setEndValue(0.0)
            self.media_player.stop()
            
        self.video_extra_anim_group.start()

    def _delayed_video_setup(self):
        """Perform video scanning and loading after animation completes."""
        if self.chk_video.isChecked():
            self._scan_available_cameras()
            self._load_video_for_current_file()
            # Trigger manual sync of filmstrip on first load
            QTimer.singleShot(100, self._on_range_changed_manual)

    def _toggle_video_sync(self):
        is_checked = self.btn_sync_video.isChecked()
        if is_checked:
            self.btn_sync_video.setIcon(QIcon(resource_path("assets/icons/mouse_lock_24dp_000000_FILL1_wght400_GRAD0_opsz24.png")))
            self.btn_sync_video.setStyleSheet("background-color: #F39200; border: none; border-radius: 4px;")
            self.log_message.emit("Video sync enabled")
        else:
            self.btn_sync_video.setIcon(QIcon(resource_path("assets/icons/mouse_lock_off_24dp_FFFFFF_FILL0_wght400_GRAD0_opsz24.png")))
            self.btn_sync_video.setStyleSheet("background-color: transparent; border: none;")
            self.log_message.emit("Video sync disabled")

    def _on_video_speed_changed(self, text):
        rate = float(text.replace("x", ""))
        self.media_player.setPlaybackRate(rate)

    def _on_camera_id_changed(self, text):
        if self.chk_video.isChecked():
            self._load_video_for_current_file()

    def _scan_available_cameras(self):
        if not self.current_tracking_path:
            return
        
        mf4_dir = os.path.dirname(self.current_tracking_path)
        base_name = os.path.splitext(os.path.basename(self.current_tracking_path))[0].replace("_tracking", "")
        
        cameras = []
        import glob
        # Pattern: base_name_cam*.avi
        pattern = os.path.join(mf4_dir, f"{base_name}_cam*.avi")
        files = glob.glob(pattern)
        
        # Also check parent dir
        parent_dir = os.path.dirname(mf4_dir)
        pattern_p = os.path.join(parent_dir, f"{base_name}_cam*.avi")
        files += glob.glob(pattern_p)
        
        # Extract indices
        for f in files:
            m = re.search(r'_cam(\d+)\.avi$', f)
            if m:
                cam_id = f"cam{m.group(1)}"
                if cam_id not in cameras:
                    cameras.append(cam_id)
        
        cameras.sort()
        
        self.cb_camera_id.blockSignals(True)
        current = self.cb_camera_id.currentText()
        self.cb_camera_id.clear()
        self.cb_camera_id.addItems(cameras)
        if current in cameras:
            self.cb_camera_id.setCurrentText(current)
        elif cameras:
            self.cb_camera_id.setCurrentIndex(0)
        self.cb_camera_id.blockSignals(False)

    def _toggle_play_pause(self):
        if self.media_player.playbackState() == QMediaPlayer.PlaybackState.PlayingState:
            self.media_player.pause()
            self.btn_play_pause.set_playing(False)
        else:
            self.media_player.play()
            self.btn_play_pause.set_playing(True)

    def _seek_video(self, position):
        self.media_player.setPosition(position)

    def _update_progress(self, position):
        self.video_progress.setValue(position)
        self._update_time_label()

    def _update_duration(self, duration):
        self.video_progress.setMaximum(duration)
        if hasattr(self, 'filmstrip'):
            self.filmstrip.video_duration = duration / 1000.0
            self.filmstrip.update()
        self._update_time_label()

    def _update_time_label(self):
        position = self.media_player.position()
        duration = self.media_player.duration()
        pos_str = self._format_time(position)
        dur_str = self._format_time(duration)
        self.lbl_video_time.setText(f"{pos_str} / {dur_str}")

    def _format_time(self, ms):
        seconds = ms // 1000
        minutes = seconds // 60
        seconds %= 60
        return f"{minutes:02d}:{seconds:02d}"

    def eventFilter(self, obj, event):
        """Handle wheel events for video zoom and mouse events for pan."""
        if obj == self.video_widget and self.video_widget.isVisible() and self.chk_video.isChecked():
            
            # Handle wheel events for zoom
            if event.type() == event.Type.Wheel:
                # Get cursor position relative to video widget
                from PySide6.QtGui import QCursor
                global_cursor_pos = QCursor.pos()
                cursor_pos = self.video_widget.mapFromGlobal(global_cursor_pos)
                
                # Store old zoom factor
                old_zoom = self.video_zoom_factor
                
                # Handle zoom with mouse wheel
                delta = event.angleDelta().y()
                zoom_step = 0.1
                if delta > 0:
                    # Zoom in
                    self.video_zoom_factor = min(3.0, self.video_zoom_factor + zoom_step)
                elif delta < 0:
                    # Zoom out
                    self.video_zoom_factor = max(0.5, self.video_zoom_factor - zoom_step)
                
                # Apply zoom by changing widget size
                base_width, base_height = self.video_base_size
                new_width = int(base_width * self.video_zoom_factor)
                new_height = int(base_height * self.video_zoom_factor)
                self.video_widget.setFixedSize(new_width, new_height)
                
                # Center zoom on cursor position - keep the exact point under cursor fixed
                if old_zoom != self.video_zoom_factor:
                    # Calculate which point in the content is currently under the cursor
                    current_scroll_x = self.video_scroll_area.horizontalScrollBar().value()
                    current_scroll_y = self.video_scroll_area.verticalScrollBar().value()
                    
                    # Point in content coordinates that is under cursor
                    content_point_x = cursor_pos.x() + current_scroll_x
                    content_point_y = cursor_pos.y() + current_scroll_y
                    
                    # After zoom, adjust scroll so this content point stays under cursor
                    new_scroll_x = content_point_x - cursor_pos.x()
                    new_scroll_y = content_point_y - cursor_pos.y()
                    
                    # Ensure scroll position stays within bounds
                    scroll_area_size = self.video_scroll_area.size()
                    max_scroll_x = max(0, new_width - scroll_area_size.width())
                    max_scroll_y = max(0, new_height - scroll_area_size.height())
                    
                    new_scroll_x = max(0, min(new_scroll_x, max_scroll_x))
                    new_scroll_y = max(0, min(new_scroll_y, max_scroll_y))
                    
                    self.video_scroll_area.horizontalScrollBar().setValue(int(new_scroll_x))
                    self.video_scroll_area.verticalScrollBar().setValue(int(new_scroll_y))
                
                # Prevent event propagation
                event.accept()
                return True
            
            # Handle mouse events for pan (only when zoomed in)
            elif self.video_zoom_factor > 1.0:
                
                if event.type() == event.Type.MouseButtonPress and event.button() == Qt.LeftButton:
                    # Start pan
                    self.pan_active = True
                    self.pan_start_pos = event.pos()
                    self.pan_start_scroll = (
                        self.video_scroll_area.horizontalScrollBar().value(),
                        self.video_scroll_area.verticalScrollBar().value()
                    )
                    # Change cursor to indicate pan mode
                    self.video_widget.setCursor(Qt.ClosedHandCursor)
                    event.accept()
                    return True
                    
                elif event.type() == event.Type.MouseMove and self.pan_active:
                    # Perform pan
                    if self.pan_start_pos is not None and self.pan_start_scroll is not None:
                        delta = event.pos() - self.pan_start_pos
                        new_scroll_x = self.pan_start_scroll[0] - delta.x()
                        new_scroll_y = self.pan_start_scroll[1] - delta.y()
                        
                        # Ensure scroll position stays within bounds
                        scroll_area_size = self.video_scroll_area.size()
                        video_size = self.video_widget.size()
                        max_scroll_x = max(0, video_size.width() - scroll_area_size.width())
                        max_scroll_y = max(0, video_size.height() - scroll_area_size.height())
                        
                        new_scroll_x = max(0, min(new_scroll_x, max_scroll_x))
                        new_scroll_y = max(0, min(new_scroll_y, max_scroll_y))
                        
                        self.video_scroll_area.horizontalScrollBar().setValue(int(new_scroll_x))
                        self.video_scroll_area.verticalScrollBar().setValue(int(new_scroll_y))
                    
                    event.accept()
                    return True
                    
                elif event.type() == event.Type.MouseButtonRelease and event.button() == Qt.LeftButton:
                    # End pan
                    self.pan_active = False
                    self.pan_start_pos = None
                    self.pan_start_scroll = None
                    # Restore cursor based on zoom state
                    if self.video_zoom_factor > 1.0:
                        self.video_widget.setCursor(Qt.OpenHandCursor)
                    else:
                        self.video_widget.setCursor(Qt.ArrowCursor)
                    event.accept()
                    return True
                    
                elif event.type() == event.Type.Enter and self.video_zoom_factor > 1.0:
                    # Change cursor to indicate pan is available
                    if not self.pan_active:
                        self.video_widget.setCursor(Qt.OpenHandCursor)
                    event.accept()
                    return True
                    
                elif event.type() == event.Type.Leave:
                    # Restore cursor when leaving video area
                    self.video_widget.setCursor(Qt.ArrowCursor)
                    event.accept()
                    return True
        return super().eventFilter(obj, event)

    def _load_video_for_current_file(self):
        if not self.current_tracking_path:
            return
        
        # Reset video zoom to default
        self.video_zoom_factor = 1.0
        base_width, base_height = self.video_base_size
        self.video_widget.setFixedSize(base_width, base_height)
        # Reset scroll position to top-left
        self.video_scroll_area.horizontalScrollBar().setValue(0)
        self.video_scroll_area.verticalScrollBar().setValue(0)
        # Reset pan state
        self.pan_active = False
        self.pan_start_pos = None
        self.pan_start_scroll = None
        self.video_widget.setCursor(Qt.ArrowCursor)
        # Reset marker selection
        self.selected_marker_index = -1
        
        # Find video file using selected camera ID
        mf4_dir = os.path.dirname(self.current_tracking_path)
        base_name = os.path.splitext(os.path.basename(self.current_tracking_path))[0].replace("_tracking", "")
        
        selected_cam = self.cb_camera_id.currentText() if self.cb_camera_id.count() > 0 else "cam1"
        video_name = f"{base_name}_{selected_cam}.avi"
        video_path = os.path.join(mf4_dir, video_name)
        if not os.path.exists(video_path):
            # Try parent dir
            parent_dir = os.path.dirname(mf4_dir)
            video_path = os.path.join(parent_dir, video_name)
        if os.path.exists(video_path):
            from PySide6.QtCore import QUrl
            self.media_player.setSource(QUrl.fromLocalFile(video_path))
            # Force first frame rendering by playing and pausing immediately
            self.media_player.play()
            QTimer.singleShot(150, self.media_player.pause)
            
            # Load filmstrip thumbnails
            self.filmstrip.set_thumbnails([]) # Clear old
            # Get video duration from media_player if available, else ThumbnailWorker will estimate
            self.filmstrip.video_duration = self.media_player.duration() / 1000.0
            
            self.thumb_worker = ThumbnailWorker(video_path)
            self.thumb_worker.finished.connect(self.filmstrip.set_thumbnails)
            self.thumb_worker.start()
            
            # Reset play button state
            self.btn_play_pause.set_playing(False)
            
            self.log_message.emit(f"Loaded video: {video_name}")
        else:
            self.log_message.emit(f"Video not found: {video_name}")

    def load_tracking_file(self, fpath_or_mdf):
        """Load an MF4 (path or MDF object) and populate selectors. Does not change Logic config."""
        from asammdf import MDF
        
        # Update combo box selection without triggering reload
        if isinstance(fpath_or_mdf, str):
            self.combo_cases.blockSignals(True)
            idx = self.combo_cases.findData(fpath_or_mdf)
            if idx >= 0:
                self.combo_cases.setCurrentIndex(idx)
            self.combo_cases.blockSignals(False)
            
            # Update Navigation Labels logic is in AnalysisWidget, but we are in TimeSelectorWidget.
            # We need to notify parent. We emit 'tracking_loaded'.
            if self.tracking_loaded:
                self.tracking_loaded.emit(fpath_or_mdf)
            
        if isinstance(fpath_or_mdf, MDF):
            # Already-loaded MDF object — apply directly (extract data synchronously since already loaded)
            self.current_tracking_path = None
            self._apply_tracking_mdf(fpath_or_mdf, None, None)
            return
        
        # File path — load in background thread
        try:
            self.current_tracking_path = str(fpath_or_mdf)
        except Exception:
            self.current_tracking_path = None
        
        # Keep reference to prevent GC and stop previous worker if any
        if hasattr(self, '_mdf_loader') and self._mdf_loader is not None:
            self._mdf_loader.quit()
            self._mdf_loader.wait()
        
        self._mdf_loader = MdfLoaderWorker(str(fpath_or_mdf), numeric_only=True)
        self._mdf_loader.loaded.connect(self._on_tracking_mdf_loaded)
        self._mdf_loader.error.connect(self._on_tracking_mdf_error)
        self._mdf_loader.start()

    def _on_tracking_mdf_loaded(self, mdf, signal_names, signal_data):
        """Callback when MDF file is loaded off the main thread."""
        self._apply_tracking_mdf(mdf, signal_names, signal_data)

    def _on_tracking_mdf_error(self, error_msg):
        """Callback when MDF loading fails."""
        QMessageBox.critical(self, "Error", f"Could not load tracking MF4:\n{error_msg}")

    def _apply_tracking_mdf(self, mdf, signal_names, signal_data):
        """Apply a loaded MDF to the UI - always runs on main thread."""
        self.mdf = mdf
        
        if signal_names is None:
            # MDF object was passed directly — extract signals now (already on main thread for pre-loaded objects)
            sigs = []
            for ch in self.mdf.iter_channels():
                name = ch.name
                try:
                    if ch.samples is None:
                        continue
                    kind = getattr(ch.samples, 'dtype', None)
                    if kind is not None and getattr(ch.samples.dtype, 'kind', 'f') in 'fiudb':
                        sigs.append(name)
                except Exception:
                    try:
                        float(ch.samples[0])
                        sigs.append(name)
                    except Exception:
                        pass
            signal_names = sorted(sigs)
            # Extract data synchronously for pre-loaded MDF objects
            signal_data = {}
            for name in signal_names:
                try:
                    sig = mdf.get(name)
                    signal_data[name] = (
                        np.array(sig.timestamps),
                        np.array(sig.samples, dtype=float)
                    )
                except Exception:
                    pass

        # Store cached data for _plot_signals
        self.signal_data_cache = signal_data if signal_data else {}

        self.available_signals = signal_names
        self.cb_top.blockSignals(True)
        self.cb_bottom.blockSignals(True)
        self.cb_top.clear(); self.cb_bottom.clear()
        self.cb_top.addItems(self.available_signals)
        self.cb_bottom.addItems(self.available_signals)
        self.cb_top.blockSignals(False)
        self.cb_bottom.blockSignals(False)

        # Set default signals: Head_H_Angle/Head_V_Angle if both present, else H_Ratio/V_Ratio if both present, else fallback to first two
        idx_top = idx_bottom = -1
        if "Head_H_Angle" in self.available_signals and "Head_V_Angle" in self.available_signals:
            idx_top = self.available_signals.index("Head_H_Angle")
            idx_bottom = self.available_signals.index("Head_V_Angle")
        elif "H_Ratio" in self.available_signals and "V_Ratio" in self.available_signals:
            idx_top = self.available_signals.index("H_Ratio")
            idx_bottom = self.available_signals.index("V_Ratio")
        else:
            if len(self.available_signals) >= 1:
                idx_top = 0
            if len(self.available_signals) >= 2:
                idx_bottom = 1
        if idx_top >= 0:
            self.cb_top.setCurrentIndex(idx_top)
        if idx_bottom >= 0:
            self.cb_bottom.setCurrentIndex(idx_bottom)

        # Show file name in label
        if self.current_tracking_path:
            # Update auto-centering
            try:
                self.plot_top.enableAutoRange()
                self.plot_bottom.enableAutoRange()
            except:
                pass
        else:
            pass

        self._plot_signals()
        # Load video if checkbox is checked
        if self.chk_video.isChecked():
            self._load_video_for_current_file()
        # Reset markers for this plot/file without persisting the temporary clear
        self._clear_markers(persist=False)
        # Restore marks (always)
        self._restore_marks()
        # Restore shaded regions for restored marks
        self._restore_shaded_regions()
        # Emit signal to parent UI so it can highlight corresponding participant
        try:
            if getattr(self, 'tracking_loaded', None) is not None and self.current_tracking_path:
                try:
                    self.tracking_loaded.emit(self.current_tracking_path)
                except Exception:
                    try:
                        self.tracking_loaded(self.current_tracking_path)
                    except Exception:
                        pass
        except Exception:
            pass

    def _plot_signals(self):
        if not self.mdf:
            return
        top_name = self.cb_top.currentText()
        bot_name = self.cb_bottom.currentText()
        self.plot_top.clear(); self.plot_bottom.clear()
        # re-add vlines
        self.plot_top.addItem(self.vline_top, ignoreBounds=True)
        self.plot_bottom.addItem(self.vline_bottom, ignoreBounds=True)
        # Use cached signal data from worker thread (no mdf.get() on main thread)
        cache = getattr(self, 'signal_data_cache', {})
        try:
            if top_name and top_name in cache:
                t, y = cache[top_name]
                self.plot_top.plot(t, y, pen=pg.mkPen('#00AAFF'))
                self.plot_top.setLabel('left', top_name)
            elif top_name:
                # Fallback if not in cache
                sig = self.mdf.get(top_name)
                t = np.array(sig.timestamps)
                y = np.array(sig.samples, dtype=float)
                self.plot_top.plot(t, y, pen=pg.mkPen('#00AAFF'))
                self.plot_top.setLabel('left', top_name)
            if bot_name and bot_name in cache:
                t2, y2 = cache[bot_name]
                self.plot_bottom.plot(t2, y2, pen=pg.mkPen('#00FF88'))
                self.plot_bottom.setLabel('left', bot_name)
            elif bot_name:
                sig2 = self.mdf.get(bot_name)
                t2 = np.array(sig2.timestamps)
                y2 = np.array(sig2.samples, dtype=float)
                self.plot_bottom.plot(t2, y2, pen=pg.mkPen('#00FF88'))
                self.plot_bottom.setLabel('left', bot_name)

            # --- EVALUATE AND PLOT PASS CRITERIA ---
            try:
                analysis = self.parent()
                while analysis and not hasattr(analysis, 'tab_logic'):
                    analysis = analysis.parent()
                if analysis and hasattr(analysis, 'tab_logic') and getattr(self, 'current_tracking_path', None):
                    config = analysis.tab_logic._collect_logic_config()
                    category = analysis.tab_logic._determine_category_from_filename(self.current_tracking_path)
                    pass_params = config.get('pass_criteria', {}).get(category)
                    if pass_params:
                        sig_pass = pass_params.get('signal')
                        op1 = pass_params.get('operator1')
                        val1 = pass_params.get('value1')
                        t_p, y_p = None, None
                        
                        if sig_pass in cache:
                            t_p, y_p = cache[sig_pass]
                        elif sig_pass:
                            try:
                                sig_obj = self.mdf.get(sig_pass)
                                t_p = np.array(sig_obj.timestamps)
                                y_p = np.array(sig_obj.samples, dtype=float)
                            except:
                                pass
                        
                        if t_p is not None and y_p is not None and len(y_p) > 0 and op1 not in [None, "None", ""]:
                            import operator
                            ops = {'>': operator.gt, '<': operator.lt, '>=': operator.ge, '<=': operator.le, '==': operator.eq, '!=': operator.ne}
                            op_func = ops.get(op1)
                            if op_func:
                                met_idx = np.where(op_func(y_p, float(val1)))[0]
                                if len(met_idx) > 0:
                                    pass_t = t_p[met_idx[0]]
                                    # Create dashed red line
                                    from PySide6.QtCore import Qt
                                    pen = pg.mkPen((255, 50, 50), width=2, style=Qt.DashLine)
                                    pass_line_top = pg.InfiniteLine(pos=pass_t, angle=90, movable=False, pen=pen, label='PASS', labelOpts={'position':0.1, 'color': (255,50,50), 'fill': (50,50,50,200), 'movable': True})
                                    pass_line_bot = pg.InfiniteLine(pos=pass_t, angle=90, movable=False, pen=pen, label='PASS', labelOpts={'position':0.1, 'color': (255,50,50), 'fill': (50,50,50,200), 'movable': True})
                                    self.plot_top.addItem(pass_line_top, ignoreBounds=True)
                                    self.plot_bottom.addItem(pass_line_bot, ignoreBounds=True)
            except Exception as e:
                print(f"Error plotting PASS criteria: {e}")

            # link X
            try:
                self.plot_bottom.getPlotItem().vb.setXLink(self.plot_top.getPlotItem().vb)
            except Exception:
                pass
        except Exception as e:
            print(f"Plot error: {e}")

    def _on_mouse_moved(self, evt):
        # evt is a QPointF in scene coordinates
        try:
            pos = evt
            # map to top viewbox
            vb = self.plot_top.getPlotItem().vb
            mousePoint = vb.mapSceneToView(pos)
            x = mousePoint.x()
            self.vline_top.setPos(x); self.vline_bottom.setPos(x)
            if hasattr(self, 'filmstrip'):
                self.filmstrip.set_cursor(x)
            # tooltip
            from PySide6.QtWidgets import QToolTip
            from PySide6.QtGui import QCursor
            QToolTip.showText(QCursor.pos(), f"{x:.3f}")
            # Sync video only when enabled
            if self.chk_video.isChecked() and self.btn_sync_video.isChecked() and self.media_player.hasVideo():
                self.media_player.setPosition(int(x * 1000))  # x in seconds, position in ms
        except Exception:
            pass

    def _on_mouse_clicked(self, evt):
        """Add new marker on left click."""
        try:
            ev = evt
            if ev.button() == Qt.LeftButton:
                pos = ev.scenePos()
                vb = self.plot_top.getPlotItem().vb
                x = vb.mapSceneToView(pos).x()
                self._add_marker(x)
        except Exception:
            pass

    def _add_marker(self, x):
        """Create new marker line in both plots at position x."""
        import pyqtgraph as pg
        pen = pg.mkPen('#ff9800', width=2)
        hover_pen = pg.mkPen('#ffb74d', width=3)
        line1 = pg.InfiniteLine(pos=x, angle=90, movable=True, pen=pen, hoverPen=hover_pen)
        line2 = pg.InfiniteLine(pos=x, angle=90, movable=True, pen=pen, hoverPen=hover_pen)
        
        line1.setCursor(Qt.SizeHorCursor)
        line2.setCursor(Qt.SizeHorCursor)
        
        line1.sigClicked.connect(self._on_line_clicked)
        line2.sigClicked.connect(self._on_line_clicked)
        
        self.plot_top.addItem(line1, ignoreBounds=True)
        self.plot_bottom.addItem(line2, ignoreBounds=True)
        self.markers.append((line1, line2, x))
        line1.sigPositionChanged.connect(lambda *args, l=line1: self._on_marker_moved(l))
        line2.sigPositionChanged.connect(lambda *args, l=line2: self._on_marker_moved(l))
        
        self.log_message.emit(f"Mark added at T={x:.4f}s")
        
        num_markers = len(self.markers)
        if num_markers >= 2 and num_markers % 2 == 0:
            prev_x = self.markers[-2][2]
            curr_x = x
            x1, x2 = sorted((prev_x, curr_x))
            from pyqtgraph import LinearRegionItem
            region_top = LinearRegionItem(
                values=[x1, x2],
                movable=False,
                brush=pg.mkBrush(255, 152, 0, 50)
            )
            region_bottom = LinearRegionItem(
                values=[x1, x2],
                movable=False,
                brush=pg.mkBrush(255, 152, 0, 50)
            )
            self.plot_top.addItem(region_top)
            self.plot_bottom.addItem(region_bottom)
            self.shaded_regions.append((region_top, region_bottom))
            self.log_message.emit(f"Interval created: {x1:.4f}s to {x2:.4f}s (Duration: {x2-x1:.4f}s)")
        
        self._update_marks_file()
        self.undo_stack.append({'type': 'add_marker'})
        self.btn_undo.setEnabled(True)


    def _find_marker_index_at_x(self, x, tolerance=None):
        if tolerance is None:
            try:
                x_range = self.plot_top.getPlotItem().vb.viewRange()[0]
                tolerance = max(1e-6, (x_range[1] - x_range[0]) * 0.01)
            except Exception:
                tolerance = 1e-3
        for idx, (_, _, mx) in enumerate(self.markers):
            if abs(mx - x) <= tolerance:
                return idx
        return None

    def _rebuild_shaded_regions(self):
        for region_top, region_bottom in self.shaded_regions:
            try:
                self.plot_top.removeItem(region_top)
            except Exception:
                pass
            try:
                self.plot_bottom.removeItem(region_bottom)
            except Exception:
                pass
        self.shaded_regions = []
        from pyqtgraph import LinearRegionItem
        for i in range(1, len(self.markers), 2):
            x1, x2 = sorted((self.markers[i-1][2], self.markers[i][2]))
            region_top = LinearRegionItem(
                values=[x1, x2],
                movable=False,
                brush=pg.mkBrush(255, 152, 0, 50)
            )
            region_bottom = LinearRegionItem(
                values=[x1, x2],
                movable=False,
                brush=pg.mkBrush(255, 152, 0, 50)
            )
            self.plot_top.addItem(region_top)
            self.plot_bottom.addItem(region_bottom)
            self.shaded_regions.append((region_top, region_bottom))

    def _remove_marker_by_index(self, index):
        if index < 0 or index >= len(self.markers):
            return
        l1, l2, x = self.markers.pop(index)
        try:
            self.plot_top.removeItem(l1)
        except Exception:
            pass
        try:
            self.plot_bottom.removeItem(l2)
        except Exception:
            pass
        
        # Adjust selection index
        if self.selected_marker_index == index:
            # Removed marker was selected
            self.selected_marker_index = -1
        elif self.selected_marker_index > index:
            # Adjust index if marker was after the removed one
            self.selected_marker_index -= 1
        
        self._rebuild_shaded_regions()
        self._update_marks_file()
        self.log_message.emit(f"Marker removed at T={x:.4f}s")

    def _on_marker_moved(self, line):
        # Update position in markers list and mirror the paired line
        x = line.value()
        for i, (l1, l2, old_x) in enumerate(self.markers):
            if l1 is line or l2 is line:
                # Move the paired line to the same x if one was dragged
                if l1 is line:
                    l2.setPos(x)
                else:
                    l1.setPos(x)
                self.markers[i] = (l1, l2, x)
                break
        self._rebuild_shaded_regions()
        self._update_marks_file()

    def _is_line_over_delete(self, line):
        # Check if line is dragged over the delete button
        try:
            # Get line position in scene coordinates
            scene_pos = self.plot_top.plotItem.vb.mapViewToScene(line.value(), 0)
            # Map to widget coordinates
            widget_pos = self.plot_top.mapFromScene(scene_pos)
            # Map to global coordinates
            global_pos = self.plot_top.mapToGlobal(widget_pos)
            # Get delete button global rect
            delete_global_rect = QRect(self.btn_delete.mapToGlobal(QPoint(0, 0)), self.btn_delete.size())
            # Check if global_pos is inside the delete button rect
            return delete_global_rect.contains(global_pos)
        except Exception as e:
            print(f"Error in _is_line_over_delete: {e}")
            return False

    def _clear_markers(self, persist=True):
        for l1, l2, _ in self.markers:
            try:
                self.plot_top.removeItem(l1)
            except Exception:
                pass
            try:
                self.plot_bottom.removeItem(l2)
            except Exception:
                pass
        self.markers = []
        
        # Also clear all shaded regions
        for region_top, region_bottom in self.shaded_regions:
            try:
                self.plot_top.removeItem(region_top)
            except Exception:
                pass
            try:
                self.plot_bottom.removeItem(region_bottom)
            except Exception:
                pass
        self.shaded_regions = []
        
        # Reset marker selection
        self.selected_marker_index = -1
        
        if persist:
            # Persist cleared marks so marks.json stays in sync with the view
            self._update_marks_file()
    
    # Stub methods for compatibility with AnalysisWidget
    def update_oem_logo(self, oem_name):
        """Compatibility stub - logos are handled in matplotlib report."""
        pass
    
    def update_vehicle_text(self, text):
        """Compatibility stub - vehicle text is handled in matplotlib report."""
        pass
    
    def update_audio_params(self, min_f, max_f, thresh):
        self.audio_min_freq = min_f
        self.audio_max_freq = max_f
        self.audio_threshold = thresh

    def reset_graph_view(self):
        """Reset the plot view to auto-range."""
        try:
             self.plot_top.autoRange()
             self.plot_bottom.autoRange()
             self.log_message.emit("Graph view reset.")
        except Exception:
             pass

class AnalysisWidget(QWidget):
    # Signal to notify MainWindow about busy state (for global spinner)
    busy_changed = Signal(bool)
    
    def __init__(self, parent=None):
        super().__init__(parent)
        self.layout_main = QHBoxLayout(self)
        self.layout_main.setContentsMargins(10, 10, 10, 10)
        
        self.splitter = QSplitter(Qt.Horizontal)
        
        self.left_widget = QWidget()
        self.left_layout = QVBoxLayout(self.left_widget)
        self.left_layout.setContentsMargins(0, 0, 0, 0)
        self.left_layout.setSpacing(10)
        
        self.tabs_left = QTabWidget()
        self.tabs_left.tabBar().setCursor(Qt.PointingHandCursor)
        
        self.create_tab_source()
        self.create_tab_micro()
        self.create_tab_report()
        
        self.create_group_participants()
        
        self.left_layout.addWidget(self.tabs_left)
        self.left_layout.addWidget(self.grp_participants, 1)
        
        self.right_widget = QWidget()
        self.right_layout = QVBoxLayout(self.right_widget)
        self.right_layout.setContentsMargins(0, 0, 0, 0)
        
        self.create_tabs_right()

        self.splitter.addWidget(self.left_widget)
        self.splitter.addWidget(self.right_widget)
        self.splitter.setStretchFactor(0, 1)
        self.splitter.setStretchFactor(1, 2)
        
        self.layout_main.addWidget(self.splitter)
        
        setup_tab_icon_switching(self.tabs_left, [
            ("hub_16dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png", "hub_16dp_FFFFFF_FILL1_wght400_GRAD0_opsz20.png"),
            ("mic_16dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png", "mic_16dp_FFFFFF_FILL1_wght400_GRAD0_opsz20.png"),
            ("file_png_18dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png", "file_png_18dp_FFFFFF_FILL1_wght400_GRAD0_opsz20.png")
        ])
        
        setup_tab_icon_switching(self.tabs_right, [
            ("ar_on_you_16dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png", "ar_on_you_16dp_FFFFFF_FILL1_wght400_GRAD0_opsz20.png"),
            ("point_scan_16dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png", "point_scan_16dp_FFFFFF_FILL1_wght400_GRAD0_opsz20.png"),
            ("flowchart_16dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png", "flowchart_16dp_FFFFFF_FILL1_wght400_GRAD0_opsz20.png"),
            ("terminal_2_16dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png", "terminal_2_16dp_FFFFFF_FILL1_wght400_GRAD0_opsz20.png"),
            ("brand_awareness_16dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png", "brand_awareness_16dp_FFFFFF_FILL1_wght400_GRAD0_opsz20.png")
        ])
        
        self.scanner_thread = None
        self.autodetect_thread = None
        self.chronos_worker = None
        self._chronos_stop_requested = False
        self._report_queue = []
        self._report_wait_for_marks = False
        self._stop_report_after_current = False

    def log(self, msg):
        ts = datetime.now().strftime("%H:%M:%S")
        self.txt_log.append(f"[{ts}] {msg}")
        sb = self.txt_log.verticalScrollBar()
        sb.setValue(sb.maximum())

    def create_tabs_right(self):
        self.tabs_right = QTabWidget()
        self.tabs_right.tabBar().setCursor(Qt.PointingHandCursor)
        self.tabs_right.setIconSize(QSize(20, 20))
        
        # === 1. Tracking (was ChronosID — merged with old left Tracking tab) ===
        self.tab_chronos = QWidget()
        self.l_chronos = QVBoxLayout(self.tab_chronos)
        self.l_chronos.setContentsMargins(5, 5, 5, 5)
        self.l_chronos.setSpacing(5)
        
        # Top bar: [Engine | File | Frames]  ---  [Camera ID | Spinbox]  ---  [Start/Stop]
        self.chronos_top_bar = QHBoxLayout()
        
        # Left group: Engine + File + Frames
        self.lbl_engine = QLabel("ENGINE")
        self.lbl_engine.setStyleSheet("background-color: #333; color: #0f0; font-weight: bold; padding: 4px 10px; border-radius: 4px;")
        self.lbl_engine.setFixedWidth(80)
        self.lbl_engine.setAlignment(Qt.AlignCenter)
        self.lbl_file = QLabel("No file")
        self.lbl_file.setStyleSheet("color: #aaa; font-size: 10pt;")
        self.lbl_frames = QLabel("0 / 0")
        self.lbl_frames.setStyleSheet("color: #888; font-size: 10pt;")
        self.lbl_frames.setFixedWidth(100)
        self.lbl_frames.setAlignment(Qt.AlignRight)
        self.chronos_top_bar.addWidget(self.lbl_engine)
        self.chronos_top_bar.addWidget(self.lbl_file, 1)
        self.chronos_top_bar.addWidget(self.lbl_frames)
        self.chronos_top_bar.addSpacing(20)
        
        # Center group: Camera ID
        lbl_cam = QLabel("Camera ID:")
        lbl_cam.setStyleSheet("font-weight: bold;")
        self.spin_cam = QSpinBox()
        self.spin_cam.setStyleSheet(LOGIC_INPUT_STYLE)
        self.spin_cam.setRange(1, 4)
        self.chronos_top_bar.addWidget(lbl_cam)
        self.chronos_top_bar.addWidget(self.spin_cam)
        self.chronos_top_bar.addSpacing(20)
        
        # Right group: Dynamic Start/Stop button
        self._tracking_running = False
        self.btn_tracking_toggle = QPushButton("  Start")
        self.btn_tracking_toggle.setIcon(QIcon(resource_path("assets/icons/play_circle_16dp_000000_FILL0_wght400_GRAD0_opsz20.png")))
        self.btn_tracking_toggle.setCursor(Qt.PointingHandCursor)
        self.btn_tracking_toggle.setMinimumHeight(32)
        self.btn_tracking_toggle.setMinimumWidth(100)
        self.btn_tracking_toggle.setStyleSheet(f"background-color: {IDIADA_ORANGE}; color: black; font-weight: bold; border-radius: 4px; padding: 4px 12px;")
        self.btn_tracking_toggle.clicked.connect(self.toggle_tracking)
        self.chronos_top_bar.addWidget(self.btn_tracking_toggle)
        self.l_chronos.addLayout(self.chronos_top_bar)
        
        # Video display
        self.lbl_video = QLabel("Waiting for video...")
        self.lbl_video.setAlignment(Qt.AlignCenter)
        self.lbl_video.setStyleSheet("background-color: #000; color: #555; border: 1px solid #333;")
        self.lbl_video.setMinimumSize(640, 360)
        self.l_chronos.addWidget(self.lbl_video, 1)
        
        # Bottom stats bar: H value | V value | FPS | Progress
        self.chronos_stats_bar = QHBoxLayout()
        self.lbl_h_val = QLabel("H: --")
        self.lbl_h_val.setStyleSheet("color: #ff9800; font-weight: bold; font-size: 11pt;")
        self.lbl_v_val = QLabel("V: --")
        self.lbl_v_val.setStyleSheet("color: #2196f3; font-weight: bold; font-size: 11pt;")
        self.lbl_proc_fps = QLabel("0 fps")
        self.lbl_proc_fps.setStyleSheet("color: #4caf50; font-size: 10pt;")
        self.lbl_proc_fps.setFixedWidth(80)
        self.progress_chronos = QProgressBar()
        self.progress_chronos.setValue(0)
        self.progress_chronos.setTextVisible(True)
        self.progress_chronos.setStyleSheet("QProgressBar { height: 16px; }")
        self.chronos_stats_bar.addWidget(self.lbl_h_val)
        self.chronos_stats_bar.addWidget(self.lbl_v_val)
        self.chronos_stats_bar.addStretch()
        self.chronos_stats_bar.addWidget(self.lbl_proc_fps)
        self.chronos_stats_bar.addWidget(self.progress_chronos, 1)
        self.l_chronos.addLayout(self.chronos_stats_bar)
        
        icon_tracking = QIcon(resource_path("assets/icons/ar_on_you_16dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png"))
        self.tabs_right.addTab(self.tab_chronos, icon_tracking, "Tracking")
        
        # === 2. Time Selector ===
        icon_selector = QIcon(resource_path("assets/icons/point_scan_16dp_FFFFFF_FILL1_wght400_GRAD0_opsz20.png"))
        from PySide6.QtWidgets import QToolTip
        try:
            self.tab_time = TimeSelectorWidget()
            self.tabs_right.addTab(self.tab_time, icon_selector, "Time Selector")
            try:
                self.tab_time.btn_next_plot.clicked.connect(self.next_plot)
                self.tab_time.btn_prev_plot.clicked.connect(self.previous_plot)
                
                from PySide6.QtGui import QShortcut, QKeySequence
                self.shortcut_next = QShortcut(QKeySequence(Qt.Key_Tab), self.tab_time)
                self.shortcut_next.setContext(Qt.WidgetWithChildrenShortcut)
                self.shortcut_next.activated.connect(self.next_plot)
                
                self.shortcut_prev = QShortcut(QKeySequence(Qt.Key_Backtab), self.tab_time)
                self.shortcut_prev.setContext(Qt.WidgetWithChildrenShortcut)
                self.shortcut_prev.activated.connect(self.previous_plot)
            except Exception:
                pass
        except Exception as e:
            error_widget = QWidget()
            l = QVBoxLayout(error_widget)
            lbl = QLabel(f"Error loading Time Selector: {e}")
            lbl.setStyleSheet("color: red; font-weight: bold; padding: 20px;")
            l.addWidget(lbl)
            self.tabs_right.addTab(error_widget, icon_selector, "Time Selector")
        
        # === 4. Logic ===
        icon_logic = QIcon(resource_path("assets/icons/flowchart_16dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png"))
        self.tab_logic = LogicTabWidget()
        self.tab_logic.btn_generate_report.clicked.connect(self.toggle_generate_reports)
        self.tabs_right.addTab(self.tab_logic, icon_logic, "Logic")
        # Connect TimeSelectorWidget tracking_loaded signal
        try:
            if hasattr(self.tab_time, 'tracking_loaded') and getattr(self.tab_time, 'tracking_loaded', None) is not None:
                try:
                    self.tab_time.tracking_loaded.connect(self._on_tracking_loaded)
                    if hasattr(self.tab_time, 'log_message'):
                         self.tab_time.log_message.connect(self.log)
                except Exception:
                    try:
                        self.tab_time.tracking_loaded = lambda p: self._on_tracking_loaded(p)
                    except Exception:
                        pass
        except Exception:
            pass
        
        # === 5. AI Brain Training ===
        icon_brain = QIcon(resource_path("assets/icons/brand_awareness_16dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png"))
        self.tab_brain = AIBrainWidget()
        self.tabs_right.addTab(self.tab_brain, icon_brain, "AI Brain")

        # === 6. Log ===
        self.tab_log = QWidget()
        self.l_log = QVBoxLayout(self.tab_log)
        self.l_log.setContentsMargins(0,0,0,0)
        self.txt_log = QTextEdit()
        self.txt_log.setReadOnly(True)
        self.txt_log.setStyleSheet("background-color: #1e1e1e; color: #ccc; font-family: Consolas; font-size: 10pt;")
        self.l_log.addWidget(self.txt_log)
        
        icon_log = QIcon(resource_path("assets/icons/terminal_2_16dp_FFFFFF_FILL1_wght400_GRAD0_opsz20.png"))
        self.tabs_right.addTab(self.tab_log, icon_log, "Log")
        
        self.right_layout.addWidget(self.tabs_right)

    # ... Group Creations (Sources, Micro, Report) ...
    def create_tab_source(self):
        tab = QWidget()
        l_src = QVBoxLayout()
        l_src.setContentsMargins(15, 15, 15, 15)
        l_src.setSpacing(10)
        
        self.txt_source = QLineEdit()
        self.txt_source.setPlaceholderText("Select data source...")
        self.txt_source.textChanged.connect(self.on_source_changed)
        
        btn_browse = QPushButton(" Select Folder")
        btn_browse.setIcon(QIcon(resource_path("assets/icons/folder_16dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png")))
        btn_browse.setFixedHeight(35)
        btn_browse.setCursor(Qt.PointingHandCursor)
        btn_browse.setToolTip("Browse for folder")
        btn_browse.setStyleSheet("font-weight: normal;")
        btn_browse.clicked.connect(self.browse_source)
        
        self.btn_refresh = QPushButton(" Refresh Path")
        self.btn_refresh.setIcon(QIcon(resource_path("assets/icons/refresh_16dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png")))
        self.btn_refresh.setFixedHeight(35)
        self.btn_refresh.setCursor(Qt.PointingHandCursor)
        self.btn_refresh.setToolTip("Refresh participant status")
        self.btn_refresh.setStyleSheet("font-weight: normal;")
        self.btn_refresh.clicked.connect(self.refresh_participants)
        
        l_src.addWidget(self.txt_source)
        l_src.addWidget(btn_browse)
        l_src.addWidget(self.btn_refresh)
        
        l_src.addStretch()
        tab.setLayout(l_src)
        self.tabs_left.addTab(tab, QIcon(resource_path("assets/icons/hub_16dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png")), "Source")
        
    def _notify(self, text, notif_type="info", duration=3000):
        mw = self.window()
        if hasattr(mw, "show_global_notification"):
            mw.show_global_notification(text, notif_type, duration)

    def toggle_tracking(self):
        """Toggle between Start and Stop tracking states."""
        if self._tracking_running:
            self.stop_chronos_tracking()
        else:
            self.start_chronos_tracking()
    
    def _set_tracking_button_start(self):
        """Set button to idle Start state."""
        self._tracking_running = False
        self.btn_tracking_toggle.setText("  Start")
        self.btn_tracking_toggle.setIcon(QIcon(resource_path("assets/icons/play_circle_16dp_000000_FILL0_wght400_GRAD0_opsz20.png")))
        self.btn_tracking_toggle.setStyleSheet(f"background-color: {IDIADA_ORANGE}; color: black; font-weight: bold; border-radius: 4px; padding: 4px 12px;")
        self.btn_tracking_toggle.setEnabled(True)
    
    def _set_tracking_button_stop(self):
        """Set button to running Stop state."""
        self._tracking_running = True
        self.btn_tracking_toggle.setText("  Stop")
        self.btn_tracking_toggle.setIcon(QIcon(resource_path("assets/icons/stop_circle_16dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png")))
        self.btn_tracking_toggle.setStyleSheet("background-color: #d1242f; color: white; font-weight: bold; border-radius: 4px; padding: 4px 12px;")
        self.btn_tracking_toggle.setEnabled(True)

    def create_tab_micro(self):
        tab = QWidget()
        l = QVBoxLayout()
        l.setContentsMargins(15, 15, 15, 15)
        
        row1 = QHBoxLayout()
        self.spin_min_freq = QSpinBox()
        self.spin_min_freq.setStyleSheet(LOGIC_INPUT_STYLE)
        self.spin_min_freq.setRange(0, 20000)
        self.spin_min_freq.setValue(20)
        self.spin_min_freq.setSuffix(" Hz")
        self.spin_min_freq.setPrefix("Min: ")
        self.spin_max_freq = QSpinBox()
        self.spin_max_freq.setStyleSheet(LOGIC_INPUT_STYLE)
        self.spin_max_freq.setRange(0, 20000)
        self.spin_max_freq.setValue(20000)
        self.spin_max_freq.setSuffix(" Hz")
        self.spin_max_freq.setPrefix("Max: ")
        row1.addWidget(self.spin_min_freq)
        row1.addWidget(self.spin_max_freq)
        
        row2 = QHBoxLayout()
        self.spin_threshold = QDoubleSpinBox()
        self.spin_threshold.setStyleSheet(LOGIC_INPUT_STYLE)
        self.spin_threshold.setRange(0.0, 100.0)
        self.spin_threshold.setSingleStep(0.1)
        self.spin_threshold.setPrefix("Thresh: ")
        self.btn_autodetect = QPushButton("Autodetect")
        icon_path = resource_path("assets/icons/hdr_auto_16dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png")
        if os.path.exists(icon_path): self.btn_autodetect.setIcon(QIcon(icon_path))
        self.btn_autodetect.setCursor(Qt.PointingHandCursor)
        self.btn_autodetect.clicked.connect(self.on_autodetect_click)
        
        self.spin_min_freq.valueChanged.connect(self.on_audio_params_changed)
        self.spin_max_freq.valueChanged.connect(self.on_audio_params_changed)
        self.spin_threshold.valueChanged.connect(self.on_audio_params_changed)
        
        row2.addWidget(self.spin_threshold)
        row2.addWidget(self.btn_autodetect)
        row2.addSpacing(10)
        
        l.addLayout(row1)
        l.addLayout(row2)
        l.addStretch()
        tab.setLayout(l)
        self.tabs_left.addTab(tab, QIcon(resource_path("assets/icons/mic_16dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png")), "Micro")

    def on_audio_params_changed(self):
        """Propagate audio settings to Logic Tab for visualization."""
        if hasattr(self, 'tab_logic'):
            self.tab_logic.update_audio_params(
                self.spin_min_freq.value(),
                self.spin_max_freq.value(),
                self.spin_threshold.value()
            )

    def create_tab_report(self):
        tab = QWidget()
        main_layout = QVBoxLayout(tab)
        main_layout.setContentsMargins(15, 15, 15, 15)
        main_layout.setSpacing(10)

        # Use FadeStackedWidget for premium feel transition
        self.report_stack = FadeStackedWidget()
        
        # --- PAGE 1: OEM, Vehicle, Track ---
        page1 = QWidget()
        l1 = QVBoxLayout(page1)
        l1.setContentsMargins(0, 0, 0, 0)
        
        f1 = QFormLayout()
        f1.setVerticalSpacing(15)
        f1.setHorizontalSpacing(10)
        f1.setLabelAlignment(Qt.AlignRight | Qt.AlignVCenter)

        self.combo_oem = QComboBox()
        self.populate_oem_combo()
        self.combo_oem.setPlaceholderText("Select OEM")
        self.combo_oem.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        self.txt_vehicle = QLineEdit()
        self.txt_vehicle.setPlaceholderText("VW Golf 8")
        self.txt_vehicle.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        self.combo_track = QComboBox()
        self.combo_track.setEditable(True)
        self.populate_track_combo()
        self.combo_track.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        for text, field in [("OEM:", self.combo_oem), 
                            ("Vehicle:", self.txt_vehicle), 
                            ("Track:", self.combo_track)]:
            lbl = QLabel(text)
            lbl.setStyleSheet("font-weight: bold; color: #bbb;")
            f1.addRow(lbl, field)
            
        l1.addLayout(f1)
        l1.addStretch()

        # --- PAGE 2: Engineer, Analyst, Euro NCAP ---
        page2 = QWidget()
        l2 = QVBoxLayout(page2)
        l2.setContentsMargins(0, 0, 0, 0)
        
        f2 = QFormLayout()
        f2.setVerticalSpacing(15)
        f2.setHorizontalSpacing(10)
        f2.setLabelAlignment(Qt.AlignRight | Qt.AlignVCenter)

        self.txt_engineer = QLineEdit()
        self.txt_engineer.setPlaceholderText("Firstname Lastname")
        self.txt_engineer.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        self.txt_analyst = QLineEdit()
        self.txt_analyst.setPlaceholderText("Firstname Lastname")
        self.txt_analyst.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        self.toggle_ncap = AnimatedToggle()

        for text, field in [("Engineer:", self.txt_engineer), 
                            ("Analyst:", self.txt_analyst), 
                            ("Euro NCAP:", self.toggle_ncap)]:
            lbl = QLabel(text)
            lbl.setStyleSheet("font-weight: bold; color: #bbb;")
            f2.addRow(lbl, field)
            
        l2.addLayout(f2)
        l2.addStretch()

        self.report_stack.addWidget(page1)
        self.report_stack.addWidget(page2)
        
        main_layout.addWidget(self.report_stack)

        # --- Navigation Controls (Visible always) ---
        h_nav = QHBoxLayout()
        
        self.btn_report_prev = QPushButton(" Previous")
        self.btn_report_next = QPushButton(" Next")
        
        for btn in [self.btn_report_prev, self.btn_report_next]:
            btn.setCursor(Qt.PointingHandCursor)
            btn.setMinimumSize(100, 32)
            btn.setStyleSheet("""
                QPushButton { 
                    background-color: #333; 
                    border: 1px solid #555;
                    border-radius: 6px; 
                    padding: 4px 10px; 
                    font-weight: bold;
                    color: white;
                } 
                QPushButton:hover { background-color: #444; border-color: #777; }
                QPushButton:pressed { background-color: #222; }
                QPushButton:disabled { background-color: #222; color: #555; border-color: #333; }
            """)

        icon_path = resource_path("assets/icons/keyboard_arrow_up_16dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png")
        if os.path.exists(icon_path):
            pix = QPixmap(icon_path)
            # Next: Rotate 90
            trans_next = QTransform().rotate(90)
            self.btn_report_next.setIcon(QIcon(pix.transformed(trans_next, Qt.SmoothTransformation)))
            # Prev: Rotate -90
            trans_prev = QTransform().rotate(-90)
            self.btn_report_prev.setIcon(QIcon(pix.transformed(trans_prev, Qt.SmoothTransformation)))
            
            self.btn_report_next.setIconSize(QSize(16, 16))
            self.btn_report_prev.setIconSize(QSize(16, 16))

        h_nav.addWidget(self.btn_report_prev)
        h_nav.addStretch()
        h_nav.addWidget(self.btn_report_next)
        main_layout.addLayout(h_nav)

        self.tabs_left.addTab(tab, QIcon(resource_path("assets/icons/file_png_18dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png")), "Report")
        
        # Connections
        self.btn_report_next.clicked.connect(lambda: self.report_stack.setCurrentIndex(1))
        self.btn_report_prev.clicked.connect(lambda: self.report_stack.setCurrentIndex(0))
        
        # Update button states
        def update_report_nav_buttons():
            idx = self.report_stack.currentIndex()
            self.btn_report_prev.setEnabled(idx > 0)
            self.btn_report_next.setEnabled(idx < self.report_stack.count() - 1)
            
        self.report_stack.currentChanged.connect(update_report_nav_buttons)
        update_report_nav_buttons() # Initial call
        
        self.combo_oem.currentTextChanged.connect(self.on_oem_changed)
        self.txt_vehicle.textChanged.connect(self.on_vehicle_changed)
        self.toggle_ncap.stateChanged.connect(self.on_ncap_toggled)
    
    def on_oem_changed(self, oem_name):
        """Propagate OEM selection to Logic Tab."""
        if hasattr(self, 'tab_logic'):
            self.tab_logic.update_oem_logo(oem_name)
    
    def on_vehicle_changed(self, text):
        """Propagate vehicle text to Logic Tab."""
        if hasattr(self, 'tab_logic'):
            self.tab_logic.update_vehicle_text(text)
    
    def on_ncap_toggled(self, state):
        # When Euro NCAP toggle is enabled, auto-select EuroNCAP as OEM and change report type.
        if state == Qt.Checked:
            # Buscar "EuroNCAP" en el combo OEM
            idx = self.combo_oem.findText("EuroNCAP")
            if idx >= 0:
                self.combo_oem.setCurrentIndex(idx)
            # Also set report type to Euro NCAP in Logic tab
            if hasattr(self, 'tab_logic') and hasattr(self.tab_logic, 'combo_report_type'):
                self.tab_logic.combo_report_type.setCurrentText("Euro NCAP")

    def create_group_participants(self):
        self.grp_participants = IconGroupBox("Participant Status", "group_search_16dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png", title_color="white", title_weight="500")
        
        # Add Collapse All button to header
        self.btn_collapse_all = QPushButton()
        collapse_icon = resource_path("assets/icons/collapse_all_16dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png")
        if os.path.exists(collapse_icon):
            self.btn_collapse_all.setIcon(QIcon(collapse_icon))
        else:
            self.btn_collapse_all.setText("▲")
        self.btn_collapse_all.setToolTip("Collapse All")
        self.btn_collapse_all.setFixedSize(22, 22)
        self.btn_collapse_all.setCursor(Qt.PointingHandCursor)
        self.btn_collapse_all.setStyleSheet("QPushButton { background: transparent; border: none; padding: 2px; } QPushButton:hover { background: rgba(255,255,255,0.1); border-radius: 4px; }")
        
        self.grp_participants.header_layout.addWidget(self.btn_collapse_all)
        
        l = QVBoxLayout()
        l.setContentsMargins(0, 5, 0, 0)
        self.grp_radios = QButtonGroup(self)
        self.row_radios_layout = QHBoxLayout()
        row_radios = self.row_radios_layout
        self.rb_all = QRadioButton("All")
        self.rb_none = QRadioButton("None")
        self.rb_inc_tracking = QRadioButton("Tracking")
        self.rb_inc_marks = QRadioButton("Marks")
        self.rb_inc_reports = QRadioButton("Report")
        grp_radios = self.grp_radios
        self.rb_all.setChecked(True)
        grp_radios.addButton(self.rb_all)
        grp_radios.addButton(self.rb_none)
        grp_radios.addButton(self.rb_inc_tracking)
        grp_radios.addButton(self.rb_inc_marks)
        grp_radios.addButton(self.rb_inc_reports)
        self.rb_all.clicked.connect(lambda: self.set_tree_checked(Qt.Checked))
        self.rb_none.clicked.connect(lambda: self.set_tree_checked(Qt.Unchecked))
        self.rb_inc_tracking.clicked.connect(self.select_incomplete_tracking)
        self.rb_inc_marks.clicked.connect(self.select_incomplete_marks)
        self.rb_inc_reports.clicked.connect(self.select_incomplete_reports)
        row_radios.addWidget(self.rb_all)
        row_radios.addWidget(self.rb_none)
        row_radios.addWidget(self.rb_inc_tracking)
        row_radios.addWidget(self.rb_inc_marks)
        row_radios.addWidget(self.rb_inc_reports)
        l.addLayout(row_radios)
        self.tree_participants = QTreeWidget()
        self.btn_collapse_all.clicked.connect(self.tree_participants.collapseAll)
        self.tree_participants.setHeaderLabels(["Structure", "Tracking", "Marks", "Report"])
        hdr = self.tree_participants.header()
        hdr.setSectionResizeMode(0, QHeaderView.Interactive)   # user can drag col 0
        hdr.setSectionResizeMode(1, QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(2, QHeaderView.ResizeToContents)
        hdr.setSectionResizeMode(3, QHeaderView.ResizeToContents)
        hdr.setStretchLastSection(False)          # don't force-stretch Report col
        self.tree_participants.setHorizontalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        self.tree_participants.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
        # Connect signal to detect when user checks/unchecks items
        self.tree_participants.itemChanged.connect(self._on_participant_check_changed)
        # Connect signal for expanding/collapsing to auto-resize col 0
        self.tree_participants.itemExpanded.connect(self.on_tree_item_expanded)
        self.tree_participants.itemCollapsed.connect(self.on_tree_item_expanded)
        
        # Permit drag and drop
        self.tree_participants.setDragEnabled(True)
        self.tree_participants.setSelectionMode(QAbstractItemView.SingleSelection)
        self.tree_participants.setDragDropMode(QAbstractItemView.DragOnly)
        
        l.addWidget(self.tree_participants)
        # Controls under participant tree: Next Plot / Auto-load
        # row_controls = QHBoxLayout()
        # self.btn_next_plot = QPushButton("Next Plot")
        # self.btn_next_plot.setToolTip("Load next participant _tracking file (Tab)")
        # self.btn_next_plot.setFixedHeight(28)
        # row_controls.addWidget(self.btn_next_plot)
        # row_controls.addStretch()
        # l.addLayout(row_controls)
        self.grp_participants.setLayout(l)
        # Internal tracking iteration state
        self._tracking_list = []
        self._tracking_index = 0


    # --- Populators --- 
    def populate_oem_combo(self):
        self.combo_oem.clear()
        logos_dir = resource_path("assets/logos")
        if os.path.exists(logos_dir):
            files = [f for f in os.listdir(logos_dir) if f.lower().endswith('.png')]
            files.sort(key=lambda x: x.lower())
            for f in files: self.combo_oem.addItem(os.path.splitext(f)[0])

    def populate_track_combo(self):
        self.combo_track.clear()
        hq = ["(0) Highway Loop", 
            "(0A) Highway Loop A", 
            "(0B) Highway Lopp B", 
            "(1) High-Speed Circuit", 
            "(2) External Noise Track",
            "(3) Fatigue/Comfort A",
            "(4) Dynamic Platform A", 
            "(5) Dry Handling Circuit",
            "(5B) Dynamic Platform C",  
            "(6) Test Hills", 
            "(7) Straight Line Braking", 
            "(7B) Comfort B & Sim City",
            "(8) Urban Area ADAS/CAV 2", 
            "(9) Dynamic Platform B", 
            "(10) Off-Road Track", 
            "(11) Wet Circle",
            "(12) Wet Handling Circuit", 
            "(13) Misuse Area", 
            "(14) ADAS/CAV 1",
            "(15) ADAS/CAV 3"]
        
        icpg = ["(1) High Speed Circuit", 
                    "(2) External Noise Track", 
                    "(3) Dynamic Platform", 
                    "(4), Straight Line Braking", 
                    "(4B) SLB Dry", 
                    "(5) NVH and Comfort", 
                    "(6) Multipurpose", 
                    "(7) Off-road", 
                    "(8) Dry Handling", 
                    "(9) Wet Handling", 
                    "(10) Wet Circle", 
                    "(11) Drift and pull", 
                    "(12) KERBS", 
                    "(13) Durability & Fatigue", 
                    "(14) General Road", 
                    "(15) Test Hills", 
                    "(16) SLB B", 
                    "(17) Bend Line Braking"]
        
        def natural_sort_key(s):
            return [int(text) if text.isdigit() else text.lower() for text in re.split('([0-9]+)', s)]

        hq.sort(key=natural_sort_key)
        icpg.sort(key=natural_sort_key)
        
        self.combo_track.addItem("---- HQ ----")
        for t in hq: self.combo_track.addItem(t)
        self.combo_track.addItem("---- ICPG ----")
        for t in icpg: self.combo_track.addItem(t)
        self.combo_track.setCurrentIndex(-1)

    def browse_source(self):
        # Browse for source folder.
        d = QFileDialog.getExistingDirectory(self, "Select Analysis Source")
        if d: self.txt_source.setText(d)
    
    def on_source_changed(self, text):
        if not text: return
        # Ensure marks.json exists in the project directory
        marks_path = os.path.join(text, 'marks.json')
        if not os.path.exists(marks_path):
            import json
            try:
                with open(marks_path, 'w', encoding='utf-8') as f:
                    json.dump({}, f, indent=2)
                self.log(f"Created new marks.json in {text}")
            except Exception as e:
                self.log(f"Could not create marks.json: {e}")
        self._current_project_source = text
        # Notify ReportingWidget of new project source via MainWindow
        try:
            mw = self.window()
            if mw and hasattr(mw, 'reporting_tab'):
                mw.reporting_tab.set_project_source(text)
        except Exception as e:
            self.log(f"Reporting update failed: {e}")
            
        self._notify(f"Data source configured: {os.path.basename(text)}", "success")
        self.log(f"Project source set to: {text}")
        self.scan_participants(text)

    def refresh_participants(self):
        # Refresh the participant tree.
        source = self.txt_source.text()
        if source and os.path.isdir(source):
            self.log("Refreshing participant status...")
            self.scan_participants(source)
        else:
            self.log("No valid source to refresh.")

    def on_tree_item_expanded(self, item=None):
        """Auto-resize col 0 to fit content after expand/collapse."""
        self.tree_participants.resizeColumnToContents(0)
        # Ensure col 0 is at least 120px wide so header is always readable
        if self.tree_participants.columnWidth(0) < 120:
            self.tree_participants.setColumnWidth(0, 120)


    # --- Autodetect Logic ---
    def on_autodetect_click(self):
        self.log("Opening file dialog for audio autodetection...")
        f, _ = QFileDialog.getOpenFileName(self, "Select MF4 for Audio Analysis", "", "MF4 Files (*.mf4)")
        if not f: return
        self.log(f"Selected file: {os.path.basename(f)}")
        self.btn_autodetect.setEnabled(False)
        self.autodetect_thread = AutodetectWorker(f)
        self.autodetect_thread.finished.connect(self.on_autodetect_result)
        self.autodetect_thread.error.connect(self.on_autodetect_error)
        self.autodetect_thread.log.connect(self.log)
        self.autodetect_thread.start()
        
    def on_autodetect_result(self, freq):
        self.btn_autodetect.setEnabled(True)
        freq = float(freq)
        self.log(f"✅ Peak Frequency Detected: {freq:.2f} Hz")
        self._notify(f"Autodetection finished ({freq:.2f} Hz)", "success")
        try:
            min_f = max(0, int(freq - 15))
            max_f = int(freq + 15)
            self.spin_min_freq.setValue(min_f)
            self.spin_max_freq.setValue(max_f)
            QMessageBox.information(self, "Autodetection Result", f"Detected Frequency: {freq:.2f} Hz\n\nFilter range set to {min_f} - {max_f} Hz")
        except: pass
            
    def on_autodetect_error(self, err):
        self.btn_autodetect.setEnabled(True)
        self.log(f"❌ Autodetection Error: {err}")
        self._notify("Autodetection failed.", "error")
        QMessageBox.critical(self, "Autodetection Failed", f"Error: {err}")

    # --- ChronosID Logic ---
    def start_chronos_tracking(self):
        # Collects selected files, resolves video paths, and starts ChronosWorker.
        camera_id = self.spin_cam.value()
        source_dir = self.txt_source.text()
        
        if not source_dir or not os.path.exists(source_dir):
            self.log("Error: No valid source directory selected.")
            return
        
        self._set_tracking_button_stop()
        self._chronos_stop_requested = False
        self.log("ChronosID: Collecting selected cases...")
        
        tasks = []
        iterator = QTreeWidgetItemIterator(self.tree_participants)
        
        while iterator.value():
            item = iterator.value()
            # Process only checked leaf items (files)
            if item.checkState(0) == Qt.Checked and item.childCount() == 0:
                mf4_path = item.data(0, Qt.UserRole)
                
                mf4_path_str = str(mf4_path) if mf4_path else ""
                if mf4_path_str.lower().endswith(".mf4"):
                    fname_clean = os.path.basename(mf4_path).replace("_tracking.mf4", ".mf4")
                    logic = ChronosManager.get_logic_for_file(fname_clean)
                    
                    if logic:
                        # Build video filename: D8_1.mf4 -> D8_1_cam2.avi
                        base_name = os.path.splitext(fname_clean)[0]
                        video_name = f"{base_name}_cam{camera_id}.avi"
                        
                        # Search for video in same directory as MF4
                        mf4_dir = os.path.dirname(mf4_path)
                        video_path = os.path.join(mf4_dir, video_name)
                        
                        # If not found, try parent directory (Pxx level)
                        if not os.path.exists(video_path):
                            parent_dir = os.path.dirname(mf4_dir)
                            video_path = os.path.join(parent_dir, video_name)
                        
                        # If still not found, try source_dir/Pxx structure
                        if not os.path.exists(video_path):
                            # Extract participant (P01, P02, etc.) from path
                            path_parts = mf4_path.replace("\\", "/").split("/")
                            for part in path_parts:
                                if re.match(r'^[A-Z]\d{2}$', part):
                                    video_path = os.path.join(source_dir, part, video_name)
                                    break
                        
                        # Log and add if found
                        if os.path.exists(video_path):
                            self.log(f"  ✓ Found: {video_name} -> {logic}")
                            tasks.append({'file_path': video_path, 'logic': logic})
                        else:
                            self.log(f"  ✗ Video not found: {video_name}")
                    else:
                        self.log(f"  ? No tracking logic for: {fname_clean}")
                        self.log(f"  ? No tracking logic for: {fname_clean}")
            
            iterator += 1
        
        if not tasks:
            self.log("ChronosID: No valid video files found. Check Camera ID and file structure.")
            self._set_tracking_button_start()
            return
        
        self.log(f"ChronosID: Starting processing of {len(tasks)} video(s)...")
        
        # Switch to Tracking tab and notify MainWindow
        self.tabs_right.setCurrentIndex(0)
        self.progress_chronos.setValue(0)
        self.busy_changed.emit(True)  # Start global spinner
        
        # Create and start worker with QueuedConnection for thread safety
        self.chronos_worker = ChronosWorker(tasks, camera_id)
        self.chronos_worker.new_frame.connect(self.update_chronos_video, Qt.QueuedConnection)
        self.chronos_worker.log.connect(self.log, Qt.QueuedConnection)
        self.chronos_worker.progress.connect(self.progress_chronos.setValue, Qt.QueuedConnection)
        self.chronos_worker.all_finished.connect(self.on_chronos_finished, Qt.QueuedConnection)
        self.chronos_worker.error.connect(self.on_chronos_error, Qt.QueuedConnection)
        self.chronos_worker.stats.connect(self.update_chronos_stats, Qt.QueuedConnection)
        self.chronos_worker.finished_task.connect(self.on_tracking_task_done, Qt.QueuedConnection)
        self.chronos_worker.start()

    def stop_chronos_tracking(self):
        # Stop ChronosID processing without saving partial outputs.
        if not self.chronos_worker:
            return
        try:
            self._chronos_stop_requested = True
            self.btn_tracking_toggle.setEnabled(False)
            self.log("ChronosID: Stop requested. Current video will be skipped.")
            self.chronos_worker.stop()
        except Exception:
            pass

    def update_chronos_video(self, qimg):
        # Update the video label with the new frame, scaled to fit.
        try:
            pixmap = QPixmap.fromImage(qimg)
            scaled = pixmap.scaled(self.lbl_video.size(), Qt.KeepAspectRatio, Qt.SmoothTransformation)
            self.lbl_video.setPixmap(scaled)
        except Exception as e:
            print(f"Frame display error: {e}")

    def on_chronos_finished(self):
        if self._chronos_stop_requested:
            self.log("ChronosID stopped by user.")
        else:
            self.log("ChronosID Analysis Completed.")
        self._set_tracking_button_start()
        if not self._chronos_stop_requested:
            self.progress_chronos.setValue(100)
        self.busy_changed.emit(False)  # Stop global spinner
        self.reset_chronos_overlay()
        if not self._chronos_stop_requested:
            QMessageBox.information(self, "ChronosID", "Tracking Analysis Completed!")
            # Auto-refresh participant tree to show updated status
            self.refresh_participants()

    def on_chronos_error(self, error_msg):
        # Handle errors from ChronosWorker with log and MessageBox.
        self.log(f"❌ Error: {error_msg}")
        # Don't stop processing, just notify - app continues with next file
        QMessageBox.warning(self, "ChronosID Error", f"An error occurred:\n\n{error_msg}\n\nProcessing will continue.")

    def update_chronos_stats(self, stats):
        # Update the ChronosID overlay with real-time stats.
        engine = stats.get('engine', 'N/A')
        # Color-code engine badge
        engine_colors = {'OWL': '#4caf50', 'LIZARD': '#ff9800', 'EYE': '#2196f3'}
        color = engine_colors.get(engine, '#888')
        self.lbl_engine.setText(engine)
        self.lbl_engine.setStyleSheet(f"background-color: #333; color: {color}; font-weight: bold; padding: 4px 10px; border-radius: 4px;")
        
        self.lbl_file.setText(stats.get('file', 'N/A'))
        frame = stats.get('frame', 0)
        total = stats.get('total_frames', 0)
        self.lbl_frames.setText(f"{frame} / {total}")
        
        h_val = stats.get('h_val', 0)
        v_val = stats.get('v_val', 0)
        # Format based on engine type
        if engine == 'OWL':
            self.lbl_h_val.setText(f"H: {h_val:.1f}°")
            self.lbl_v_val.setText(f"V: {v_val:.1f}°")
        else:
            self.lbl_h_val.setText(f"H: {h_val:.2f}")
            self.lbl_v_val.setText(f"V: {v_val:.2f}")
        
        proc_fps = stats.get('fps', 0)
        self.lbl_proc_fps.setText(f"{proc_fps:.0f} fps")

    def _on_tracking_loaded(self, tracking_path):
        # Highlight/select the tree item corresponding to a loaded _tracking.mf4 file.
        try:
            if not tracking_path: return
            # Convert tracking path to original mf4 path used in tree items
            candidate = tracking_path
            if candidate.lower().endswith('_tracking.mf4'):
                candidate_mf4 = candidate[:-len('_tracking.mf4')] + '.mf4'
            else:
                # Fallback: try replacing suffix
                candidate_mf4 = candidate

            # Walk tree and find matching item by UserRole path
            iterator = QTreeWidgetItemIterator(self.tree_participants)
            found_item = None
            while iterator.value():
                item = iterator.value()
                data = item.data(0, Qt.UserRole)
                try:
                    if data and isinstance(data, str):
                        # compare normalized paths lowercased
                        if os.path.normcase(os.path.abspath(data)) == os.path.normcase(os.path.abspath(candidate_mf4)):
                            found_item = item
                            break
                except Exception:
                    pass
                iterator += 1

            if found_item:
                # ensure parents expanded
                p = found_item.parent()
                while p:
                    p.setExpanded(True)
                    p = p.parent()
                # select and scroll to item
                self.tree_participants.setCurrentItem(found_item)
                found_item.setSelected(True)
                try:
                    self.tree_participants.scrollToItem(found_item)
                except Exception:
                    pass
                # flash highlight the item
                try:
                    self._highlight_tracking_item(tracking_path)
                except Exception:
                    pass
        except Exception:
            pass

    def _highlight_tracking_item(self, tracking_path, duration_ms=1200):
        # Set a persistent background highlight on the tree item matching tracking_path.
        # Also auto-expands parents and scrolls to the item.
        try:
            if not tracking_path: return
            # derive mf4 path
            tracking_path_str = str(tracking_path)
            if tracking_path_str.lower().endswith('_tracking.mf4'):
                target = tracking_path_str[:-len('_tracking.mf4')] + '.mf4'
            else:
                target = tracking_path

            iterator = QTreeWidgetItemIterator(self.tree_participants)
            found_item = None
            while iterator.value():
                item = iterator.value()
                data = item.data(0, Qt.UserRole)
                try:
                    if data and isinstance(data, str):
                        if os.path.normcase(os.path.abspath(data)) == os.path.normcase(os.path.abspath(target)):
                            found_item = item
                            break
                except Exception:
                    pass
                iterator += 1

            if not found_item:
                return

            # clear ALL previous persistent highlights (fix trailing bug)
            try:
                from PySide6.QtGui import QBrush
                it2 = QTreeWidgetItemIterator(self.tree_participants)
                while it2.value():
                    old = it2.value()
                    for c in range(self.tree_participants.columnCount()):
                        old.setBackground(c, QBrush())
                    it2 += 1
                self._last_highlighted_item = found_item
            except Exception:
                pass

            # --- Auto-expand parents ---
            parent = found_item.parent()
            while parent:
                if not parent.isExpanded():
                    parent.setExpanded(True)
                parent = parent.parent()

            # --- Auto-scroll to item ---
            self.tree_participants.scrollToItem(
                found_item,
                QAbstractItemView.PositionAtCenter
            )

            # --- Temporary flash (yellow) on top of persistent ---
            from PySide6.QtGui import QBrush, QColor
            brush = QBrush(QColor('#fff3cd'))
            for c in range(self.tree_participants.columnCount()):
                try:
                    found_item.setBackground(c, brush)
                except Exception:
                    pass

            # schedule restore to persistent grey
            try:
                from PySide6.QtCore import QTimer
                QTimer.singleShot(duration_ms, lambda: self._restore_persistent_highlight())
            except Exception:
                pass
        except Exception:
            pass

    def _restore_persistent_highlight(self):
        """After yellow flash, restore the persistent grey highlight."""
        try:
            it = getattr(self, '_last_highlighted_item', None)
            if it and isinstance(it, QTreeWidgetItem):
                from PySide6.QtGui import QBrush, QColor
                persistent_brush = QBrush(QColor('#3a3a3a'))
                for c in range(self.tree_participants.columnCount()):
                    it.setBackground(c, persistent_brush)
        except Exception:
            pass

    def on_tracking_task_done(self, video_path):
        # Update tree icon when a tracking task completes.
        # Extract base name: video_path is like D8_1_cam2.avi -> D8_1
        base = os.path.basename(video_path).replace(".avi", "")
        base = re.sub(r'_cam\d+$', '', base)  # Remove _camX suffix
        
        self._notify(f"Tracking completed for: {base}", "success")
        
        mf4_name = base + ".mf4"
        
        # Find and update tree item
        icon_done = QIcon(resource_path("assets/icons/check_16dp_75FB4C_FILL0_wght400_GRAD0_opsz20.png"))
        iterator = QTreeWidgetItemIterator(self.tree_participants)
        while iterator.value():
            item = iterator.value()
            if item.childCount() == 0:  # Leaf item
                if item.text(0) == mf4_name:
                    item.setIcon(1, icon_done)  # Column 1 is Tracking
                    break
            iterator += 1
    
    def reset_chronos_overlay(self):
        # Reset the overlay to default state.
        self.lbl_engine.setText("ENGINE")
        self.lbl_engine.setStyleSheet("background-color: #333; color: #0f0; font-weight: bold; padding: 4px 10px; border-radius: 4px;")
        self.lbl_file.setText("No file")
        self.lbl_frames.setText("0 / 0")
        self.lbl_h_val.setText("H: --")
        self.lbl_v_val.setText("V: --")
        self.lbl_proc_fps.setText("0 fps")

    # --- Scanner Logic ---
    from PySide6.QtWidgets import QTreeWidgetItemIterator # Import helper
    
    def scan_participants(self, folder):
        self.tree_participants.clear()
        self.busy_changed.emit(True)  # Notify MainWindow to start spinner
        self.tree_participants.setEnabled(False)
        
        if hasattr(self, 'scanner_thread') and self.scanner_thread is not None:
             self.scanner_thread.quit()
             self.scanner_thread.wait()
        
        marks_path = os.path.join(folder, 'marks.json')
        self.scanner_thread = AnalysisScanner(folder, marks_path=marks_path)
        self.scanner_thread.finished.connect(self.on_scan_finished)
        self.scanner_thread.log.connect(self.log)
        self.scanner_thread.start()
        
    def on_scan_finished(self, results):
        print(f"[DEBUG] on_scan_finished called with {len(results)} results")
        self.tree_participants.setUpdatesEnabled(False)  # Batch updates for speed
        self.tree_participants.clear()
        
        for res in results:
            item = QTreeWidgetItem(self.tree_participants)
            item.setText(0, res["name"])
            
            # Tracking stats in column 1
            tracking_done, total = res["tracking_stats"]
            item.setText(1, f"{tracking_done}/{total}")
            item.setForeground(1, QColor(res["color"]))
            item.setFont(1, self.tree_participants.font())
            
            # Marks stats in column 2
            marks_done, _ = res.get("marks_stats", (0, total))
            item.setText(2, f"{marks_done}/{total}")
            # Color for marks: green if all done, orange if some, red if none
            if total > 0:
                if marks_done == total: 
                    item.setForeground(2, QColor("#2da44e"))
                elif marks_done > 0: 
                    item.setForeground(2, QColor(IDIADA_ORANGE))
                else: 
                    item.setForeground(2, QColor("#d1242f"))

            # Analysis stats in column 3
            analysis_done, _ = res["analysis_stats"]
            item.setText(3, f"{analysis_done}/{total}")
            # Color for analysis: green if all done, orange if some, red if none
            if total > 0:
                if analysis_done == total: 
                    item.setForeground(3, QColor("#2da44e"))
                elif analysis_done > 0: 
                    item.setForeground(3, QColor(IDIADA_ORANGE))
                else: 
                    item.setForeground(3, QColor("#d1242f"))
            
            item.setCheckState(0, Qt.Checked)
            item.setExpanded(True)
            
            item.setData(0, Qt.UserRole, res.get("path")) # Store path
            
            if res.get("children"):
                self._add_tree_items(item, res["children"])
        
        self.tree_participants.setUpdatesEnabled(True)  # Re-enable updates
        # Col 0: resize to contents initially (user can adjust after via Interactive mode)
        self.tree_participants.resizeColumnToContents(0)
        if self.tree_participants.columnWidth(0) < 120:
            self.tree_participants.setColumnWidth(0, 120)
        self.tree_participants.setEnabled(True)
        self.busy_changed.emit(False)  # Notify MainWindow to stop spinner
        self.scanner_thread = None
        print("[DEBUG] on_scan_finished: tree populated, spinner stopped")
        # Auto-load first tracking file (if any checked) into Time Selector
        try:
            print("[DEBUG] calling auto_load_first_tracking...")
            self.auto_load_first_tracking()
            print("[DEBUG] auto_load_first_tracking completed")
        except Exception as e:
            print(f"[DEBUG] auto_load_first_tracking EXCEPTION: {e}")
            import traceback
            traceback.print_exc()



    def _add_tree_items(self, parent_item, children_data):
        for data in children_data:
            fpath = data.get("path", "")
            
            item = QTreeWidgetItem(parent_item)
            item.setText(0, data["name"])
            item.setCheckState(0, Qt.Checked)
            item.setData(0, Qt.UserRole, fpath)
            
            if data["type"] == "folder":
                item.setIcon(0, QIcon(resource_path("assets/icons/folder_16dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png")))
                # Tracking stats in column 1
                if "tracking_stats" in data:
                    tracking_done, tracking_total = data["tracking_stats"]
                    item.setText(1, f"{tracking_done}/{tracking_total}")
                    if tracking_total > 0 and tracking_done == tracking_total: 
                        item.setForeground(1, QColor("#2da44e"))
                    elif tracking_done > 0: 
                        item.setForeground(1, QColor(IDIADA_ORANGE))
                    else: 
                        item.setForeground(1, QColor("#d1242f"))
                
                # Marks stats in column 2
                if "marks_stats" in data:
                    marks_done, marks_total = data["marks_stats"]
                    item.setText(2, f"{marks_done}/{marks_total}")
                    if marks_total > 0 and marks_done == marks_total: 
                        item.setForeground(2, QColor("#2da44e"))
                    elif marks_done > 0: 
                        item.setForeground(2, QColor(IDIADA_ORANGE))
                    else: 
                        item.setForeground(2, QColor("#d1242f"))

                # Analysis stats in column 3
                if "analysis_stats" in data:
                    analysis_done, analysis_total = data["analysis_stats"]
                    item.setText(3, f"{analysis_done}/{analysis_total}")
                    if analysis_total > 0 and analysis_done == analysis_total: 
                        item.setForeground(3, QColor("#2da44e"))
                    elif analysis_done > 0: 
                        item.setForeground(3, QColor(IDIADA_ORANGE))
                    else: 
                        item.setForeground(3, QColor("#d1242f"))
                
            elif data["type"] == "file":
                # Tracking status (column 1)
                has_tracking = data.get("has_tracking", False)
                tracking_icon = "icons/check_16dp_75FB4C_FILL0_wght400_GRAD0_opsz20.png" if has_tracking else "icons/hourglass_16dp_F39200_FILL0_wght400_GRAD0_opsz20.png"
                item.setIcon(1, QIcon(resource_path(f"assets/{tracking_icon}")))
                
                # Marks status (column 2)
                # We need to check if marks.json contains entry for this file
                # This is tricky because marks.json is global. Assume scanner updated 'has_marks'
                has_marks = data.get("has_marks", False)
                marks_icon = "icons/check_16dp_75FB4C_FILL0_wght400_GRAD0_opsz20.png" if has_marks else "icons/hourglass_16dp_F39200_FILL0_wght400_GRAD0_opsz20.png"
                item.setIcon(2, QIcon(resource_path(f"assets/{marks_icon}")))
                
                # Report status (column 3)
                has_report = data.get("has_report", False)
                report_icon = "icons/check_16dp_75FB4C_FILL0_wght400_GRAD0_opsz20.png" if has_report else "icons/hourglass_16dp_F39200_FILL0_wght400_GRAD0_opsz20.png"
                item.setIcon(3, QIcon(resource_path(f"assets/{report_icon}")))
                
            if data.get("children"):
                self._add_tree_items(item, data["children"])
            
            # Cache has_tracking status in UserRole+1 to avoid os.path.exists on main thread later
            if data["type"] == "file":
                has_tracking = data.get("has_tracking", False)
                item.setData(0, Qt.UserRole + 1, has_tracking)
    
    def set_tree_checked(self, state):
        def recursive_check(item):
            item.setCheckState(0, state)
            for i in range(item.childCount()):
                recursive_check(item.child(i))
        
        for i in range(self.tree_participants.topLevelItemCount()):
            recursive_check(self.tree_participants.topLevelItem(i))

    def select_incomplete_tracking(self):
        # Select only participants/folders with incomplete tracking.
        def recursive_incomplete(item):
            is_file = (item.childCount() == 0 and ".mf4" in item.text(0).lower())
            if is_file: 
                # For files, check if they have tracking (icon in column 1)
                pass  # Files get checked/unchecked based on parent folder
            else:
                # For folders, check tracking stats in column 1
                txt = item.text(1)
                if "/" in txt:
                    try:
                        parts = txt.split("/")
                        d = int(parts[0])
                        t = int(parts[1])
                        if d == t and t > 0: 
                            item.setCheckState(0, Qt.Unchecked)
                        else: 
                            item.setCheckState(0, Qt.Checked)
                    except: 
                        pass
            
            for i in range(item.childCount()):
                recursive_incomplete(item.child(i))
                
        for i in range(self.tree_participants.topLevelItemCount()):
            recursive_incomplete(self.tree_participants.topLevelItem(i))
    
    def select_incomplete_marks(self):
        # Select only participants/folders with incomplete marks.
        def recursive_incomplete(item):
            is_file = (item.childCount() == 0 and ".mf4" in item.text(0).lower())
            if not is_file:
                txt = item.text(2)
                if "/" in txt:
                    try:
                        parts = txt.split("/")
                        d = int(parts[0])
                        t = int(parts[1])
                        if d == t and t > 0: 
                            item.setCheckState(0, Qt.Unchecked)
                        else: 
                            item.setCheckState(0, Qt.Checked)
                    except: 
                        pass
            
            for i in range(item.childCount()):
                recursive_incomplete(item.child(i))
                
        for i in range(self.tree_participants.topLevelItemCount()):
            recursive_incomplete(self.tree_participants.topLevelItem(i))
    
    def select_incomplete_reports(self):
        # Select only participants/folders with incomplete reports.
        def recursive_incomplete(item):
            is_file = (item.childCount() == 0 and ".mf4" in item.text(0).lower())
            if not is_file:
                txt = item.text(3)
                if "/" in txt:
                    try:
                        parts = txt.split("/")
                        d = int(parts[0])
                        t = int(parts[1])
                        if d == t and t > 0: 
                            item.setCheckState(0, Qt.Unchecked)
                        else: 
                            item.setCheckState(0, Qt.Checked)
                    except: 
                        pass
            
            for i in range(item.childCount()):
                recursive_incomplete(item.child(i))
                
        for i in range(self.tree_participants.topLevelItemCount()):
            recursive_incomplete(self.tree_participants.topLevelItem(i))

    def gather_checked_tracking_files(self):
        # Return list of existing _tracking.mf4 paths for checked leaf items, in traversal order.
        files = []
        iterator = QTreeWidgetItemIterator(self.tree_participants)
        while iterator.value():
            item = iterator.value()
            if item.checkState(0) == Qt.Checked and item.childCount() == 0:
                mf4_path = item.data(0, Qt.UserRole)
                # Optimization: Use cached has_tracking status (UserRole+1) instead of os.path.exists
                has_tracking = item.data(0, Qt.UserRole + 1)
                
                if mf4_path and isinstance(mf4_path, str) and mf4_path.lower().endswith('.mf4'):
                    mf4_path_str = str(mf4_path) # Ensure it's a string
                    # If we have cached status, use it. Otherwise fallback to disk check (slower)
                    if has_tracking is not None:
                         if has_tracking:
                             if mf4_path_str.lower().endswith('_tracking.mf4'):
                                 # We need the non-tracking one for AnalysisWidget? No, AnalysisWidget likes tracking one
                                 tracking_path = mf4_path_str
                             else:
                                 tracking_path = mf4_path_str[:-4] + '_tracking.mf4'
                             files.append(tracking_path)
                    else:
                        # Fallback for legacy/other items
                        mf4_path_str = str(mf4_path)
                        if mf4_path_str.lower().endswith('_tracking.mf4'):
                            tracking_path = mf4_path_str
                        else:
                            tracking_path = mf4_path_str[:-4] + '_tracking.mf4'
                        if os.path.exists(tracking_path):
                            files.append(tracking_path)
            iterator += 1
        
        return files

    def gather_checked_mf4_files(self):
        # Return list of checked mf4 paths (non-tracking) for checked leaf items.
        files = []
        iterator = QTreeWidgetItemIterator(self.tree_participants)
        while iterator.value():
            item = iterator.value()
            if item.checkState(0) == Qt.Checked and item.childCount() == 0:
                mf4_path = item.data(0, Qt.UserRole)
                if mf4_path and isinstance(mf4_path, str) and mf4_path.lower().endswith('.mf4'):
                    files.append(mf4_path)
            iterator += 1
        return files

    def toggle_generate_reports(self):
        if hasattr(self, '_report_queue') and self._report_queue:
            self.stop_report_generation()
        else:
            self.start_report_generation()

    def stop_report_generation(self):
        self._stop_report_after_current = True
        self.log("Stopping report generation...")
        if hasattr(self.tab_logic, 'btn_generate_report'):
            self.tab_logic.btn_generate_report.setText(" Stopping...")
            self.tab_logic.btn_generate_report.setEnabled(False)

    def start_report_generation(self):
        files = self.gather_checked_mf4_files()
        if not files:
            QMessageBox.information(self, "No Files", "No checked MF4 files found.")
            return

        # Upfront Missing Marks Validation
        missing_marks_files = []
        for mf4_path in files:
            marks = self._get_marks_for_mf4(mf4_path)
            if marks is None:
                missing_marks_files.append(os.path.basename(mf4_path))
                
        if missing_marks_files:
            files_list = "\n".join(missing_marks_files[:10])
            if len(missing_marks_files) > 10:
                files_list += f"\n... and {len(missing_marks_files) - 10} more."
                
            resp = QMessageBox.question(
                self,
                "Marks Missing",
                f"The following cases are missing marks:\n\n{files_list}\n\nDo you want to generate reports anyway (without marks)?\nClick 'No' to cancel and add them manually.",
                QMessageBox.Yes | QMessageBox.No
            )
            if resp == QMessageBox.No:
                return

        self._report_queue = files
        self._stop_report_after_current = False
        self._report_wait_for_marks = False
        self._total_reports_to_generate = len(files)
        import time
        self._report_start_time = time.time()
        self._reports_generated = 0
        self._current_batch_participant = None

        if hasattr(self.tab_logic, 'btn_generate_report'):
            self.tab_logic.btn_generate_report.setText(" Stop Generation")
            self.tab_logic.btn_generate_report.setStyleSheet("background-color: #d1242f; color: white; font-weight: bold; padding: 0 15px; border-radius: 3px;")
            icon_stop = QIcon(resource_path("assets/icons/stop_circle_16dp_FFFFFF_FILL0_wght400_GRAD0_opsz20.png"))
            if not icon_stop.isNull():
                self.tab_logic.btn_generate_report.setIcon(icon_stop)

        if hasattr(self, 'main_window') and hasattr(self.main_window, 'lbl_stats'):
            self.main_window.lbl_stats.setText(f"Starting generation of {self._total_reports_to_generate} reports...")
            
        self._process_next_report()

    def _get_marks_path(self, context_file=None):
        source = getattr(self, '_current_project_source', None)
        if source and os.path.isdir(source):
            return os.path.join(source, 'marks.json')
        
        # Fallback: search upwards from context_file if provided
        if context_file:
            search_dir = os.path.dirname(os.path.abspath(context_file))
            while len(search_dir) > 3: # Stop at drive root
                p = os.path.join(search_dir, 'marks.json')
                if os.path.exists(p):
                    return p
                # Also check for _FUSION_RESULTS/marks.json
                p_fusion = os.path.join(search_dir, '_FUSION_RESULTS', 'marks.json')
                if os.path.exists(p_fusion):
                    return p_fusion
                search_dir = os.path.dirname(search_dir)

        # Final fallback to app root
        return os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', '..', 'marks.json')

    def _load_marks_map(self, context_file=None):
        marks_path = self._get_marks_path(context_file=context_file)
        if not os.path.exists(marks_path):
            return {}
        try:
            import json
            with open(marks_path, 'r', encoding='utf-8') as f:
                return json.load(f)
        except Exception:
            return {}

    def _get_marks_key_for_path(self, tracking_path: str) -> str:
        p = os.path.normpath(tracking_path)
        parts = p.split(os.sep)
        if len(parts) < 3:
            return os.path.basename(p)
        return '/'.join(parts[-3:])

    def _get_tracking_path_for_mf4(self, mf4_path) -> str:
        mf4_path = str(mf4_path)
        if mf4_path.lower().endswith('_tracking.mf4'):
            return mf4_path
        return mf4_path[:-4] + '_tracking.mf4'

    def _get_marks_for_mf4(self, mf4_path: str):
        tracking_path = self._get_tracking_path_for_mf4(mf4_path)
        marks_map = self._load_marks_map(context_file=mf4_path)
        key = self._get_marks_key_for_path(tracking_path)
        
        # 1. Try exact structured key match
        if key in marks_map:
            return marks_map.get(key, [])
            
        # 2. Robust fallback: try to find a key that ends with the same filename
        file_base = os.path.splitext(os.path.basename(tracking_path))[0]
        # Try matching base name (handle cases with or without _tracking)
        clean_base = file_base.replace("_tracking", "")
        for k in marks_map:
            # Check if clean_base is in the filename part of the key
            k_filename = os.path.basename(k)
            if clean_base in k_filename:
                return marks_map[k]
                
        return None

    def _ensure_reports_dir(self, mf4_path: str):
        base_dir = os.path.dirname(mf4_path)
        if not base_dir:
            base_dir = os.getcwd()
        reports_dir = os.path.join(base_dir, "Reports")
        os.makedirs(reports_dir, exist_ok=True)
        return reports_dir

    def _build_report_config_for_mf4(self, mf4_path: str, driver_marks: list):
        from asammdf import MDF
        
        protocol = "Euro NCAP"
        if hasattr(self, 'tab_logic') and hasattr(self.tab_logic, 'combo_protocol'):
            protocol = self.tab_logic.combo_protocol.currentText()
        target_category = None
        if hasattr(self, 'tab_logic'):
            try:
                target_category = self.tab_logic._determine_category_from_filename(os.path.basename(mf4_path))
            except Exception:
                target_category = None

        oem_name = self.combo_oem.currentText() if hasattr(self, 'combo_oem') else ""
        vehicle = self.txt_vehicle.text() if hasattr(self, 'txt_vehicle') else ""
        engineer = self.txt_engineer.text() if hasattr(self, 'txt_engineer') else ""
        analyst = self.txt_analyst.text() if hasattr(self, 'txt_analyst') else ""
        track = self.combo_track.currentText() if hasattr(self, 'combo_track') else ""

        signals = {}
        
        try:
            with MDF(mf4_path) as mdf:
                if hasattr(self, 'tab_logic'):
                    for category, table in self.tab_logic.category_tables.items():
                        if target_category and category != target_category:
                            continue
                        for row in range(table.rowCount()):
                            chk = table.item(row, 0)
                            name_item = table.item(row, 1)
                            if chk and name_item and chk.checkState() == Qt.Checked:
                                sig_name = name_item.text()
                                try:
                                    sig = mdf.get(sig_name)
                                except Exception:
                                    continue
        
                                op_widget = table.cellWidget(row, 2)
                                operator = op_widget.currentText() if isinstance(op_widget, QComboBox) else "None"
        
                                val_widget = table.cellWidget(row, 3)
                                value = None
                                if isinstance(val_widget, QDoubleSpinBox):
                                    value = val_widget.value()
                                elif isinstance(val_widget, QComboBox):
                                    value = val_widget.currentText()
                                elif isinstance(val_widget, QLineEdit):
                                    try:
                                        value = float(val_widget.text()) if val_widget.text() else None
                                    except Exception:
                                        value = val_widget.text()
        
                                units_item = table.item(row, 4)
                                units = units_item.text() if units_item else ""
        
                                alias_item = table.item(row, 5)
                                alias = alias_item.text() if alias_item else sig_name
        
                                signals[sig_name] = {
                                    'timestamps': list(sig.timestamps),
                                    'samples': list(sig.samples),
                                    'threshold': value,
                                    'operator': operator,
                                    'unit': units or getattr(sig, 'unit', 'Value'),
                                    'category': category,
                                    'alias': alias
                                }
        except Exception as e:
            raise RuntimeError(f"Could not load MF4: {e}")

        # Compute metrics
        metrics = self._compute_times_and_events(signals, target_category, driver_marks, parent_scope=self)

        # Calculate relative path for the report banner
        relative_path = os.path.basename(mf4_path)
        source_root = getattr(self, '_current_project_source', None)
        if source_root:
            try:
                # Find the 'Pxx' or 'Exx' participant folder level
                abs_mf4 = os.path.abspath(mf4_path)
                parts = abs_mf4.split(os.sep)
                # Look for a part that looks like a participant ID
                for i, part in enumerate(parts):
                    if re.match(r'^[PE]\d+', part, re.IGNORECASE):
                        relative_path = os.sep.join(parts[i:])
                        break
            except Exception:
                pass

        # Resolve GSR image
        camera_image_path = None
        if (protocol == "GSR ADDW" or protocol == "2023/2590") and hasattr(self, 'tab_logic'):
            camera_image_path = self.tab_logic._resolve_gsr_image_path(os.path.basename(mf4_path))

        return {
            'filename': os.path.basename(mf4_path),
            'relative_path': relative_path,
            'target_category': target_category,
            'pass_signal_name': metrics['pass_signal_name'],
            'show_thresholds': False,
            'oem_name': oem_name,
            'vehicle': vehicle,
            'protocol': protocol,
            'engineer': engineer,
            'analyst': analyst,
            'track': track,
            'test_date': datetime.now(),
            'signals': signals,
            'signal_times': metrics['signal_times'],
            'camera_image_path': camera_image_path,
            'tgaze': metrics['tgaze'],
            't_event': metrics['t_event'],
            't_event_color': metrics['t_event_color'],
            'mask': metrics.get('mask', 6.0),
            'audio_params': {
                'min_freq': getattr(self.tab_logic, 'audio_min_freq', 0) if hasattr(self, 'tab_logic') else 0,
                'max_freq': getattr(self.tab_logic, 'audio_max_freq', 0) if hasattr(self, 'tab_logic') else 0,
                'threshold': getattr(self.tab_logic, 'audio_threshold', 0) if hasattr(self, 'tab_logic') else 0
            },
            'gauge_rules_path': getattr(self.tab_logic, 'active_gauge_rules_path', None) if hasattr(self, 'tab_logic') else None,
            'driver_marks': driver_marks
        }



    def _compute_times_and_events(self, signals, target_category, driver_marks, parent_scope=None):
        """Shared logic to compute first match times and T_event / T_gaze across contexts."""
        signal_times = {}
        target_logic_tab = self.tab_logic if hasattr(self, 'tab_logic') else (parent_scope.tab_logic if parent_scope and hasattr(parent_scope, 'tab_logic') else None)
        
        # 1. Get Mask value for this category (used for signal detection and as tgaze fallback)
        mask_start = 6.0
        if target_logic_tab and hasattr(target_logic_tab, 'pass_criteria_tables') and target_category:
            pass_table = target_logic_tab.pass_criteria_tables.get(target_category)
            if pass_table and pass_table.rowCount() > 0:
                mask_widget = pass_table.cellWidget(0, 5)
                if isinstance(mask_widget, QDoubleSpinBox):
                    mask_start = mask_widget.value()

        # 2. Calculate signal times (first match for each signal)
        for sig_name, sig_info in signals.items():
            timestamps = sig_info['timestamps']
            samples = sig_info['samples']
            operator = sig_info['operator']
            threshold = sig_info['threshold']
            
            first_match_time = None

            if sig_name == "SoundPressure":
                try:
                    # Retrieve audio params
                    audio_thresh = 0
                    if target_logic_tab:
                        audio_thresh = getattr(target_logic_tab, 'audio_threshold', 0)
                    
                    if len(samples) > 0 and len(timestamps) > 0:
                        samples_numeric = samples
                        # Only filter if we have enough frequency info
                        if target_logic_tab:
                            try:
                                min_f = getattr(target_logic_tab, 'audio_min_freq', 230)
                                max_f = getattr(target_logic_tab, 'audio_max_freq', 2000)
                                if min_f > 0 and max_f > 0:
                                    samples_np = np.array(samples, dtype=float)
                                    # Calc Fs
                                    dur = timestamps[-1] - timestamps[0]
                                    fs = len(samples) / dur if dur > 0 else 44100
                                    nyq = 0.5 * fs
                                    low = min_f / nyq
                                    high = max_f / nyq
                                    from scipy.signal import butter, filtfilt
                                    b, a = butter(4, [low, high], btype='band')
                                    samples_numeric = list(filtfilt(b, a, samples_np))
                            except Exception as e:
                                print(f"[DEBUG signal_times] Filter error for SoundPressure: {e}")
                        
                        op = operator if operator and operator != 'None' else '>='
                        first_match_time = find_first_valid_event(
                            np.array(samples_numeric),
                            np.array(timestamps),
                            float(audio_thresh),
                            op,
                            mask_start=mask_start
                        )
                except Exception as e:
                    print(f"[DEBUG signal_times] Error calculating SoundPressure: {e}")
            elif threshold is not None and operator and operator != 'None':
                threshold_numeric = None
                is_numeric_threshold = False
                try:
                    threshold_numeric = float(threshold)
                    is_numeric_threshold = True
                except (ValueError, TypeError):
                    is_numeric_threshold = False
                
                try:
                    samples_numeric = [float(s) for s in samples]
                    is_numeric_signal = True
                except:
                    is_numeric_signal = False
                
                if is_numeric_threshold and is_numeric_signal:
                    for t, val in zip(timestamps, samples_numeric):
                        if t < mask_start: continue
                        match = False
                        if operator == '>': match = val > threshold_numeric
                        elif operator == '<': match = val < threshold_numeric
                        elif operator == '>=': match = val >= threshold_numeric
                        elif operator == '<=': match = val <= threshold_numeric
                        elif operator == '==': match = abs(val - threshold_numeric) < 1e-6
                        elif operator == '!=': match = abs(val - threshold_numeric) >= 1e-6
                        if match:
                            first_match_time = t
                            break
                else:
                    samples_str = [str(s) for s in samples]
                    threshold_str = str(threshold)
                    for t, val_str in zip(timestamps, samples_str):
                        if t < mask_start: continue
                        match = False
                        if operator == '==': match = val_str == threshold_str
                        elif operator == '!=': match = val_str != threshold_str
                        if match:
                            first_match_time = t
                            break
            
            signal_times[sig_name] = first_match_time
            print(f"[DEBUG signal_times] {sig_name}: {first_match_time:.3f}s" if first_match_time else f"[DEBUG signal_times] {sig_name}: None")

        # 2. Determine pass_signal_name and conditions from PASS criteria table
        pass_signal_name = None
        conditions = []
        if target_logic_tab and hasattr(target_logic_tab, 'pass_criteria_tables') and target_category:
            pass_table = target_logic_tab.pass_criteria_tables.get(target_category)
            if pass_table and pass_table.rowCount() > 0:
                # 2.1 Get Signal Name
                # Check for direct attribute first (most reliable if table exists)
                if hasattr(pass_table, 'signal_combo'):
                    combo = pass_table.signal_combo
                    if combo and hasattr(combo, 'currentText'):
                        sig = combo.currentText()
                        if sig and sig != "-- Select Signal --":
                            pass_signal_name = sig
                else:
                    # Fallback to searching container
                    container = pass_table.cellWidget(0, 0)
                    if isinstance(container, QWidget):
                        combo = container.findChild(QComboBox)
                        if combo:
                            sig = combo.currentText()
                            if sig and sig != "-- Select Signal --":
                                pass_signal_name = sig

                # 2.2 Get Evaluation Conditions (Operator 1, Value 1, Operator 2, Value 2)
                op1_w = pass_table.cellWidget(0, 1)
                val1_w = pass_table.cellWidget(0, 2)
                op2_w = pass_table.cellWidget(0, 3)
                val2_w = pass_table.cellWidget(0, 4)

                if isinstance(op1_w, QComboBox):
                    op1 = op1_w.currentText()
                    val1 = val1_w.value() if isinstance(val1_w, QDoubleSpinBox) else 0
                    if op1 and op1 != "None":
                        conditions.append((op1, val1))

                if isinstance(op2_w, QComboBox):
                    op2 = op2_w.currentText()
                    val2 = val2_w.value() if isinstance(val2_w, QDoubleSpinBox) else 0
                    if op2 and op2 != "None":
                        conditions.append((op2, val2))

        # 3. Compute tgaze (first mark or configurable mask_start default)
        tgaze = mask_start
        if driver_marks and len(driver_marks) > 0:
            try:
                tgaze = float(driver_marks[0])
            except (ValueError, TypeError):
                pass

        # 4. Compute t_event and evaluate result color
        t_event = "No warn"
        t_event_color = "red"
        
        warn_time = signal_times.get(pass_signal_name) if pass_signal_name else None
        
        if warn_time is not None:
            # Determine if we use accumulated effective time (Scenario 2: Short Distractions / Phone Use)
            # or simple delta from first distraction (Scenario 1: Long Distraction, Sleep, etc.)
            is_scenario2 = any(kw in (target_category or "") for kw in ["Short Distraction", "Phone Use"])
            
            if is_scenario2 and driver_marks and len(driver_marks) >= 2:
                # Sum durations of all previous glances + partial duration of current glance until warn_time
                marks_sorted = sorted([float(m) for m in driver_marks])
                accumulated = 0.0
                for i in range(0, len(marks_sorted) - 1, 2):
                    start = marks_sorted[i]
                    end = marks_sorted[i+1]
                    
                    if warn_time < start:
                        # Warning happened between glances? (System delay)
                        # We stop at previous accumulated value
                        break
                    elif warn_time <= end:
                        # Warning happened during this glance
                        accumulated += (warn_time - start)
                        break
                    else:
                        # Warning happened after this glance finished
                        accumulated += (end - start)
                t_event = accumulated
            else:
                # Scenario 1 (Long Distraction) or fallback: T_event = WarningTime - Tgaze (First distraction)
                t_event = warn_time - tgaze
            
            # Evaluate against conditions
            if conditions:
                all_met = True
                for op, limit in conditions:
                    if op == '>': match = t_event > limit
                    elif op == '<': match = t_event < limit
                    elif op == '>=': match = t_event >= limit
                    elif op == '<=': match = t_event <= limit
                    elif op == '==': match = abs(t_event - limit) < 1e-6
                    elif op == '!=': match = abs(t_event - limit) >= 1e-6
                    else: match = True # Unknown operator
                    
                    if not match:
                        all_met = False
                        break
                t_event_color = "green" if all_met else "red"
            else:
                # No specific pass criteria conditions? 
                # NCAP fallback: usually < 3.0s is green
                t_event_color = "green" if t_event < 3.0 else "red"

        return {
            'tgaze': tgaze,
            't_event': t_event,
            't_event_color': t_event_color,
            'signal_times': signal_times,
            'pass_signal_name': pass_signal_name,
            'mask': mask_start
        }

    def _process_next_report(self):
        if getattr(self, '_stop_report_after_current', False) and self._report_queue:
            skipped = len(self._report_queue)
            self._report_queue = []
            self.log(f"Report generation stopped by user. Skipped {skipped} pending report(s).")

        # Update progress label
        if hasattr(self, '_total_reports_to_generate') and self._total_reports_to_generate > 0:
            if hasattr(self, 'main_window') and hasattr(self.main_window, 'lbl_stats'):
                processed = self._total_reports_to_generate - len(self._report_queue)
                # Cap the percentage logic to ensure we don't divide cleanly by zero down the line
                perc = int((processed / self._total_reports_to_generate) * 100)
                
                eta_str = ""
                if processed > 0 and hasattr(self, '_report_start_time'):
                    import time
                    elapsed = time.time() - self._report_start_time
                    avg_time = elapsed / processed
                    remaining_secs = (self._total_reports_to_generate - processed) * avg_time
                    eta_mins, eta_secs = int(remaining_secs // 60), int(remaining_secs % 60)
                    eta_str = f" - ETA: {eta_mins}m {eta_secs}s" if eta_mins > 0 else f" - ETA: {eta_secs}s"
                    
                self.main_window.lbl_stats.setText(f"Generated {processed}/{self._total_reports_to_generate} reports ({perc}%){eta_str}")

        if not self._report_queue:
            # Finished all reports — show summary
            if hasattr(self, 'main_window') and hasattr(self.main_window, 'lbl_stats'):
                self.main_window.lbl_stats.setText("Ready")
                
            if hasattr(self.tab_logic, 'btn_generate_report'):
                self.tab_logic.btn_generate_report.setText(" Generate Reports")
                self.tab_logic.btn_generate_report.setStyleSheet(f"background-color: {IDIADA_ORANGE}; color: black; font-weight: bold; padding: 0 15px; border-radius: 3px;")
                icon_report = QIcon(resource_path("assets/icons/play_circle_16dp_000000_FILL0_wght400_GRAD0_opsz20.png"))
                if not icon_report.isNull():
                    self.tab_logic.btn_generate_report.setIcon(icon_report)
                self.tab_logic.btn_generate_report.setEnabled(True)
                
            dirs = getattr(self, '_report_generated_dirs', set())
            if dirs:
                dirs_text = "\n".join(sorted(dirs))
                # Do not show MessageBox, just app notification
                self._notify(f"Reports successfully saved to {len(dirs)} directories.", "success")
            else:
                self._notify("No reports were generated.", "warning")
            self._report_generated_dirs = set()
            self._stop_report_after_current = False
            self.busy_changed.emit(False)
            return

        if not hasattr(self, '_report_generated_dirs'):
            self._report_generated_dirs = set()

        mf4_path = self._report_queue[0]
        
        # Traverse up directory tree to find participant folder (P01, E12, etc.)
        import re
        curr_dir = os.path.dirname(mf4_path)
        participant_name = os.path.basename(curr_dir)
        for _ in range(4): # Limit search to prevent infinite loop
            bname = os.path.basename(curr_dir)
            if re.match(r'^[PE]\d+', bname, re.IGNORECASE):
                participant_name = bname
                break
            curr_dir = os.path.dirname(curr_dir)
            if not curr_dir or curr_dir == os.path.dirname(curr_dir): break
            
        if getattr(self, '_current_batch_participant', None) != participant_name:
            self._current_batch_participant = participant_name
            self.log(f"----- {participant_name} -----")
            
        self.log(f"Processing report for: {os.path.basename(mf4_path)}")
        
        marks = self._get_marks_for_mf4(mf4_path)

        try:
            config = self._build_report_config_for_mf4(mf4_path, marks)
            # Update results excel in participant folder before generating PNG
            self._update_excel_results(config, mf4_path)
        except Exception as e:
            QMessageBox.warning(self, "Report Error", f"Failed to build report for {os.path.basename(mf4_path)}:\n{e}")
            self._report_queue.pop(0)
            self._process_next_report()  # recurse to next
            return

        reports_dir = self._ensure_reports_dir(mf4_path)
        self._report_generated_dirs.add(reports_dir)
        base_name = os.path.splitext(os.path.basename(mf4_path))[0]
        output_path = os.path.join(reports_dir, f"{base_name}.png")

        # Generate this report in background thread
        self.busy_changed.emit(True)
        if hasattr(self, '_batch_report_worker') and self._batch_report_worker is not None:
            self._batch_report_worker.quit()
            self._batch_report_worker.wait()

        self._batch_report_worker = self._make_report_worker(config, output_path, dpi=300)
        self._batch_report_worker.finished.connect(self._on_batch_report_done)
        self._batch_report_worker.error.connect(self._on_batch_report_error)
        self._batch_report_worker.start()

    def _on_batch_report_done(self, output_path):
        # Callback when one batch report is done - advance to next.
        self._notify(f"Report generated: {os.path.basename(output_path)}", "success")
        self._report_queue.pop(0)
        self._process_next_report()

    def _update_excel_results(self, config: dict, mf4_path: str):
        """
        Updates Analysis_Results.xlsx in the participant's tracking directory.
        Appends or updates rows for analyzed files.
        """
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        
        try:
            # Determine directory and file path
            participant_dir = os.path.dirname(mf4_path)
            excel_path = os.path.join(participant_dir, "Analysis_Results.xlsx")
            
            # Load or create workbook
            if os.path.exists(excel_path):
                wb = openpyxl.load_workbook(excel_path)
                ws = wb.active
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "Results"
                # Add headers
                headers = ["Folder Name", "File Name", "Distraction Start", "Warning Start", "Warning Timer", "Score"]
                ws.append(headers)
                
                # Style headers (Light Blue background, Bold)
                header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                header_font = Font(bold=True)
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
            
            # Extract metrics from config
            folder_name = config.get('target_category', '--')
            file_name_full = config.get('filename', os.path.basename(mf4_path))
            # Strip extension for Excel "File Name" column if needed? 
            # User snapshot shows name without ext or truncated. Let's keep full for now.
            file_name = file_name_full
            
            dist_start = config.get('tgaze')
            
            pass_sig = config.get('pass_signal_name')
            warn_start = config.get('signal_times', {}).get(pass_sig) if pass_sig else None
            
            warn_timer = config.get('t_event')
            score = "PASS" if config.get('t_event_color') == "green" else "FAIL"
            
            # Prepare row data
            row_data = [
                folder_name,
                file_name,
                dist_start if dist_start is not None else "",
                warn_start if warn_start is not None else "nan",
                warn_timer if isinstance(warn_timer, (int, float)) else "",
                score
            ]
            
            # Check for existing row (File Name match)
            found_row = -1
            for r in range(2, ws.max_row + 1):
                if ws.cell(row=r, column=2).value == file_name:
                    found_row = r
                    break
            
            if found_row != -1:
                # Update existing row
                for idx, val in enumerate(row_data, 1):
                    ws.cell(row=found_row, column=idx).value = val
            else:
                # Append new row
                ws.append(row_data)
            
            # Apply styling to the new/updated row (center alignment)
            target_row = found_row if found_row != -1 else ws.max_row
            for cell in ws[target_row]:
                cell.alignment = Alignment(horizontal="center")
            
            # Adjust column widths (simple heuristic)
            for i, col in enumerate(ws.columns, 1):
                max_length = 0
                for cell in col:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = max_length + 4
                
            wb.save(excel_path)
            print(f"[EXCEL] Result for {file_name} saved/updated in {excel_path}")
            
        except Exception as e:
            print(f"[EXCEL ERROR] Failed to update results: {e}")
            import traceback
            traceback.print_exc()

    def _on_batch_report_error(self, error_msg):
        # Callback when one batch report fails - advance to next.
        if self._report_queue:
            base_name = os.path.basename(self._report_queue[0])
            QMessageBox.warning(self, "Report Error", f"Failed to generate report for {base_name}:\n{error_msg}")
            self._report_queue.pop(0)
        self._process_next_report()

    def _make_report_worker(self, config, output_path, dpi=300):
        """Factory method for batch reports. Override in subclasses to swap the builder."""
        return ReportGeneratorWorker(config, output_path, dpi)

    def _on_participant_check_changed(self, item, column):
        # Called when user checks/unchecks an item in participant tree.
        if column == 0:  # Only react to checkbox changes in first column
            # If item has children, propagate the state to all descendants
            if item.childCount() > 0:
                new_state = item.checkState(0)
                self._propagate_check_state(item, new_state)
            
            # Update the list silently
            # Optimization: distinct from original code, we do NOT call gather_checked_tracking_files() here
            # as it is expensive and the result is not stored or used immediately.
            # The list is gathered on-demand when "Next Plot" or "Start Chronos" is clicked.
            pass
    
    def _propagate_check_state(self, parent_item, state):
        # Recursively set check state for all children of parent_item.
        for i in range(parent_item.childCount()):
            child = parent_item.child(i)
            child.setCheckState(0, state)
            # Recursively propagate to nested children
            if child.childCount() > 0:
                self._propagate_check_state(child, state)

    def auto_load_first_tracking(self):
        # Auto-load the first tracking file from checked participants into Time Selector.
        try:
            self._tracking_list = self.gather_checked_tracking_files()
            # Populate case list in Time Selector
            if hasattr(self, 'tab_time') and hasattr(self.tab_time, 'set_case_list'):
                 self.tab_time.set_case_list(self._tracking_list)
                 
            self._tracking_index = 0
            if self._tracking_list and hasattr(self, 'tab_time') and hasattr(self.tab_time, 'load_tracking_file'):
                self.tab_time.load_tracking_file(self._tracking_list[0])
                # highlight corresponding participant in tree
                try:
                    self._highlight_tracking_item(self._tracking_list[0])
                except Exception:
                    pass
        except Exception:
            pass

    def next_plot(self):
        # Go to next case. to the next tracking plot. When exhausted, ask to execute analysis.
        try:
            # Validate even number of markers before changing plot
            if hasattr(self, 'tab_time') and hasattr(self.tab_time, 'markers'):
                num_markers = len(self.tab_time.markers)
                if num_markers % 2 != 0:
                    QMessageBox.warning(
                        self, 
                        "Invalid Selection", 
                        f"Number of selections must be even (pairs).\nCurrent selections: {num_markers}\n\nPlease add or remove a marker to create pairs."
                    )
                    return
            
            if self._report_wait_for_marks:
                self._report_wait_for_marks = False
                self._process_next_report()
                return

            files = self.gather_checked_tracking_files()
            if not files:
                QMessageBox.information(self, "No Tracking Files", "No checked participants with tracking MF4 found.")
                return

            # determine current path loaded in TimeSelector (if any)
            current = getattr(self.tab_time, 'current_tracking_path', None)

            # find next file after current; if current not present, start at first
            next_path = None
            
            if current and current in files:
                idx = files.index(current)
                if idx + 1 < len(files):
                    next_path = files[idx + 1]
            else:
                next_path = files[0] if files else None

            if next_path:
                self._navigate_to_file(next_path)
                return

            # if we reach here, current was the last item
            resp = QMessageBox.question(self, "Generate reports?", "No more tracking plots. Generate reports now?", QMessageBox.Yes | QMessageBox.No)
            if resp == QMessageBox.Yes:
                try:
                    self.generate_reports_from_time_selector()
                except Exception:
                    pass
            return
        except Exception:
            pass

    def previous_plot(self):
        # Go to previous case. previous tracking plot.
        try:
            # Validate even number of markers before changing plot
            if hasattr(self, 'tab_time') and hasattr(self.tab_time, 'markers'):
                num_markers = len(self.tab_time.markers)
                if num_markers % 2 != 0:
                    QMessageBox.warning(
                        self, 
                        "Invalid Selection", 
                        f"Number of selections must be even (pairs).\nCurrent selections: {num_markers}\n\nPlease add or remove a marker to create pairs."
                    )
                    return
            
            files = self.gather_checked_tracking_files()
            if not files:
                QMessageBox.information(self, "No Tracking Files", "No checked participants with tracking MF4 found.")
                return

            current = getattr(self.tab_time, 'current_tracking_path', None)

            prev_path = None
            if current and current in files:
                idx = files.index(current)
                if idx - 1 >= 0:
                    prev_path = files[idx - 1]
                else:
                    QMessageBox.information(self, "First Plot", "Already at the first tracking plot.")
                    return
            else:
                prev_path = files[-1] if files else None
            
            if prev_path:
                self._navigate_to_file(prev_path)
        except Exception:
            pass
    
    def _navigate_to_file(self, fpath):
        """Navigate to a specific tracking file, updating subject/case combos and highlighting."""
        try:
            if not hasattr(self, 'tab_time'):
                return
            
            # Determine which subject this file belongs to
            subj = self.tab_time._extract_subject(fpath)
            current_subj_idx = self.tab_time.combo_subject.currentIndex()
            current_subj = self.tab_time.combo_subject.itemText(current_subj_idx) if current_subj_idx >= 0 else ""
            
            if subj != current_subj:
                # Switch subject (this will auto-populate cases)
                idx = self.tab_time.combo_subject.findText(subj)
                if idx >= 0:
                    self.tab_time.combo_subject.blockSignals(True)
                    self.tab_time.combo_subject.setCurrentIndex(idx)
                    self.tab_time._populate_cases_for_subject(subj)
                    self.tab_time.combo_subject.blockSignals(False)
            
            # Find file in case combo and select it
            for i in range(self.tab_time.combo_cases.count()):
                if self.tab_time.combo_cases.itemData(i) == fpath:
                    self.tab_time.combo_cases.blockSignals(True)
                    self.tab_time.combo_cases.setCurrentIndex(i)
                    self.tab_time.combo_cases.blockSignals(False)
                    break
            
            # Load the file
            self.tab_time.load_tracking_file(fpath)
            self.tab_time.update_navigation_labels()
            
            # Highlight in participant tree
            try:
                self._highlight_tracking_item(fpath)
            except Exception:
                pass
            try:
                self.tabs_right.setCurrentWidget(self.tab_time)
            except Exception:
                pass
        except Exception:
            pass
