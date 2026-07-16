import asyncio
import json
import logging
import os
from datetime import datetime
from threading import Thread

from fastapi import APIRouter, WebSocket
from pydantic import BaseModel

from backend.core.utils import resource_path
from backend.ws.manager import manager_reporting

logger = logging.getLogger("fusionstudio.reporting")

router = APIRouter()

TEMPLATE_OPTIONS: dict[str, list[tuple[str, str, bool]]] = {
    "Driver_Engagement.xlsx": [
        ("Distractions", "Distractions", True),
        ("Fatigue", "Fatigue", True),
        ("Occlusions", "Occlusions", True),
        ("Behaviours", "Behaviours", False),
    ],
}

_active_worker = None
_worker_thread: Thread | None = None


class RunRequest(BaseModel):
    template_name: str
    root_folder: str
    output_folder: str
    output_filename: str = "Report_Results.xlsx"
    selected_folders: list[str] = []


class PreviewRequest(BaseModel):
    file_path: str
    sheet_name: str = "DISTRACTION"


def _list_templates() -> list[dict]:
    templates_dir = resource_path("assets/templates")
    if not os.path.exists(templates_dir):
        return []
    results = []
    for file in sorted(os.listdir(templates_dir)):
        if file.endswith(".xlsx"):
            full_path = os.path.join(templates_dir, file)
            options = TEMPLATE_OPTIONS.get(file, [])
            results.append({
                "name": file,
                "path": full_path,
                "options": [{"label": o[0], "folder": o[1], "default": o[2]} for o in options],
            })
    return results


def _preview_excel(file_path: str, sheet_name: str) -> dict:
    try:
        import pandas as pd
        try:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
        except Exception:
            df = pd.read_excel(file_path, sheet_name=0, nrows=500)

        # Convert to dict, limit rows
        df = df.head(200).fillna("")
        return {
            "columns": df.columns.astype(str).tolist(),
            "rows": df.values.tolist(),
            "row_count": len(df),
            "sheet": sheet_name,
        }
    except Exception as e:
        return {"error": str(e)}


@router.get("/templates")
async def list_templates():
    templates = _list_templates()
    return {"templates": templates}


@router.get("/gauge_rules")
async def gauge_rules():
    from backend.core.utils import user_data_path
    rules_path = user_data_path("config/gauge_rules.json")
    if not os.path.exists(rules_path):
        rules_path = resource_path("config/gauge_rules.json")
    if os.path.exists(rules_path):
        with open(rules_path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


@router.post("/preview")
async def preview_report(req: PreviewRequest):
    loop = asyncio.get_event_loop()
    if not os.path.exists(req.file_path):
        return {"error": "File not found"}
    result = await loop.run_in_executor(None, _preview_excel, req.file_path, req.sheet_name)
    return result


@router.post("/run")
async def run_report(req: RunRequest):
    global _active_worker, _worker_thread

    if _active_worker is not None:
        return {"status": "already_running"}

    if not req.output_filename.endswith(".xlsx"):
        req.output_filename += ".xlsx"

    output_path = os.path.join(req.output_folder, req.output_filename)

    # Find template path
    templates_dir = resource_path("assets/templates")
    template_path = os.path.join(templates_dir, req.template_name)
    if not os.path.exists(template_path):
        return {"status": "error", "message": f"Template not found: {req.template_name}"}

    if not os.path.exists(req.root_folder):
        return {"status": "error", "message": "Root folder not found"}

    loop = asyncio.get_event_loop()

    def on_progress(msg):
        asyncio.run_coroutine_threadsafe(
            manager_reporting.broadcast({"type": "progress", "message": msg}), loop
        )

    class _ReportingWorker:
        def __init__(self):
            self.is_running = True

        def stop(self):
            self.is_running = False

        def run(self):
            global _active_worker
            from backend.core.dsm_processor import DSMProcessor
            try:
                processor = DSMProcessor(callback=on_progress)
                processor.process_dsm_data(
                    template_path, output_path, req.root_folder,
                    req.selected_folders
                )
                asyncio.run_coroutine_threadsafe(
                    manager_reporting.broadcast({
                        "type": "finished",
                        "output_path": output_path,
                    }), loop
                )
            except Exception as e:
                asyncio.run_coroutine_threadsafe(
                    manager_reporting.broadcast({
                        "type": "error",
                        "message": str(e),
                    }), loop
                )
            finally:
                _active_worker = None

    worker = _ReportingWorker()
    _active_worker = worker

    def _run():
        worker.run()

    _worker_thread = Thread(target=_run, daemon=True)
    _worker_thread.start()

    return {"status": "started"}


@router.post("/stop")
async def stop_report():
    global _active_worker
    if _active_worker is not None:
        _active_worker.stop()
        return {"status": "stopping"}
    return {"status": "no_worker"}


@router.websocket("/ws")
async def ws_reporting(ws: WebSocket):
    await manager_reporting.connect(ws)
    try:
        while True:
            await ws.receive_text()
    except Exception:
        await manager_reporting.disconnect(ws)


# --- Gaze Logic Tab Controls & Report Preview Endpoints ---

class GazePreviewRequest(BaseModel):
    file_path: str
    protocol: str = "Euro NCAP"
    metadata: dict = {}
    category_configs: dict = {}
    gauge_rules: dict = {}
    micro: dict = {}
    source_dir: str = ""
    report_camera_settings: dict | None = None


class GazeGenerateRequest(BaseModel):
    files: list[str]
    protocol: str = "Euro NCAP"
    metadata: dict = {}
    category_configs: dict = {}
    gauge_rules: dict = {}
    micro: dict = {}
    source_dir: str = ""
    report_camera_settings: dict | None = None


def _resolve_om_video_paths(file_path: str, camera_left: str, camera_right: str) -> tuple[str | None, str | None]:
    if not file_path or not os.path.exists(file_path):
        return None, None
    base_dir = os.path.dirname(file_path)
    left_path = None
    right_path = None
    try:
        files = os.listdir(base_dir)
        for f in files:
            if f.lower().endswith(".avi"):
                import re
                if camera_left and (
                    re.search(rf'_cam{camera_left}\.avi$', f, re.I) or 
                    re.search(rf'_{camera_left}\.avi$', f, re.I) or
                    re.search(rf'_CAM_{camera_left}\.avi$', f, re.I)
                ):
                    left_path = os.path.join(base_dir, f)
                if camera_right and (
                    re.search(rf'_cam{camera_right}\.avi$', f, re.I) or 
                    re.search(rf'_{camera_right}\.avi$', f, re.I) or
                    re.search(rf'_CAM_{camera_right}\.avi$', f, re.I)
                ):
                    right_path = os.path.join(base_dir, f)
    except Exception as e:
        logger.error(f"Error resolving OM video paths: {e}")
    return left_path, right_path


def _resolve_gsr_image_path(filename: str) -> str | None:
    import re
    match = re.search(r'(ADDW\d+)', filename, re.IGNORECASE)
    if match:
        case_key = match.group(1).upper()
        try:
            num = int(case_key.replace("ADDW", ""))
            image_name = f"{num}.png"
            # Check local first (user custom)
            local_path = resource_path(os.path.join("assets/gsr/local", image_name))
            if os.path.exists(local_path):
                return local_path
            # Check core
            core_path = resource_path(os.path.join("assets/gsr", image_name))
            if os.path.exists(core_path):
                return core_path
        except:
            pass
    return None


def build_report_config(file_path: str, protocol: str, metadata: dict, category_configs: dict, gauge_rules: dict, driver_marks: list, micro: dict = None, show_thresholds: bool = False) -> dict:
    import re
    import numpy as np
    from asammdf import MDF
    from backend.core.audio_analysis import find_first_valid_event

    basename = os.path.splitext(os.path.basename(file_path))[0]
    target_category = None

    # Detect Occupant Monitoring (OM) / Misuse categories based on folder path
    path_lower = file_path.replace('\\', '/').lower()
    if "out of position" in path_lower or "oop" in path_lower:
        if "initial phase" in path_lower or "initial_phase" in path_lower:
            target_category = "OoP \u2014 Initial Phase"
        elif "change of status" in path_lower or "change_of_status" in path_lower:
            target_category = "OoP \u2014 Change of Status"
        elif "15 minutes" in path_lower or "15 min" in path_lower:
            target_category = "OoP \u2014 15 min Warning"
    elif "correct belt" in path_lower or "correct_belt" in path_lower or "csr" in path_lower or "seatbelt" in path_lower or "belt routing" in path_lower:
        if "initial phase" in path_lower or "initial_phase" in path_lower:
            target_category = "CSR \u2014 Initial Phase"
        elif "change of status" in path_lower or "change_of_status" in path_lower:
            target_category = "CSR \u2014 Change of Status"

    if target_category is None:
        if "ADDW" in basename.upper():
            if "high speed" in path_lower:
                target_category = "High Speed"
            elif "low speed" in path_lower:
                target_category = "Low Speed"
            else:
                target_category = "High Speed"
        else:
            match_d = re.match(r'^D(\d+)', basename)
            if match_d:
                num = int(match_d.group(1))
                if 1 <= num <= 9:
                    target_category = "Long Distraction (NDT)"
                elif 10 <= num <= 15:
                    target_category = "Long Distraction (DT)"
                elif (16 <= num <= 19) or num == 28 or (29 <= num <= 42):
                    target_category = "Short Distraction (NDT)"
                elif 20 <= num <= 27:
                    target_category = "Short Distraction (DT)"
            else:
                match_f = re.match(r'^F(\d+)', basename)
                if match_f:
                    num = int(match_f.group(1))
                    if num == 1:
                        target_category = "Microsleep"
                    elif num == 2:
                        target_category = "Sleep"
                    elif num == 3:
                        target_category = "Drowsiness"
                    elif num == 4:
                        target_category = "Unresponsive driver (SLE)"
                    elif num == 5:
                        target_category = "Unresponsive driver (DTR)"

    if not target_category:
        target_category = "Long Distraction (NDT)"

    cat_conf = category_configs.get(target_category, {})
    signals_conf = cat_conf.get('signals', {})
    pass_signal_name = cat_conf.get('pass_signal_name')
    mask_start = float(cat_conf.get('mask_start', 6.0))
    conditions = []
    
    op1 = cat_conf.get('operator1')
    val1 = cat_conf.get('value1')
    if op1 and op1 != "None":
        try:
            conditions.append((op1, float(val1)))
        except:
            pass
            
    op2 = cat_conf.get('operator2')
    val2 = cat_conf.get('value2')
    if op2 and op2 != "None":
        try:
            conditions.append((op2, float(val2)))
        except:
            pass

    target_signals_conf = dict(signals_conf)
    unresponsive_phases = cat_conf.get('unresponsive_phases', [])
    for phase in unresponsive_phases:
        sig_name = phase.get('signal')
        if sig_name and sig_name not in target_signals_conf:
            is_audio = sig_name.lower().find("sound") >= 0 or sig_name.lower().find("audio") >= 0 or sig_name.lower().find("buzzer") >= 0
            target_signals_conf[sig_name] = {
                'checked': True,
                'operator': phase.get('operator', '==') if not is_audio else '>=',
                'threshold': phase.get('threshold') if is_audio else phase.get('value', 0.0),
                'alias': sig_name
            }

    signals = {}
    with MDF(file_path) as mdf:
        for sig_name, sig_info in target_signals_conf.items():
            if not sig_info.get('checked', True):
                continue
            mdf_sig = None
            lookup_names = [sig_name]
            if sig_name == "SoundPressure":
                lookup_names = ["SoundPressure", "MySound PressureTask.Sound Pressure"]
            elif sig_name == "MySound PressureTask.Sound Pressure":
                lookup_names = ["MySound PressureTask.Sound Pressure", "SoundPressure"]

            for name_to_try in lookup_names:
                actual_name = None
                if name_to_try in mdf.channels_db:
                    actual_name = name_to_try
                else:
                    for k in mdf.channels_db.keys():
                        if k.lower() == name_to_try.lower():
                            actual_name = k
                            break

                if actual_name is not None:
                    try:
                        gp, idx = mdf.channels_db[actual_name][0]
                        mdf_sig = mdf.get(actual_name, group=gp, index=idx)
                        break
                    except Exception as ex:
                        logger.warning(f"Failed to get channel '{actual_name}' via group/index: {ex}")

                # Fallback to fuzzy matching on iter_channels
                for ch in mdf.iter_channels():
                    if ch.name.lower() == name_to_try.lower():
                        mdf_sig = ch
                        break
                if mdf_sig is not None:
                    break
                    
                norm_sig = re.sub(r'_[cde]3v_', '_', name_to_try.lower())
                for ch in mdf.iter_channels():
                    norm_ch = re.sub(r'_[cde]3v_', '_', ch.name.lower())
                    if norm_ch == norm_sig:
                        mdf_sig = ch
                        logger.info(f"Fuzzy matched config signal '{name_to_try}' to MDF channel '{ch.name}'")
                        break
                if mdf_sig is not None:
                    break
                    
                for ch in mdf.iter_channels():
                    if name_to_try.lower() in ch.name.lower() or ch.name.lower() in name_to_try.lower():
                        mdf_sig = ch
                        logger.info(f"Substring fuzzy matched config signal '{name_to_try}' to MDF channel '{ch.name}'")
                        break
                if mdf_sig is not None:
                    break
                        
            if mdf_sig is None:
                continue

            # Safe conversion of samples to list of float or list of string
            raw_samples = mdf_sig.samples
            processed_samples = []
            if raw_samples is not None:
                if not isinstance(raw_samples, np.ndarray):
                    raw_samples = np.array(raw_samples)
                if np.issubdtype(raw_samples.dtype, np.number) or np.issubdtype(raw_samples.dtype, np.bool_):
                    processed_samples = list(np.asarray(raw_samples, dtype=float))
                else:
                    for v in raw_samples:
                        if isinstance(v, bytes):
                            processed_samples.append(v.decode('utf-8', errors='ignore'))
                        elif v is None:
                            processed_samples.append("")
                        else:
                            processed_samples.append(str(v))
            
            thresh_val = sig_info.get('threshold', 0.0)
            try:
                if isinstance(thresh_val, (int, float)):
                    thresh_val = float(thresh_val)
                else:
                    thresh_val = float(thresh_val)
            except (ValueError, TypeError):
                thresh_val = str(thresh_val)

            signals[sig_name] = {
                'timestamps': list(np.asarray(mdf_sig.timestamps, dtype=float)),
                'samples': processed_samples,
                'operator': sig_info.get('operator', '>='),
                'threshold': micro.get('threshold') if (sig_name == "SoundPressure" and micro and micro.get('threshold') is not None) else thresh_val,
                'unit': getattr(mdf_sig, 'unit', 'Value'),
                'category': target_category,
                'alias': sig_info.get('alias') or sig_name
            }

    # For OM scenarios (OoP/CSR), default tgaze is 0.0 (not mask_start from gaze)
    is_om_category = "OoP" in target_category or "CSR" in target_category
    tgaze = 0.0 if is_om_category else mask_start
    if driver_marks and len(driver_marks) > 0:
        try:
            tgaze = float(driver_marks[0])
        except:
            pass

    signal_times = {}
    for sig_name, sig_info in signals.items():
        timestamps = sig_info['timestamps']
        samples = sig_info['samples']
        operator = sig_info['operator']
        threshold = sig_info['threshold']
        first_match_time = None

        eval_start = mask_start
        if "Unresponsive" in target_category or "OoP" in target_category or "CSR" in target_category:
            eval_start = tgaze

        if sig_name == "SoundPressure":
            try:
                audio_thresh = micro.get('threshold') if (micro and micro.get('threshold') is not None) else threshold
                if len(samples) > 0 and len(timestamps) > 0:
                    samples_numeric = samples
                    try:
                        min_f = micro.get('min_freq') if (micro and micro.get('min_freq') is not None) else 230
                        max_f = micro.get('max_freq') if (micro and micro.get('max_freq') is not None) else 2000
                        samples_np = np.array(samples, dtype=float)
                        dur = timestamps[-1] - timestamps[0]
                        fs = len(samples) / dur if dur > 0 else 44100
                        nyq = 0.5 * fs
                        
                        low = max(1e-6, float(min_f) / nyq)
                        high = min(1.0 - 1e-6, float(max_f) / nyq)
                        if low >= high:
                            high = low + 0.1
                            if high >= 1.0:
                                low = 0.1
                                high = 0.9
                        
                        from scipy.signal import butter, filtfilt
                        b, a = butter(4, [low, high], btype='band')
                        samples_numeric = list(filtfilt(b, a, samples_np))
                    except Exception as fe:
                        logger.error(f"Error filtering SoundPressure: {fe}")
                    
                    op = operator if operator and operator != 'None' else '>='
                    first_match_time, cluster_duration = find_first_valid_event(
                        np.array(samples_numeric),
                        np.array(timestamps),
                        float(audio_thresh) if audio_thresh is not None else 0.0,
                        op,
                        mask_start=eval_start
                    )
            except Exception as e:
                logger.error(f"Error calculating SoundPressure: {e}")
        elif threshold is not None and operator and operator != 'None':
            try:
                threshold_num = float(threshold)
                for t, val in zip(timestamps, samples):
                    if t < eval_start:
                        continue
                    val_num = float(val)
                    match = False
                    if operator == '>': match = val_num > threshold_num
                    elif operator == '<': match = val_num < threshold_num
                    elif operator == '>=': match = val_num >= threshold_num
                    elif operator == '<=': match = val_num <= threshold_num
                    elif operator == '==': match = abs(val_num - threshold_num) < 1e-6
                    elif operator == '!=': match = abs(val_num - threshold_num) >= 1e-6
                    if match:
                        first_match_time = t
                        break
            except (ValueError, TypeError):
                threshold_str = str(threshold)
                for t, val in zip(timestamps, samples):
                    if t < eval_start:
                        continue
                    val_str = str(val)
                    match = False
                    if operator == '==': match = val_str == threshold_str
                    elif operator == '!=': match = val_str != threshold_str
                    elif operator == '>': match = val_str > threshold_str
                    elif operator == '<': match = val_str < threshold_str
                    elif operator == '>=': match = val_str >= threshold_str
                    elif operator == '<=': match = val_str <= threshold_str
                    if match:
                        first_match_time = t
                        break
        signal_times[sig_name] = first_match_time

    if "Unresponsive" in target_category or "OoP" in target_category or "CSR" in target_category:
        for idx, phase in enumerate(unresponsive_phases):
            sig_name = phase.get('signal')
            
            # Detect audio phase by presence of frequency fields (not just signal name)
            has_audio_fields = phase.get('min_freq') is not None or phase.get('max_freq') is not None
            is_audio_by_name = sig_name and (sig_name.lower().find("sound") >= 0 or sig_name.lower().find("audio") >= 0 or sig_name.lower().find("buzzer") >= 0)
            is_audio = has_audio_fields or is_audio_by_name
            
            logger.info(f"Phase {idx} ({phase.get('phaseName')}): sig_name={sig_name}, has_audio_fields={has_audio_fields}, is_audio_by_name={is_audio_by_name}, is_audio={is_audio}")
            
            # For audio phases, use SoundPressure if available
            if is_audio and "SoundPressure" in signals:
                sig_name = "SoundPressure"
            
            if phase.get('alertType') == 'visual':
                first_match_time = None
                if driver_marks and len(driver_marks) > 1:
                    try:
                        first_match_time = float(driver_marks[1])
                    except:
                        pass
                signal_times[f"phase_{idx}"] = first_match_time
                logger.info(f"Visual phase {idx} resolved to time={first_match_time}")
                continue

            if not sig_name or sig_name not in signals:
                signal_times[f"phase_{idx}"] = None
                continue
            
            sig_info = signals[sig_name]
            timestamps = sig_info['timestamps']
            samples = sig_info['samples']
            
            # Detect audio phase by presence of frequency fields (not just signal name)
            has_audio_fields = phase.get('min_freq') is not None or phase.get('max_freq') is not None
            is_audio_by_name = sig_name.lower().find("sound") >= 0 or sig_name.lower().find("audio") >= 0 or sig_name.lower().find("buzzer") >= 0
            is_audio = has_audio_fields or is_audio_by_name
            
            if is_audio:
                operator = ">="
                threshold = phase.get('threshold') if phase.get('threshold') is not None else 0.5
            else:
                operator = phase.get('operator', '==')
                threshold = phase.get('value', 0)
                
            first_match_time = None
            cluster_duration = None
            
            # Determine evaluation start time (eval_start)
            mask_val = phase.get('mask')
            is_custom_mask = False
            custom_mask_time = None
            if mask_val is not None and str(mask_val).strip().lower() != 'previous' and str(mask_val).strip() != '':
                try:
                    custom_mask_time = float(mask_val)
                    is_custom_mask = True
                except ValueError:
                    pass
            
            if is_custom_mask:
                eval_start = custom_mask_time
            else:
                # Default: from previous phase
                if idx == 0:
                    eval_start = tgaze
                else:
                    # Find last enabled phase before idx
                    prev_t = None
                    found_active_prev = False
                    for p_idx in range(idx - 1, -1, -1):
                        p_prev = unresponsive_phases[p_idx]
                        if p_prev.get('enabled', True):
                            found_active_prev = True
                            prev_t = signal_times.get(f"phase_{p_idx}")
                            break
                    if found_active_prev:
                        if prev_t is not None:
                            eval_start = prev_t
                        else:
                            # Previous active phase didn't trigger, so this one can't trigger
                            eval_start = None
                    else:
                        eval_start = tgaze

            if eval_start is None:
                signal_times[f"phase_{idx}"] = None
                continue
            
            if is_audio:
                try:
                    audio_thresh = threshold
                    if len(samples) > 0 and len(timestamps) > 0:
                        samples_numeric = samples
                        try:
                            min_f = phase.get('min_freq') or 230
                            max_f = phase.get('max_freq') or phase.get('frequency') or 2000
                            samples_np = np.array(samples, dtype=float)
                            dur = timestamps[-1] - timestamps[0]
                            fs = len(samples) / dur if dur > 0 else 44100
                            nyq = 0.5 * fs
                            
                            low = max(1e-6, float(min_f) / nyq)
                            high = min(1.0 - 1e-6, float(max_f) / nyq)
                            if low >= high:
                                high = low + 0.1
                                if high >= 1.0:
                                    low = 0.1
                                    high = 0.9
                            
                            from scipy.signal import butter, filtfilt
                            b, a = butter(4, [low, high], btype='band')
                            samples_numeric = list(filtfilt(b, a, samples_np))
                        except Exception as fe:
                            logger.error(f"Error filtering audio for phase {idx}: {fe}")
                        
                        op = operator if operator and operator != 'None' else '>='
                        first_match_time, cluster_duration = find_first_valid_event(
                            np.array(samples_numeric),
                            np.array(timestamps),
                            float(audio_thresh) if audio_thresh is not None else 0.0,
                            op,
                            mask_start=eval_start
                        )
                        
                        # Calculate total event duration using envelope of filtered audio
                        if first_match_time is not None:
                            samples_arr = np.array(samples_numeric)
                            timestamps_arr = np.array(timestamps)
                            
                            # Use envelope (absolute value smoothed) for duration calculation
                            envelope = np.abs(samples_arr)
                            # Smooth with moving average (100ms window)
                            window_size = max(1, int(0.1 * len(samples_arr) / (timestamps_arr[-1] - timestamps_arr[0])))
                            if window_size > 1:
                                kernel = np.ones(window_size) / window_size
                                envelope = np.convolve(envelope, kernel, mode='same')
                            
                            above_mask = envelope >= float(audio_thresh)
                            above_mask = above_mask & (timestamps_arr >= eval_start)
                            above_indices = np.flatnonzero(above_mask)
                            
                            if len(above_indices) > 0:
                                first_idx = above_indices[0]
                                last_idx = above_indices[-1]
                                event_start = timestamps_arr[first_idx]
                                event_end = timestamps_arr[last_idx]
                                total_duration = event_end - event_start
                                signal_times[f"phase_{idx}_duration"] = total_duration
                except Exception as e:
                    logger.error(f"Error calculating audio for phase {idx}: {e}")
            elif threshold is not None and operator and operator != 'None':
                try:
                    threshold_num = float(threshold)
                    samples_np = np.array(samples, dtype=float)
                    first_match_time, cluster_duration = find_first_valid_event(
                        samples_np,
                        np.array(timestamps),
                        threshold_num,
                        operator,
                        min_cluster_duration=0.05,
                        mask_start=eval_start
                    )
                except (ValueError, TypeError, Exception):
                    threshold_str = str(threshold)
                    for t, val in zip(timestamps, samples):
                        if t < eval_start:
                            continue
                        val_str = str(val)
                        match = False
                        if operator == '==': match = val_str == threshold_str
                        elif operator == '!=': match = val_str != threshold_str
                        elif operator == '>': match = val_str > threshold_str
                        elif operator == '<': match = val_str < threshold_str
                        elif operator == '>=': match = val_str >= threshold_str
                        elif operator == '<=': match = val_str <= threshold_str
                        if match:
                            first_match_time = t
                            break
                            
            if first_match_time is not None:
                signal_times[f"phase_{idx}"] = first_match_time
                # Only use cluster_duration if total duration wasn't already calculated (for audio)
                if cluster_duration is not None and f"phase_{idx}_duration" not in signal_times:
                    signal_times[f"phase_{idx}_duration"] = cluster_duration
            else:
                is_unresponsive_pass = False

    t_event = "No warn"
    t_event_color = "red"
    
    if "Unresponsive" in target_category or "OoP" in target_category or "CSR" in target_category:
        # 1. Compute t_event
        active_phases_with_trigger = []
        for idx, phase in enumerate(unresponsive_phases):
            enabled = phase.get('enabled', True)
            if enabled:
                t_trig = signal_times.get(f"phase_{idx}")
                if t_trig is not None:
                    active_phases_with_trigger.append((idx, phase, t_trig))
                    
        if len(active_phases_with_trigger) > 0:
            last_idx, last_phase, last_trig_time = active_phases_with_trigger[-1]
            t_event = last_trig_time - tgaze
            
            # 2. Compute t_event_color (PASS/FAIL validation based on active milestones)
            is_unresponsive_pass = True
            is_dtr = "DTR" in target_category
            is_sle = "SLE" in target_category
            num_milestones = len(unresponsive_phases) + 1
            
            if is_dtr or is_sle:
                for i in range(num_milestones - 1):
                    step_enabled = unresponsive_phases[i].get('enabled', True)
                    if not step_enabled:
                        continue
                    
                    t_next = signal_times.get(f"phase_{i}")
                    if i == 0:
                        t_curr = tgaze
                    else:
                        t_curr = signal_times.get(f"phase_{i-1}") if unresponsive_phases[i-1].get('enabled', True) else None
                    
                    delta = None
                    if t_curr is not None and t_next is not None:
                        delta = t_next - t_curr
                    
                    ok = False
                    if is_dtr:
                        if i == 0:
                            ok = delta is not None and 3.0 <= delta <= 4.0
                        elif i == 1:
                            ok = delta is not None and delta <= 4.0
                        elif i == 2:
                            if num_milestones == 4:
                                ok = delta is not None and delta <= 5.0
                            else:
                                ok = delta is not None and delta < 1.0
                        elif i == 3:
                            ok = delta is not None and delta <= 5.0
                    else:
                        if i == 0:
                            ok = delta is not None and delta <= 7.0
                        elif i == 1:
                            ok = delta is not None and delta <= 5.0
                    
                    if not ok:
                        is_unresponsive_pass = False
                
                # Evaluate compound brackets
                if is_dtr:
                    if len(unresponsive_phases) >= 2 and unresponsive_phases[1].get('enabled', True):
                        t0 = tgaze
                        t2 = signal_times.get("phase_1")
                        ok_m02 = False
                        if t0 is not None and t2 is not None:
                            ok_m02 = 6.0 <= (t2 - t0) <= 8.0
                        if not ok_m02:
                            is_unresponsive_pass = False
                    
                    target_idx = 2 if len(unresponsive_phases) == 3 else 3
                    if len(unresponsive_phases) > target_idx and unresponsive_phases[target_idx].get('enabled', True):
                        t0 = tgaze
                        t_end = signal_times.get(f"phase_{target_idx}")
                        ok_end = False
                        if t0 is not None and t_end is not None:
                            if len(unresponsive_phases) == 3:
                                ok_end = (t_end - t0) <= 13.0
                            else:
                                ok_end = 13.0 <= (t_end - t0) <= 14.0
                        if not ok_end:
                            is_unresponsive_pass = False
                else:
                    # SLE: compound M0->M_last must be <= 12s
                    target_idx = len(unresponsive_phases) - 1
                    if len(unresponsive_phases) > target_idx and target_idx >= 0 and unresponsive_phases[target_idx].get('enabled', True):
                        t0 = tgaze
                        t_end = signal_times.get(f"phase_{target_idx}")
                        ok_sle_total = False
                        if t0 is not None and t_end is not None:
                            ok_sle_total = (t_end - t0) <= 12.0
                        if not ok_sle_total:
                            is_unresponsive_pass = False
            else:
                # Dynamic validation for Misuse categories (OoP, CSR)
                def check_time_constraint(delta: float, constraint_str: str, unit: str = "s") -> bool:
                    if not constraint_str:
                        return True
                    s = constraint_str.strip().replace(" ", "")
                    val = delta
                    if unit == "min":
                        val = delta / 60.0
                    
                    import re
                    m = re.match(r'^([<>=≤≥!=]+)?([\d.]+)', s)
                    if not m:
                        return True
                    op, num_str = m.groups()
                    try:
                        limit = float(num_str)
                    except ValueError:
                        return True
                        
                    if not op:
                        return val <= limit
                        
                    if op in (">", "≥", ">="):
                        return val >= limit
                    elif op in ("<", "≤", "<="):
                        return val <= limit
                    elif op == "==":
                        return abs(val - limit) < 1e-6
                    elif op == "!=":
                        return abs(val - limit) >= 1e-6
                    return True

                for i in range(len(unresponsive_phases)):
                    phase = unresponsive_phases[i]
                    if not phase.get('enabled', True):
                        continue
                    t_next = signal_times.get(f"phase_{i}")
                    if t_next is None:
                        is_unresponsive_pass = False
                        break
                    
                    # Compute delta from previous active phase
                    t_curr = tgaze
                    if i > 0:
                        prev_t = None
                        for p_idx in range(i - 1, -1, -1):
                            if unresponsive_phases[p_idx].get('enabled', True):
                                prev_t = signal_times.get(f"phase_{p_idx}")
                                break
                        if prev_t is not None:
                            t_curr = prev_t
                    
                    delta = t_next - t_curr
                    stored_duration = signal_times.get(f"phase_{i}_duration")
                    
                    # For duration-based constraints (≥), use event duration if available
                    tc = phase.get('timeConstraint')
                    tcu = phase.get('timeConstraintUnit', 's')
                    
                    if stored_duration is not None and tc and ('≥' in tc or '>=' in tc or '>' in tc):
                        # Use event duration for minimum duration constraints
                        delta = stored_duration
                    elif stored_duration is not None:
                        delta = stored_duration
                        
                    tc = phase.get('timeConstraint')
                    tcu = phase.get('timeConstraintUnit', 's')
                    if tc:
                        if not check_time_constraint(delta, tc, tcu):
                            is_unresponsive_pass = False
                            break
            
            t_event_color = "green" if is_unresponsive_pass else "red"
        else:
            t_event = "No warn"
            t_event_color = "red"
    else:
        # Standard scenario calculation
        warn_time = signal_times.get(pass_signal_name) if pass_signal_name else None
        if warn_time is not None:
            is_scenario2 = any(kw in target_category for kw in ["Short Distraction", "Phone Use"])
            if is_scenario2 and driver_marks and len(driver_marks) >= 2:
                marks_sorted = sorted([float(m) for m in driver_marks])
                accumulated = 0.0
                for i in range(0, len(marks_sorted) - 1, 2):
                    start = marks_sorted[i]
                    end = marks_sorted[i+1]
                    if warn_time < start:
                        break
                    elif warn_time <= end:
                        accumulated += (warn_time - start)
                        break
                    else:
                        accumulated += (end - start)
                t_event = accumulated
            else:
                t_event = warn_time - tgaze

            if conditions:
                all_met = True
                for op, limit in conditions:
                    if op == '>': match = t_event > limit
                    elif op == '<': match = t_event < limit
                    elif op == '>=': match = t_event >= limit
                    elif op == '<=': match = t_event <= limit
                    elif op == '==': match = abs(t_event - limit) < 1e-6
                    elif op == '!=': match = abs(t_event - limit) >= 1e-6
                    else: match = True
                    if not match:
                        all_met = False
                        break
                t_event_color = "green" if all_met else "red"
            else:
                t_event_color = "green" if t_event < 3.0 else "red"

    relative_path = os.path.basename(file_path)
    try:
        abs_path = os.path.abspath(file_path)
        parts = abs_path.split(os.sep)
        for i, part in enumerate(parts):
            if re.match(r'^[PE]\d+', part, re.IGNORECASE):
                relative_path = os.sep.join(parts[i:])
                break
    except:
        pass

    camera_image_path = None
    if protocol == "GSR ADDW" or protocol == "2023/2590":
        camera_image_path = _resolve_gsr_image_path(os.path.basename(file_path))

    # Normalize gauge_rules to ensure green_range and ticks exist
    normalized_gauge_rules = {}
    for cat_key, rule in gauge_rules.items():
        r = dict(rule) if isinstance(rule, dict) else {}
        if 'green_range' not in r:
            r['green_range'] = [r.get('green_min', 0), r.get('green_max', 3)]
        if 'ticks' not in r:
            min_v = r.get('min', 0)
            max_v = r.get('max', 10)
            num_ticks = r.get('num_ticks', None)
            if num_ticks and int(num_ticks) > 1:
                count = int(num_ticks)
                diff = max_v - min_v
                r['ticks'] = [round(min_v + i * (diff / count), 2) for i in range(count + 1)]
            else:
                diff = max_v - min_v
                count = int(diff) if 0 < diff <= 10 else 5
                r['ticks'] = [round(min_v + i * (diff / count), 2) for i in range(count + 1)]
        normalized_gauge_rules[cat_key] = r

    try:
        mtime = os.path.getmtime(file_path)
        test_date = datetime.fromtimestamp(mtime).astimezone()
    except Exception:
        test_date = datetime.now().astimezone()

    return {
        'filename': os.path.basename(file_path),
        'relative_path': relative_path,
        'target_category': target_category,
        'pass_signal_name': pass_signal_name,
        'show_thresholds': show_thresholds,
        'oem_name': metadata.get('oem_name', ''),
        'vehicle': metadata.get('vehicle', ''),
        'protocol': protocol,
        'engineer': metadata.get('engineer', ''),
        'analyst': metadata.get('analyst', ''),
        'track': metadata.get('track', ''),
        'test_date': test_date,
        'signals': signals,
        'signal_times': signal_times,
        'camera_image_path': camera_image_path,
        'tgaze': tgaze,
        't_event': t_event,
        't_event_color': t_event_color,
        'mask': mask_start,
        'audio_params': {
            'min_freq': float(micro.get('min_freq')) if (micro and micro.get('min_freq') is not None) else 230.0,
            'max_freq': float(micro.get('max_freq')) if (micro and micro.get('max_freq') is not None) else 2000.0,
            'threshold': float(micro.get('threshold')) if (micro and micro.get('threshold') is not None) else (float(signals_conf.get('SoundPressure', {}).get('threshold')) if ('SoundPressure' in signals_conf and signals_conf.get('SoundPressure', {}).get('threshold') is not None) else 0.0)
        },
        'gauge_rules': normalized_gauge_rules,
        'driver_marks': driver_marks,
        'unresponsive_phases': unresponsive_phases
    }


def update_excel_results(config: dict, file_path: str):
    """Update Analysis_Results.xlsx with 6-column format matching PySide6 original."""
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment

    participant_dir = os.path.dirname(file_path)
    excel_path = os.path.join(participant_dir, "Analysis_Results.xlsx")
    file_name = os.path.basename(file_path)

    folder_name = config.get('target_category', '--')
    dist_start = config.get('tgaze')

    pass_sig = config.get('pass_signal_name')
    warn_start = config.get('signal_times', {}).get(pass_sig) if pass_sig else None

    warn_timer = config.get('t_event')
    score = "PASS" if config.get('t_event_color') == "green" else "FAIL"

    row_data = [
        folder_name,
        file_name,
        dist_start if dist_start is not None else "",
        warn_start if warn_start is not None else "nan",
        warn_timer if isinstance(warn_timer, (int, float)) else "",
        score
    ]

    try:
        if os.path.exists(excel_path):
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Results"
            headers = ["Folder Name", "File Name", "Distraction Start",
                       "Warning Start", "Warning Timer", "Score"]
            ws.append(headers)

            header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            header_font = Font(bold=True)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center")

        # Find existing row by File Name (column 2)
        found_row = -1
        for row in range(2, ws.max_row + 1):
            cell_val = ws.cell(row=row, column=2).value
            if cell_val and str(cell_val) == file_name:
                found_row = row
                break

        if found_row != -1:
            for col_idx, val in enumerate(row_data, 1):
                ws.cell(row=found_row, column=col_idx, value=val)
            target_row = found_row
        else:
            ws.append(row_data)
            target_row = ws.max_row

        for cell in ws[target_row]:
            cell.alignment = Alignment(horizontal="center")

        for i, col in enumerate(ws.columns, 1):
            max_length = 0
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = max_length + 4

        wb.save(excel_path)
        logger.info(f"Excel result saved/updated in {excel_path}")
    except Exception as e:
        logger.error(f"Failed to update excel results: {e}")


@router.post("/gaze/preview")
async def gaze_preview(req: GazePreviewRequest):
    import traceback
    from backend.routers.analysis import _load_marks_dict, _get_marks_key
    from backend.core.ga_report_builder import GAReportBuilder
    try:
        if not os.path.exists(req.file_path):
            return {"status": "error", "message": f"File not found: {req.file_path}"}
        
        source_dir = req.source_dir
        if not source_dir:
            curr = os.path.dirname(req.file_path)
            while curr and curr != os.path.dirname(curr):
                if (os.path.exists(os.path.join(curr, "GA_marks.json")) or 
                    os.path.exists(os.path.join(curr, "OM_marks.json")) or 
                    os.path.exists(os.path.join(curr, "marks.json"))):
                    source_dir = curr
                    break
                curr = os.path.dirname(curr)
            if not source_dir:
                source_dir = os.path.dirname(req.file_path)

        # Detect if this is an OM (Misuse) scenario or Gaze scenario
        file_path_lower = req.file_path.lower()
        is_om_scenario = any(kw in file_path_lower for kw in ["oop", "out of position", "csr", "correct belt", "seatbelt", "belt routing"])
        marks_type = "OM" if is_om_scenario else "GA"
        
        marks_dict = _load_marks_dict(source_dir, marks_type)
        key = _get_marks_key(req.file_path)
        # Try exact match first, then fallback to filename-based match
        driver_marks = marks_dict.get(key) or marks_dict.get(key.replace(".mf4", "_tracking.mf4")) or []
        if not driver_marks:
            # Fallback: find key ending with the filename
            filename = os.path.basename(req.file_path).lower()
            for mk, mv in marks_dict.items():
                if mk.lower().endswith(filename) or mk.lower().endswith(filename.replace(".mf4", "_tracking.mf4")):
                    driver_marks = mv
                    break
        if not isinstance(driver_marks, list):
            driver_marks = []
            
        config = build_report_config(
            file_path=req.file_path,
            protocol=req.protocol,
            metadata=req.metadata,
            category_configs=req.category_configs,
            gauge_rules=req.gauge_rules,
            driver_marks=driver_marks,
            micro=req.micro,
            show_thresholds=True
        )
        
        rc_settings = req.report_camera_settings or {}
        cam_l = rc_settings.get("left", "")
        cam_r = rc_settings.get("right", "")
        if cam_l or cam_r:
            vl, vr = _resolve_om_video_paths(req.file_path, cam_l, cam_r)
            config["om_video_left"] = vl
            config["om_video_right"] = vr
            config["report_camera_settings"] = rc_settings
        
        import tempfile
        temp_dir = tempfile.gettempdir()
        preview_output_path = os.path.join(temp_dir, "gaze_preview.png")
        
        if config.get("target_category", "").startswith("OoP") or config.get("target_category", "").startswith("CSR"):
            builder = OMReportBuilder(config)
        else:
            builder = GAReportBuilder(config)
        builder.generate(preview_output_path, dpi=300)
        
        return {
            "status": "success",
            "preview_path": preview_output_path
        }
    except Exception as e:
        logger.error(f"Error generating gaze preview: {e}")
        traceback.print_exc()
        return {"status": "error", "message": str(e)}


@router.post("/gaze/generate")
async def gaze_generate(req: GazeGenerateRequest):
    global _active_worker, _worker_thread

    if _active_worker is not None:
        return {"status": "already_running"}

    loop = asyncio.get_event_loop()

    def on_progress(msg):
        asyncio.run_coroutine_threadsafe(
            manager_reporting.broadcast({"type": "progress", "message": msg}), loop
        )

    class _GazeReportingWorker:
        def __init__(self):
            self.is_running = True

        def stop(self):
            self.is_running = False

        def run(self):
            global _active_worker
            from backend.routers.analysis import _load_marks_dict, _get_marks_key
            from backend.core.ga_report_builder import GAReportBuilder
                        
            try:
                total_files = len(req.files)
                success_count = 0
                for idx, file_path in enumerate(req.files):
                    if not self.is_running:
                        on_progress("Gaze batch generation stopped by user.")
                        break
                    
                    base_name = os.path.splitext(os.path.basename(file_path))[0]
                    asyncio.run_coroutine_threadsafe(
                        manager_reporting.broadcast({
                            "type": "progress_update",
                            "current": idx + 1,
                            "total": total_files,
                            "message": f"Processing ({idx+1}/{total_files}): {base_name}..."
                        }), loop
                    )
                    
                    try:
                        if not os.path.exists(file_path):
                            on_progress(f"[ERROR] File not found: {file_path}")
                            continue
                        
                        source_dir = req.source_dir
                        if not source_dir:
                            curr = os.path.dirname(file_path)
                            while curr and curr != os.path.dirname(curr):
                                if (os.path.exists(os.path.join(curr, "GA_marks.json")) or 
                                    os.path.exists(os.path.join(curr, "OM_marks.json")) or 
                                    os.path.exists(os.path.join(curr, "marks.json"))):
                                    source_dir = curr
                                    break
                                curr = os.path.dirname(curr)
                            if not source_dir:
                                source_dir = os.path.dirname(file_path)

                        # Detect if this is an OM (Misuse) scenario or Gaze scenario
                        file_path_lower = file_path.lower()
                        is_om_scenario = any(kw in file_path_lower for kw in ["oop", "out of position", "csr", "correct belt", "seatbelt", "belt routing"])
                        marks_type = "OM" if is_om_scenario else "GA"
                        
                        marks_dict = _load_marks_dict(source_dir, marks_type)
                        key = _get_marks_key(file_path)
                        # Try exact match first, then fallback to filename-based match
                        driver_marks = marks_dict.get(key) or marks_dict.get(key.replace(".mf4", "_tracking.mf4")) or []
                        if not driver_marks:
                            # Fallback: find key ending with the filename
                            filename = os.path.basename(file_path).lower()
                            for mk, mv in marks_dict.items():
                                if mk.lower().endswith(filename) or mk.lower().endswith(filename.replace(".mf4", "_tracking.mf4")):
                                    driver_marks = mv
                                    break
                        if not isinstance(driver_marks, list):
                            driver_marks = []
                            
                        config = build_report_config(
                            file_path=file_path,
                            protocol=req.protocol,
                            metadata=req.metadata,
                            category_configs=req.category_configs,
                            gauge_rules=req.gauge_rules,
                            driver_marks=driver_marks,
                            micro=req.micro,
                            show_thresholds=False
                        )
                        
                        rc_settings = req.report_camera_settings or {}
                        cam_l = rc_settings.get("left", "")
                        cam_r = rc_settings.get("right", "")
                        if cam_l or cam_r:
                            vl, vr = _resolve_om_video_paths(file_path, cam_l, cam_r)
                            config["om_video_left"] = vl
                            config["om_video_right"] = vr
                            config["report_camera_settings"] = rc_settings
                        
                        base_dir = os.path.dirname(file_path)
                        reports_dir = os.path.join(base_dir, "Reports")
                        os.makedirs(reports_dir, exist_ok=True)
                        output_png = os.path.join(reports_dir, f"{base_name}.png")
                        
                        if config.get("target_category", "").startswith("OoP") or config.get("target_category", "").startswith("CSR"):
                            builder = OMReportBuilder(config)
                        else:
                            builder = GAReportBuilder(config)
                        builder.generate(output_png, dpi=300)
                        
                        update_excel_results(config=config, file_path=file_path)
                        
                        success_count += 1
                    except Exception as e:
                        logger.error(f"Failed to generate report for {file_path}: {e}")
                        on_progress(f"[ERROR] {base_name}: {str(e)}")

                asyncio.run_coroutine_threadsafe(
                    manager_reporting.broadcast({
                        "type": "finished",
                        "message": f"Successfully generated {success_count} reports."
                    }), loop
                )
            except Exception as e:
                asyncio.run_coroutine_threadsafe(
                    manager_reporting.broadcast({
                        "type": "error",
                        "message": str(e)
                    }), loop
                )
            finally:
                _active_worker = None

    worker = _GazeReportingWorker()
    _active_worker = worker

    def _run():
        worker.run()

    _worker_thread = Thread(target=_run, daemon=True)
    _worker_thread.start()

    return {"status": "started"}


class GaugeFileReadRequest(BaseModel):
    file_path: str


class GaugeFileWriteRequest(BaseModel):
    file_path: str
    rules: dict


class GaugeFileExistsRequest(BaseModel):
    file_path: str


@router.post("/gauge_rules/read_file")
async def read_gauge_rules_file(req: GaugeFileReadRequest):
    file_path = req.file_path
    if file_path == "config/gauge_rules.json":
        from backend.core.utils import user_data_path
        file_path = user_data_path("config/gauge_rules.json")
        if not os.path.exists(file_path):
            file_path = resource_path("config/gauge_rules.json")
    if not os.path.exists(file_path):
        return {"error": "File not found"}
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            content = json.load(f)
            if isinstance(content, dict):
                return {"rules": content}
            return {"error": "Invalid format, must be a JSON object"}
    except Exception as e:
        return {"error": str(e)}


@router.post("/gauge_rules/write_file")
async def write_gauge_rules_file(req: GaugeFileWriteRequest):
    try:
        file_path = req.file_path
        if file_path == "config/gauge_rules.json":
            from backend.core.utils import user_data_path
            file_path = user_data_path("config/gauge_rules.json")
        parent_dir = os.path.dirname(file_path)
        if parent_dir:
            os.makedirs(parent_dir, exist_ok=True)
        with open(file_path, "w", encoding="utf-8") as f:
            json.dump(req.rules, f, indent=2)
        return {"status": "success"}
    except Exception as e:
        return {"error": str(e)}


@router.post("/gauge_rules/exists")
async def check_gauge_rules_exists(req: GaugeFileExistsRequest):
    file_path = req.file_path
    if file_path == "config/gauge_rules.json":
        from backend.core.utils import user_data_path
        if os.path.exists(user_data_path("config/gauge_rules.json")):
            return {"exists": True}
        file_path = resource_path("config/gauge_rules.json")
    return {"exists": os.path.exists(file_path) and os.path.isfile(file_path)}


class OpenFileRequest(BaseModel):
    file_path: str


@router.post("/open_file")
async def open_file_in_os_viewer(req: OpenFileRequest):
    import os
    import sys
    import subprocess
    
    if not os.path.exists(req.file_path):
        return {"status": "error", "message": f"File not found: {req.file_path}"}
        
    try:
        if os.name == 'nt':
            os.startfile(req.file_path)
        elif sys.platform == 'darwin':
            subprocess.Popen(['open', req.file_path])
        else:
            subprocess.Popen(['xdg-open', req.file_path])
        return {"status": "success"}
    except Exception as e:
        logger.error(f"Failed to open file externally: {e}")
        return {"status": "error", "message": str(e)}
