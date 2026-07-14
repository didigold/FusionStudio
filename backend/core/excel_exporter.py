
import os
import json
import re
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter
from copy import copy
from asammdf import MDF

class ExcelReportGenerator:
    def __init__(self, template_path, root_dir=None):
        self.template_path = template_path
        self.root_dir = root_dir
        
    def _determine_category_from_filename(self, filename):
        """
        Determine the distraction category based on filename pattern.
        """
        # Extract base name without extension
        basename = os.path.splitext(os.path.basename(filename))[0]
        
        # Try to match D + number pattern
        match_d = re.match(r'^D(\d+)', basename)
        if match_d:
            num = int(match_d.group(1))
            if 1 <= num <= 15:
                return "Long Distraction"
            elif 16 <= num <= 28:
                return "Short Distractions"
            elif 29 <= num <= 42:
                return "Phone Use"
        
        # Try to match F + number pattern
        match_f = re.match(r'^F(\d+)', basename)
        if match_f:
            num = int(match_f.group(1))
            if num in [1, 2]:
                return "Microsleep & Sleep"
            elif num == 3:
                return "Drowsiness"
            elif num in [4, 5]:
                return "Unresponsive driver"
                
        return None
        
    def _apply_operator(self, value, operator, threshold):
        """
        Apply a comparison operator between value and threshold.
        Returns True if the condition is satisfied.
        """
        ops = {
            '>': lambda v, t: v > t,
            '<': lambda v, t: v < t,
            '>=': lambda v, t: v >= t,
            '<=': lambda v, t: v <= t,
            '==': lambda v, t: v == t,
            '!=': lambda v, t: v != t,
        }
        fn = ops.get(operator)
        if fn is None:
            return True  # No valid operator = condition ignored
        return fn(value, threshold)

    def _calculate_metrics(self, file_path, category, logic_config, marks_timestamps=None):
        """
        Calculates the Distraction Timer and PASS/FAIL score using marks timestamps only.
        No MDF file access required.

        Algorithm (mirrors analysis_widget._build_report_config_for_mf4):
          1. Compute total accumulated active distraction time from mark pairs.
          2. Find t_event = the accumulated time at the point the pass_criteria
             threshold is first crossed (e.g. if operator1='>=' and value1=3.0,
             t_event=3.0 once 3 accumulated seconds of distraction have elapsed).
          3. If the total duration never reaches the threshold: t_event = total_duration.
          4. PASS/FAIL = apply pass_criteria (operator1/value1 AND operator2/value2) to t_event.

        Returns: (t_event, score)  - t_event is float (seconds), score is "PASS" or "FAIL"
        """
        if not marks_timestamps or len(marks_timestamps) < 2:
            return 0.0, "PASS"  # No distraction data

        marks_sorted = sorted([float(t) for t in marks_timestamps])

        # --- Get pass_criteria for this category ---
        pass_criteria = logic_config.get('pass_criteria', {}).get(category, {}) if logic_config else {}

        op1  = pass_criteria.get('operator1', '') if pass_criteria else ''
        val1 = pass_criteria.get('value1', 0)     if pass_criteria else 0
        op2  = pass_criteria.get('operator2', '') if pass_criteria else ''
        val2 = pass_criteria.get('value2', 0)     if pass_criteria else 0

        # --- Accumulate distraction time from mark pairs ---
        # marks are toggling: ON, OFF, ON, OFF, ...
        total_duration = 0.0
        for i in range(0, len(marks_sorted) - 1, 2):
            total_duration += marks_sorted[i + 1] - marks_sorted[i]

        # --- Find t_event: accumulated time when pass_criteria is first crossed ---
        # We walk through the mark pairs and accumulate until we reach val1
        # (the threshold that represents when the warning fires).
        # This matches calculate_accumulated_time in analysis_widget.py.
        scenario1_categories = ["Long Distraction", "Microsleep & Sleep", "Unresponsive driver"]
        is_scenario1 = category in scenario1_categories

        t_event = total_duration  # default: use full duration if threshold never reached

        if pass_criteria and op1 and isinstance(val1, (int, float)):
            # Determine the threshold value to check: for Scenario 1 use val1 directly,
            # for Scenario 2 (accumulated across multiple short distractions) also val1.
            threshold = float(val1)

            accumulated = 0.0
            found = False
            for i in range(0, len(marks_sorted) - 1, 2):
                start = marks_sorted[i]
                end   = marks_sorted[i + 1]
                period = end - start

                if accumulated + period >= threshold:
                    # Threshold is crossed within this period
                    t_event = threshold
                    found = True
                    break
                else:
                    accumulated += period

            if not found:
                # Threshold was never reached; t_event = full accumulated duration
                t_event = total_duration

        # --- Apply pass_criteria to t_event ---
        if pass_criteria and op1:
            cond1 = self._apply_operator(t_event, op1, val1) if op1 else True
            cond2 = self._apply_operator(t_event, op2, val2) if op2 else True
            score = "PASS" if (cond1 and cond2) else "FAIL"
        else:
            # Fallback: no pass_criteria configured for this category
            score = "PASS" if total_duration < 2.0 else "FAIL"

        return t_event, score


    def _structure_marks_data(self, marks_data):
        """
        Restructures the flat marks dictionary into a nested one:
        {
            "P01": {
                "D1_1": {"start": x, "end": y},
                ...
            },
            ...
        }
        """
        structured = {}
        for file_path, timestamps in marks_data.items():
            # file_path example: "P01/Distractions/D1_1_tracking.mf4"
            parts = file_path.split('/')
            if len(parts) < 3:
                continue
            
            p_name = parts[0]
            filename = parts[-1]
            
            # Extract mark key (remove .mf4 or _tracking.mf4)
            mark_key = filename.replace('_tracking.mf4', '').replace('.mf4', '')
            
            if p_name not in structured:
                structured[p_name] = {}
                
            start = timestamps[0] if len(timestamps) > 0 else 0
            end = timestamps[-1] if len(timestamps) > 0 else 0
            
            structured[p_name][mark_key] = {
                'start': start,
                'end': end,
                'duration': end - start,
                'file_path': file_path,
                'timestamps': timestamps
            }
        return structured

    def _copy_cell_style(self, source_cell, target_cell):
        """
        Copy formatting from source_cell to target_cell (font, fill, border, alignment, number_format).
        """
        if source_cell.has_style:
            target_cell.font = copy(source_cell.font)
            target_cell.fill = copy(source_cell.fill)
            target_cell.border = copy(source_cell.border)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.number_format = source_cell.number_format
            target_cell.protection = copy(source_cell.protection)

    def _copy_column_formatting(self, ws, template_start_col, target_start_col, width, max_row):
        """
        Copy formatting from template participant block to a new participant block.
        Copies cell styles and column widths for WIDTH columns.
        """
        for col_offset in range(width):
            src_col = template_start_col + col_offset
            dst_col = target_start_col + col_offset
            
            # Copy column width
            src_letter = get_column_letter(src_col)
            dst_letter = get_column_letter(dst_col)
            if src_letter in ws.column_dimensions:
                ws.column_dimensions[dst_letter].width = ws.column_dimensions[src_letter].width
            
            # Copy cell styles for all rows
            for row in range(1, max_row + 1):
                src_cell = ws.cell(row=row, column=src_col)
                dst_cell = ws.cell(row=row, column=dst_col)
                self._copy_cell_style(src_cell, dst_cell)

    def generate_report(self, marks_data, logic_config, output_path=None):
        """
        Generates the Excel report.
        """
        if not os.path.exists(self.template_path):
            raise FileNotFoundError(f"Template not found at: {self.template_path}")
            
        # Load workbook
        wb = openpyxl.load_workbook(self.template_path)
        if 'DISTRACTION' not in wb.sheetnames:
            raise ValueError("Template missing 'DISTRACTION' sheet")
            
        ws = wb['DISTRACTION']
        
        # Structure Data
        structured_data = self._structure_marks_data(marks_data)
        participants = sorted(structured_data.keys())
        
        # Define column mapping
        START_COL = 8
        WIDTH = 9
        
        # Determine max row with data in column G (scenario codes)
        max_row = 100
        for row in range(100, 3, -1):
            cell_val = ws.cell(row=row, column=7).value
            if cell_val and isinstance(cell_val, str):
                max_row = row
                break
        
        for i, p_name in enumerate(participants):
            p_marks = structured_data.get(p_name, {})
            current_start_col = START_COL + (i * WIDTH)
            
            # Copy formatting from template (first block) to additional participant blocks
            if i > 0:
                self._copy_column_formatting(ws, START_COL, current_start_col, WIDTH, max_row)
            
            # Header
            header_cell = ws.cell(row=1, column=current_start_col)
            if not header_cell.value or header_cell.value != p_name:
                header_cell.value = p_name
                
            # Iterate Rows
            for row in range(4, max_row + 1):
                scenario_code_cell = ws.cell(row=row, column=7) # Col G
                scenario_code = scenario_code_cell.value
                
                if not scenario_code or not isinstance(scenario_code, str):
                    continue
                    
                # Match marks
                # D01 -> D1_
                if scenario_code.startswith('D0'):
                    prefix = f"{scenario_code[0]}{scenario_code[2]}_" 
                else:
                    prefix = f"{scenario_code}_"
                    
                relevant_marks = []
                for mark_key, mark_val in p_marks.items():
                    if mark_key.startswith(prefix):
                        suffix = mark_key[len(prefix):]
                        if suffix and suffix[0].isdigit():
                            relevant_marks.append((mark_key, mark_val))
                
                # Sort by repetition
                try:
                    relevant_marks.sort(key=lambda x: int(x[0].split('_')[-1]))
                except:
                    pass
                    
                # Fill Repetitions
                for rep_idx in range(min(3, len(relevant_marks))):
                    mark_key, mark_val = relevant_marks[rep_idx]
                    
                    file_path = mark_val.get('file_path')
                    timestamps = mark_val.get('timestamps', [])
                    category = self._determine_category_from_filename(file_path)
                    
                    distraction_timer, score = self._calculate_metrics(file_path, category, logic_config, timestamps)
                    
                    timer_col = current_start_col + (rep_idx * 2) 
                    ws.cell(row=row, column=timer_col).value = distraction_timer if distraction_timer is not None else ""
                    
                    score_col = timer_col + 1
                    ws.cell(row=row, column=score_col).value = score

        if output_path:
            wb.save(output_path)
            return output_path
        return wb

    def generate_preview_data(self, marks_data, logic_config):
        """
        Generates a Pandas DataFrame preview by running the full generation in memory.
        """
        # Generate full report in memory (no output path)
        wb = self.generate_report(marks_data, logic_config, output_path=None)
        
        if 'DISTRACTION' not in wb.sheetnames:
            return pd.DataFrame(["Error: DISTRACTION sheet missing"])
            
        # Save to BytesIO to read with pandas
        from io import BytesIO
        virtual_file = BytesIO()
        wb.save(virtual_file)
        virtual_file.seek(0)
        
        # Read with Pandas
        # Use header=None to treat everything as data (since template has complex headers)
        try:
            df = pd.read_excel(virtual_file, sheet_name='DISTRACTION', header=None)
            # Fill NaN with empty string for better display
            df = df.fillna("")
            return df
        except Exception as e:
             return pd.DataFrame([f"Error reading excel: {str(e)}"])

